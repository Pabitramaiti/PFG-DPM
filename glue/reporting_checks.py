"""
AWS Glue Job: JSON-Driven Report Generator
==========================================

This script has been converted to run as an AWS Glue job while maintaining
all existing business logic and report formatting.

Key AWS Glue Features:
- S3 integration for input/output files
- Splunk logging for job monitoring
- Parallel S3 operations using ThreadPoolExecutor
- Email notifications for job failures
- Environment variable parameter handling

Required Environment Variables:
- s3BucketInput: S3 bucket for input files
- s3BucketOutput: S3 bucket for output files
- inputFileName: Name of the input XML file
- inputPath: S3 path prefix for input files
- reportPath: S3 path prefix for output reports
- configKey: S3 key for configuration file
- transactionId: Transaction identifier for logging
- clientName, clientID: Client information
- teamsId: Email for failure notifications

Glue Job Name: br_icsdev_dpmjayant_glue_reporting_checks
Step Function: report
"""

import json
import xml.etree.ElementTree as ET
import os
import sys
import re
import zipfile
import io
import ast
import time
import traceback
from datetime import datetime, timezone
from typing import Dict, List, Any, Optional
import csv
from concurrent.futures import ThreadPoolExecutor, as_completed
from botocore.exceptions import ClientError
import boto3
import splunk

try:
     import openpyxl
     from openpyxl import Workbook
     from openpyxl.styles import Font, Alignment
     XLSX_AVAILABLE = True
except ImportError:
     XLSX_AVAILABLE = False

s3 = boto3.client('s3')

# AWS Glue utility functions
def checkFile_isPresent(bucket_name, object_key, file_name, context):
          """Check if a file exists in S3 bucket"""
          try:
                    print(f"INFO: Checking if file exists - Bucket: {bucket_name}, Key: {object_key}, File: {file_name}")
                    s3.head_object(Bucket=bucket_name, Key=object_key)
                    print(f"INFO: File {file_name} exists in the bucket {bucket_name}")
                    return True
          except ClientError as e:
                    if e.response['Error']['Code'] == 'NoSuchKey':
                              print(f"WARNING: File {file_name} not found in the bucket {bucket_name}")
                              return False
                    print(f"ERROR: Error checking file {file_name}: {e}")
                    print(f"ERROR: Traceback: {traceback.format_exc()}")
                    return False
          except Exception as e:
                    print(f"ERROR: Unexpected error checking file {file_name}: {e}")
                    print(f"ERROR: Traceback: {traceback.format_exc()}")
                    return False

def create_sample_config(bucket_name, config_key, file_type):
          """Create a sample configuration file for the given file type"""
          try:
                    print(f"INFO: Creating sample configuration for file type: {file_type}")
                   
                    sample_config = {
                              "report_types": {
                                        file_type: {
                                                  "data_sources": {
                                                            "input_xml_path": "",
                                                            "output_path": ""
                                                  },
                                                  "report_settings": {
                                                            "file_type": file_type,
                                                            "output_format": "csv",
                                                            "include_headers": True
                                                  },
                                                  "processing": {
                                                            "extract_all_elements": True,
                                                            "create_summary": True
                                                  },
                                                  "layout": {
                                                            "header_section": {
                                                                      "enabled": True,
                                                                      "fields": ["report_name", "generation_date"]
                                                            },
                                                            "data_section": {
                                                                      "enabled": True,
                                                                      "extract_all": True
                                                            },
                                                            "footer_section": {
                                                                      "enabled": True,
                                                                      "fields": ["total_records", "generation_time"]
                                                            }
                                                  }
                                        }
                              }
                    }
                   
                    # Upload sample config to S3
                    config_content = json.dumps(sample_config, indent=2)
                    s3.put_object(
                              Bucket=bucket_name,
                              Key=config_key,
                              Body=config_content,
                              ContentType='application/json'
                    )
                   
                    print(f"INFO: Successfully created sample configuration: s3://{bucket_name}/{config_key}")
                    log_message("INFO", f"Successfully created sample configuration: s3://{bucket_name}/{config_key}")
                    return True
                   
          except Exception as e:
                    print(f"ERROR: Failed to create sample configuration: {str(e)}")
                    print(f"ERROR: Traceback: {traceback.format_exc()}")
                    log_message("ERROR", f"Failed to create sample configuration: {str(e)}")
                    return False

def get_run_id():
     """Generate account_id using function"""
     try:
          print("INFO: Getting AWS account ID for run ID generation")
          account_id = boto3.client("sts").get_caller_identity()["Account"]
          run_id = "arn:dpm:glue:br_icsdev_dpmjayant_glue_reporting_checks:" + account_id
          print(f"INFO: Successfully generated run ID: {run_id}")
          return run_id
     except Exception as e:
          print(f"ERROR: Failed to get account ID: {str(e)}")
          print(f"ERROR: Traceback: {traceback.format_exc()}")
          raise

def log_message(status, message):
     """Function to log messages in splunk"""
     try:
          print(f"INFO: Logging message - Status: {status}, Message: {message}")
          splunk.log_message({'FileName': os.getenv('inputFileName', ''), 'Status': status, 'Message': message}, get_run_id())
          print(f"INFO: Successfully logged message to Splunk")
     except Exception as e:
          print(f"ERROR: Failed to log message to Splunk: {str(e)}")
          print(f"ERROR: Traceback: {traceback.format_exc()}")
          # Don't raise here as logging failure shouldn't stop the main process

def set_job_params_as_env_vars():
     """Loop through all command-line arguments starting from the second argument (skip the script name)"""
     try:
          print("INFO: Starting to set job parameters as environment variables")
          for i in range(1, len(sys.argv), 2):
               if sys.argv[i].startswith('--'):
                    key = sys.argv[i][2:] # Remove the leading '--'
                    value = sys.argv[i + 1]
                    os.environ[key] = value
                    print(f'INFO: Set environment variable {key} to {value}')
          print("INFO: Successfully set all job parameters as environment variables")
     except Exception as e:
          print(f"ERROR: Failed to set job parameters as environment variables: {str(e)}")
          print(f"ERROR: Traceback: {traceback.format_exc()}")
          raise

def is_valid_email(email: str) -> bool:
     """Return True if email looks like a valid address."""
     try:
          print(f"INFO: Validating email address: {email}")
          if not email:
               print("WARNING: Empty email address provided")
               return False
          email_regex = r"^[\w\.-]+@[\w\.-]+\.\w+$"
          is_valid = re.match(email_regex, email) is not None
          print(f"INFO: Email validation result: {is_valid}")
          return is_valid
     except Exception as e:
          print(f"ERROR: Failed to validate email address: {str(e)}")
          print(f"ERROR: Traceback: {traceback.format_exc()}")
          return False

def upload_to_s3_parallel(s3_operations: List[Dict]) -> List[bool]:
     """Upload multiple files to S3 in parallel using ThreadPoolExecutor"""
     def upload_single_file(operation):
          try:
               print(f"INFO: Starting upload to S3: s3://{operation['bucket']}/{operation['key']}")
               s3_client = boto3.client('s3')
               s3_client.put_object(
                    Bucket=operation['bucket'],
                    Key=operation['key'],
                    Body=operation['body'],
                    ContentType=operation.get('content_type', 'application/octet-stream')
               )
               print(f"INFO: Successfully uploaded to S3: s3://{operation['bucket']}/{operation['key']}")
               log_message("SUCCESS", f"Uploaded to S3: s3://{operation['bucket']}/{operation['key']}")
               return True
          except Exception as e:
               print(f"ERROR: Failed to upload to S3: s3://{operation['bucket']}/{operation['key']} - {str(e)}")
               print(f"ERROR: Traceback: {traceback.format_exc()}")
               log_message("ERROR", f"failed to upload to S3: {str(e)}")
               return False
    
     try:
          print(f"INFO: Starting parallel upload of {len(s3_operations)} files to S3")
          results = []
          with ThreadPoolExecutor(max_workers=16) as executor:
               futures = {executor.submit(upload_single_file, op): op for op in s3_operations}
               for future in as_completed(futures):
                    results.append(future.result())
         
          success_count = sum(results)
          print(f"INFO: Parallel upload completed. {success_count}/{len(s3_operations)} files uploaded successfully")
          return results
     except Exception as e:
          print(f"ERROR: Failed to execute parallel S3 upload: {str(e)}")
          print(f"ERROR: Traceback: {traceback.format_exc()}")
          return [False] * len(s3_operations)


class ExcelFormatter:
     """Formats data as Excel-compatible text with proper column alignment"""

     def __init__(self):
          self.tab_char = '\t'# Tab character for Excel column separation

     def format_currency(self, value: float) -> str:
          """Format value as currency"""
          try:
                   return f"{float(value):.2f}"
          except:
                   return "0.00"

     def clean_masked_value(self, value: str) -> str:
          """Clean masked values (remove asterisks and extract numeric part)"""
          if not value:
                   return "0.00"
         
          # Remove asterisks and extract numeric part
          cleaned = str(value).replace('*', '')
         
          # If the cleaned value is empty or just whitespace, return 0.00
          if not cleaned.strip():
                   return "0.00"
         
          return cleaned

     def safe_float_convert(self, value: str, default: float = 0.0) -> float:
          """Safely convert a value to float, handling masked values"""
          if not value:
                   return default
         
          # Clean masked values first
          cleaned = self.clean_masked_value(value)
         
          try:
                   return float(cleaned)
          except (ValueError, TypeError):
                   return default

     def format_number(self, value: Any, width: int = 0) -> str:
          """Format number with specified width"""
          try:
                   num = float(value)
                   if width > 0:
                        return f"{num:>{width}.2f}"
                   return f"{num:.2f}"
          except:
                   return "0.00"

     def align_text(self, text: str, width: int, align: str = 'left') -> str:
          """Align text within specified width"""
          if align == 'right':
                   return text.rjust(width)
          elif align == 'center':
                   return text.center(width)
          else:# left
                   return text.ljust(width)

     def create_excel_row(self, columns: List[Dict], data: Dict) -> str:
          """Create an Excel-formatted row"""
          row_parts = []

          for col in columns:
                   value = self._get_column_value(col, data)
                   formatted_value = self._format_column_value(col, value)
                   row_parts.append(formatted_value)

          return self.tab_char.join(row_parts)

     def _get_column_value(self, col: Dict, data: Dict) -> Any:
          """Get value for a column from data"""
          source_field = col.get('source_field', '')

          if source_field == 'static':
                   return col.get('value', '')
          elif source_field == 'calculated':
                   return self._calculate_value(col, data)
          else:
                   value = data.get(source_field, '')
                   # For CUSTOM_FUND_CODE, pad with leading zeros to 7 characters
                   if source_field == 'CUSTOM_FUND_CODE' and value:
                        return value.zfill(7)
                   # For CUSTOM_CHECK_AMOUNT, clean masked values
                   elif source_field == 'CUSTOM_CHECK_AMOUNT' and value:
                        return self.clean_masked_value(value)
                   return value

     def _calculate_value(self, col: Dict, data: Dict) -> Any:
          """Calculate value based on column configuration"""
          calculation = col.get('calculation', '')

          if calculation == 'count_records':
                   return 1# Will be summed up
          elif calculation == 'sum_amounts':
                   # For Redemption_SWP reports, use CUSTOM_CHECK_AMOUNT
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'sum_sheets':
                   return float(data.get('DOC_SHEET_COUNT', 0))
          elif calculation == 'sum_images':
                   return float(data.get('DOC_IMAGE_COUNT', 0))
          elif calculation == 'sum_field':
                   field_name = col.get('field', '')
                   return float(data.get(field_name, 0))
          elif calculation == 'count_email_delivery':
                   return 1 if data.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'E' else 0
          elif calculation == 'count_print_delivery':
                   return 1 if data.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'P' else 0
          elif calculation == 'count_suppressed':
                   return 1 if data.get('DOC_SUPPRESSION_FLAG') == 'Y' else 0
          elif calculation == 'count_suppressed_amount':
                   # Count records with amount < $5.00
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   amount = self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 1 if amount < 5.00 else 0
          elif calculation == 'count_zero_check':
                   # Count records with zero check number
                   check_number = data.get('FN_CHECK_NUMBER', '')
                   return 1 if check_number == '0' or check_number == '' else 0
          elif calculation == 'count_suppressed_total':
                   # Count total suppressed records
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   amount = self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   check_number = data.get('CUSTOM_CHECK_NUMBER', '')
                   return 1 if amount < 5.00 or check_number == '0' or check_number == '' else 0
          elif calculation == 'count_packages':
                   return 1# Each record is a package
          elif calculation == 'count_check_pages':
                   return 2# Each check has 2 pages
          elif calculation == 'extract_group':
                   # Extract group from account number or other field
                   account = data.get('UC_ACCOUNT_NUMBER', '')
                   if len(account) >= 7:
                        return account[:7]
                   return account
          elif calculation == 'count_redemption_records':
                   # Count redemption records based on FN_CHECK_TYPE_DESC
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   return 1 if 'REDEMPTION' in check_type.upper() else 0
          elif calculation == 'sum_redemption_amounts':
                   # Sum amounts for redemption records
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   if 'REDEMPTION' in check_type.upper():
                        return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 0
          elif calculation == 'count_swp_records':
                   # Count SWP records based on FN_CHECK_TYPE_DESC
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   return 1 if 'SWP' in check_type.upper() else 0
          elif calculation == 'sum_swp_amounts':
                   # Sum amounts for SWP records
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   if 'SWP' in check_type.upper():
                        return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 0
          elif calculation == 'concatenate_fund_code_name':
                   # Concatenate CUSTOM_FUND_CODE and CUSTOM_FUND_DESC
                   fund_code = data.get('CUSTOM_FUND_CODE', '')
                   fund_desc = data.get('CUSTOM_FUND_DESC', '')
                   if fund_code and fund_desc:
                        return f"{fund_code} {fund_desc}"
                   elif fund_code:
                        return fund_code
                   elif fund_desc:
                        return fund_desc
                   return ""
          elif calculation == 'red_begin_check_number':
                   # Get the first RED check number for this fund
                   return self._get_red_begin_check_number(data)
          elif calculation == 'red_end_check_number':
                   # Get the last RED check number for this fund
                   return self._get_red_end_check_number(data)
          elif calculation == 'red_check_count':
                   # Count RED records for this fund
                   return self._get_red_check_count(data)
          elif calculation == 'total_red_dollar_value':
                   # Sum RED amounts for this fund
                   return self._get_total_red_dollar_value(data)
          elif calculation == 'swp_begin_check_number':
                   # Get the first SWP check number for this fund
                   return self._get_swp_begin_check_number(data)
          elif calculation == 'swp_end_check_number':
                   # Get the last SWP check number for this fund
                   return self._get_swp_end_check_number(data)
          elif calculation == 'swp_check_count':
                   # Count SWP records for this fund
                   return self._get_swp_check_count(data)
          elif calculation == 'total_swp_dollar_value':
                   # Sum SWP amounts for this fund
                   return self._get_total_swp_dollar_value(data)
          elif calculation == 'calculate_total_input':
                   # Count total number of input rows in the report
                   return 1 # Will be summed up for each record
          elif calculation == 'calculate_total_packages':
                   # Count total number of package entries (rows)
                   return 1 # Each record is a package
          elif calculation == 'calculate_total_check_page':
                   # Sum of Total Input + Total Packages
                   return 2 # Total Input (1) + Total Packages (1) = 2
          elif calculation == 'calculate_total_dollar_amount':
                   # Sum of all check amounts present in the report
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'calculate_total_input':
                   # Count total number of input rows in the report
                   return 1 # Will be summed up for each record
          elif calculation == 'calculate_total_packages':
                   # Count total number of package entries (rows)
                   return 1 # Each record is a package
          elif calculation == 'calculate_total_check_page':
                   # Sum of Total Input + Total Packages
                   return 2 # Total Input (1) + Total Packages (1) = 2
          elif calculation == 'calculate_total_dollar_amount':
                   # Sum of all check amounts present in the report
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'static':
                   return col.get('value', 0)
          else:
                   return 0

     def _get_red_begin_check_number(self, data: Dict) -> str:
          """Get the first RED check number for this fund"""
          return data.get('RED_BEGIN_CHECK_NUMBER', '')

     def _get_red_end_check_number(self, data: Dict) -> str:
          """Get the last RED check number for this fund"""
          return data.get('RED_END_CHECK_NUMBER', '')

     def _get_red_check_count(self, data: Dict) -> int:
          """Count RED records for this fund"""
          return data.get('RED_CHECK_COUNT', 0)

     def _get_total_red_dollar_value(self, data: Dict) -> float:
          """Sum RED amounts for this fund"""
          return data.get('TOTAL_RED_DOLLAR_VALUE', 0.0)

     def _get_swp_begin_check_number(self, data: Dict) -> str:
          """Get the first SWP check number for this fund"""
          return data.get('SWP_BEGIN_CHECK_NUMBER', '')

     def _get_swp_end_check_number(self, data: Dict) -> str:
          """Get the last SWP check number for this fund"""
          return data.get('SWP_END_CHECK_NUMBER', '')

     def _get_swp_check_count(self, data: Dict) -> int:
          """Count SWP records for this fund"""
          return data.get('SWP_CHECK_COUNT', 0)

     def _get_total_swp_dollar_value(self, data: Dict) -> float:
          """Sum SWP amounts for this fund"""
          return data.get('TOTAL_SWP_DOLLAR_VALUE', 0.0)

     def _format_column_value(self, col: Dict, value: Any) -> str:
          """Format column value according to column configuration"""
          width = col.get('width', 0)
          align = col.get('align', 'left')
          format_type = col.get('format', 'text')

          if format_type == 'currency':
                   formatted = self.format_currency(value)
          elif format_type == 'number':
                   formatted = self.format_number(value, width)
          else:
                   formatted = str(value)

          if width > 0:
                   return self.align_text(formatted, width, align)
          else:
                   return formatted

class CSVFormatter:
     """Formats data as CSV with proper column alignment and headers"""

     def __init__(self, config: Dict):
          self.config = config.get('export_settings', {}).get('csv', {})
          self.delimiter = self.config.get('delimiter', ',')
          self.quote_char = self.config.get('quote_char', '"')
          self.encoding = self.config.get('encoding', 'utf-8')
          self.include_headers = self.config.get('include_headers', True)

     def format_currency(self, value: float) -> str:
          """Format value as currency"""
          try:
                   return f"{float(value):.2f}"
          except:
                   return "0.00"

     def clean_masked_value(self, value: str) -> str:
          """Clean masked values (remove asterisks and extract numeric part)"""
          if not value:
                   return "0.00"
         
          # Remove asterisks and extract numeric part
          cleaned = str(value).replace('*', '')
         
          # If the cleaned value is empty or just whitespace, return 0.00
          if not cleaned.strip():
                   return "0.00"
         
          return cleaned

     def safe_float_convert(self, value: str, default: float = 0.0) -> float:
          """Safely convert a value to float, handling masked values"""
          if not value:
                   return default
         
          # Clean masked values first
          cleaned = self.clean_masked_value(value)
         
          try:
                   return float(cleaned)
          except (ValueError, TypeError):
                   return default

     def create_csv_headers(self, columns: List[Dict]) -> List[str]:
          """Create CSV headers from column configuration with single column layout and consistent alignment"""
          headers = []
          for col in columns:
                   header = col.get('csv_header', col.get('name', ''))
                   # Single column layout - labels and values in one column with consistent alignment
                   headers.append(header)
          return headers

     def create_csv_row(self, columns: List[Dict], data: Dict) -> List[str]:
          """Create a CSV row from data with single column layout and consistent alignment"""
          row_values = []

          for col in columns:
                   value = self._get_column_value(col, data)
                   formatted_value = self._format_column_value(col, value)
              
                   row_values.append(formatted_value)

          return row_values

     def _get_column_value(self, col: Dict, data: Dict) -> Any:
          """Get value for a column from data"""
          source_field = col.get('source_field', '')

          if source_field == 'static':
                   return col.get('value', '')
          elif source_field == 'calculated':
                   return self._calculate_value(col, data)
          else:
                   value = data.get(source_field, '')
                   # For CUSTOM_FUND_CODE, pad with leading zeros to 7 characters
                   if source_field == 'CUSTOM_FUND_CODE' and value:
                        return value.zfill(7)
                   # For CUSTOM_CHECK_AMOUNT, clean masked values
                   elif source_field == 'CUSTOM_CHECK_AMOUNT' and value:
                        return self.clean_masked_value(value)
                   return value

     def _calculate_value(self, col: Dict, data: Dict) -> Any:
          """Calculate value based on column configuration"""
          calculation = col.get('calculation', '')

          if calculation == 'count_records':
                   return 1# Will be summed up
          elif calculation == 'sum_amounts':
                   # For Redemption_SWP reports, use CUSTOM_CHECK_AMOUNT
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'sum_sheets':
                   return float(data.get('DOC_SHEET_COUNT', 0))
          elif calculation == 'sum_images':
                   return float(data.get('DOC_IMAGE_COUNT', 0))
          elif calculation == 'sum_field':
                   field_name = col.get('field', '')
                   return float(data.get(field_name, 0))
          elif calculation == 'count_email_delivery':
                   return 1 if data.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'E' else 0
          elif calculation == 'count_print_delivery':
                   return 1 if data.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'P' else 0
          elif calculation == 'count_suppressed':
                   return 1 if data.get('DOC_SUPPRESSION_FLAG') == 'Y' else 0
          elif calculation == 'count_suppressed_amount':
                   # Count records with amount < $5.00
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   amount = self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 1 if amount < 5.00 else 0
          elif calculation == 'count_zero_check':
                   # Count records with zero check number
                   check_number = data.get('FN_CHECK_NUMBER', '')
                   return 1 if check_number == '0' or check_number == '' else 0
          elif calculation == 'count_suppressed_total':
                   # Count total suppressed records
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   amount = self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   check_number = data.get('CUSTOM_CHECK_NUMBER', '')
                   return 1 if amount < 5.00 or check_number == '0' or check_number == '' else 0
          elif calculation == 'count_packages':
                   return 1# Each record is a package
          elif calculation == 'count_check_pages':
                   return 2# Each check has 2 pages
          elif calculation == 'extract_group':
                   # Extract group from account number or other field
                   account = data.get('UC_ACCOUNT_NUMBER', '')
                   if len(account) >= 7:
                        return account[:7]
                   return account
          elif calculation == 'count_redemption_records':
                   # Count redemption records based on FN_CHECK_TYPE_DESC
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   return 1 if 'REDEMPTION' in check_type.upper() else 0
          elif calculation == 'sum_redemption_amounts':
                   # Sum amounts for redemption records
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   if 'REDEMPTION' in check_type.upper():
                        return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 0
          elif calculation == 'count_swp_records':
                   # Count SWP records based on FN_CHECK_TYPE_DESC
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   return 1 if 'SWP' in check_type.upper() else 0
          elif calculation == 'sum_swp_amounts':
                   # Sum amounts for SWP records
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   if 'SWP' in check_type.upper():
                        return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 0
          elif calculation == 'concatenate_fund_code_name':
                   # Concatenate CUSTOM_FUND_CODE and CUSTOM_FUND_DESC
                   fund_code = data.get('CUSTOM_FUND_CODE', '')
                   fund_desc = data.get('CUSTOM_FUND_DESC', '')
                   if fund_code and fund_desc:
                        return f"{fund_code} {fund_desc}"
                   elif fund_code:
                        return fund_code
                   elif fund_desc:
                        return fund_desc
                   return ""
          elif calculation == 'calculate_total_input':
                   # Count total number of input rows in the report
                   return 1 # Will be summed up for each record
          elif calculation == 'calculate_total_packages':
                   # Count total number of package entries (rows)
                   return 1 # Each record is a package
          elif calculation == 'calculate_total_check_page':
                   # Sum of Total Input + Total Packages
                   return 2 # Total Input (1) + Total Packages (1) = 2
          elif calculation == 'calculate_total_dollar_amount':
                   # Sum of all check amounts present in the report
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'static':
                   return col.get('value', 0)
          else:
                   return 0

     def _get_red_begin_check_number(self, data: Dict) -> str:
          """Get the first RED check number for this fund"""
          return data.get('RED_BEGIN_CHECK_NUMBER', '')

     def _get_red_end_check_number(self, data: Dict) -> str:
          """Get the last RED check number for this fund"""
          return data.get('RED_END_CHECK_NUMBER', '')

     def _get_red_check_count(self, data: Dict) -> int:
          """Count RED records for this fund"""
          return data.get('RED_CHECK_COUNT', 0)

     def _get_total_red_dollar_value(self, data: Dict) -> float:
          """Sum RED amounts for this fund"""
          return data.get('TOTAL_RED_DOLLAR_VALUE', 0.0)

     def _get_swp_begin_check_number(self, data: Dict) -> str:
          """Get the first SWP check number for this fund"""
          return data.get('SWP_BEGIN_CHECK_NUMBER', '')

     def _get_swp_end_check_number(self, data: Dict) -> str:
          """Get the last SWP check number for this fund"""
          return data.get('SWP_END_CHECK_NUMBER', '')

     def _get_swp_check_count(self, data: Dict) -> int:
          """Count SWP records for this fund"""
          return data.get('SWP_CHECK_COUNT', 0)

     def _get_total_swp_dollar_value(self, data: Dict) -> float:
          """Sum SWP amounts for this fund"""
          return data.get('TOTAL_SWP_DOLLAR_VALUE', 0.0)

     def _format_column_value(self, col: Dict, value: Any) -> str:
          """Format column value according to column configuration"""
          format_type = col.get('format', 'text')

          if format_type == 'currency':
                   return self.format_currency(value)
          else:
                   return str(value)

class XLSXFormatter:
     """Formats data as XLSX with proper styling and formatting"""

     def __init__(self, config: Dict):
          self.config = config.get('export_settings', {}).get('xlsx', {})
          self.sheet_name = self.config.get('sheet_name', 'Report')
          self.header_row = self.config.get('header_row', 1)
          self.data_start_row = self.config.get('data_start_row', 2)
          self.auto_fit_columns = self.config.get('auto_fit_columns', True)

     def create_workbook(self) -> Workbook:
          """Create a new workbook"""
          if not XLSX_AVAILABLE:
                   raise ImportError("openpyxl is required for XLSX export. Install with: pip install openpyxl")

          wb = Workbook()
          ws = wb.active
          ws.title = self.sheet_name
          return wb, ws

     def format_currency(self, value: float) -> str:
          """Format value as currency"""
          try:
                   return f"{float(value):.2f}"
          except:
                   return "0.00"

     def clean_masked_value(self, value: str) -> str:
          """Clean masked values (remove asterisks and extract numeric part)"""
          if not value:
                   return "0.00"
         
          # Remove asterisks and extract numeric part
          cleaned = str(value).replace('*', '')
         
          # If the cleaned value is empty or just whitespace, return 0.00
          if not cleaned.strip():
                   return "0.00"
         
          return cleaned

     def safe_float_convert(self, value: str, default: float = 0.0) -> float:
          """Safely convert a value to float, handling masked values"""
          if not value:
                   return default
         
          # Clean masked values first
          cleaned = self.clean_masked_value(value)
         
          try:
                   return float(cleaned)
          except (ValueError, TypeError):
                   return default

     def create_xlsx_headers(self, ws, columns: List[Dict], row: int = 1):
          """Create XLSX headers with single column layout and consistent alignment"""
          # Single column layout - labels and values in one column
          for col_idx, col in enumerate(columns, 1):
                   header = col.get('name', '')
                   cell = ws.cell(row=row, column=col_idx, value=header)
                   cell.font = Font(bold=True, size=10)
                   # All labels and titles consistently aligned to the left
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

     def create_xlsx_row(self, ws, columns: List[Dict], data: Dict, row: int):
          """Create an XLSX row from data with single column layout and consistent alignment"""
          for col_idx, col in enumerate(columns, 1):
                   value = self._get_column_value(col, data)
                   formatted_value = self._format_column_value(col, value)
                   cell = ws.cell(row=row, column=col_idx, value=formatted_value)
                   # All labels and titles consistently aligned to the left
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                   # Apply bold formatting to labels and titles
                   if col_idx == 1:# First column typically contains labels
                        cell.font = Font(bold=True, size=10)
                   else:
                        cell.font = Font(size=10)

     def _get_column_value(self, col: Dict, data: Dict) -> Any:
          """Get value for a column from data"""
          source_field = col.get('source_field', '')

          if source_field == 'static':
                   return col.get('value', '')
          elif source_field == 'calculated':
                   return self._calculate_value(col, data)
          else:
                   value = data.get(source_field, '')
                   # For CUSTOM_FUND_CODE, pad with leading zeros to 7 characters
                   if source_field == 'CUSTOM_FUND_CODE' and value:
                        return value.zfill(7)
                   # For CUSTOM_CHECK_AMOUNT, clean masked values
                   elif source_field == 'CUSTOM_CHECK_AMOUNT' and value:
                        return self.clean_masked_value(value)
                   return value

     def _calculate_value(self, col: Dict, data: Dict) -> Any:
          """Calculate value based on column configuration"""
          calculation = col.get('calculation', '')

          if calculation == 'count_records':
                   return 1# Will be summed up
          elif calculation == 'sum_amounts':
                   # For Redemption_SWP reports, use CUSTOM_CHECK_AMOUNT
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'sum_sheets':
                   return float(data.get('DOC_SHEET_COUNT', 0))
          elif calculation == 'sum_images':
                   return float(data.get('DOC_IMAGE_COUNT', 0))
          elif calculation == 'sum_field':
                   field_name = col.get('field', '')
                   return float(data.get(field_name, 0))
          elif calculation == 'count_email_delivery':
                   return 1 if data.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'E' else 0
          elif calculation == 'count_print_delivery':
                   return 1 if data.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'P' else 0
          elif calculation == 'count_suppressed':
                   return 1 if data.get('DOC_SUPPRESSION_FLAG') == 'Y' else 0
          elif calculation == 'count_suppressed_amount':
                   # Count records with amount < $5.00
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   amount = self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 1 if amount < 5.00 else 0
          elif calculation == 'count_zero_check':
                   # Count records with zero check number
                   check_number = data.get('FN_CHECK_NUMBER', '')
                   return 1 if check_number == '0' or check_number == '' else 0
          elif calculation == 'count_suppressed_total':
                   # Count total suppressed records
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   amount = self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   check_number = data.get('CUSTOM_CHECK_NUMBER', '')
                   return 1 if amount < 5.00 or check_number == '0' or check_number == '' else 0
          elif calculation == 'count_packages':
                   return 1# Each record is a package
          elif calculation == 'count_check_pages':
                   return 2# Each check has 2 pages
          elif calculation == 'extract_group':
                   # Extract group from account number or other field
                   account = data.get('UC_ACCOUNT_NUMBER', '')
                   if len(account) >= 7:
                        return account[:7]
                   return account
          elif calculation == 'count_redemption_records':
                   # Count redemption records based on FN_CHECK_TYPE_DESC
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   return 1 if 'REDEMPTION' in check_type.upper() else 0
          elif calculation == 'sum_redemption_amounts':
                   # Sum amounts for redemption records
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   if 'REDEMPTION' in check_type.upper():
                        return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 0
          elif calculation == 'count_swp_records':
                   # Count SWP records based on FN_CHECK_TYPE_DESC
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   return 1 if 'SWP' in check_type.upper() else 0
          elif calculation == 'sum_swp_amounts':
                   # Sum amounts for SWP records
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   if 'SWP' in check_type.upper():
                        return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 0
          elif calculation == 'concatenate_fund_code_name':
                   # Concatenate CUSTOM_FUND_CODE and CUSTOM_FUND_DESC
                   fund_code = data.get('CUSTOM_FUND_CODE', '')
                   fund_desc = data.get('CUSTOM_FUND_DESC', '')
                   if fund_code and fund_desc:
                        return f"{fund_code} {fund_desc}"
                   elif fund_code:
                        return fund_code
                   elif fund_desc:
                        return fund_desc
                   return ""
          elif calculation == 'calculate_total_input':
                   # Count total number of input rows in the report
                   return 1 # Will be summed up for each record
          elif calculation == 'calculate_total_packages':
                   # Count total number of package entries (rows)
                   return 1 # Each record is a package
          elif calculation == 'calculate_total_check_page':
                   # Sum of Total Input + Total Packages
                   return 2 # Total Input (1) + Total Packages (1) = 2
          elif calculation == 'calculate_total_dollar_amount':
                   # Sum of all check amounts present in the report
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'static':
                   return col.get('value', 0)
          else:
                   return 0

     def _get_red_begin_check_number(self, data: Dict) -> str:
          """Get the first RED check number for this fund"""
          return data.get('RED_BEGIN_CHECK_NUMBER', '')

     def _get_red_end_check_number(self, data: Dict) -> str:
          """Get the last RED check number for this fund"""
          return data.get('RED_END_CHECK_NUMBER', '')

     def _get_red_check_count(self, data: Dict) -> int:
          """Count RED records for this fund"""
          return data.get('RED_CHECK_COUNT', 0)

     def _get_total_red_dollar_value(self, data: Dict) -> float:
          """Sum RED amounts for this fund"""
          return data.get('TOTAL_RED_DOLLAR_VALUE', 0.0)

     def _get_swp_begin_check_number(self, data: Dict) -> str:
          """Get the first SWP check number for this fund"""
          return data.get('SWP_BEGIN_CHECK_NUMBER', '')

     def _get_swp_end_check_number(self, data: Dict) -> str:
          """Get the last SWP check number for this fund"""
          return data.get('SWP_END_CHECK_NUMBER', '')

     def _get_swp_check_count(self, data: Dict) -> int:
          """Count SWP records for this fund"""
          return data.get('SWP_CHECK_COUNT', 0)

     def _get_total_swp_dollar_value(self, data: Dict) -> float:
          """Sum SWP amounts for this fund"""
          return data.get('TOTAL_SWP_DOLLAR_VALUE', 0.0)

     def _format_column_value(self, col: Dict, value: Any) -> str:
          """Format column value according to column configuration"""
          format_type = col.get('format', 'text')

          if format_type == 'currency':
                   return self.format_currency(value)
          else:
                   return str(value)

class XMLDataExtractor:
     """Extracts data from XML files based on configuration"""

     def __init__(self):
          pass

     def extract_data(self, xml_file_path: str, config: Dict) -> List[Dict]:
          """Extract data from XML file based on configuration"""
          try:
                    # Try multiple encoding strategies
                    xml_content = None

                    # Initialize S3 client
                    s3_client = boto3.client('s3')

                    # Parse S3 path
                    if xml_file_path.startswith('s3://'):
                              # Extract bucket and key from S3 URL
                              s3_path_parts = xml_file_path[5:].split('/', 1)
                              bucket = s3_path_parts[0]
                              key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
                    else:
                              # Use environment variables for S3 bucket and key
                              bucket = os.getenv('s3BucketInput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
                              key = xml_file_path

                    # Read file from S3
                    try:
                              response = s3_client.get_object(Bucket=bucket, Key=key)
                              raw_content = response['Body'].read()
                              log_message("INFO", f"Successfully read XML file from S3: s3://{bucket}/{key}")
                    except ClientError as e:
                              log_message("ERROR", f"failed to read XML file from S3: {str(e)}")
                              return []

                    # Remove BOM if present
                    if raw_content.startswith(b'\xef\xbb\xbf'):
                              raw_content = raw_content[3:]
                              print("Removed UTF-8 BOM")
                    elif raw_content.startswith(b'\xff\xfe'):
                              raw_content = raw_content[2:]
                              print("Removed UTF-16 LE BOM")
                    elif raw_content.startswith(b'\xfe\xff'):
                              raw_content = raw_content[2:]
                              print("Removed UTF-16 BE BOM")

                    # Fix non-breaking spaces (common issue)
                    if b'\xc2\xa0' in raw_content[:200]:
                              print("Fixing non-breaking spaces in XML...")
                              raw_content = raw_content.replace(b'\xc2\xa0', b' ')

                    # Try to decode with different encodings
                    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                    for encoding in encodings:
                              try:
                                        xml_content = raw_content.decode(encoding)
                                        print(f"Successfully decoded with {encoding} encoding")
                                        break
                              except UnicodeDecodeError:
                                        continue

                    if xml_content is None:
                              print("Could not decode file with any encoding")
                              return []

                    # Clean the content
                    xml_content = xml_content.lstrip('\ufeff\ufffe\u200b')     # Remove BOM characters
                    xml_content = xml_content.replace('\xa0', ' ')                         # Replace non-breaking spaces
                    xml_content = xml_content.replace('\u00a0', ' ')
                   
                    # Remove control characters (except \n, \r, \t)
                    import re
                    xml_content = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]', '', xml_content)

                    # Strip leading/trailing whitespace
                    xml_content = xml_content.strip()

                    # Validate XML structure
                    if not xml_content.startswith('<?xml') and not xml_content.startswith('<'):
                              print("File does not appear to be valid XML")
                              print(f"First 100 characters: {xml_content[:100]}")
                              return []

                    # Parse the cleaned XML content
                    try:
                              root = ET.fromstring(xml_content)
                              print("✓ XML parsed successfully")
                    except ET.ParseError as e:
                              print(f"XML Parse Error: {e}")
                              print(f"First 200 characters of content: {repr(xml_content[:200])}")

                              # Try to fix common XML issues
                              print("Attempting to fix XML structure...")
                              try:
                                        # Try to add missing closing tags
                                        if '<Statements>' in xml_content and '</Statements>' not in xml_content:
                                                  xml_content = xml_content.replace('</File>', '</Statements>\n</File>')
                                                  print("Added missing </Statements> tag")

                                        # Try parsing again
                                        root = ET.fromstring(xml_content)
                                        print("✓ XML parsed successfully after fixing structure")
                              except ET.ParseError as e2:
                                        print(f"Still unable to parse XML after fixes: {e2}")
                                        return []

                    statements = []
                    statement_path = config['xml_mapping']['statement_path']

                    print(f"Extracting from {xml_file_path}")
                    print(f"Using statement path: {statement_path}")

                    # Find all statement elements
                    statement_elements = root.findall(statement_path)
                    print(f"Found {len(statement_elements)} statement elements")

                    if len(statement_elements) == 0:
                              # Try alternative paths
                              print("Trying alternative paths...")
                              alt_paths = ['Statement', '//Statement', './/Statement']
                              for alt_path in alt_paths:
                                        alt_elements = root.findall(alt_path)
                                        print(f"Path '{alt_path}': {len(alt_elements)} elements")
                                        if len(alt_elements) > 0:
                                                  statement_elements = alt_elements
                                                  break

                    # Extract data for each statement
                    for i, statement in enumerate(statement_elements):
                              data = {}
                              properties_path = config['xml_mapping']['properties_path']

                              # Find all property elements within this statement
                              prop_elements = statement.findall(properties_path)
                              print(f"Statement {i + 1}: Found {len(prop_elements)} property elements")

                              for prop in prop_elements:
                                        name = prop.get('Name')
                                        value = prop.get('Value')
                                        if name:
                                                  data[name] = value

                              # Add data if found
                              if data:
                                        statements.append(data)
                                        print(f"Added statement {i + 1} with {len(data)} properties")
                              else:
                                        print(f"No data found in statement {i + 1}")

                    print(f"Total statements extracted: {len(statements)}")
                    return statements

          except Exception as e:
                    print(f"Error extracting data from {xml_file_path}: {e}")
                    print('-----------------------------------------')
                    import traceback
                    traceback.print_exc()
                    return []


     def filter_data(self, data: List[Dict], filters: Dict) -> List[Dict]:
          """Filter data based on criteria"""
          if not filters:
                   return data

          filtered = []
          for record in data:
                   match = True
                   for field, expected_value in filters.items():
                        if record.get(field) != expected_value:
                             match = False
                             break
                   if match:
                        filtered.append(record)

          return filtered

class ReportGenerator:
     """Generates reports based on JSON configuration"""

     def __init__(self):
          self.formatter = ExcelFormatter()
          self.extractor = XMLDataExtractor()
         
          # Dynamic mapping table for DOC_SPECIAL_HANDLING_CODE to SHCode
          # This will be loaded from configuration for Redemption_SWP reports
          self.special_handling_mapping = {}

     def safe_float_convert(self, value: str, default: float = 0.0) -> float:
          """Safely convert a value to float, handling masked values"""
          if not value:
                   return default
         
          # Clean masked values first
          cleaned = str(value).replace('*', '')
         
          try:
                   return float(cleaned)
          except (ValueError, TypeError):
                   return default

     def clean_masked_value(self, value: str) -> str:
          """Clean masked values (remove asterisks and extract numeric part)"""
          if not value:
                   return "0.00"
         
          # Remove asterisks and extract numeric part
          cleaned = str(value).replace('*', '')
         
          # If the cleaned value is empty or just whitespace, return 0.00
          if not cleaned.strip():
                   return "0.00"
         
          return cleaned

     def format_currency(self, value: float) -> str:
          """Format value as currency"""
          try:
                   return f"{float(value):.2f}"
          except:
                   return "0.00"
         
     def load_config(self, config_file_path: str, file_type: str) -> Dict:
          """Load specific report configuration from unified config file"""
          try:
                    # Initialize S3 client
                    s3_client = boto3.client('s3')

                    # Use S3 path for configuration
                    if config_file_path.startswith('s3://'):
                              # Extract bucket and key from S3 URL
                              s3_path_parts = config_file_path[5:].split('/', 1)
                              bucket = s3_path_parts[0]
                              key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
                    else:
                              # Use environment variables for S3 bucket and key
                              bucket = os.getenv('s3BucketInput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
                              key = os.getenv('configKey', 'jhi/checks/config/jhi_checks_reporting_config.json')

                    # Read configuration from S3
                    try:
                              print(f"INFO: Attempting to load configuration from S3: s3://{bucket}/{key}")
                              response = s3_client.get_object(Bucket=bucket, Key=key)
                              config_content = response['Body'].read().decode('utf-8')
                              full_config = json.loads(config_content)
                              print(f"INFO: Successfully loaded configuration from S3: s3://{bucket}/{key}")
                              log_message("INFO", f"Successfully loaded configuration from S3: s3://{bucket}/{key}")
                    except ClientError as e:
                              print(f"ERROR: Failed to read configuration from S3: s3://{bucket}/{key} - {str(e)}")
                              log_message("ERROR", f"failed to read configuration from S3: {str(e)}")
                             
                              # Try alternative configuration paths
                              print("INFO: Attempting to find alternative configuration files...")
                              alternative_configs = [
                                        f"jhi/checks/config/jhi_checks_reporting_config.json",
                                        f"jhi/checks/config/{file_type}_config.json",
                                        f"jhi/checks/config/default_config.json"
                              ]
                             
                              for alt_key in alternative_configs:
                                        try:
                                                  print(f"INFO: Trying alternative configuration: s3://{bucket}/{alt_key}")
                                                  response = s3_client.get_object(Bucket=bucket, Key=alt_key)
                                                  config_content = response['Body'].read().decode('utf-8')
                                                  full_config = json.loads(config_content)
                                                  print(f"INFO: Successfully loaded alternative configuration: s3://{bucket}/{alt_key}")
                                                  log_message("INFO", f"Successfully loaded alternative configuration: s3://{bucket}/{alt_key}")
                                                  break
                                        except ClientError as alt_e:
                                                  print(f"WARNING: Alternative configuration not found: s3://{bucket}/{alt_key} - {str(alt_e)}")
                                                  continue
                              else:
                                        print("ERROR: No configuration files found in any alternative locations")
                                        log_message("ERROR", "No configuration files found in any alternative locations")
                                        return {}

                    config = full_config.get('report_types', {}).get(file_type, {})

                    # Load special handling mapping for Redemption_SWP reports
                    if file_type == 'redemption_swp' and 'layout' in config:
                              footer_config = config['layout'].get('footer_section', {})
                              if 'special_handling_mapping' in footer_config:
                                        self.special_handling_mapping = footer_config['special_handling_mapping']
                                        print(f"Loaded special handling mapping for {file_type}: "
                                                       f"{len(self.special_handling_mapping)} mappings")

                    return config

          except Exception as e:
                    print(f"Error loading config {config_file_path}: {e}")
                    return {}

     def generate_report(self, config_file_path: str) -> bool:
          """Generate report based on configuration file"""
          config = self.load_config(config_file_path)
          if not config:
                   return False

          return self.generate_report_from_config(config)
     
     # def send_email_notification(self,from_email, to_emails, subject, report_path):
     #    print("INSIDE send_email_notification") 
     #    try:
     #        ses_client = boto3.client('ses')
     #        body_html = f"""
     #        <html>
     #        <body>
     #            <p>Report successfully generated and uploaded to:</p>
     #            <p><strong>{report_path}</strong></p>
     #            <p>Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
     #        </body>
     #        </html>
     #        """
     #        ses_client.send_email(
     #            Source=from_email,
     #            Destination={'ToAddresses': to_emails},
     #            Message={
     #                'Subject': {'Data': subject},
     #                'Body': {
     #                    'Html': {'Data': body_html},
     #                    'Text': {'Data': ''}
     #                }
     #            }
     #        )
     #        print(f"✓ Email notification sent to: {', '.join(to_emails)}")
     #    except Exception as e:
     #        print(f"Error sending email: {e}")


#      def generate_report_from_config(self, config: Dict) -> bool:
#           """Generate report based on configuration dictionary"""
#           # Load special handling mapping for Redemption_SWP reports
#           report_type = config.get('workflow_metadata', {}).get('report_type')
#           if report_type == 'redemption_swp' and 'layout' in config:
#                footer_config = config['layout'].get('footer_section', {})
#                if 'special_handling_mapping' in footer_config:
#                         self.special_handling_mapping = footer_config['special_handling_mapping']
#                         print(f"Loaded special handling mapping for {report_type}: {len(self.special_handling_mapping)} mappings")
         
#           # Extract data from XML
#           print("EXtract XML")
#           xml_path = config['data_sources']['input_xml_path']
#           print("XML PATH - ", xml_path)

#           data = self.extractor.extract_data(xml_path, config)
#           if not data:
#                print(f"No data extracted from {xml_path}")
#                return False

#           # Get supported output formats with is_type control
#           output_formats_config = config.get('workflow_metadata', {}).get('output_formats', {'txt': {'is_type': True}})
#           output_paths = config['data_sources']['output_paths']

#           # Filter formats based on is_type attribute
#           active_formats = [fmt for fmt, settings in output_formats_config.items()
#                          if settings.get('is_type', False)]

#           success_count = 0
#           total_formats = len(active_formats)

#           if total_formats == 0:
#                print("No output formats enabled (all is_type: false)")
#                return False

#           # Generate reports for each active format
#           for format_type in active_formats:
#                format_settings = output_formats_config.get(format_type, {})
         
#                if format_type == 'txt':
#                         success = self._generate_txt_report(config, data, output_paths.get('txt'), format_settings)
#                elif format_type == 'csv':
#                         success = self._generate_csv_report(config, data, output_paths.get('csv'), format_settings)
#                elif format_type == 'xlsx':
#                         success = self._generate_xlsx_report(config, data, output_paths.get('xlsx'), format_settings)
#                else:
#                         print(f"Unsupported format: {format_type}")
#                         continue
         
#                if success:
#                         success_count += 1

#           return success_count == total_formats


    
     def generate_report_from_config(self, config: Dict) -> bool:
          """Generate report based on configuration dictionary"""

          # Construct input XML path from environment
          bucket = os.getenv("s3BucketInput")
          input_path = os.getenv("inputPath")
          input_file_name = os.getenv("inputFileName")
          xml_path = f"s3://{bucket}/{input_path}{input_file_name}"

          print(f"Using input XML path: {xml_path}")

          # Extract data
          data = self.extractor.extract_data(xml_path, config)
          if not data:
                    print(f"No data extracted from {xml_path}")
                    return False

          # Determine output formats
          output_formats_config = config.get("workflow_metadata", {}).get("output_formats", {"txt": {"is_type": True}})
          active_formats = [fmt for fmt, settings in output_formats_config.items() if settings.get("is_type", False)]
          print(f"Active formats: {active_formats}")

          if not active_formats:
                    print("No output formats enabled (all is_type: false)")
                    return False

          success_count = 0
          total_formats = len(active_formats)

          for format_type in active_formats:
                    format_settings = output_formats_config.get(format_type, {})
                    report_postfix = format_settings.get("report_postfix", "_report")
                    replace_ext = format_settings.get("replace_existing_ext", True)

                    # Construct output file name
                    base_name = os.path.splitext(input_file_name)[0]
                    output_file_name = f"{base_name}{report_postfix}.{format_type}" if replace_ext else f"{input_file_name}{report_postfix}.{format_type}"

                    # Construct full output path
                    report_path = os.getenv("reportPath", "jhi/checks/wip/wipid/report/")
                    output_path = f"s3://{bucket}/{report_path}{output_file_name}"
                   
                    print(output_path)

                    # Generate report
                    if format_type == "txt":
                              print(f"Calling TXT report generation for: {output_path}")
                              success = self._generate_txt_report(config, data, output_path, format_settings)
                    elif format_type == "csv":
                              success = self._generate_csv_report(config, data, output_path, format_settings)
                    elif format_type == "xlsx":
                              success = self._generate_xlsx_report(config, data, output_path, format_settings)
                    else:
                              print(f"Unsupported format: {format_type}")
                              continue

                    if success:
                              success_count += 1
                              
          # if success_count == total_formats:
          #   print("---------- 1 -------------")  
          #   email_config = config.get("emailConfig", {})
          #   print("---------- 2 -------------")
          #   from_email = email_config.get("fromEmail")
          #   print("---------- 3 -------------")
          #   to_emails = email_config.get("toEmails", [])
          #   print("---------- 4 -------------")
          #   subject = email_config.get("subject", "Report Generation Successful")
          #   print("---------- 5 -------------")

          #   if from_email and to_emails:
          #       print("---------- 6 -------------")
          #       self.send_email_notification(from_email, to_emails, subject, output_path)
          #       print("---------- 7 -------------")

          return success_count == total_formats

     def _generate_dynamic_output_path(self, xml_path: str, format_type: str, format_settings: Dict) -> str:
          """Generate dynamic output path based on new naming logic"""
          import os

          # Get the base filename from XML path
          base_filename = os.path.splitext(os.path.basename(xml_path))[0]

          # Get configuration settings
          replace_existing_ext = format_settings.get("replace_existing_ext", True)
          report_postfix = format_settings.get("report_postfix", "_report")

          # Build the new filename
          if replace_existing_ext:
                    # Replace the extension
                    new_filename = f"{base_filename}{report_postfix}.{format_type}"
          else:
                    # Append the new extension
                    new_filename = f"{base_filename}{report_postfix}.{format_type}"

          # Create the full output path in the result directory
          output_dir = "result"
          os.makedirs(output_dir, exist_ok=True)

          return os.path.join(output_dir, new_filename)

#      def _generate_txt_report(self, config: Dict, data: List[Dict], output_path: str, format_settings: Dict = None) -> bool:
#           """Generate TXT report with preserved formatting"""
#           if not output_path:
#                return False

#           if format_settings is None:
#                format_settings = {}

#           # Check if formatting should be preserved
#           preserve_formatting = format_settings.get('preserve_formatting', True)

#           # Generate report content
#           if preserve_formatting:
#                report_content = self._generate_report_content(config, data)
#           else:
#                # Generate simplified content if formatting preservation is disabled
#                report_content = self._generate_simplified_report_content(config, data)

#           # Upload to S3
#           try:
#                s3_client = boto3.client('s3')
              
#                # Parse S3 path or use environment variables
#                if output_path.startswith('s3://'):
#                         s3_path_parts = output_path[5:].split('/', 1)
#                         bucket = s3_path_parts[0]
#                         key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
#                else:
#                         bucket = os.getenv('s3BucketOutput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
#                         report_path = os.getenv('reportPath', 'jhi/checks/wip/wipid/report/')
#                         key = f"{report_path}{os.path.basename(output_path)}"
              
#                # Upload to S3
#                s3_client.put_object(
#                         Bucket=bucket,
#                         Key=key,
#                         Body=report_content.encode('utf-8'),
#                         ContentType='text/plain'
#                )
              
#                log_message("SUCCESS", f"TXT Report uploaded to S3: s3://{bucket}/{key}")
#                print(f"✓ TXT Report generated and uploaded: s3://{bucket}/{key}")
#                return True
#           except Exception as e:
#                log_message("ERROR", f"Error uploading TXT report to S3: {str(e)}")
#                print(f"Error uploading TXT report to S3: {e}")
#                return False

     def _generate_txt_report(self, config: Dict, data: List[Dict], output_path: str, format_settings: Dict = None) -> bool:
         """Generate TXT report with preserved formatting"""
         print("INSIDE _generate_txt_report")
         if not output_path:
              print("INSIDE 1")
              return False
         if format_settings is None:
              print("INSIDE 2") 
              format_settings = {}
         # Check if this is a Redemption_SWP report and handle specially
         print("INSIDE 3")
         report_type = config.get('workflow_metadata', {}).get('report_type')
         print("INSIDE 4")
         if report_type == 'redemption_swp':
              print("INSIDE 5")
              return self._generate_redemption_swp_txt_report(config, data, output_path)
         # Check if formatting should be preserved
         print("INSIDE 6")
         preserve_formatting = format_settings.get('preserve_formatting', True)
         # Generate report content
         print("INSIDE 7")
         if preserve_formatting:
              print("INSIDE 8")
              report_content = self._generate_report_content(config, data)
         else:
              # Generate simplified content if formatting preservation is disabled
              print("INSIDE 9")
              print("In else ")
              report_content = self._generate_simplified_report_content(config, data)
         # Save to output file
         #  os.makedirs(os.path.dirname(output_path), exist_ok=True)
                  
         s3_client = boto3.client('s3')
         # Parse S3 path
         print("INSIDE 11")
         if output_path.startswith('s3://'):
            print("INSIDE 12")
            s3_path_parts = output_path[5:].split('/', 1)
            print("INSIDE 13")
            bucket = s3_path_parts[0]
            print("INSIDE 14")
            key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
            print("INSIDE 15")
         else:
            print("INSIDE 16") 
            bucket = os.getenv('s3BucketOutput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
            print("INSIDE 17")
            report_path = os.getenv('reportPath', 'jhi/checks/wip/wipid/report/')
            print("INSIDE 18")
            key = f"{report_path}{os.path.basename(output_path)}"

          # Upload to S3
         print("INSIDE 19")
         s3_client.put_object(
            Bucket=bucket,
            Key=key,
            Body=report_content.encode('utf-8'),
            ContentType='text/plain'
          )
         print("INSIDE 21")
        
         try:
             print(f"✓ TXT Report generated: {output_path}")
             return True
         except Exception as e:
              print(f"Error saving TXT report to {output_path}: {e}")
              return False
     
    #  def _generate_redemption_swp_txt_report(self, config: Dict, data: List[Dict], output_path: str) -> bool:
    #      """Generate specialized TXT report for Redemption_SWP with pipeline separators"""
    #      print("INSIDE _generate_redemption_swp_txt_report")
    #      try:
    #           # Group data by fund and type
    #           grouped_data = self._group_data_by_fund_and_type_enhanced(data)
             
    #           # Create output lines
    #           lines = []
             
    #           # Add header information
    #           current_date = datetime.now().strftime("%m/%d/%Y")
    #           trade_date = "01/15/2025"
             
    #           lines.append(f"{current_date}|Remaining fields blank|This is the date of the report||||||||")
    #           lines.append(f"Trade Date: {trade_date}|Remaining fields blank|See Map Worksheet||||||||")
    #           lines.append("Fund Code and Fund Name from RRSU0224&RRSU0001.|||||||||")
    #           lines.append("Fund Code length: BR needs to add leading zeros - add leading zeros so fund code is always 7 in length. All of JHI our fund codes are 2 digits.|||||||||")
    #           lines.append("always blank. JHI would like BR to populate this|||||||||")
    #           lines.append("always blank. JHI would like BR to populate this|||||||||")
    #           lines.append("Specific order - list in Fund code order; All Redemption first in Fund code order, then all SWP in Fund code order.|Total of both both Redemption and SWP checks in the Fund|always blank. JHI would like BR to populate this||This is a total of both Redemption and SWP checks in the Fund|always blank. JHI would like BR to populate this||Total of both both Redemption and SWP checks in the Fund")
             
    #           # Process each fund in sorted order - Redemptions first, then SWPs
    #           for fund_code in sorted(grouped_data.keys()):
    #               fund_data = grouped_data[fund_code]
                 
    #               # Get fund description
    #               fund_desc = ""
    #               if fund_data.get('RED'):
    #                   fund_desc = fund_data['RED'][0].get('CUSTOM_FUND_DESC', '')
    #               elif fund_data.get('SWP'):
    #                   fund_desc = fund_data['SWP'][0].get('CUSTOM_FUND_DESC', '')
                 
    #               # Format fund code with leading zeros (7 digits total)
    #               formatted_fund_code = f"000{fund_code}"
                 
    #               # Calculate totals for this fund
    #               red_count = len(fund_data.get('RED', []))
    #               swp_count = len(fund_data.get('SWP', []))
    #               total_count = red_count + swp_count
                 
    #               red_total_amount = sum(self.safe_float_convert(r.get('CUSTOM_CHECK_AMOUNT', 0)) for r in fund_data.get('RED', []))
    #               swp_total_amount = sum(self.safe_float_convert(r.get('CUSTOM_CHECK_AMOUNT', 0)) for r in fund_data.get('SWP', []))
    #               total_amount = red_total_amount + swp_total_amount
                 
    #               # Get check number ranges
    #               red_begin_check = ""
    #               red_end_check = ""
    #               if fund_data.get('RED'):
    #                   red_check_numbers = [r.get('CUSTOM_CHECK_NUMBER', '') for r in fund_data['RED'] if r.get('CUSTOM_CHECK_NUMBER')]
    #                   if red_check_numbers:
    #                       red_check_numbers.sort()
    #                       red_begin_check = red_check_numbers[0]
    #                       red_end_check = red_check_numbers[-1]
                 
    #               swp_begin_check = ""
    #               swp_end_check = ""
    #               if fund_data.get('SWP'):
    #                   swp_check_numbers = [r.get('CUSTOM_CHECK_NUMBER', '') for r in fund_data['SWP'] if r.get('CUSTOM_CHECK_NUMBER')]
    #                   if swp_check_numbers:
    #                       swp_check_numbers.sort()
    #                       swp_begin_check = swp_check_numbers[0]
    #                       swp_end_check = swp_check_numbers[-1]
                 
    #               # Add Redemption entries first (if any)
    #               if red_count > 0:
    #                   fund_display_red = f"{formatted_fund_code} {fund_desc}"
    #                   line_red = f"{fund_display_red}|{red_count}|{red_total_amount:.2f}|{red_begin_check}|{red_end_check}|{red_count}|{red_total_amount:.2f}||||"
    #                   lines.append(line_red)
                 
    #               # Add SWP entries next (if any)
    #               if swp_count > 0:
    #                   fund_display_swp = f"{formatted_fund_code} {fund_desc}"
    #                   line_swp = f"{fund_display_swp}||||||{swp_begin_check}|{swp_end_check}|{swp_count}|{swp_total_amount:.2f}"
    #                   lines.append(line_swp)
                 
    #               # Add combined totals row for this fund
    #               if red_count > 0 or swp_count > 0:
    #                   fund_display_total = f"{formatted_fund_code} {fund_desc} - TOTAL"
    #                   line_total = f"{fund_display_total}|{total_count}|{total_amount:.2f}|{red_begin_check}|{red_end_check}|{red_count}|{red_total_amount:.2f}|{swp_begin_check}|{swp_end_check}|{swp_count}|{swp_total_amount:.2f}"
    #                   lines.append(line_total)
             
    #           # Add totals row
    #           total_records = len(data)
    #           total_amount_all = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
    #           lines.append(f"TOTAL|{total_records}|{total_amount_all:.2f}||||||||")
             
            #   s3_client = boto3.client('s3')
            #   # Parse S3 path
            #   print("INSIDE 11")
            #   if output_path.startswith('s3://'):
            #     print("INSIDE 12")
            #     s3_path_parts = output_path[5:].split('/', 1)
            #     print("INSIDE 13")
            #     bucket = s3_path_parts[0]
            #     print("INSIDE 14")
            #     key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
            #     print("INSIDE 15")
            #   else:
            #     print("INSIDE 16") 
            #     bucket = os.getenv('s3BucketOutput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
            #     print("INSIDE 17")
            #     report_path = os.getenv('reportPath', 'jhi/checks/wip/wipid/report/')
            #     print("INSIDE 18")
            #     key = f"{report_path}{os.path.basename(output_path)}"

            #  # Upload to S3
            #   print("INSIDE 19")
            #   s3_client.put_object(
            #      Bucket=bucket,
            #      Key=key,
            #      Body='\n'.join(lines).encode('utf-8'),
            #      ContentType='text/plain'
            #  )
            #   print("INSIDE 21")

             
    #           print(f"✓ Redemption_SWP TXT Report generated: {output_path}")
    #           return True
             
    #      except Exception as e:
    #           print(f"Error generating Redemption_SWP TXT report: {e}")
    #           return False
    
     def _generate_redemption_swp_txt_report(self, config: Dict, data: List[Dict], output_path: str) -> bool:
         """Generate specialized TXT report for Redemption_SWP matching XLSX logic exactly"""
         try:
              # Use the same logic as XLSX: filter and sort data exactly like XLSX version
              # Filter records to only include those with CUSTOM_CHECK_TYPE1 = "RED" or "SWP"
              filtered_data = [record for record in data
                             if record.get('CUSTOM_CHECK_TYPE1') in ['RED', 'SWP']]
             
              # Sort by fund code, then by check type (RED first, then SWP) - same as XLSX
              filtered_data.sort(key=lambda x: (x.get('CUSTOM_FUND_CODE', ''),
                                          x.get('CUSTOM_CHECK_TYPE1', '')))
             
              # Create output lines
              lines = []
             
              # Add header information - match XLSX layout exactly
              current_date = datetime.now().strftime("%m/%d/%Y")
              trade_date = "01/15/2025"
             
              lines.append(f"{current_date}|Remaining fields blank|This is the date of the report||||||||")
              lines.append(f"Trade Date: {trade_date}|Remaining fields blank|See Map Worksheet||||||||")
              lines.append("Fund Code and Fund Name from RRSU0224&RRSU0001.|||||||||")
              lines.append("Fund Code length: BR needs to add leading zeros - add leading zeros so fund code is always 7 in length. All of JHI our fund codes are 2 digits.|||||||||")
              lines.append("always blank. JHI would like BR to populate this|||||||||")
              lines.append("always blank. JHI would like BR to populate this|||||||||")
              lines.append("Specific order - list in Fund code order; All Redemption first in Fund code order, then all SWP in Fund code order.|Total of both both Redemption and SWP checks in the Fund|always blank. JHI would like BR to populate this||This is a total of both Redemption and SWP checks in the Fund|always blank. JHI would like BR to populate this||Total of both both Redemption and SWP checks in the Fund")
             
              # Get column configuration from config - same as XLSX
              layout = config.get('layout', {})
              data_section = layout.get('data_section', {})
              columns = data_section.get('columns', [])
             
              # Process individual records exactly like XLSX version
              for record in filtered_data:
                  # Create enhanced record with individual record data - same as XLSX
                  enhanced_record = self._create_individual_record(record)
                 
                  # Generate row using same column logic as XLSX
                  row_values = []
                  for col in columns:
                      value = self._get_column_value_for_xlsx(col, enhanced_record)
                      formatted_value = self._format_column_value_for_xlsx(col, value)
                      row_values.append(formatted_value)
                 
                  # Join with pipe separator for TXT format
                  line = '|'.join(str(v) for v in row_values)
                  lines.append(line)
             
              # Add totals row using same column structure
              if columns:
                  total_values = []
                  for col in columns:
                      if col.get('name') == 'GROWTHFUND':
                          total_values.append('TOTAL')
                      elif col.get('name') == 'CHECKS PRINTED TODAY':
                          total_values.append(len(filtered_data))
                      elif col.get('name') == 'DOLLAR VALUE TODAY':
                          total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in filtered_data)
                          total_values.append(f"{total_amount:.2f}")
                      else:
                          total_values.append('')
                 
                  total_line = '|'.join(str(v) for v in total_values)
                  lines.append(total_line)
             
              # Add footer section - use specialized format for Redemption_SWP
              footer_lines = self._generate_redemption_swp_footer_section(config, data)
              if footer_lines:
                  lines.append('')  # Add blank line before footer
                  lines.extend(footer_lines)
             
              s3_client = boto3.client('s3')
              # Parse S3 path
              print("INSIDE 11")
              if output_path.startswith('s3://'):
                print("INSIDE 12")
                s3_path_parts = output_path[5:].split('/', 1)
                print("INSIDE 13")
                bucket = s3_path_parts[0]
                print("INSIDE 14")
                key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
                print("INSIDE 15")
              else:
                print("INSIDE 16") 
                bucket = os.getenv('s3BucketOutput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
                print("INSIDE 17")
                report_path = os.getenv('reportPath', 'jhi/checks/wip/wipid/report/')
                print("INSIDE 18")
                key = f"{report_path}{os.path.basename(output_path)}"

             # Upload to S3
              print("INSIDE 19")
              s3_client.put_object(
                 Bucket=bucket,
                 Key=key,
                 Body='\n'.join(lines).encode('utf-8'),
                 ContentType='text/plain'
             )
              print("INSIDE 21")
             
              print(f"✓ Redemption_SWP TXT Report generated: {output_path}")
              return True
             
         except Exception as e:
              print(f"Error generating Redemption_SWP TXT report: {e}")
              return False
              
     def _generate_redemption_swp_footer_section(self, config: Dict, data: List[Dict]) -> List[str]:
         """Generate footer section for Redemption_SWP reports with pipe-separated format"""
         layout = config['layout']
         footer_config = layout.get('footer_section', {})
         
         lines = []
         
         # Categories from configuration
         categories = footer_config.get('categories', [])
         
         # Calculate dynamic footer data
         category_data = self._calculate_dynamic_footer_data(data)
         
         # Add the special header lines first
         lines.append("RETURN TO JANUS||||||||||                                                                                                                                                                                                                                                                                   ")
         lines.append("RETURN TO JANUS|NUMBER RECEIVED|||||||||                                                                                                                                                                                                                                                                    ")
         
         # Generate footer lines with pipe-separated format
         for i, category in enumerate(categories):
             data_value = category_data[i] if i < len(category_data) else 0
             
             # Format the category name and value with pipe separators
             # Format the value as 7-digit zero-padded number
             formatted_value = f"{data_value:07d}"
             
             # Map category names to the expected format
             display_name = category
             if category == "RETURNTOJANUS":
                 display_name = "JANUS"
             elif category == "STATESTREETBANK":
                 display_name = "STATE STREET BANK"
             elif category == "DONOTMAIL":
                 display_name = "DO NOT MAIL"
             elif category == "PULLREPORTCHECKS":
                 display_name = "PULL REPORT CHECKS"
             elif category == "SUB-TOTAL":
                 display_name = "SUB-TOTAL"
             elif category == "SPECIALHANDLING":
                 display_name = "SPECIAL HANDLING"
             elif category == "OVERNIGHTW-SIG":
                 display_name = "OVERNIGHT W-SIG"
             elif category == "OVERNIGHTWO-SIG":
                 display_name = "OVERNIGHT WO-SIG"
             elif category == "OVERNIGHTSAT-W-SIG":
                 display_name = "OVERNIGHT SAT-W-SIG"
             elif category == "OVERNIGHTSAT-WO-SIG":
                 display_name = "OVERNIGHT SAT-WO-SIG"
             elif category == "SUPPRESSEDTEST":
                 display_name = "SUPPRESSED TEST"
             
             # Create the line with pipe separators and proper spacing
             # Each line should have the category name, value, and then many pipes with spaces
             line = f"{display_name}|{formatted_value}|||||||||                                                                                                                                                                                                                                                                                        "
             lines.append(line)
         
         return lines

              
     def _generate_csv_report(self, config: Dict, data: List[Dict], output_path: str, format_settings: Dict = None) -> bool:
          """Generate CSV report with structured layout matching TXT format"""
          if not output_path:
                   return False

          if format_settings is None:
                   format_settings = {}

          os.makedirs(os.path.dirname(output_path), exist_ok=True)

          try:
                   # Check if this is a 12b1_compensation report and handle specially
                   report_type = config.get('workflow_metadata', {}).get('report_type')
                   if report_type == '12b1_compensation':
                        return self._generate_12b1_compensation_csv(config, data, output_path)
              
                   # Generate structured content like TXT format
                   report_content = self._generate_report_content(config, data)
         
                   # Convert structured content to CSV format with single column layout and consistent alignment
                   lines = report_content.split('\n')
                   csv_lines = []
         
                   for line in lines:
                        if line.strip() == "":
                             # Empty lines become empty CSV rows
                             csv_lines.append([])
                        elif '\t' in line:
                             # Tab-separated lines - convert to single column layout with consistent alignment
                             cells = line.split('\t')
                             # Create label-value pairs in single column with consistent alignment
                             csv_row = []
                             for i, cell_value in enumerate(cells):
                                  if i == 0:
                                        # First column contains the label - consistently aligned
                                        csv_row.append(cell_value)
                                  else:
                                        # Subsequent columns contain values - consistently aligned
                                        if cell_value.strip():# Only add non-empty values
                                             csv_row.append(cell_value)
                             csv_lines.append(csv_row)
                        else:
                             # Regular lines become single-column CSV rows with consistent alignment
                             csv_lines.append([line])
              
                   # Enhanced Logic for Redemption_SWP Report - CSV
                   if report_type == 'redemption_swp':
                        csv_lines = self._apply_redemption_swp_enhanced_logic_csv(csv_lines, data)
         
                   # Upload CSV to S3
                   s3_client = boto3.client('s3')
              
                   # Parse S3 path or use environment variables
                   if output_path.startswith('s3://'):
                        s3_path_parts = output_path[5:].split('/', 1)
                        bucket = s3_path_parts[0]
                        key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
                   else:
                        bucket = os.getenv('s3BucketOutput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
                        report_path = os.getenv('reportPath', 'jhi/checks/wip/wipid/report/')
                        key = f"{report_path}{os.path.basename(output_path)}"
              
                   # Convert CSV data to string
                   csv_content = io.StringIO()
                   writer = csv.writer(csv_content, delimiter=',', quotechar='"')
                   for row in csv_lines:
                        writer.writerow(row)
                   csv_string = csv_content.getvalue()
              
                   # Upload to S3
                   s3_client.put_object(
                        Bucket=bucket,
                        Key=key,
                        Body=csv_string.encode('utf-8'),
                        ContentType='text/csv'
                   )
              
                   log_message("SUCCESS", f"CSV Report uploaded to S3: s3://{bucket}/{key}")
                   print(f"✓ CSV Report generated and uploaded: s3://{bucket}/{key}")
                   return True
          except Exception as e:
                   log_message("ERROR", f"Error uploading CSV report to S3: {str(e)}")
                   print(f"Error uploading CSV report to S3: {e}")
                   return False

     def _generate_xlsx_report(self, config: Dict, data: List[Dict], output_path: str, format_settings: Dict = None) -> bool:
          """Generate XLSX report with proper cell separation"""
          if not output_path:
                   return False

          if not XLSX_AVAILABLE:
                   print("XLSX export requires openpyxl. Install with: pip install openpyxl")
                   return False

          if format_settings is None:
                   format_settings = {}

          os.makedirs(os.path.dirname(output_path), exist_ok=True)

          try:
                   wb = Workbook()
                   ws = wb.active
                   ws.title = config.get('export_settings', {}).get('xlsx', {}).get('sheet_name', 'Report')
         
                   current_row = 1
         
                   # For Redemption_SWP reports, ensure static text is placed FIRST in row 3 (A3, B3, C3)
                   report_type = config.get('workflow_metadata', {}).get('report_type')
                   if report_type == 'redemption_swp':
                        # Place static text in row 3 (A3, B3, C3) - these cells must remain fixed
                        self._apply_redemption_swp_enhanced_logic_xlsx(ws, data, current_row)
         
                   # Generate header only if header section exists
                   if 'header' in config.get('layout', {}):
                        header = self._generate_header(config)
                        header_lines = header.split('\n')
                        for line in header_lines:
                             if line.strip():
                                  cell = ws.cell(row=current_row, column=1, value=line)
                                  # Check if this is a subtitle (second non-empty line after title)
                                  header_config = config.get('layout', {}).get('header', {})
                                  subtitle = header_config.get('subtitle', '')
                                  if subtitle and line.strip() == subtitle:
                                        # Subtitle formatting - slightly smaller font
                                        cell.font = Font(bold=True, size=9)
                                  else:
                                        # Title formatting - normal size
                                        cell.font = Font(bold=True, size=10)
                                  cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
         
                   # Generate summary section with proper cell separation
                   current_row = self._generate_xlsx_summary_section(ws, config, data, current_row)
         
                   # Generate totals section with proper cell separation - position depends on report type
                   if 'totals_section' in config.get('layout', {}):
                        report_type = config.get('workflow_metadata', {}).get('report_type')
                        if report_type == 'replacement':
                             # For replacement report, totals appear before data section
                             current_row = self._generate_xlsx_totals_section(ws, config, data, current_row)
                        elif report_type == 'redemption_swp':
                             pass
                        else:
                             # Default behavior - totals before data
                             current_row = self._generate_xlsx_totals_section(ws, config, data, current_row)
         
                   # Generate data section with proper column separation
                   current_row = self._generate_xlsx_data_section(ws, config, data, current_row)

                   # Generate totals section for redemption_swp report (after data sections)
                   if 'totals_section' in config.get('layout', {}) and config.get('workflow_metadata', {}).get('report_type') == 'redemption_swp':
                        current_row = self._generate_xlsx_totals_section(ws, config, data, current_row)
         
                   # Generate footer if exists with proper cell separation
                   if 'footer_section' in config['layout']:
                        current_row = self._generate_xlsx_footer_section(ws, config, data, current_row)
         
                   # Auto-fit columns
                   for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                             try:
                                  if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                             except:
                                  pass
                        # Set appropriate width for each column
                        if column_letter == 'A':# First column (labels) - wider for readability
                             adjusted_width = min(max_length + 8, 70)
                        else:# Data columns - appropriate width for values
                             adjusted_width = min(max_length + 5, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
              
                        # Set row height to accommodate wrapped text
                        for cell in column:
                             if cell.value and len(str(cell.value)) > 40:# Long values need more height
                                  ws.row_dimensions[cell.row].height = 35
                             elif cell.value and len(str(cell.value)) > 20:# Medium values
                                  ws.row_dimensions[cell.row].height = 25
                             elif cell.value:# Regular rows with proper spacing
                                  ws.row_dimensions[cell.row].height = 20
         
                   # Upload XLSX to S3
                   s3_client = boto3.client('s3')
              
                   # Parse S3 path or use environment variables
                   if output_path.startswith('s3://'):
                        s3_path_parts = output_path[5:].split('/', 1)
                        bucket = s3_path_parts[0]
                        key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
                   else:
                        bucket = os.getenv('s3BucketOutput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
                        report_path = os.getenv('reportPath', 'jhi/checks/wip/wipid/report/')
                        key = f"{report_path}{os.path.basename(output_path)}"
              
                   # Save workbook to bytes
                   xlsx_buffer = io.BytesIO()
                   wb.save(xlsx_buffer)
                   xlsx_buffer.seek(0)
              
                   # Upload to S3
                   s3_client.put_object(
                        Bucket=bucket,
                        Key=key,
                        Body=xlsx_buffer.getvalue(),
                        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                   )
              
                   log_message("SUCCESS", f"XLSX Report uploaded to S3: s3://{bucket}/{key}")
                   print(f"✓ XLSX Report generated and uploaded: s3://{bucket}/{key}")
                   return True
          except Exception as e:
                   log_message("ERROR", f"Error uploading XLSX report to S3: {str(e)}")
                   print(f"Error uploading XLSX report to S3: {e}")
                   return False

     def _generate_report_content(self, config: Dict, data: List[Dict]) -> str:
          """Generate the complete report content"""
          content_parts = []
         
          # Check if this is a 12b1_compensation report and handle specially
          report_type = config.get('workflow_metadata', {}).get('report_type')
          if report_type == '12b1_compensation':
                   return self._generate_12b1_compensation_content(config, data)

          # Generate header only if header section exists
          if 'header' in config.get('layout', {}):
                   header = self._generate_header(config)
                   content_parts.append(header)

          # Generate summary section if exists
          if 'summary_section' in config.get('layout', {}):
                   summary = self._generate_summary_section(config, data)
                   content_parts.append(summary)

          # Generate totals section if exists - position depends on report type
          if 'totals_section' in config.get('layout', {}):
                   if report_type == 'replacement':
                        # For replacement report, totals appear before data section
                        totals = self._generate_totals_section(config, data)
                        content_parts.append(totals)
                   elif report_type == 'redemption_swp':
                        # For Redemption report, totals should appear at the bottom after all tables
                        # This will be handled after data sections
                        pass
                   elif report_type == 'dividend_checks':
                        # For dividend report, totals should appear after data section
                        # This will be handled after data sections
                        pass
                   else:
                        # Default behavior - totals before data
                        totals = self._generate_totals_section(config, data)
                        content_parts.append(totals)

          # Generate data section if exists
          if 'data_section' in config.get('layout', {}):
                   data_section = self._generate_data_section(config, data)
                   content_parts.append(data_section)

          # Generate matrix sections if exist
          if 'matrix_sections' in config.get('layout', {}):
                   matrix_content = self._generate_matrix_text_section(config, data)
                   content_parts.append(matrix_content)

          # Generate totals section for redemption_swp and dividend_checks reports (after data sections)
          if 'totals_section' in config.get('layout', {}) and report_type in ['redemption_swp', 'dividend_checks']:
                   # Add blank line before totals section
                   content_parts.append("")# One blank line above totals
                   totals = self._generate_totals_section(config, data)
                   content_parts.append(totals)

          # Generate footer if exists
          if 'footer_section' in config.get('layout', {}):
                   footer = self._generate_footer_section(config, data)
                   content_parts.append(footer)

          return '\n'.join(content_parts)

     def _generate_12b1_compensation_content(self, config: Dict, data: List[Dict]) -> str:
          """Generate specialized content for 12b1_compensation reports to match XLSX layout"""
          content_parts = []
         
          # Generate header
          if 'header' in config.get('layout', {}):
                   header = self._generate_header(config)
                   content_parts.append(header)
         
          # Generate summary section
          if 'summary_section' in config.get('layout', {}):
                   summary = self._generate_summary_section(config, data)
                   content_parts.append(summary)
         
          # Generate totals section
          if 'totals_section' in config.get('layout', {}):
                   totals = self._generate_totals_section(config, data)
                   content_parts.append(totals)
         
          # Generate suppression sections with proper formatting
          layout = config.get('layout', {})
          suppression_sections = layout.get('suppression_sections', [])
         
          for section in suppression_sections:
                   section_content = self._generate_12b1_suppression_section(section, data)
                   content_parts.append(section_content)
         
          return '\n'.join(content_parts)

     def _generate_12b1_suppression_section(self, section: Dict, data: List[Dict]) -> str:
          """Generate a suppression section for 12b1_compensation reports with proper formatting"""
          lines = []
         
          # Add section title
          title = section.get('title', '')
          if title:
                   lines.append(title)
         
          # Add subtitle if provided
          subtitle = section.get('subtitle', '')
          if subtitle:
                   lines.append("") # Empty line separator
                   lines.append(subtitle)
         
          # Add date if provided
          date = section.get('date', '')
          if date:
                   lines.append("") # Empty line separator
                   lines.append(date)
         
          # Generate column headers with proper alignment
          columns = section.get('columns', [])
          if columns:
                   header_parts = []
              
                   for col in columns:
                        name = col.get('name', '')
                        width = col.get('width', 0)
                        align = col.get('align', 'left')
                   
                        if width > 0:
                             formatted = self.formatter.align_text(name, width, align)
                        else:
                             formatted = name
                   
                        header_parts.append(formatted)
              
                   lines.append(self.formatter.tab_char.join(header_parts))
              
                   # Apply filtering based on CUSTOM_CHECK_NUMBER for 12b1_compensation reports
                   filtered_data = self._filter_12b1_data_by_check_number(data, section)
              
                   # Generate data rows for this section
                   for record in filtered_data:
                        row = self.formatter.create_excel_row(columns, record)
                        lines.append(row)
         
          # Add dynamic summary if configured
          summary_config = section.get('summary', {})
          if summary_config:
                   summary_lines = self._generate_12b1_section_summary(section, filtered_data)
                   lines.extend(summary_lines)
         
          # Add spacing between sections
          lines.append("")
         
          return '\n'.join(lines)

     def _generate_12b1_section_summary(self, section: Dict, data: List[Dict]) -> List[str]:
          """Generate summary for 12b1_compensation suppression sections"""
          lines = []
          summary_config = section.get('summary', {})
         
          if not summary_config:
                   return lines
         
          # Calculate summary values
          calculation = summary_config.get('calculation', '')
          if calculation == 'sum_amounts':
                   # Use CUSTOM_CHECK_AMOUNT for 12b1_compensation report to match the column display
                   # For 12b1_compensation report, sum all amounts in the filtered data (not just < $5.00)
                   total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
                   formatted_amount = self.formatter.format_currency(total_amount)
              
                   # Generate summary line
                   summary_line = summary_config.get('line', '')
                   if '{sum_amount}' in summary_line:
                        summary_line = summary_line.replace('{sum_amount}', formatted_amount)
              
                   lines.append(summary_line)
              
          elif calculation == 'count_records':
                   total_count = len(data)
              
                   # Generate summary line
                   summary_line = summary_config.get('line', '')
                   if '{count_rows}' in summary_line:
                        summary_line = summary_line.replace('{count_rows}', str(total_count))
              
                   lines.append(summary_line)
         
          # Add title after if configured
          title_after = summary_config.get('title_after', '')
          if title_after:
                   lines.append("")
                   lines.append(title_after)
         
          return lines

     def _filter_12b1_data_by_check_number(self, data: List[Dict], section: Dict) -> List[Dict]:
          """Filter data based on CUSTOM_CHECK_NUMBER for 12b1_compensation suppression sections"""
          # Determine filter type based on section title or filter configuration
          title = section.get('title', '').lower()
          subtitle = section.get('subtitle', '').lower()
          filter_type = section.get('filter', '')
         
          # Table 1: CUSTOM_CHECK_NUMBER ≠ 0 (non-zero check numbers)
          # Table 2: CUSTOM_CHECK_NUMBER = 0 (zero check numbers)
         
          if 'zero' in title or 'zero' in subtitle or filter_type == 'zero_check_number':
                   # Table 2: Show only records where CUSTOM_CHECK_NUMBER = 0
                   return [record for record in data if self._is_zero_check_number(record.get('CUSTOM_CHECK_NUMBER', ''))]
          elif filter_type == 'non_zero_check_number' or filter_type == 'amount_less_than_5':
                   # Table 1: Show only records where CUSTOM_CHECK_NUMBER ≠ 0
                   return [record for record in data if not self._is_zero_check_number(record.get('CUSTOM_CHECK_NUMBER', ''))]
          else:
                   # Default: Show only records where CUSTOM_CHECK_NUMBER ≠ 0
                   return [record for record in data if not self._is_zero_check_number(record.get('CUSTOM_CHECK_NUMBER', ''))]

     def _is_zero_check_number(self, check_number: str) -> bool:
          """Check if a check number is considered zero"""
          if not check_number:
                   return True
          try:
                   # Convert to float to handle numeric strings
                   num_value = float(check_number)
                   return num_value == 0.0
          except (ValueError, TypeError):
                   # If it's not a number, check if it's '0' or empty
                   return check_number.strip() == '0' or check_number.strip() == ''

     def _generate_12b1_compensation_csv(self, config: Dict, data: List[Dict], output_path: str) -> bool:
          """Generate CSV report for 12b1_compensation with proper column structure matching XLSX layout"""
          try:
                   csv_lines = []
              
                   # Generate header section
                   if 'header' in config.get('layout', {}):
                        header = self._generate_header(config)
                        header_lines = header.split('\n')
                        for line in header_lines:
                             if line.strip():
                                  csv_lines.append([line])
              
                   # Generate summary section
                   if 'summary_section' in config.get('layout', {}):
                        summary = self._generate_summary_section(config, data)
                        summary_lines = summary.split('\n')
                        for line in summary_lines:
                             if line.strip():
                                  csv_lines.append([line])
              
                   # Generate totals section
                   if 'totals_section' in config.get('layout', {}):
                        totals = self._generate_totals_section(config, data)
                        totals_lines = totals.split('\n')
                        for line in totals_lines:
                             if line.strip():
                                  csv_lines.append([line])
              
                   # Generate suppression sections with proper CSV structure
                   layout = config.get('layout', {})
                   suppression_sections = layout.get('suppression_sections', [])
              
                   for section in suppression_sections:
                        # Add section title
                        title = section.get('title', '')
                        if title:
                             csv_lines.append([title])
                   
                        # Add subtitle if provided
                        subtitle = section.get('subtitle', '')
                        if subtitle:
                             csv_lines.append([]) # Empty row
                             csv_lines.append([subtitle])
                   
                        # Add date if provided
                        date = section.get('date', '')
                        if date:
                             csv_lines.append([]) # Empty row
                             csv_lines.append([date])
                   
                        # Generate column headers
                        columns = section.get('columns', [])
                        if columns:
                             header_row = []
                             for col in columns:
                                  header_row.append(col.get('name', ''))
                             csv_lines.append(header_row)
                        
                             # Apply filtering based on CUSTOM_CHECK_NUMBER for 12b1_compensation reports
                             filtered_data = self._filter_12b1_data_by_check_number(data, section)
                        
                             # Generate data rows for this section
                             for record in filtered_data:
                                  data_row = []
                                  for col in columns:
                                        value = self._get_column_value_for_csv(col, record)
                                        data_row.append(value)
                                  csv_lines.append(data_row)
                   
                        # Add dynamic summary if configured
                        summary_config = section.get('summary', {})
                        if summary_config:
                             summary_lines = self._generate_12b1_section_summary(section, filtered_data)
                             for line in summary_lines:
                                  csv_lines.append([line])
                   
                        # Add spacing between sections
                        csv_lines.append([])
              
                   # Write CSV file
                   with open(output_path, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f, delimiter=',', quotechar='"')
                        for row in csv_lines:
                             writer.writerow(row)
              
                   print(f"✓ CSV Report generated: {output_path}")
                   return True
          except Exception as e:
                   print(f"Error saving 12b1_compensation CSV report to {output_path}: {e}")
                   return False

     def _get_column_value_for_csv(self, col: Dict, data: Dict) -> str:
          """Get column value for CSV generation"""
          source_field = col.get('source_field', '')
          calculation = col.get('calculation', '')
          format_type = col.get('format', 'text')
         
          if calculation == 'extract_group':
                   # Extract group from dealer number (first 2 digits)
                   dealer_number = data.get('FN_DEALER_NUMBER', '')
                   if dealer_number and len(dealer_number) >= 2:
                        return dealer_number[:2]
                   return ''
          elif source_field == 'calculated':
                   # Handle calculated fields
                   if calculation == 'extract_group':
                        dealer_number = data.get('FN_DEALER_NUMBER', '')
                        if dealer_number and len(dealer_number) >= 2:
                             return dealer_number[:2]
                        return ''
                   else:
                        return ''
          else:
                   # Direct field mapping
                   value = data.get(source_field, '')
              
                   if format_type == 'currency':
                        try:
                             float_value = float(value) if value else 0.0
                             return self.formatter.format_currency(float_value)
                        except (ValueError, TypeError):
                             return str(value)
                   else:
                        return str(value)

     def _generate_header(self, config: Dict) -> str:
          """Generate report header with exact current design formatting"""
          layout = config['layout']
          header_config = layout.get('header', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          lines = []
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks':
                   # Title without parentheses
                   title = header_config.get('title', 'JANUS FUNDS: DIVIDEND CHECKS')
                   lines.append(title)
              
                   # Centered subtitle
                   subtitle = header_config.get('subtitle', 'AUDIT REPORT')
                   if subtitle:
                        # Center the subtitle with proper spacing
                        lines.append("")
                        lines.append(f"      {subtitle}")
              
                   # Fields with specific alignment
                   fields = header_config.get('fields', [])
                   for field in fields:
                        name = field.get('name', '')
                        position = field.get('position', 'left')
                        alignment = field.get('alignment', 'left')
                   
                        # Handle dynamic date replacement
                        if '{CURRENT_DATE}' in name:
                             current_date = datetime.now().strftime('%m/%d/%Y')
                             name = name.replace('{CURRENT_DATE}', current_date)
                   
                        if position == 'left' and alignment == 'left':
                             lines.append(f"     {name}")
                        elif position == 'right' and alignment == 'right':
                             # Right-align the date field
                             lines.append(f"     {name}")
              
                   # Add spacing
                   spacing = header_config.get('spacing', 1)
                   for _ in range(spacing):
                        lines.append("")
              
                   return '\n'.join(lines)

          # Original logic for other report types
          # Title and date - preserve exact current design
          title = header_config.get('title', config.get('workflow_metadata', {}).get('report_name', 'Report'))
          subtitle = header_config.get('subtitle', '')
         
          # Handle dynamic dates for Redemption_SWP report
          if header_config.get('date') == 'dynamic':
                   date = datetime.now().strftime('%m/%d/%Y')
          else:
                   date = header_config.get('date', datetime.now().strftime('%m/%d/%Y'))
              
          if header_config.get('trade_date') == 'dynamic':
                   # For dynamic trade date, we'll need to extract from data
                   # For now, use current date - this will be enhanced to extract from data source
                   trade_date = f"Trade Date: {datetime.now().strftime('%m/%d/%Y')}"
          else:
                   trade_date = header_config.get('trade_date', datetime.now().strftime('%m/%d/%Y'))

          # Check if this report should have a single header line
          single_header = header_config.get('single_header', True)# Default to single header

          if header_config.get('date_position') == 'right':
                   # Exact formatting: title left-aligned to 50 chars, date right-aligned to 30 chars
                   line = f"{title:<50} {date:>30}"
                   lines.append(line)
         
                   # Add subtitle if provided, separated by a line
                   if subtitle:
                        lines.append("")# Empty line separator
                        lines.append(subtitle)
         
                   # Only add second line if single_header is false
                   if not single_header and trade_date and trade_date != date:
                        lines.append(trade_date)
          else:
                   # For reports without date_position (like Redemption), use the original logic
                   line = f"{date}"
                   lines.append(line)
         
                   # Add subtitle if provided, separated by a line
                   if subtitle:
                        lines.append("")# Empty line separator
                        lines.append(subtitle)
         
                   # Only add trade_date if single_header is false and trade_date exists
                   if not single_header and trade_date and trade_date != date:
                        lines.append(trade_date)

          # Add spacing based on configuration
          spacing = header_config.get('spacing', 2)# Default to 2 if not specified
          for _ in range(spacing):
                   lines.append("")

          return '\n'.join(lines)

     def _generate_matrix_text_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate matrix section for TXT/CSV output"""
          layout = config['layout']
          matrix_sections = layout.get('matrix_sections', [])

          all_lines = []

          for section in matrix_sections:
                   lines = []
         
                   # Section title
                   title = section.get('title', '')
                   if title:
                        lines.append(title)
                        lines.append("")
         
                   # Headers
                   headers = section.get('headers', [])
                   if headers:
                        header_line = self.formatter.tab_char.join([f"{h:30s}" if i == 0 else f"{h:15s}" for i, h in enumerate(headers)])
                        lines.append(header_line)
         
                   # Rows
                   rows_config = section.get('rows', [])
                   for row_config in rows_config:
                        row_label = row_config.get('label', '')
                        cell_values = row_config.get('values', [])
                        indent_level = row_config.get('indent', 0)
              
                        # Apply indentation
                        if indent_level > 0:
                             row_label = '' * indent_level + row_label
              
                        # Build row
                        row_parts = [f"{row_label:30s}"]
              
                        for value_config in cell_values:
                             if isinstance(value_config, dict):
                                  calc_type = value_config.get('calculation', '')
                                  value = self._calculate_matrix_value(calc_type, data, value_config)
                                  if isinstance(value, (int, float)):
                                        row_parts.append(f"{value:>15}")
                                  else:
                                        row_parts.append(f"{str(value):>15}")
                             else:
                                  row_parts.append(f"{str(value_config):>15}")
              
                        lines.append(self.formatter.tab_char.join(row_parts))
         
                   lines.append("")
                   all_lines.extend(lines)

          return '\n'.join(all_lines)

     def _generate_data_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate data section with exact current design formatting"""
          layout = config.get('layout', {})
          data_config = layout.get('data_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          lines = []
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks' and data_config and 'columns' in data_config:
                   # Group data by fund code
                   fund_groups = {}
                   for record in data:
                        fund_code = record.get('CUSTOM_FUND_CODE', '')
                        if fund_code not in fund_groups:
                             fund_groups[fund_code] = []
                        fund_groups[fund_code].append(record)
              
                   # Column headers with proper formatting
                   columns = data_config['columns']
              
                   # Create header row with proper alignment
                   header_line = ""
                   for col in columns:
                        name = col.get('name', '')
                        width = col.get('width', 0)
                        align = col.get('align', 'left')
                   
                        if width > 0:
                             formatted = self.formatter.align_text(name, width, align)
                        else:
                             formatted = name
                   
                        header_line += formatted + " "
              
                   lines.append(header_line.rstrip())
              
                   # Add separator line
                   separator_line = ""
                   for col in columns:
                        width = col.get('width', 0)
                        if width > 0:
                             separator_line += "-" * width + " "
                        else:
                             separator_line += "-" * len(col.get('name', '')) + " "
              
                   lines.append(separator_line.rstrip())
              
                   # Data rows for each fund
                   for fund_code in sorted(fund_groups.keys()):
                        fund_records = fund_groups[fund_code]
                   
                        # Calculate totals for this fund
                        total_statements = len(fund_records)
                        total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', '0')) for record in fund_records)
                   
                        # Create fund record
                        fund_record = {
                             'CUSTOM_FUND_CODE': fund_code,
                             'Client Statements': total_statements,
                             'Client Amount': total_amount
                        }
                   
                        # Generate row with proper formatting
                        row_line = ""
                        for col in columns:
                             source_field = col.get('source_field', '')
                             calculation = col.get('calculation', '')
                        
                             if calculation == 'count_client_statements':
                                  value = fund_record.get('Client Statements', 0)
                             elif calculation == 'sum_client_amounts':
                                  value = fund_record.get('Client Amount', 0)
                             elif calculation == 'sum_custom_check_amounts':
                                  value = fund_record.get('Client Amount', 0)
                             else:
                                  value = fund_record.get(source_field, '')
                        
                             # Format the value
                             if col.get('format') == 'currency':
                                  try:
                                       value = f"{float(value):.2f}"
                                  except (ValueError, TypeError):
                                       value = "0.00"
                             elif col.get('format') == 'number':
                                  value = str(value)
                             else:
                                  value = str(value)
                        
                             width = col.get('width', 0)
                             align = col.get('align', 'left')
                        
                             if width > 0:
                                  formatted = self.formatter.align_text(value, width, align)
                             else:
                                  formatted = value
                        
                             row_line += formatted + " "
                   
                        lines.append(row_line.rstrip())
              
                   # Note: Totals section is handled separately, not in data section
              
                   return '\n'.join(lines)

          # Handle regular data_section if it exists
          if data_config and 'columns' in data_config:
                   # Column headers - exact current design formatting
                   columns = data_config['columns']
                   header_parts = []
         
                   for col in columns:
                        name = col.get('name', '')
                        width = col.get('width', 0)
                        align = col.get('align', 'left')
              
                        if width > 0:
                             formatted = self.formatter.align_text(name, width, align)
                        else:
                             formatted = name
              
                        header_parts.append(formatted)
         
                   # Use tab character for Excel-compatible formatting (current design)
                   lines.append(self.formatter.tab_char.join(header_parts))
              
                   # Add informational row for Redemption_SWP report
                   info_config = layout.get('informational_row', {})
                   if info_config.get('enabled', False):
                        info_row = self._generate_informational_row(info_config, data)
                        if info_row:
                             lines.append(info_row)

                   # Data rows - exact current design formatting
                   report_type = config.get('workflow_metadata', {}).get('report_type')
              
                   if report_type == 'redemption_swp':
                        # For Redemption_SWP reports, display individual records with enhanced calculations
                        # Filter records to only include those with CUSTOM_CHECK_TYPE1 = "RED" or "SWP"
                        filtered_data = [record for record in data
                                   if record.get('CUSTOM_CHECK_TYPE1') in ['RED', 'SWP']]
                   
                        # Sort by fund code, then by check type (RED first, then SWP)
                        filtered_data.sort(key=lambda x: (x.get('CUSTOM_FUND_CODE', ''),
                                                      x.get('CUSTOM_CHECK_TYPE1', '')))
                   
                        # Generate rows for each individual record
                        for record in filtered_data:
                             # Create enhanced record with individual record data
                             enhanced_record = self._create_individual_record(record)
                        
                             # Generate row for this individual record
                             row = self.formatter.create_excel_row(data_config['columns'], enhanced_record)
                             lines.append(row)
                   else:
                        # For other reports, use original logic
                        for record in data:
                             row = self.formatter.create_excel_row(data_config['columns'], record)
                             lines.append(row)
              
                   # Enhanced Logic for Redemption_SWP Report
                   if report_type == 'redemption_swp':
                        lines = self._apply_redemption_swp_enhanced_logic(lines, data)

          # Handle suppression_sections (used in 12b1 compensation reports)
          suppression_sections = layout.get('suppression_sections', [])
          for section in suppression_sections:
                   if 'columns' in section:
                        # Add section title
                        title = section.get('title', '')
                        if title:
                             lines.append(title)
              
                        # Add subtitle if provided
                        subtitle = section.get('subtitle', '')
                        if subtitle:
                             lines.append("")# Empty line separator
                             lines.append(subtitle)
              
                        # Generate column headers
                        columns = section['columns']
                        header_parts = []
              
                        for col in columns:
                             name = col.get('name', '')
                             width = col.get('width', 0)
                             align = col.get('align', 'left')
                   
                             if width > 0:
                                  formatted = self.formatter.align_text(name, width, align)
                             else:
                                  formatted = name
                   
                             header_parts.append(formatted)
              
                        lines.append(self.formatter.tab_char.join(header_parts))
              
                        # Generate data rows for this section
                        for record in data:
                             row = self.formatter.create_excel_row(section['columns'], record)
                             lines.append(row)
              
                        # Add dynamic summary if configured
                        summary_config = section.get('summary', {})
                        if summary_config:
                             summary_lines = self._generate_txt_dynamic_summary(section, data)
                             lines.extend(summary_lines)
              
                        lines.append("")# Add spacing between sections

          # Add spacing - always 1 blank line after data (current design)
          if lines:
                   lines.append("")

          return '\n'.join(lines)

     def _generate_informational_row(self, info_config: Dict, data: List[Dict]) -> str:
          """Generate informational row for Redemption_SWP report"""
          columns = info_config.get('columns', [])
          if not columns:
                   return ""
         
          row_parts = []
         
          for col in columns:
                   name = col.get('name', '')
                   calculation = col.get('calculation', '')
                   format_type = col.get('format', 'text')
              
                   if calculation == 'format_fund_display':
                        # Format fund code and name with proper padding
                        fund_info = self._format_fund_display(data)
                        row_parts.append(fund_info)
                   elif calculation == 'count_total_checks':
                        # Count total checks (redemption + SWP)
                        total_checks = len(data)
                        row_parts.append(str(total_checks))
                   elif calculation == 'sum_total_amounts':
                        # Sum total amounts (currently blank as per requirements)
                        # JHI wants this populated but currently blank
                        row_parts.append("") # Currently blank as per requirements
                   else:
                        row_parts.append("")
         
          return self.formatter.tab_char.join(row_parts)

     def _format_fund_display(self, data: List[Dict]) -> str:
          """Format fund code and name display for informational row"""
          # For Redemption_SWP reports, we need to show individual fund codes per row
          # not concatenated values. This function should return a single fund code.
          # The actual fund display logic should be handled in the data section generation.
          return ""

     def _apply_redemption_swp_enhanced_logic(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Apply enhanced logic for Redemption_SWP report: move 11 lines to the top of the table"""
          if not lines:
                   return lines
         
          # 1. Move the 11 lines to the top of the table
          enhanced_lines = self._move_11_lines_to_top(lines, data)
         
          return enhanced_lines

     def _group_data_by_fund_and_type(self, data: List[Dict]) -> Dict[str, Dict[str, List[Dict]]]:
          """Group data by CUSTOM_FUND_CODE and CUSTOM_CHECK_TYPE for Redemption_SWP report"""
          grouped_data = {}
         
          for record in data:
                   fund_code = record.get('CUSTOM_FUND_CODE', '')
                   check_type = record.get('CUSTOM_CHECK_TYPE', '')
              
                   # Determine if it's Redemption or SWP
                   if 'REDEMPTION' in check_type.upper():
                        type_key = 'REDEMPTION'
                   elif 'SWP' in check_type.upper():
                        type_key = 'SWP'
                   else:
                        type_key = 'OTHER'
              
                   if fund_code not in grouped_data:
                        grouped_data[fund_code] = {}
              
                   if type_key not in grouped_data[fund_code]:
                        grouped_data[fund_code][type_key] = []
              
                   grouped_data[fund_code][type_key].append(record)
         
          return grouped_data

     def _group_data_by_fund_and_type_enhanced(self, data: List[Dict]) -> Dict[str, Dict[str, List[Dict]]]:
          """Enhanced grouping by CUSTOM_FUND_CODE and CUSTOM_CHECK_TYPE1 for Redemption_SWP report"""
          grouped_data = {}
         
          for record in data:
                   fund_code = record.get('CUSTOM_FUND_CODE', '')
                   check_type1 = record.get('CUSTOM_CHECK_TYPE1', '')
              
                   # Use CUSTOM_CHECK_TYPE1 for grouping (RED or SWP)
                   if check_type1 == 'RED':
                        type_key = 'RED'
                   elif check_type1 == 'SWP':
                        type_key = 'SWP'
                   else:
                        type_key = 'OTHER'
              
                   if fund_code not in grouped_data:
                        grouped_data[fund_code] = {}
              
                   if type_key not in grouped_data[fund_code]:
                        grouped_data[fund_code][type_key] = []
              
                   grouped_data[fund_code][type_key].append(record)
         
          return grouped_data

     def _create_enhanced_fund_record(self, fund_code: str, fund_data: Dict[str, List[Dict]]) -> Dict:
          """Create enhanced record with calculated values for Redemption_SWP report"""
          # Get fund description from any record
          fund_desc = ""
          if fund_data.get('RED'):
                   fund_desc = fund_data['RED'][0].get('CUSTOM_FUND_DESC', '')
          elif fund_data.get('SWP'):
                   fund_desc = fund_data['SWP'][0].get('CUSTOM_FUND_DESC', '')
         
          # Calculate RED values
          red_records = fund_data.get('RED', [])
          red_begin_check = ""
          red_end_check = ""
          red_count = len(red_records)
          red_total_amount = 0.0
         
          if red_records:
                   # Sort by check number to get begin and end
                   red_check_numbers = [r.get('CUSTOM_CHECK_NUMBER', '') for r in red_records if r.get('CUSTOM_CHECK_NUMBER')]
                   if red_check_numbers:
                        red_check_numbers.sort()
                        red_begin_check = red_check_numbers[0]
                        red_end_check = red_check_numbers[-1]
              
                   # Calculate total amount
                   red_total_amount = sum(self.safe_float_convert(r.get('CUSTOM_CHECK_AMOUNT', 0)) for r in red_records)
         
          # Calculate SWP values
          swp_records = fund_data.get('SWP', [])
          swp_begin_check = ""
          swp_end_check = ""
          swp_count = len(swp_records)
          swp_total_amount = 0.0
         
          if swp_records:
                   # Sort by check number to get begin and end
                   swp_check_numbers = [r.get('CUSTOM_CHECK_NUMBER', '') for r in swp_records if r.get('CUSTOM_CHECK_NUMBER')]
                   if swp_check_numbers:
                        swp_check_numbers.sort()
                        swp_begin_check = swp_check_numbers[0]
                        swp_end_check = swp_check_numbers[-1]
              
                   # Calculate total amount
                   swp_total_amount = sum(self.safe_float_convert(r.get('CUSTOM_CHECK_AMOUNT', 0)) for r in swp_records)
         
          # Create enhanced record
          enhanced_record = {
                   'CUSTOM_FUND_CODE': fund_code,
                   'CUSTOM_FUND_DESC': fund_desc,
                   'CUSTOM_CHECK_TYPE': 'COMBINED',
                   'CUSTOM_CHECK_AMOUNT': red_total_amount + swp_total_amount,
                   # RED values
                   'RED_BEGIN_CHECK_NUMBER': red_begin_check,
                   'RED_END_CHECK_NUMBER': red_end_check,
                   'RED_CHECK_COUNT': red_count,
                   'TOTAL_RED_DOLLAR_VALUE': red_total_amount,
                   # SWP values
                   'SWP_BEGIN_CHECK_NUMBER': swp_begin_check,
                   'SWP_END_CHECK_NUMBER': swp_end_check,
                   'SWP_CHECK_COUNT': swp_count,
                   'TOTAL_SWP_DOLLAR_VALUE': swp_total_amount
          }
         
          return enhanced_record

     def _create_individual_record(self, record: Dict) -> Dict:
          """Create enhanced record for individual Redemption_SWP record"""
          fund_code = record.get('CUSTOM_FUND_CODE', '')
          fund_desc = record.get('CUSTOM_FUND_DESC', '')
          check_type1 = record.get('CUSTOM_CHECK_TYPE1', '')
          check_number = record.get('CUSTOM_CHECK_NUMBER', '')
          check_amount = self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0))
         
          # Initialize all values
          red_begin_check = ""
          red_end_check = ""
          red_count = 0
          red_total_amount = 0.0
          swp_begin_check = ""
          swp_end_check = ""
          swp_count = 0
          swp_total_amount = 0.0
         
          # Set values based on the individual record's check type
          if check_type1 == 'RED':
                   red_begin_check = check_number
                   red_end_check = check_number
                   red_count = 1
                   red_total_amount = check_amount
          elif check_type1 == 'SWP':
                   swp_begin_check = check_number
                   swp_end_check = check_number
                   swp_count = 1
                   swp_total_amount = check_amount
         
          # Create enhanced record
          enhanced_record = {
                   'CUSTOM_FUND_CODE': fund_code,
                   'CUSTOM_FUND_DESC': fund_desc,
                   'CUSTOM_CHECK_TYPE': check_type1,
                   'CUSTOM_CHECK_AMOUNT': check_amount,
                   # RED values
                   'RED_BEGIN_CHECK_NUMBER': red_begin_check,
                   'RED_END_CHECK_NUMBER': red_end_check,
                   'RED_CHECK_COUNT': red_count,
                   'TOTAL_RED_DOLLAR_VALUE': red_total_amount,
                   # SWP values
                   'SWP_BEGIN_CHECK_NUMBER': swp_begin_check,
                   'SWP_END_CHECK_NUMBER': swp_end_check,
                   'SWP_CHECK_COUNT': swp_count,
                   'TOTAL_SWP_DOLLAR_VALUE': swp_total_amount
          }
         
          return enhanced_record

     def _calculate_group_totals(self, grouped_data: Dict[str, Dict[str, List[Dict]]]) -> Dict[str, Dict[str, Dict]]:
          """Calculate totals for each group in Redemption_SWP report"""
          group_totals = {}
         
          for fund_code, type_data in grouped_data.items():
                   group_totals[fund_code] = {}
              
                   for check_type, records in type_data.items():
                        total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in records)
                        count = len(records)
                   
                        group_totals[fund_code][check_type] = {
                             'count': count,
                             'total_amount': total_amount,
                             'records': records
                        }
         
          return group_totals

     def _get_dynamic_special_handling_counts(self, data: List[Dict]) -> Dict[str, int]:
          """Dynamically extract and count DOC_SPECIAL_HANDLING_CODE values from XML data"""
          special_handling_counts = {}
         
          for record in data:
                   handling_code = record.get('DOC_SPECIAL_HANDLING_CODE', '')
                   if handling_code:
                        # Count occurrences of each handling code
                        special_handling_counts[handling_code] = special_handling_counts.get(handling_code, 0) + 1
         
          return special_handling_counts

     def _map_special_handling_to_shcode(self, handling_code: str) -> Dict:
          """Map DOC_SPECIAL_HANDLING_CODE to corresponding SHCode and description"""
          if handling_code in self.special_handling_mapping:
                   return self.special_handling_mapping[handling_code]
          else:
                   # Return unknown mapping for codes not in our table
                   return {
                        'code': f'UNKNOWN-{handling_code}',
                        'description': f'Unknown Special Handling Code: {handling_code}',
                        'shcode': handling_code
                   }

     def _get_special_handling_summary(self, data: List[Dict]) -> Dict:
          """Get detailed summary of special handling codes found in the data"""
          special_handling_counts = self._get_dynamic_special_handling_counts(data)
          summary = {
                   'total_records': len(data),
                   'special_handling_codes_found': list(special_handling_counts.keys()),
                   'code_details': {},
                   'unknown_codes': []
          }
         
          for handling_code, count in special_handling_counts.items():
                   mapping = self._map_special_handling_to_shcode(handling_code)
                   summary['code_details'][handling_code] = {
                        'count': count,
                        'mapping': mapping
                   }
              
                   if 'UNKNOWN' in mapping['code']:
                        summary['unknown_codes'].append({
                             'code': handling_code,
                             'count': count
                        })
         
          return summary

     def add_special_handling_mapping(self, shcode: str, code: str, description: str) -> None:
          """Dynamically add a new special handling code mapping"""
          self.special_handling_mapping[shcode] = {
                   'code': code,
                   'description': description,
                   'shcode': shcode
          }
          print(f"Added new special handling mapping: {shcode} -> {code} ({description})")

     def get_special_handling_mappings(self) -> Dict:
          """Get all current special handling mappings"""
          return self.special_handling_mapping.copy()

     def _calculate_dynamic_footer_data(self, data: List[Dict]) -> List:
          """Calculate dynamic footer data based on CUSTOM_PULL_TYPE and DOC_SPECIAL_HANDLING_CODE for Redemption_SWP report"""
          # Count occurrences of each CUSTOM_PULL_TYPE
          pull_type_counts = {}
          for record in data:
                   pull_type = record.get('CUSTOM_PULL_TYPE', '')
                   if pull_type:
                        pull_type_counts[pull_type] = pull_type_counts.get(pull_type, 0) + 1
         
          # Get dynamic special handling counts
          special_handling_counts = self._get_dynamic_special_handling_counts(data)
         
          # Define the categories in order
          categories = [
                   "RETURNTOJANUS",
                   "JANUS",
                   "STATESTREETBANK",
                   "DONOTMAIL",
                   "VOID",
                   "PULLREPORTCHECKS",
                   "SUB-TOTAL",
                   "SPECIALHANDLING",
                   "OVERNIGHTW-SIG",
                   "OVERNIGHTWO-SIG",
                   "OVERNIGHTSAT-W-SIG",
                   "OVERNIGHTSAT-WO-SIG",
                   "SUB-TOTAL",
                   "NORMAL",
                   "TOTAL",
                   "SUPPRESSEDTEST"
          ]
         
          # Initialize category data with zeros
          category_data = [0] * len(categories)
         
          # Map individual pull types to their counts (handle actual XML values)
          category_data[1] = pull_type_counts.get('JANUS', 0) # JANUS
          # category_data[2] = pull_type_counts.get('STATW STREET BANK', 0) # STATESTREETBANK (actual XML value)
          category_data[2] = pull_type_counts.get('STATW STREET BANK', pull_type_counts.get('STANDARD', 0))

          category_data[3] = pull_type_counts.get('DO NOT MAIL', 0) # DONOTMAIL (actual XML value)
          category_data[4] = pull_type_counts.get('VOID', 0) # VOID
          category_data[5] = pull_type_counts.get('PULLREPORTCHECKS', 0) # PULLREPORTCHECKS
          category_data[6] = pull_type_counts.get('SUB-TOTAL', 0) # SUB-TOTAL
         
          # Map special handling types dynamically based on DOC_SPECIAL_HANDLING_CODE
          # Initialize special handling counts to 0
          overnight_w_sig_count = 0
          overnight_wo_sig_count = 0
          overnight_sat_w_sig_count = 0
          overnight_sat_wo_sig_count = 0
         
          # Process each special handling code found in the data
          for handling_code, count in special_handling_counts.items():
                   mapping = self._map_special_handling_to_shcode(handling_code)
              
                   # Map to the appropriate category based on the code
                   if mapping['code'] == 'OVERNIGHTW-SIG':
                        overnight_w_sig_count = count
                   elif mapping['code'] == 'OVERNIGHTWO-SIG':
                        overnight_wo_sig_count = count
                   elif mapping['code'] == 'OVERNIGHTSAT-W-SIG':
                        overnight_sat_w_sig_count = count
                   elif mapping['code'] == 'OVERNIGHTSAT-WO-SIG':
                        overnight_sat_wo_sig_count = count
                   # For unknown codes, we could add them to a separate category or log them
                   else:
                        print(f"Warning: Unknown special handling code '{handling_code}' with count {count}")
         
          # Assign the dynamically calculated counts
          category_data[8] = overnight_w_sig_count # OVERNIGHTW-SIG
          category_data[9] = overnight_wo_sig_count # OVERNIGHTWO-SIG
          category_data[10] = overnight_sat_w_sig_count # OVERNIGHTSAT-W-SIG
          category_data[11] = overnight_sat_wo_sig_count # OVERNIGHTSAT-WO-SIG
          category_data[12] = pull_type_counts.get('SUB-TOTAL', 0) # SUB-TOTAL (second occurrence)
          category_data[13] = pull_type_counts.get('NORMAL', 0) # NORMAL
          category_data[15] = pull_type_counts.get('SUPPRESSEDTEST', 0) # SUPPRESSEDTEST
         
          # Calculate RETURNTOJANUS (sum of JANUS, STATESTREETBANK, DONOTMAIL, VOID, PULLREPORTCHECKS, SUB-TOTAL)
          return_to_janus = (category_data[1] + category_data[2] + category_data[3] +
                         category_data[4] + category_data[5] + category_data[6])
          category_data[0] = return_to_janus
         
          # Calculate SPECIALHANDLING dynamically (sum of all special handling codes found)
          special_handling = (category_data[8] + category_data[9] + category_data[10] + category_data[11])
          category_data[7] = special_handling
         
          # Calculate TOTAL (sum of all individual counts)
          total = sum(category_data[1:6]) + sum(category_data[8:12]) + category_data[13] + category_data[15]
          category_data[14] = total
         
          return category_data

     def _move_11_lines_to_top(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Move the 11 lines to the top of the table in Redemption_SWP report"""
          if not lines:
                   return lines
         
          # Create the 11 lines that need to be moved to the top
          top_lines = self._create_11_top_lines(data)
         
          # Find where the actual data table starts (after headers)
          data_start_index = self._find_data_table_start(lines)
         
          if data_start_index == -1:
                   # If we can't find the data table start, just prepend the 11 lines
                   return top_lines + lines
         
          # Split the lines into header section and data section
          header_section = lines[:data_start_index]
          data_section = lines[data_start_index:]
         
          # Combine: header section + 11 top lines + data section
          return header_section + top_lines + data_section

     def _create_11_top_lines(self, data: List[Dict]) -> List[str]:
          """Create the 11 lines that need to be moved to the top"""
          lines = []
         
          # Line 1: Date
          current_date = datetime.now().strftime('%m/%d/%Y')
          lines.append(current_date)
         
          # Line 2: Trade Date
          trade_date = f"Trade Date: {datetime.now().strftime('%m/%d/%Y')}"
          lines.append(trade_date)
         
          # Line 3: Fund Code and Fund Name information
          lines.append("Fund Code and Fund Name from RRSU0224 & RRSU0001.")
         
          # Line 4: Fund Code length information
          lines.append("Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits.")
         
          # Line 5: Blank field instruction
          lines.append("always blank. JHI would like BR to populate this")
         
          # Line 6: Another blank field instruction
          lines.append("always blank. JHI would like BR to populate this")
         
          # Line 7: Specific order information
          lines.append("Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order).")
         
          # Line 8: Total checks information
          total_checks = len(data)
          lines.append(f"This is a total of both Redemption and SWP checks in the Fund. Total: {total_checks}")
         
          # Line 9: Total dollar value information
          total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
          lines.append(f"Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field. Calculated Total: {total_amount:.2f}")
         
          # Line 10: Empty line for spacing
          lines.append("")
         
          # Line 11: Table headers (this will be the consolidated header line)
          consolidated_header = self._create_consolidated_static_line(data)
          lines.append(consolidated_header)
         
          return lines

     def _find_data_table_start(self, lines: List[str]) -> int:
          """Find where the actual data table starts"""
          for i, line in enumerate(lines):
                   # Look for table headers or data rows
                   if line.strip() and ('GROWTH' in line.upper() or 'FUND' in line.upper() or
                                  'CHECKS' in line.upper() or 'TOTAL' in line.upper()):
                        return i
          return -1

     def _insert_static_informational_lines(self, data: List[Dict]) -> List[str]:
          """Insert static informational lines for Redemption_SWP report"""
          static_lines = []
         
          # Above "GROWTH FUND" - Fund Code and Fund Name information
          fund_info_line = "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order)."
          static_lines.append(fund_info_line)
         
          # Above "CHECKS PRINTED TODAY" - Total checks information
          total_checks = len(data)
          checks_info_line = f"This is a total of both Redemption and SWP checks in the Fund. Total: {total_checks}"
          static_lines.append(checks_info_line)
         
          # Above "DOLLAR VALUE TODAY" - Total dollar value information
          total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
          dollar_info_line = f"Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field. Calculated Total: {total_amount:.2f}"
          static_lines.append(dollar_info_line)
         
          return static_lines

     def _insert_static_informational_lines_aligned(self, data: List[Dict], lines: List[str], header_start_index: int) -> List[str]:
          """Insert static informational lines above specific headers with proper column alignment"""
          static_lines = []
         
          # Find the header line to determine column structure
          header_line = None
          for i in range(header_start_index, len(lines)):
                   if lines[i].strip() and ('GROWTHFUND' in lines[i].upper() or 'GROWTH FUND' in lines[i].upper()):
                        header_line = lines[i]
                        break
         
          if header_line:
                   # Determine column structure from header line
                   if '\t' in header_line:
                        # Tab-separated format
                        columns = header_line.split('\t')
                        num_columns = len(columns)
                   else:
                        # Single column or space-separated
                        num_columns = 1
          else:
                   # Default to single column if no header found
                   num_columns = 1
         
          # Create consolidated static line with text distributed across columns A, B, C
          if num_columns >= 3:
                   # Multi-column format - distribute text across columns A, B, C
                   column_a_text = "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order)."
                   column_b_text = "This is a total of both Redemption and SWP checks in the Fund."
                   column_c_text = "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
              
                   # Create consolidated line with tab separation
                   consolidated_line = f"{column_a_text}\t{column_b_text}\t{column_c_text}"
              
                   # Add empty columns if there are more than 3 columns
                   for _ in range(num_columns - 3):
                        consolidated_line += '\t'
              
                   static_lines.append(consolidated_line)
          else:
                   # Fallback to single column format
                   static_texts = [
                        "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order).",
                        "This is a total of both Redemption and SWP checks in the Fund.",
                        "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
                   ]
              
                   for text in static_texts:
                        static_lines.append(text)
         
          return static_lines

     def _apply_row_reordering(self, lines: List[str]) -> List[str]:
          """Apply row reordering: exchange row 4 and row 5"""
          if len(lines) < 5:
                   return lines
         
          # Create a copy of the lines
          reordered_lines = lines.copy()
         
          # Exchange row 4 (index 3) and row 5 (index 4)
          reordered_lines[3], reordered_lines[4] = reordered_lines[4], reordered_lines[3]
         
          return reordered_lines

     def _position_consolidated_text_after_row_4(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Position consolidated static text immediately after row 4"""
          if len(lines) < 4:
                   return lines
         
          # Create consolidated static text line
          consolidated_line = self._create_consolidated_static_line(data)
         
          # Find the position after row 4 (index 4)
          # Insert the consolidated line at position 5 (index 4)
          positioned_lines = lines[:4] + [consolidated_line] + lines[4:]
         
          return positioned_lines

     def _create_consolidated_static_line(self, data: List[Dict]) -> str:
          """Create consolidated static line with text distributed across columns A, B, C"""
          column_a_text = "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order)."
          column_b_text = "This is a total of both Redemption and SWP checks in the Fund."
          column_c_text = "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
         
          # Create consolidated line with tab separation for TXT format
          return f"{column_a_text}\t{column_b_text}\t{column_c_text}"

     def _apply_redemption_swp_enhanced_logic_xlsx(self, ws, data: List[Dict], start_row: int) -> int:
          """Apply enhanced logic for Redemption_SWP report in XLSX format"""
          # For Redemption_SWP reports, we need to ensure the static text is placed in row 3 (A3, B3, C3)
          # These cells must remain fixed and static
         
          # Cell A3: Fund Code and Fund Name information
          cell_a3 = ws.cell(row=3, column=1, value="Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order).")
          cell_a3.font = Font(bold=True, size=10)
          cell_a3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
          # Cell B3: Total checks information
          cell_b3 = ws.cell(row=3, column=2, value="This is a total of both Redemption and SWP checks in the Fund.")
          cell_b3.font = Font(bold=True, size=10)
          cell_b3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
          # Cell C3: Total dollar value information
          cell_c3 = ws.cell(row=3, column=3, value="Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field.")
          cell_c3.font = Font(bold=True, size=10)
          cell_c3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
          # Return the current row (no additional rows added)
          return start_row

     def _apply_redemption_swp_enhanced_logic_csv(self, csv_lines: List[List[str]], data: List[Dict]) -> List[List[str]]:
          """Apply enhanced logic for Redemption_SWP report in CSV format - move 11 lines to top"""
          if not csv_lines:
                   return csv_lines
         
          # Convert CSV lines to string format for processing
          lines = []
          for row in csv_lines:
                   if row:
                        lines.append(','.join(str(cell) for cell in row))
                   else:
                        lines.append("")
         
          # Apply the enhanced logic to move 11 lines to top
          enhanced_lines = self._move_11_lines_to_top_csv(lines, data)
         
          # Convert back to CSV format
          enhanced_csv_lines = []
          for line in enhanced_lines:
                   if line.strip():
                        enhanced_csv_lines.append(line.split(','))
                   else:
                        enhanced_csv_lines.append([])
         
          return enhanced_csv_lines

     def _move_11_lines_to_top_csv(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Move the 11 lines to the top of the table in CSV format"""
          if not lines:
                   return lines
         
          # Create the 11 lines that need to be moved to the top (CSV format)
          top_lines = self._create_11_top_lines_csv(data)
         
          # Find where the actual data table starts (after headers)
          data_start_index = self._find_data_table_start(lines)
         
          if data_start_index == -1:
                   # If we can't find the data table start, just prepend the 11 lines
                   return top_lines + lines
         
          # Split the lines into header section and data section
          header_section = lines[:data_start_index]
          data_section = lines[data_start_index:]
         
          # Combine: header section + 11 top lines + data section
          return header_section + top_lines + data_section

     def _create_11_top_lines_csv(self, data: List[Dict]) -> List[str]:
          """Create the 11 lines that need to be moved to the top (CSV format)"""
          lines = []
         
          # Line 1: Date
          current_date = datetime.now().strftime('%m/%d/%Y')
          lines.append(current_date)
         
          # Line 2: Trade Date
          trade_date = f"Trade Date: {datetime.now().strftime('%m/%d/%Y')}"
          lines.append(trade_date)
         
          # Line 3: Fund Code and Fund Name information
          lines.append("Fund Code and Fund Name from RRSU0224 & RRSU0001.")
         
          # Line 4: Fund Code length information
          lines.append("Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits.")
         
          # Line 5: Blank field instruction
          lines.append("always blank. JHI would like BR to populate this")
         
          # Line 6: Another blank field instruction
          lines.append("always blank. JHI would like BR to populate this")
         
          # Line 7: Specific order information
          lines.append("Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order).")
         
          # Line 8: Total checks information
          total_checks = len(data)
          lines.append(f"This is a total of both Redemption and SWP checks in the Fund. Total: {total_checks}")
         
          # Line 9: Total dollar value information
          total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
          lines.append(f"Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field. Calculated Total: {total_amount:.2f}")
         
          # Line 10: Empty line for spacing
          lines.append("")
         
          # Line 11: Table headers (this will be the consolidated header line for CSV)
          consolidated_header = self._create_consolidated_static_line_csv(data)
          lines.append(consolidated_header)
         
          return lines

     def _apply_redemption_swp_enhanced_logic_csv_strings(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Apply enhanced logic for Redemption_SWP report in CSV string format"""
          if not lines:
                   return lines
         
          # 1. Detect three consecutive blank lines and insert static informational lines above headers
          enhanced_lines = []
          i = 0
          while i < len(lines):
                   current_line = lines[i]
              
                   # Check if current line is blank/empty
                   if not current_line.strip():
                        # Count consecutive blank lines starting from current position
                        blank_count = 0
                        j = i
                        while j < len(lines) and not lines[j].strip():
                             blank_count += 1
                             j += 1
                   
                        # If we found 3 or more consecutive blank lines, insert static text above headers
                        if blank_count >= 3:
                             # Insert static informational lines above the relevant headers with proper alignment
                             enhanced_lines.extend(self._insert_static_informational_lines_csv_aligned(data, lines, j))
                             # Skip the blank lines we just processed
                             i = j
                             continue
              
                   enhanced_lines.append(current_line)
                   i += 1
         
          # 2. Apply row reordering (exchange row 4 and row 5)
          enhanced_lines = self._apply_row_reordering(enhanced_lines)
         
          # 3. Position consolidated static text immediately after row 4
          enhanced_lines = self._position_consolidated_text_after_row_4_csv(enhanced_lines, data)
         
          return enhanced_lines

     def _position_consolidated_text_after_row_4_csv(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Position consolidated static text immediately after row 4 for CSV format"""
          if len(lines) < 4:
                   return lines
         
          # Create consolidated static text line for CSV
          consolidated_line = self._create_consolidated_static_line_csv(data)
         
          # Find the position after row 4 (index 4)
          # Insert the consolidated line at position 5 (index 4)
          positioned_lines = lines[:4] + [consolidated_line] + lines[4:]
         
          return positioned_lines

     def _create_consolidated_static_line_csv(self, data: List[Dict]) -> str:
          """Create consolidated static line with text distributed across columns A, B, C for CSV format"""
          column_a_text = "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order)."
          column_b_text = "This is a total of both Redemption and SWP checks in the Fund."
          column_c_text = "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
         
          # Create consolidated line with comma separation for CSV format
          return f"{column_a_text},{column_b_text},{column_c_text}"

     def _insert_static_informational_lines_csv_aligned(self, data: List[Dict], lines: List[str], header_start_index: int) -> List[str]:
          """Insert static informational lines above specific headers with proper CSV column alignment"""
          static_lines = []
         
          # Find the header line to determine column structure
          header_line = None
          for i in range(header_start_index, len(lines)):
                   if lines[i].strip() and ('GROWTHFUND' in lines[i].upper() or 'GROWTH FUND' in lines[i].upper()):
                        header_line = lines[i]
                        break
         
          if header_line:
                   # Determine column structure from header line
                   if ',' in header_line:
                        # Comma-separated format
                        columns = header_line.split(',')
                        num_columns = len(columns)
                   else:
                        # Single column
                        num_columns = 1
          else:
                   # Default to single column if no header found
                   num_columns = 1
         
          # Create consolidated static line with text distributed across columns A, B, C
          if num_columns >= 3:
                   # Multi-column format - distribute text across columns A, B, C
                   column_a_text = "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order)."
                   column_b_text = "This is a total of both Redemption and SWP checks in the Fund."
                   column_c_text = "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
              
                   # Create consolidated line with comma separation
                   consolidated_line = f"{column_a_text},{column_b_text},{column_c_text}"
              
                   # Add empty columns if there are more than 3 columns
                   for _ in range(num_columns - 3):
                        consolidated_line += ','
              
                   static_lines.append(consolidated_line)
          else:
                   # Fallback to single column format
                   static_texts = [
                        "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order).",
                        "This is a total of both Redemption and SWP checks in the Fund.",
                        "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
                   ]
              
                   for text in static_texts:
                        static_lines.append(text)
         
          return static_lines

     def _generate_simplified_report_content(self, config: Dict, data: List[Dict]) -> str:
          """Generate simplified report content without complex formatting"""
          content_parts = []

          # Generate basic header
          report_name = config.get('workflow_metadata', {}).get('report_name', 'Report')
          content_parts.append(report_name)
          content_parts.append("")

          # Generate data section only
          data_section = config['layout'].get('data_section', {})
          if 'columns' in data_section:
                   columns = data_section['columns']
         
                   # Simple column headers
                   headers = [col.get('name', '') for col in columns]
                   content_parts.append('\t'.join(headers))
         
                   # Data rows
                   for record in data:
                        row_values = []
                        for col in columns:
                             source_field = col.get('source_field', '')
                             if source_field == 'static':
                                  value = col.get('value', '')
                             elif source_field == 'calculated':
                                  value = self._calculate_simple_value(col, record)
                             else:
                                  value = record.get(source_field, '')
                             row_values.append(str(value))
                        content_parts.append('\t'.join(row_values))

          return '\n'.join(content_parts)

     def _calculate_simple_value(self, col: Dict, data: Dict) -> Any:
          """Calculate simple value for simplified reports"""
          calculation = col.get('calculation', '')

          if calculation == 'count_records':
                   return 1
          elif calculation == 'sum_amounts':
                   # For Redemption_SWP reports, use CUSTOM_CHECK_AMOUNT
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          else:
                   return 0

     def _generate_xlsx_data_section(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX data section with proper cell separation - dynamically adapts to JSON structure"""
          layout = config['layout']
          current_row = start_row
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          # Check for data_section first
          data_config = layout.get('data_section', {})
          if data_config and 'columns' in data_config:
                   # Special handling for dividend reports
                   if report_type == 'dividend_checks':
                        # Group data by fund code
                        fund_groups = {}
                        for record in data:
                             fund_code = record.get('CUSTOM_FUND_CODE', '')
                             if fund_code not in fund_groups:
                                  fund_groups[fund_code] = []
                             fund_groups[fund_code].append(record)
                   
                        # Generate column headers
                        columns = data_config.get('columns', [])
                        for col_idx, col in enumerate(columns, 1):
                             name = col.get('name', '')
                             cell = ws.cell(row=current_row, column=col_idx, value=name)
                             cell.font = Font(bold=True, size=10)
                             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1
                   
                        # Generate data rows for each fund
                        for fund_code in sorted(fund_groups.keys()):
                             fund_records = fund_groups[fund_code]
                        
                             # Calculate totals for this fund
                             total_statements = len(fund_records)
                             total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', '0')) for record in fund_records)
                        
                             # Create fund record
                             fund_record = {
                                  'CUSTOM_FUND_CODE': fund_code,
                                  'Client Statements': total_statements,
                                  'Client Amount': total_amount
                             }
                        
                             # Generate row
                             for col_idx, col in enumerate(columns, 1):
                                  source_field = col.get('source_field', '')
                                  calculation = col.get('calculation', '')
                             
                                  if calculation == 'count_client_statements':
                                        value = fund_record.get('Client Statements', 0)
                                  elif calculation == 'sum_client_amounts':
                                        value = fund_record.get('Client Amount', 0)
                                  elif calculation == 'sum_custom_check_amounts':
                                        value = fund_record.get('Client Amount', 0)
                                  else:
                                        value = fund_record.get(source_field, '')
                             
                                  # Format the value
                                  if col.get('format') == 'currency':
                                        try:
                                             value = f"{float(value):.2f}"
                                        except (ValueError, TypeError):
                                             value = "0.00"
                                  elif col.get('format') == 'number':
                                        value = str(value)
                                  else:
                                        value = str(value)
                             
                                  cell = ws.cell(row=current_row, column=col_idx, value=value)
                                  cell.font = Font(size=10)
                             
                                  # Set alignment based on column configuration
                                  align = col.get('align', 'left')
                                  if align == 'right':
                                        cell.alignment = Alignment(horizontal='right', vertical='top')
                                  else:
                                        cell.alignment = Alignment(horizontal='left', vertical='top')
                        
                             current_row += 1
                   
                        # Add note if specified
                        table_style = data_config.get('table_style', {})
                        note = table_style.get('note', '')
                        if note:
                             current_row += 1
                             cell = ws.cell(row=current_row, column=1, value=note)
                             cell.font = Font(size=10, italic=True)
                             cell.alignment = Alignment(horizontal='right', vertical='top')
                             current_row += 1
                   
                        return current_row
              
                   # Original logic for other report types
                   # Generate column headers
                   columns = data_config.get('columns', [])
                   for col_idx, col in enumerate(columns, 1):
                        name = col.get('name', '')
                        cell = ws.cell(row=current_row, column=col_idx, value=name)
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                   current_row += 1
              
                   # Add informational row for Redemption_SWP report
                   info_config = layout.get('informational_row', {})
                   if info_config.get('enabled', False):
                        current_row = self._generate_xlsx_informational_row(ws, info_config, data, current_row)
              
                   # Generate data rows
                   if report_type == 'redemption_swp':
                        # For Redemption_SWP reports, display individual records with enhanced calculations
                        # Filter records to only include those with CUSTOM_CHECK_TYPE1 = "RED" or "SWP"
                        filtered_data = [record for record in data
                                   if record.get('CUSTOM_CHECK_TYPE1') in ['RED', 'SWP']]
                   
                        # Sort by fund code, then by check type (RED first, then SWP)
                        filtered_data.sort(key=lambda x: (x.get('CUSTOM_FUND_CODE', ''),
                                                      x.get('CUSTOM_CHECK_TYPE1', '')))
                   
                        # Generate rows for each individual record
                        for record in filtered_data:
                             # Create enhanced record with individual record data
                             enhanced_record = self._create_individual_record(record)
                        
                             # Generate row for this individual record
                             for col_idx, col in enumerate(columns, 1):
                                  value = self._get_column_value_for_xlsx(col, enhanced_record)
                                  formatted_value = self._format_column_value_for_xlsx(col, value)
                                  cell = ws.cell(row=current_row, column=col_idx, value=formatted_value)
                                  cell.font = Font(size=10)
                                  cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
                   else:
                        # For other reports, use original logic
                        for record in data:
                             for col_idx, col in enumerate(columns, 1):
                                  value = self._get_column_value_for_xlsx(col, record)
                                  formatted_value = self._format_column_value_for_xlsx(col, value)
                                  cell = ws.cell(row=current_row, column=col_idx, value=formatted_value)
                                  cell.font = Font(size=10)
                                  cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
              
                   # Enhanced Logic for Redemption_SWP Report - XLSX
                   # Note: Static text placement is handled earlier in the generation process

          # Check for suppression_sections (used in 12b1 compensation reports)
          suppression_sections = layout.get('suppression_sections', [])
          for section in suppression_sections:
                   if 'columns' in section:
                        # Add section title
                        title = section.get('title', '')
                        if title:
                             cell = ws.cell(row=current_row, column=1, value=title)
                             cell.font = Font(bold=True, size=10)
                             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
              
                        # Add subtitle if provided
                        subtitle = section.get('subtitle', '')
                        if subtitle:
                             cell = ws.cell(row=current_row, column=1, value=subtitle)
                             cell.font = Font(bold=True, size=9)
                             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
              
                        # Apply filtering based on CUSTOM_CHECK_NUMBER for 12b1_compensation reports
                        filtered_data = self._filter_12b1_data_by_check_number(data, section)
                   
                        # Generate columns for this section
                        current_row = self._generate_xlsx_columns_section(ws, section, filtered_data, current_row)
              
                        # Add dynamic summary if configured
                        summary_config = section.get('summary', {})
                        if summary_config:
                             current_row = self._generate_xlsx_dynamic_summary(ws, section, filtered_data, current_row)
              
                        current_row += 1# Add spacing between sections

          # NEW: Check for matrix_sections (for complex layouts like the image)
          matrix_sections = layout.get('matrix_sections', [])
          for section in matrix_sections:
                   current_row = self._generate_xlsx_matrix_section(ws, section, data, current_row)
                   current_row += 1# Add spacing between sections

          return current_row

     def _generate_xlsx_columns_section(self, ws, section_config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX columns section with proper cell separation"""
          current_row = start_row
          columns = section_config.get('columns', [])

          if not columns:
                   return current_row

          # Generate column headers
          for col_idx, col in enumerate(columns, 1):
                   name = col.get('name', '')
                   cell = ws.cell(row=current_row, column=col_idx, value=name)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

          current_row += 1
         
          # Add informational row for Redemption_SWP report
          # Check if this is a data_section and if informational row is enabled
          layout = section_config.get('layout', {}) if hasattr(section_config, 'get') else {}
          if not layout: # If section_config doesn't have layout, try to get it from parent
                   # This is a workaround - we need to pass the full config to access layout
                   pass
         
          # For now, we'll add the informational row logic in the main data section method
          # This will be handled in _generate_xlsx_data_section
         
          # Generate data rows with proper cell separation
          for record in data:
                   for col_idx, col in enumerate(columns, 1):
                        value = self._get_column_value_for_xlsx(col, record)
                        formatted_value = self._format_column_value_for_xlsx(col, value)
              
                        cell = ws.cell(row=current_row, column=col_idx, value=formatted_value)
                        cell.font = Font(size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   current_row += 1

          return current_row

     def _generate_xlsx_informational_row(self, ws, info_config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX informational row for Redemption_SWP report"""
          current_row = start_row
          columns = info_config.get('columns', [])
         
          if not columns:
                   return current_row
         
          for col_idx, col in enumerate(columns, 1):
                   name = col.get('name', '')
                   calculation = col.get('calculation', '')
                   format_type = col.get('format', 'text')
              
                   if calculation == 'format_fund_display':
                        # Format fund code and name with proper padding
                        fund_info = self._format_fund_display(data)
                        cell = ws.cell(row=current_row, column=col_idx, value=fund_info)
                   elif calculation == 'count_total_checks':
                        # Count total checks (redemption + SWP)
                        total_checks = len(data)
                        cell = ws.cell(row=current_row, column=col_idx, value=total_checks)
                   elif calculation == 'sum_total_amounts':
                        # Sum total amounts (currently blank as per requirements)
                        # JHI wants this populated but currently blank
                        cell = ws.cell(row=current_row, column=col_idx, value="") # Currently blank
                   else:
                        cell = ws.cell(row=current_row, column=col_idx, value="")
              
                   # Format the cell
                   cell.font = Font(size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
          return current_row + 1

     def _generate_xlsx_summary_section(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX summary section with proper cell separation"""
          layout = config.get('layout', {})
          summary_config = layout.get('summary_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          if not summary_config:
                   return start_row

          current_row = start_row
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks':
                   # Title with proper alignment
                   title = summary_config.get('title', 'GENERAL AUDIT COUNTS')
                   if title:
                        cell = ws.cell(row=current_row, column=1, value=f"{title}")
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1
              
                   # Subtitle with proper alignment
                   subtitle = summary_config.get('subtitle', '')
                   if subtitle:
                        # Use current date for document date
                        current_date = datetime.now().strftime('%m/%d/%Y')
                        cell = ws.cell(row=current_row, column=1, value=f"{subtitle}     {current_date}")
                        cell.font = Font(size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1
              
                   # Data block with proper alignment
                   data_block = summary_config.get('data_block', {})
                   if data_block:
                        current_row += 1
                        label = data_block.get('label', 'R06686 Data')
                        # Use current date for record date
                        record_date = datetime.now().strftime('%m/%d/%Y')
                        cell = ws.cell(row=current_row, column=1, value=f"{label}                    Record Date:     {record_date}")
                        cell.font = Font(size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1
                   
                        # Left-aligned fields with right-aligned dates
                        fields = data_block.get('fields', [])
                        right_aligned = data_block.get('right_aligned', [])
                   
                        for i, field in enumerate(fields):
                             name = field.get('name', '')
                             calculation = field.get('calculation', '')
                             
                             # Calculate value based on calculation type
                             if calculation == 'count_input_records':
                                  value = len(data)
                             elif calculation == 'count_total_checks':
                                  value = len(data) # Total checks = number of records
                             elif calculation == 'sum_total_check_amounts':
                                  value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', '0')) for record in data)
                             else:
                                  value = field.get('value', '')
                             
                             # Format the value
                             if field.get('format') == 'currency':
                                  try:
                                       value = f"{float(value):.2f}"
                                  except (ValueError, TypeError):
                                       value = "0.00"
                             else:
                                  value = str(value)
                             
                             cell = ws.cell(row=current_row, column=1, value=f"{name} {value}")
                             cell.font = Font(size=10)
                             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
                        
                             # Add right-aligned field if available
                             if i < len(right_aligned):
                                  right_field = right_aligned[i]
                                  right_name = right_field.get('name', '')
                                  # Replace placeholder with current date
                                  if '{CUSTOM_RECORD_DATE}' in right_name:
                                       right_name = right_name.replace('{CUSTOM_RECORD_DATE}', record_date)
                                  cell = ws.cell(row=current_row, column=1, value=f"                     {right_name}")
                                  cell.font = Font(size=10)
                                  cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                                  current_row += 1
              
                   # Add spacing
                   spacing = summary_config.get('spacing', 2)
                   for _ in range(spacing):
                        current_row += 1
              
                   return current_row

          # Original logic for other report types
          # Title
          title = summary_config.get('title', '')
          if title:
                   cell = ws.cell(row=current_row, column=1, value=title)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                   current_row += 1

          # Fields with proper cell separation
          fields = summary_config.get('fields', [])
          for field in fields:
                   if "row" in field:
                        # Handle grouped fields in one row
                        for col_idx, subfield in enumerate(field["row"], start=1):
                             name = subfield.get('name', '')
                             value = subfield.get('value', '')
                             cell_a = ws.cell(row=current_row, column=col_idx * 2 - 1, value=name)
                             cell_a.font = Font(bold=True, size=10)
                             cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             cell_b = ws.cell(row=current_row, column=col_idx * 2, value=value)
                             cell_b.font = Font(size=10)
                             cell_b.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
                   else:
                        name = field.get('name', '')
                   
                        # Handle dynamic field fetching
                        if 'source_field' in field:
                             # Get value from first record in data
                             source_field = field.get('source_field', '')
                             value = data[0].get(source_field, '') if data else ''
                        else:
                             # Fallback to static value for backward compatibility
                             value = field.get('value', '')
                   
                        # Put name in column A, value in column B
                        cell_a = ws.cell(row=current_row, column=1, value=name)
                        cell_a.font = Font(bold=True, size=10)
                        cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)               
                        cell_b = ws.cell(row=current_row, column=2, value=value)
                        cell_b.font = Font(size=10)
                        cell_b.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1

          # Add blank row after section if it had content
          if summary_config.get('fields'):
                   current_row += 1

          return current_row

     def _generate_xlsx_totals_section(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX totals section with proper cell separation"""
          layout = config.get('layout', {})
          totals_config = layout.get('totals_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          if not totals_config:
                   return start_row

          current_row = start_row
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks':
                   # Calculate totals for dividend reports
                   total_records = len(data)
                   total_amount = sum(self.safe_float_convert(record.get('FN_NET_AMOUNT', 0)) for record in data)
              
                   # Title
                   title = totals_config.get('title', 'TOTALS')
                   if title:
                        cell = ws.cell(row=current_row, column=1, value=title)
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1
              
                   # Fields
                   fields = totals_config.get('fields', [])
                   for field in fields:
                        if 'row' in field:
                             # Handle row-based fields
                             row_fields = field.get('row', [])
                             row_parts = []
                        
                             for col_idx, row_field in enumerate(row_fields, 1):
                                  name = row_field.get('name', '')
                                  calculation = row_field.get('calculation', '')
                                  position = row_field.get('position', 'left')
                             
                                  if calculation == 'sum_client_statements':
                                        value = total_records
                                  elif calculation == 'sum_client_amounts':
                                        value = total_amount
                                  elif calculation == 'sum_custom_check_amounts':
                                        value = total_amount
                                  else:
                                        value = row_field.get('value', '')
                             
                                  # Format the value
                                  if row_field.get('format') == 'currency':
                                        try:
                                             value = f"{float(value):.2f}"
                                        except (ValueError, TypeError):
                                             value = "0.00"
                                  elif row_field.get('format') == 'number':
                                        value = str(value)
                                  else:
                                        value = str(value)
                             
                                  # Set cell value and alignment
                                  cell = ws.cell(row=current_row, column=col_idx, value=f"{name} {value}")
                                  cell.font = Font(size=10)
                             
                                  if position == 'right':
                                        cell.alignment = Alignment(horizontal='right', vertical='top')
                                  else:
                                        cell.alignment = Alignment(horizontal='left', vertical='top')
                        
                             current_row += 1
              
                   # Add spacing
                   spacing = totals_config.get('spacing', 1)
                   for _ in range(spacing):
                        current_row += 1
              
                   return current_row

          # Original logic for other report types
          # Calculate totals
          total_records = len(data)
          total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
          total_sheets = sum(float(record.get('DOC_SHEET_COUNT', 0)) for record in data)
          total_images = sum(float(record.get('DOC_IMAGE_COUNT', 0)) for record in data)
          email_delivery_count = sum(1 for record in data if record.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'E')
          print_delivery_count = sum(1 for record in data if record.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'P')
          suppressed_count = sum(1 for record in data if record.get('DOC_SUPPRESSION_FLAG') == 'Y')

          # Additional calculations for different report types
          # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT and CUSTOM_CHECK_NUMBER for consistency
          suppressed_amount_count = sum(1 for record in data if self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) < 5.00)
          zero_check_count = sum(1 for record in data if record.get('CUSTOM_CHECK_NUMBER', '') == '0' or record.get('CUSTOM_CHECK_NUMBER', '') == '')
          suppressed_total_count = sum(1 for record in data if self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) < 5.00 or record.get('CUSTOM_CHECK_NUMBER', '') == '0' or record.get('CUSTOM_CHECK_NUMBER', '') == '')
          packages_count = len(data)# Each record is a package
          check_pages_count = len(data) * 2# Each check has 2 pages

          # Fields with proper cell separation
          fields = totals_config.get('fields', [])
          for field in fields:
                   name = field.get('name', '')
                   calculation = field.get('calculation', '')
                   format_type = field.get('format', 'text')
         
                   if calculation == 'count_records':
                        value = total_records
                   elif calculation == 'sum_amounts':
                        value = total_amount
                   elif calculation == 'sum_sheets':
                        value = total_sheets
                   elif calculation == 'sum_images':
                        value = total_images
                   elif calculation == 'count_email_delivery':
                        value = email_delivery_count
                   elif calculation == 'count_print_delivery':
                        value = print_delivery_count
                   elif calculation == 'count_suppressed':
                        value = suppressed_count
                   elif calculation == 'count_suppressed_amount':
                        value = suppressed_amount_count
                   elif calculation == 'count_zero_check':
                        value = zero_check_count
                   elif calculation == 'count_suppressed_total':
                        value = suppressed_total_count
                   elif calculation == 'count_packages':
                        value = packages_count
                   elif calculation == 'count_check_pages':
                        value = check_pages_count
                   elif calculation == 'count_records * 2':
                        value = total_records * 2
                   elif calculation == 'count_redemption_records':
                        # Count redemption records
                        value = sum(1 for record in data if 'REDEMPTION' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'sum_redemption_amounts':
                        # Sum redemption amounts
                        value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data
                                  if 'REDEMPTION' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'count_swp_records':
                        # Count SWP records
                        value = sum(1 for record in data if 'SWP' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'sum_swp_amounts':
                        # Sum SWP amounts
                        value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data
                                  if 'SWP' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'calculate_total_input':
                        # Count total number of input rows in the report
                        value = total_records
                   elif calculation == 'calculate_total_packages':
                        # Count total number of package entries (rows)
                        value = packages_count
                   elif calculation == 'calculate_total_check_page':
                        # Sum of Total Input + Total Packages
                        value = total_records + packages_count
                   elif calculation == 'calculate_total_dollar_amount':
                        # Sum of all check amounts present in the report
                        value = total_amount
                   else:
                        value = 0

                   if format_type == 'currency':
                        formatted_value = self.formatter.format_currency(value)
                   else:
                        formatted_value = str(value)

                   # Put name in column A, value in column B
                   cell_a = ws.cell(row=current_row, column=1, value=name)
                   cell_a.font = Font(bold=True, size=10)
                   cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   cell_b = ws.cell(row=current_row, column=2, value=formatted_value)
                   cell_b.font = Font(size=10)
                   cell_b.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   current_row += 1

          # Add blank row after section if it had content
          if totals_config.get('fields'):
                   current_row += 1

          return current_row

     def _generate_xlsx_footer_section(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX footer section with proper cell separation"""
          layout = config['layout']
          footer_config = layout.get('footer_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type')

          current_row = start_row

          # Categories and data with proper cell separation
          categories = footer_config.get('categories', [])
         
          # Use dynamic data for Redemption_SWP reports, static data for others
          if report_type == 'redemption_swp':
                   category_data = self._calculate_dynamic_footer_data(data)
          else:
                   category_data = footer_config.get('category_data', [])

          for i, category in enumerate(categories):
                   data_value = category_data[i] if i < len(category_data) else ""
         
                   # Put category name in column A, value in column B
                   cell_a = ws.cell(row=current_row, column=1, value=category)
                   cell_a.font = Font(bold=True, size=10)
                   cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   cell_b = ws.cell(row=current_row, column=2, value=data_value)
                   cell_b.font = Font(size=10)
                   cell_b.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   current_row += 1

          return current_row

     def _get_column_value_for_xlsx(self, col: Dict, data: Dict) -> Any:
          """Get value for a column from data for XLSX generation"""
          source_field = col.get('source_field', '')

          if source_field == 'static':
                   return col.get('value', '')
          elif source_field == 'calculated':
                   return self._calculate_value_for_xlsx(col, data)
          else:
                   value = data.get(source_field, '')
                   # For CUSTOM_FUND_CODE, pad with leading zeros to 7 characters
                   if source_field == 'CUSTOM_FUND_CODE' and value:
                        return value.zfill(7)
                   # For CUSTOM_CHECK_AMOUNT, clean masked values
                   elif source_field == 'CUSTOM_CHECK_AMOUNT' and value:
                        return self.clean_masked_value(value)
                   return value

     def _calculate_value_for_xlsx(self, col: Dict, data: Dict) -> Any:
          """Calculate value based on column configuration for XLSX"""
          calculation = col.get('calculation', '')

          if calculation == 'count_records':
                   return 1
          elif calculation == 'sum_amounts':
                   # For Redemption_SWP reports, use CUSTOM_CHECK_AMOUNT
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'extract_group':
                   account = data.get('UC_ACCOUNT_NUMBER', '')
                   if len(account) >= 7:
                        return account[:7]
                   return account
          elif calculation == 'concatenate_fund_code_name':
                   # Concatenate CUSTOM_FUND_CODE and CUSTOM_FUND_DESC
                   fund_code = data.get('CUSTOM_FUND_CODE', '')
                   fund_desc = data.get('CUSTOM_FUND_DESC', '')
                   if fund_code and fund_desc:
                        return f"{fund_code} {fund_desc}"
                   elif fund_code:
                        return fund_code
                   elif fund_desc:
                        return fund_desc
                   return ""
          elif calculation == 'red_begin_check_number':
                   return data.get('RED_BEGIN_CHECK_NUMBER', '')
          elif calculation == 'red_end_check_number':
                   return data.get('RED_END_CHECK_NUMBER', '')
          elif calculation == 'red_check_count':
                   return data.get('RED_CHECK_COUNT', 0)
          elif calculation == 'total_red_dollar_value':
                   return data.get('TOTAL_RED_DOLLAR_VALUE', 0.0)
          elif calculation == 'swp_begin_check_number':
                   return data.get('SWP_BEGIN_CHECK_NUMBER', '')
          elif calculation == 'swp_end_check_number':
                   return data.get('SWP_END_CHECK_NUMBER', '')
          elif calculation == 'swp_check_count':
                   return data.get('SWP_CHECK_COUNT', 0)
          elif calculation == 'total_swp_dollar_value':
                   return data.get('TOTAL_SWP_DOLLAR_VALUE', 0.0)
          else:
                   return 0

     def _format_column_value_for_xlsx(self, col: Dict, value: Any) -> str:
          """Format column value for XLSX"""
          format_type = col.get('format', 'text')

          if format_type == 'currency':
                   try:
                        return f"{float(value):.2f}"
                   except:
                        return "0.00"
          else:
                   return str(value)

     def _generate_footer_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate footer section"""
          layout = config['layout']
          footer_config = layout.get('footer_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type')

          lines = []

          # Categories
          categories = footer_config.get('categories', [])
         
          # Use dynamic data for Redemption_SWP reports, static data for others
          if report_type == 'redemption_swp':
                   category_data = self._calculate_dynamic_footer_data(data)
          else:
                   category_data = footer_config.get('category_data', [])

          for i, category in enumerate(categories):
                   data_value = category_data[i] if i < len(category_data) else ""
                   line = f"{category:<30} {data_value:>15}"
                   lines.append(line)

          return '\n'.join(lines)

     def _generate_txt_dynamic_summary(self, section: Dict, data: List[Dict]) -> List[str]:
          """Generate TXT dynamic summary section after table completion"""
          lines = []
          summary_config = section.get('summary', {})

          if not summary_config:
                   return lines

          # Calculate the summary value based on configuration
          calculation = summary_config.get('calculation', '')
          summary_line = summary_config.get('line', '')
          title_after = summary_config.get('title_after', '')

          if summary_line:
                   # Calculate the value
                   if calculation == 'sum_amounts':
                        # For 12b1_compensation report, sum all amounts in the filtered data (not just < $5.00)
                        # Use CUSTOM_CHECK_AMOUNT for consistency
                        total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
                        formatted_line = summary_line.replace('{sum_amount}', f"{total_amount:.2f}")
                   elif calculation == 'count_records':
                        # Count records with zero check number
                        count = sum(1 for record in data
                             if record.get('FN_CHECK_NUMBER', '') == '0' or record.get('FN_CHECK_NUMBER', '') == '')
                        formatted_line = summary_line.replace('{count_rows}', str(count))
                   else:
                        formatted_line = summary_line
         
                   # Add the summary line
                   lines.append(formatted_line)

          # Add title after if specified (with one-line gap)
          if title_after:
                   lines.append("")# One-line gap
                   lines.append(title_after)

          return lines

     def _generate_xlsx_dynamic_summary(self, ws, section: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX dynamic summary section after table completion"""
          current_row = start_row
          summary_config = section.get('summary', {})

          if not summary_config:
                   return current_row

          # Calculate the summary value based on configuration
          calculation = summary_config.get('calculation', '')
          summary_line = summary_config.get('line', '')
          title_after = summary_config.get('title_after', '')

          if summary_line:
                   # Calculate the value
                   if calculation == 'sum_amounts':
                        # For 12b1_compensation report, sum all amounts in the filtered data (not just < $5.00)
                        # Use CUSTOM_CHECK_AMOUNT for consistency
                        total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
                        formatted_line = summary_line.replace('{sum_amount}', f"{total_amount:.2f}")
                   elif calculation == 'count_records':
                        # Count records with zero check number
                        count = sum(1 for record in data
                             if record.get('FN_CHECK_NUMBER', '') == '0' or record.get('FN_CHECK_NUMBER', '') == '')
                        formatted_line = summary_line.replace('{count_rows}', str(count))
                   else:
                        formatted_line = summary_line
         
                   # Add the summary line
                   cell = ws.cell(row=current_row, column=1, value=formatted_line)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                   current_row += 1

          # Add title after if specified (with one-line gap)
          if title_after:
                   current_row += 1# One-line gap
                   cell = ws.cell(row=current_row, column=1, value=title_after)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                   current_row += 1

          return current_row

     def _generate_xlsx_matrix_section(self, ws, section: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX matrix section - for complex tabular layouts with calculated values"""
          current_row = start_row

          # Section title
          title = section.get('title', '')
          if title:
                   cell = ws.cell(row=current_row, column=1, value=title)
                   cell.font = Font(bold=True, size=11)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                   current_row += 1

          # Get matrix configuration
          headers = section.get('headers', [])
          rows_config = section.get('rows', [])

          # Generate header row
          for col_idx, header in enumerate(headers, 1):
                   cell = ws.cell(row=current_row, column=col_idx, value=header)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

          current_row += 1

          # Generate data rows
          for row_config in rows_config:
                   row_label = row_config.get('label', '')
                   cell_values = row_config.get('values', [])
                   is_bold = row_config.get('bold', False)
                   is_total = row_config.get('is_total', False)
                   indent_level = row_config.get('indent', 0)
         
                   # Apply indentation to label
                   if indent_level > 0:
                        row_label = '' * indent_level + row_label
         
                   # Column 1: Label
                   cell = ws.cell(row=current_row, column=1, value=row_label)
                   cell.font = Font(bold=is_bold or is_total, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   # Remaining columns: Values
                   for col_idx, value_config in enumerate(cell_values, 2):
                        # Calculate or retrieve value
                        if isinstance(value_config, dict):
                             calc_type = value_config.get('calculation', '')
                             value = self._calculate_matrix_value(calc_type, data, value_config)
                        else:
                             value = value_config
              
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.font = Font(bold=is_bold or is_total, size=10)
                        cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
         
                   current_row += 1

          return current_row

     def _calculate_matrix_value(self, calc_type: str, data: List[Dict], config: Dict) -> Any:
          """Calculate value for matrix cells based on calculation type"""
          if calc_type == 'count_records':
                   filters = config.get('filters', {})
                   return self._count_filtered_records(data, filters)

          elif calc_type == 'sum_field':
                   field_name = config.get('field', '')
                   filters = config.get('filters', {})
                   return self._sum_filtered_field(data, field_name, filters)

          elif calc_type == 'multiply':
                   base_calc = config.get('base', {})
                   multiplier = config.get('multiplier', 1)
                   base_value = self._calculate_matrix_value(base_calc.get('calculation', ''), data, base_calc)
                   return base_value * multiplier

          elif calc_type == 'subtract':
                   minuend = config.get('minuend', {})
                   subtrahend = config.get('subtrahend', {})
                   minuend_value = self._calculate_matrix_value(minuend.get('calculation', ''), data, minuend)
                   subtrahend_value = self._calculate_matrix_value(subtrahend.get('calculation', ''), data, subtrahend)
                   return minuend_value - subtrahend_value

          elif calc_type == 'static':
                   return config.get('value', 0)

          elif calc_type == 'sum_amounts':
                   filters = config.get('filters', {})
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   return self._sum_filtered_field(data, 'CUSTOM_CHECK_AMOUNT', filters)

          elif calc_type == 'sum_sheets':
                   filters = config.get('filters', {})
                   return self._sum_filtered_field(data, 'DOC_SHEET_COUNT', filters)

          elif calc_type == 'sum_images':
                   filters = config.get('filters', {})
                   return self._sum_filtered_field(data, 'DOC_IMAGE_COUNT', filters)

          else:
                   return 0

     def _count_filtered_records(self, data: List[Dict], filters: Dict) -> int:
          """Count records that match the given filters"""
          if not filters:
                   return len(data)

          count = 0
          for record in data:
                   match = True
                   for field, expected_value in filters.items():
                        record_value = record.get(field, '')
              
                        # Handle different comparison types
                        if isinstance(expected_value, dict):
                             operator = expected_value.get('operator', '==')
                             compare_value = expected_value.get('value', '')
                   
                             # Convert to appropriate types for comparison
                             try:
                                  if operator in ['<', '>', '<=', '>=']:
                                        record_num = float(record_value) if record_value else 0
                                        compare_num = float(compare_value) if compare_value else 0
                             
                                        if operator == '<':
                                             if not (record_num < compare_num):
                                                  match = False
                                                  break
                                        elif operator == '>':
                                             if not (record_num > compare_num):
                                                  match = False
                                                  break
                                        elif operator == '<=':
                                             if not (record_num <= compare_num):
                                                  match = False
                                                  break
                                        elif operator == '>=':
                                             if not (record_num >= compare_num):
                                                  match = False
                                                  break
                                  elif operator == '==':
                                        if record_value != compare_value:
                                             match = False
                                             break
                                  elif operator == '!=':
                                        if record_value == compare_value:
                                             match = False
                                             break
                                  elif operator == 'in':
                                        if record_value not in compare_value:
                                             match = False
                                             break
                             except (ValueError, TypeError):
                                  # If conversion fails, treat as string comparison
                                  if operator == '==':
                                        if record_value != compare_value:
                                             match = False
                                             break
                                  elif operator == '!=':
                                        if record_value == compare_value:
                                             match = False
                                             break
                        else:
                             if record_value != expected_value:
                                  match = False
                                  break
         
                   if match:
                        count += 1

          return count

     def _sum_filtered_field(self, data: List[Dict], field_name: str, filters: Dict) -> float:
          """Sum a field value for records that match the given filters"""
          if not field_name:
                   return 0.0

          total = 0.0
          for record in data:
                   match = True
         
                   # Check filters
                   for filter_field, expected_value in filters.items():
                        if record.get(filter_field, '') != expected_value:
                             match = False
                             break
         
                   if match:
                        try:
                             total += float(record.get(field_name, 0))
                        except:
                             pass

          return total

     def _generate_summary_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate summary section with exact current design formatting"""
          layout = config.get('layout', {})
          summary_config = layout.get('summary_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          if not summary_config:
                   return ""

          lines = []
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks':
                   # Title with proper alignment
                   title = summary_config.get('title', 'GENERAL AUDIT COUNTS')
                   lines.append(f"{title}")
              
                   # Subtitle with proper alignment - use current date
                   subtitle = summary_config.get('subtitle', '')
                   if subtitle:
                        # Use current date for document date
                        current_date = datetime.now().strftime('%m/%d/%Y')
                        lines.append(f"{subtitle}     {current_date}")
              
                   # Data block with proper alignment
                   data_block = summary_config.get('data_block', {})
                   if data_block:
                        lines.append("")
                        label = data_block.get('label', 'R06686 Data')
                   
                        # Use current date for record date
                        record_date = datetime.now().strftime('%m/%d/%Y')
                        lines.append(f"{label}                    Record Date:     {record_date}")
                   
                        # Left-aligned fields with right-aligned dates
                        fields = data_block.get('fields', [])
                        right_aligned = data_block.get('right_aligned', [])
                   
                        for i, field in enumerate(fields):
                             name = field.get('name', '')
                             calculation = field.get('calculation', '')
                             
                             # Calculate value based on calculation type
                             if calculation == 'count_input_records':
                                  value = len(data)
                             elif calculation == 'count_total_checks':
                                  value = len(data) # Total checks = number of records
                             elif calculation == 'sum_total_check_amounts':
                                  value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', '0')) for record in data)
                             else:
                                  value = field.get('value', '')
                             
                             # Format the value
                             if field.get('format') == 'currency':
                                  try:
                                       value = f"{float(value):.2f}"
                                  except (ValueError, TypeError):
                                       value = "0.00"
                             else:
                                  value = str(value)
                             
                             lines.append(f"{name} {value}")
                        
                             # Add right-aligned field if available
                             if i < len(right_aligned):
                                  right_field = right_aligned[i]
                                  right_name = right_field.get('name', '')
                                  # Replace placeholder with current date
                                  if '{CUSTOM_RECORD_DATE}' in right_name:
                                       right_name = right_name.replace('{CUSTOM_RECORD_DATE}', record_date)
                                  lines.append(f"                     {right_name}")
              
                   # Add spacing
                   spacing = summary_config.get('spacing', 2)
                   for _ in range(spacing):
                        lines.append("")
              
                   return '\n'.join(lines)

          # Original logic for other report types
          # Title - "Reconciliation Report"
          title = summary_config.get('title', '')
          if title:
                   lines.append(title)

          # Fields - exact current design formatting
          fields = summary_config.get('fields', [])
          for field in fields:
                   name = field.get('name', '')
                   position = field.get('position', 'left')
              
                   # Handle dynamic field fetching
                   if 'source_field' in field:
                        # Get value from first record in data
                        source_field = field.get('source_field', '')
                        value = data[0].get(source_field, '') if data else ''
                   else:
                        # Fallback to static value for backward compatibility
                        value = field.get('value', '')
         
                   if position == 'left':
                        # Exact formatting: name left-aligned to 20 chars, value follows
                        line = f"{name:<20} {value}"
                   else:
                        line = f"{name} {value:>20}"
         
                   lines.append(line)

          # Add spacing - always 2 blank lines after summary (current design)
          if lines:
                   lines.append("")
                   lines.append("")

          return '\n'.join(lines)

     def _generate_totals_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate totals section with exact current design formatting"""
          layout = config.get('layout', {})
          totals_config = layout.get('totals_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          if not totals_config:
                   return ""

          lines = []
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks':
                   # Calculate totals for dividend reports
                   total_records = len(data)
                   total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
              
                   # Title
                   title = totals_config.get('title', 'TOTALS')
                   lines.append(title)
              
                   # Fields
                   fields = totals_config.get('fields', [])
                   for field in fields:
                        if 'row' in field:
                             # Handle row-based fields
                             row_fields = field.get('row', [])
                             row_parts = []
                        
                             # Build the totals line in the exact format from image
                             totals_line = "Total:     Total Client Statements "
                             
                             # Add statements count
                             statements_count = total_records
                             totals_line += str(statements_count)
                             
                             # Add amount with proper spacing
                             totals_line += "     Total Client Amount "
                             try:
                                  amount_formatted = f"{float(total_amount):.2f}"
                             except (ValueError, TypeError):
                                  amount_formatted = "0.00"
                             totals_line += amount_formatted
                             
                             lines.append(totals_line)
              
                   # Add spacing
                   spacing = totals_config.get('spacing', 1)
                   for _ in range(spacing):
                        lines.append("")
              
                   return '\n'.join(lines)

          # Original logic for other report types
          # Calculate totals
          total_records = len(data)
          total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
          total_sheets = sum(float(record.get('DOC_SHEET_COUNT', 0)) for record in data)
          total_images = sum(float(record.get('DOC_IMAGE_COUNT', 0)) for record in data)
          email_delivery_count = sum(1 for record in data if record.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'E')
          print_delivery_count = sum(1 for record in data if record.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'P')
          suppressed_count = sum(1 for record in data if record.get('DOC_SUPPRESSION_FLAG') == 'Y')

          # Additional calculations for different report types
          # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT and CUSTOM_CHECK_NUMBER for consistency
          suppressed_amount_count = sum(1 for record in data if self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) < 5.00)
          zero_check_count = sum(1 for record in data if record.get('CUSTOM_CHECK_NUMBER', '') == '0' or record.get('CUSTOM_CHECK_NUMBER', '') == '')
          suppressed_total_count = sum(1 for record in data if self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) < 5.00 or record.get('CUSTOM_CHECK_NUMBER', '') == '0' or record.get('CUSTOM_CHECK_NUMBER', '') == '')
          packages_count = len(data)# Each record is a package
          check_pages_count = len(data) * 2# Each check has 2 pages

          # Fields - exact current design formatting
          fields = totals_config.get('fields', [])
          for field in fields:
                   name = field.get('name', '')
                   calculation = field.get('calculation', '')
                   format_type = field.get('format', 'text')
         
                   if calculation == 'count_records':
                        value = total_records
                   elif calculation == 'sum_amounts':
                        value = total_amount
                   elif calculation == 'sum_sheets':
                        value = total_sheets
                   elif calculation == 'sum_images':
                        value = total_images
                   elif calculation == 'count_email_delivery':
                        value = email_delivery_count
                   elif calculation == 'count_print_delivery':
                        value = print_delivery_count
                   elif calculation == 'count_suppressed':
                        value = suppressed_count
                   elif calculation == 'count_suppressed_amount':
                        value = suppressed_amount_count
                   elif calculation == 'count_zero_check':
                        value = zero_check_count
                   elif calculation == 'count_suppressed_total':
                        value = suppressed_total_count
                   elif calculation == 'count_packages':
                        value = packages_count
                   elif calculation == 'count_check_pages':
                        value = check_pages_count
                   elif calculation == 'count_records * 2':
                        value = total_records * 2
                   elif calculation == 'count_redemption_records':
                        # Count redemption records
                        value = sum(1 for record in data if 'REDEMPTION' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'sum_redemption_amounts':
                        # Sum redemption amounts
                        value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data
                                  if 'REDEMPTION' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'count_swp_records':
                        # Count SWP records
                        value = sum(1 for record in data if 'SWP' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'sum_swp_amounts':
                        # Sum SWP amounts
                        value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data
                                  if 'SWP' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'calculate_total_input':
                        # Count total number of input rows in the report
                        value = total_records
                   elif calculation == 'calculate_total_packages':
                        # Count total number of package entries (rows)
                        value = packages_count
                   elif calculation == 'calculate_total_check_page':
                        # Sum of Total Input + Total Packages
                        value = total_records + packages_count
                   elif calculation == 'calculate_total_dollar_amount':
                        # Sum of all check amounts present in the report
                        value = total_amount
                   else:
                        value = 0

                   if format_type == 'currency':
                        formatted_value = self.formatter.format_currency(value)
                   else:
                        formatted_value = str(value)

                   # Exact formatting: name left-aligned to 30 chars, value right-aligned to 15 chars
                   line = f"{name:<30} {formatted_value:>15}"
                   lines.append(line)

          # Add spacing - always 1 blank line after totals (current design)
          if lines:
                   lines.append("")

          return '\n'.join(lines)

class JSONDrivenReportManager:
     """Main manager for JSON-driven report generation"""

     def __init__(self):
          self.generator = ReportGenerator()
          self.config_dir = 'config'

     def detect_file_type(self, filename: str) -> str:
          """Detect file type using regex patterns"""
          import re
        #   filename_lower = filename.lower()

          # Regex patterns for different file types
          patterns = {
                   'redemption_swp': [r'^JHI\.RED\.RECON\.Report\.\d{8}\.xml$'],
                   'dividend': [r'^DIV\.chck\.RECON\.\d{8}\.xml$'],
                   '12b1_compensation': [r'^12b1\.chck\.RECONSUPPRESS\.\d{8}\.xml$'],
                   'replacement': [r'^JHI\.REPL\.Recon\.Report\.\d{8}\.xml$'],
                   'confirms': [r'.*confirms.*', r'.*confirm.*']
          }
          for file_type, pattern_list in patterns.items():
                   for pattern in pattern_list:
                        if re.search(pattern, filename, re.IGNORECASE):
                             return file_type

          return 'unknown'

#      def process_file_by_type(self, xml_file_path: str, file_type: str) -> bool:
#                         """Process file based on detected type"""
#                         # Use S3 configuration path
#                         s3_bucket = os.getenv('s3BucketInput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
#                         config_key = os.getenv('configKey', 'jhi/checks/config/jhi_checks_reporting_config.json')
#                         config_path = f"s3://{s3_bucket}/{config_key}"
                   
#                         config = self.generator.load_config(config_path, file_type)
#                         log_message("INFO", f"Loaded configuration for file type: {file_type}")
                   
#                         if not config:
#                          log_message("ERROR", f"No configuration found for file type {file_type}")
#                          return False
                   
#                         config['data_sources']['input_xml_path'] = xml_file_path
#                         log_message("INFO", f"Processing file: {xml_file_path}")
                   
#                         return self.generator.generate_report_from_config

     def process_file_by_type(self, xml_file_path: str, file_type: str) -> bool:
          # print("Inside procees_file_by_type_1")
          """Process file based on detected type"""
                    # Use S3 configuration path
          s3_bucket = os.getenv(
                    "s3BucketInput", "br-icsdev-dpmjayant-dataingress-us-east-1-s3"
          )
          config_key = os.getenv(
                    "configKey", "jhi/checks/config/jhi_checks_reporting_config.json"
          )
          config_path = f"s3://{s3_bucket}/{config_key}"
          # print("Inside procees_file_by_type_2")      # new unified config
          config = self.generator.load_config(config_path, file_type)
          print(config)

          if not config:
                    # print(f"Error: No configuration found for file type {file_type}")
                    return False

          # Add the input XML path to the config
          if "data_sources" not in config:
                    config["data_sources"] = {}
          config["data_sources"]["input_xml_path"] = xml_file_path
          print("error is sucessfull", config["data_sources"]["input_xml_path"])

          return self.generator.generate_report_from_config(config)
          # type: ignore #print("Inside procees_file_by_type_6")


     def generate_all_reports(self) -> Dict[str, bool]:
          """Generate all reports based on configuration files"""
          results = {}

          if not os.path.exists(self.config_dir):
                   print(f"Configuration directory not found: {self.config_dir}")
                   return results

          config_files = [f for f in os.listdir(self.config_dir) if f.endswith('.json')]

          if not config_files:
                   print(f"No configuration files found in {self.config_dir}")
                   return results

          print(f"Found {len(config_files)} configuration files")

          for config_file in config_files:
                   config_path = os.path.join(self.config_dir, config_file)
                   report_type = config_file.replace('_config.json', '').replace('_report', '')
         
                   print(f"\nProcessing {report_type} report...")
                   success = self.generator.generate_report(config_path)
                   results[report_type] = success
          return results

     def generate_single_report(self, report_type: str) -> bool:
          """Generate a single report by type"""
          # Try different naming patterns
          possible_configs = [
                   f"{report_type}_config.json",
                   f"{report_type}_report_config.json",
                   f"{report_type}.json"
          ]

          config_path = None
          for config_file in possible_configs:
                   test_path = os.path.join(self.config_dir, config_file)
                   if os.path.exists(test_path):
                        config_path = test_path
                        break

          if not config_path:
                   print(f"Configuration file not found. Tried: {possible_configs}")
                   return False

          return self.generator.generate_report(config_path)

def get_key_list(bucket_name, path, xml_regex):
          print(f"INFO: Starting to get key list from S3 bucket: {bucket_name}")
          print(f"INFO: Path: {path}")
          print(f"INFO: XML regex pattern: {xml_regex}")
          keyList = []
          valueList = []
          try:
                    print("INFO: Creating S3 paginator for list_objects_v2")
                    paginator = s3.get_paginator('list_objects_v2')
                    regex_pattern = re.compile(xml_regex)
                    print(f"INFO: Compiled regex pattern: {regex_pattern}")
                    count = 0
                    print("INFO: Starting to paginate through S3 objects")
                    for page in paginator.paginate(Bucket=bucket_name, Prefix=path):
                              if 'Contents' not in page:
                                        print("WARNING: No Contents found in current page, continuing...")
                                        continue
                              for obj in page['Contents']:
                                        count += 1
                                        filename = obj['Key'].split('/')[-1]
                                        if regex_pattern.search(filename):
                                                  print(f"INFO: Found matching file: {filename}")
                                                  keyList.append(obj['Key'])
                    print(f"INFO: Processed {count} objects total")
                    print(f"INFO: Found {len(keyList)} matching files")
                    print(f"INFO: KeyList: {keyList}")
                    return keyList
          except Exception as e:
                    print(f"ERROR: Failed to list objects in bucket {bucket_name} with prefix {path}: {str(e)}")
                    print(f"ERROR: Traceback: {traceback.format_exc()}")
                    message=f"Error listing objects in bucket {bucket_name} with prefix {path}: {str(e)}"
                    log_message("failed", message)
                    raise Exception(message)



def execute_report_generation(input_key, bucket, input_path, report_path, config_key, manager, overall_success):
          try:
                    print(f"INFO: Starting report generation for input key: {input_key}")
                    print(f"INFO: Bucket: {bucket}, Input path: {input_path}, Report path: {report_path}")

                    # Construct full S3 path to input XML
                    s3_input_path = f"s3://{bucket}/{input_key}"
                    input_file_name = input_key.split("/")[-1]
                    print(f"INFO: S3 input path: {s3_input_path}")
                    print(f"INFO: Input file name: {input_file_name}")

                    # Detect file type
                    print("INFO: Detecting file type...")
                    file_type = manager.detect_file_type(input_file_name)
                    print(f"INFO: Detected file type: {file_type}")
                    log_message("INFO", f"Detected file type: {file_type}")

                    # Check if config file exists first
                    print(f"INFO: Checking if configuration file exists: s3://{bucket}/{config_key}")
                    config_exists = checkFile_isPresent(bucket, config_key, config_key.split('/')[-1], None)
                   
                    if not config_exists:
                              print(f"WARNING: Configuration file does not exist: s3://{bucket}/{config_key}")
                              print(f"INFO: Attempting to create sample configuration for file type: {file_type}")
                             
                              # Try to create a sample configuration file
                              if create_sample_config(bucket, config_key, file_type):
                                        print(f"INFO: Sample configuration created successfully")
                              else:
                                        print(f"WARNING: Failed to create sample configuration, will use default config")
                   
                    # Load config
                    print(f"INFO: Loading configuration from: s3://{bucket}/{config_key}")
                    config = manager.generator.load_config(f"s3://{bucket}/{config_key}", file_type)
                    if not config:
                              print(f"ERROR: No configuration found for file type {file_type}")
                              print(f"INFO: Creating a default configuration for file type: {file_type}")
                             
                              # Create a default configuration for the file type
                              default_config = {
                                        "data_sources": {
                                                  "input_xml_path": f"s3://{bucket}/{input_key}",
                                                  "output_path": f"s3://{bucket}/{report_path}"
                                        },
                                        "report_settings": {
                                                  "file_type": file_type,
                                                  "output_format": "csv",
                                                  "include_headers": True
                                        },
                                        "processing": {
                                                  "extract_all_elements": True,
                                                  "create_summary": True
                                        }
                              }
                             
                              print(f"INFO: Using default configuration for file type {file_type}")
                              log_message("INFO", f"Using default configuration for file type {file_type}")
                              config = default_config

                    # Set environment variables for use in generate_report_from_config
                    print("INFO: Setting environment variables for report generation")
                    os.environ["s3BucketInput"] = bucket
                    os.environ["inputPath"] = input_path
                    os.environ["inputFileName"] = input_file_name
                    os.environ["reportPath"] = report_path
                    print("INFO: Environment variables set successfully")
                   
                    print(f"INFO: Configuration loaded: {config}")

                    # Generate report
                    print("INFO: Starting report generation from configuration")
                    success = manager.generator.generate_report_from_config(config)
                    if success:
                              print(f"INFO: Report generation completed successfully for {input_key}")
                              log_message("SUCCESS", f"Report generation completed successfully for {input_key}")
                    else:
                              print(f"ERROR: Report generation failed for {input_key}")
                              log_message("failed", f"Report generation failed for {input_key}")
                              overall_success = False
          except Exception as e:
                    print(f"ERROR: Exception in execute_report_generation: {str(e)}")
                    print(f"ERROR: Traceback: {traceback.format_exc()}")
                    log_message("ERROR", f"Exception in execute_report_generation: {str(e)}")
                    overall_success = False

def main():
          """AWS Glue job main function (list-based params)"""
          try:
                    print("INFO: Starting AWS Glue job main function")
                    # Set up Glue job parameters (still needed for other env args/logging)
                    print("INFO: Setting up Glue job parameters")
                    set_job_params_as_env_vars()
                    print("INFO: Glue job parameters set successfully")
                   
                   
                    # # Get input parameters from environment variables
                    print("INFO: Retrieving input parameters from environment variables")
                    global inputFileName
                    inputFileName = os.getenv('inputFileName')
                    bucket = os.getenv('bucket')
                    print(f"INFO: Input file name: {inputFileName}")
                    print(f"INFO: Bucket: {bucket}")
                   
                    print("INFO: Parsing JSON environment variables")
                    inputPathList = json.loads(os.getenv('inputPathList'))
                    inputPatternList = json.loads(os.getenv('inputPatternList'))
                    inputKeyList = json.loads(os.getenv('inputKeyList'))
                    reportPathList = json.loads(os.getenv('reportPathList'))
                    configKeyList = json.loads(os.getenv('configKeyList'))
                    print(f"INFO: Parsed {len(inputPathList)} input paths, {len(inputPatternList)} patterns, {len(inputKeyList)} keys, {len(reportPathList)} report paths, {len(configKeyList)} config keys")
                   
                    # Initialize manager
                    print("INFO: Initializing JSONDrivenReportManager")
                    manager = JSONDrivenReportManager()
                    print("INFO: Manager initialized successfully")

                    # ========= 1) Your new list-style inputs (hard-coded or passed-in) =========
                    # inputFileName = 'filename.zip'     # OPTIONAL: only used for logging context if you want
                    # bucket = 'br-icsdev-dpmjayant-dataingress-us-east-1-s3'
                    # inputPathList = ['jhi/checks/wip/wipid/']
                    # inputPatternList = ['^((janus_report_input_data_1_Redemption_SWP_Report_sp001.xml)|(janus_report_input_data_1_Replacement_Report_sp002.xml)|(janus_report_report_dividend_data_sp003.xml)|(report_12b1_data.xml))']
                    # inputKeyList = []     # If pre-filled, will be used as keys; else discovered via pattern
                    # reportPathList = ['jhi/checks/wip/wipid/report/']
                    # configKeyList = ['jhi/checks/config/jhi_checks_reporting_config.json']

                    # ========= 2) Helper: broadcast singletons =========
                    def pick(lst, idx):
                              """Return lst[idx] if available; else broadcast lst[0] if lst has one; else raise."""
                              if not lst:
                                        raise ValueError("A required list is empty.")
                              if len(lst) == 1:
                                        return lst[0]
                              return lst[idx]

                    # We'll iterate by the longest of the list dimensions you control (usually they are all len 1)
                    list_len = max(len(inputPathList), len(inputPatternList), len(reportPathList), len(configKeyList))
                    print(f"INFO: Processing {list_len} list groups")
                    # dont check for max, check if all the lengths are same (jk)

                    overall_success = True
                    processed_any = False
                    print("INFO: Starting report generation process")

                    # ========= 3) For each (path, pattern, report_path, config_key) set =========
                    if not inputKeyList:
                              print("INFO: No input key list provided, discovering keys using patterns")
                              for i in range(list_len):
                                        print(f"INFO: Processing group {i+1}/{list_len}")
                                        try:
                                                  input_path = pick(inputPathList, i)
                                                  input_pattern = pick(inputPatternList, i)
                                                  report_path = pick(reportPathList, i)
                                                  config_key = pick(configKeyList, i)
                                                  print(f"INFO: Group {i+1} - Input path: {input_path}, Pattern: {input_pattern}, Report path: {report_path}, Config key: {config_key}")
                                                 
                                                  discovered_keys = get_key_list(bucket, input_path, input_pattern)
                                                  print(f"INFO: Group {i+1} - Discovered {len(discovered_keys)} keys: {discovered_keys}")

                                                  if not discovered_keys:
                                                            print(f"WARNING: No keys discovered for group {i+1} - path={input_path} pattern={input_pattern}")
                                                            log_message("INFO", f"No keys discovered for path={input_path} pattern={input_pattern}")
                                                            continue

                                                  processed_any = True
                                                  # -------- 3b) Process each key in this group --------
                                                  for j, key in enumerate(discovered_keys):
                                                            print(f"INFO: Processing key {j+1}/{len(discovered_keys)} in group {i+1}: {key}")
                                                            execute_report_generation(key, bucket, input_path, report_path, config_key, manager, overall_success)
                                        except Exception as e:
                                                  print(f"ERROR: Exception processing group {i+1}: {str(e)}")
                                                  print(f"ERROR: Traceback: {traceback.format_exc()}")
                                                  log_message("ERROR", f"Exception processing group {i+1}: {str(e)}")
                                                  overall_success = False
                    else:
                              print(f"INFO: Using provided inputKeyList with {len(inputKeyList)} keys")
                              for i in range(len(inputKeyList)):
                                        try:
                                                  print(f"INFO: Processing provided key {i+1}/{len(inputKeyList)}: {inputKeyList[i]}")
                                                  execute_report_generation(inputKeyList[i], bucket, inputPathList[i], reportPathList[i], configKeyList[i], manager, overall_success)
                                                  processed_any = True
                                        except Exception as e:
                                                  print(f"ERROR: Exception processing provided key {i+1}: {str(e)}")
                                                  print(f"ERROR: Traceback: {traceback.format_exc()}")
                                                  log_message("ERROR", f"Exception processing provided key {i+1}: {str(e)}")
                                                  overall_success = False

                    if not processed_any:
                              print("ERROR: No inputs were processed across all list groups")
                              print("INFO: This could be due to:")
                              print("     - Missing configuration files in S3")
                              print("     - No files matching the specified patterns")
                              print("     - Incorrect S3 paths or bucket permissions")
                              log_message("failed", "No inputs to process across all list groups")
                              return False

                    
          except Exception as e:
                    print(f"ERROR: Glue job failed with error: {str(e)}")
                    print(f"ERROR: Traceback: {traceback.format_exc()}")
                    error_msg = f"Glue job failed with error: {str(e)}"
                    log_message("failed", error_msg)
                    raise Exception(error_msg)

         
inputFileName=""
if __name__ == "__main__":
     main()