"""
AWS Lambda Function: JSON-Driven Report Generator
=================================================

This Lambda function generates reports based on JSON configuration, preserving
all existing business logic and formatting.

Key Lambda Features:
- S3 integration for input/output files
- Splunk logging for monitoring
- Parallel S3 operations using ThreadPoolExecutor
- Email notifications for failures
- Environment variable parameter handling

Required Environment Variables:
- S3_BUCKET_INPUT: S3 bucket for input files
- S3_BUCKET_OUTPUT: S3 bucket for output file
- INPUT_FILE_NAME: Name of the input XML/JSON file
- INPUT_PATH: S3 path prefix for input files
- REPORT_PATH: S3 path prefix for output reports
- CONFIG_KEY: S3 key for configuration file
- TRANSACTION_ID: Transaction identifier for logging
- CLIENT_NAME, CLIENT_ID: Client information
- TEAMS_ID: Email/Teams ID for failure notifications

Lambda Function Name: br_icsdev_dpmjayant_lambda_reporting_checks
Step Function / Trigger: report
"""

import json
import xml.etree.ElementTree as ET
import os
import sys
import re
import zipfile
import io
import ast
import time
import traceback
from datetime import datetime, timezone
from typing import Dict, List, Any, Optional
import csv
from concurrent.futures import ThreadPoolExecutor, as_completed
from botocore.exceptions import ClientError
import boto3
# import splunk
from dpm_splunk_logger_py import splunk
from botocore.exceptions import ClientError, NoCredentialsError
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import json
import traceback
import os
from datetime import datetime
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

s3 = boto3.client('s3')
def extract_base_path(prefix):
    parts = prefix.split("/")
    # Take first 4 components: jhi / checks / wip / UUID
    base_path = "/".join(parts[:4])
    return base_path

def list_csv_files(bucket, prefix):
    """
    List all .csv files under an S3 prefix.
    """

    # Ensure prefix ends with "/"
    if not prefix.endswith("/"):
        prefix += "/"

    # Step 1: Extract base path (folder)
    base_path = extract_base_path(prefix)
    if not base_path.endswith("/"):
        base_path += "/"

    

    # Step 2: List objects under base path
    response = s3.list_objects_v2(Bucket=bucket, Prefix=base_path)

    csv_files = []

    # Step 3: If no files found, return empty list
    if "Contents" not in response:
        return csv_files

    # Step 4: Loop S3 objects and pick only .csv
    for obj in response["Contents"]:
        key = obj["Key"]
        if key.lower().endswith(".csv"):
            csv_files.append(key)

    return csv_files


def parse_trailer(line):
    """Extract package count and detail count from trailer line."""
    parts = [p.strip() for p in line.split(',')]
    # Trailer format: "Trailer , package_count , detail_count"
    try:
        package_count = int(parts[1]) if parts[1] else 0
        detail_count = int(parts[2]) if len(parts) > 2 and parts[2] else 0
        return package_count, detail_count
    except (ValueError, IndexError):
        return 0, 0


def update_package_number(line, new_package_number,pattern):
    """Update the package number in a detail row line."""
    # Format: ", number ," or ",        ,"
    # Find the pattern between first comma and second comma
    # Use regex to match and replace: ", number ," or ",        ,"
    # pattern = r'(,)\s*\d*\s*(,)'
    replacement = f'\\1 {new_package_number} \\2'
    result = re.sub(pattern, replacement, line, count=1)
    return result



def parse_package_number(line):
    """Extract package number from a detail row line."""
    # Format: ", package_number ," or ",        ," (empty)
    # Split by comma and get the second field (index 1)
    parts = line.split(',')
    if len(parts) > 1:
        pkg_str = parts[1].strip()
        try:
            return int(pkg_str) if pkg_str else None
        except ValueError:
            return None
    return None

def merge_and_rename_csv_files(files_list, bucket, config_key, manager, transaction_id=""):
    """
    Fully S3-only merging workflow.
    No /tmp/ usage.
    All processing happens in memory + WIP folder on S3.
    """
    # ---------------------------
    # Load config
    # ---------------------------

    name_list=[]
    for i in range(len(files_list)):
          name_list+=[files_list[i].split('/')[-1]]
    

    config = manager.generator.load_config(
        f"s3://{bucket}/{config_key}",
        manager.detect_file_type(name_list)
    )
    
    output_settings = config.get("workflow_metadata", {}).get("output_formats", {})
    txt_cfg = output_settings.get("csv", {})

    header = None
    detail_header = None
    all_detail_rows = []
    total_pkgs = 0
    total_details = 0
    max_pkg = 0
    wip_files = []
    for key in files_list:
        obj = s3.get_object(Bucket=bucket, Key=key)
        lines = obj['Body'].read().decode('utf-8').splitlines(True)
        # print('-1')
        detail_rows = []
        trailer = None
       
        for line in lines:
            s = line.strip()
            
            if s.startswith("Header,") and header is None:
                header = line
            elif s.startswith("Detail,") and detail_header is None:
                detail_header = line
            elif s.startswith("Trailer"):
                trailer = line
            else:
                if s:
                    detail_rows.append(line)

        # Parse trailer counts
        pkg_count, dt_count = parse_trailer(trailer)
        total_pkgs += pkg_count
        total_details += dt_count
        pattern = txt_cfg.get("pattern")
        
        # Renumber package numbers
        last_pkg = None
        updated = []
        for row in detail_rows:
            pkg = parse_package_number(row)

            if pkg is not None:
                if last_pkg is None or pkg != last_pkg:
                    max_pkg += 1
                    last_pkg = pkg

                updated.append(update_package_number(row, max_pkg,pattern))
            else:
                updated.append(row)

        all_detail_rows.extend(updated)

    # ---------------------------
    # BUILD MERGED CSV IN MEMORY
    # ---------------------------
    file_path = files_list[0]
    wip_folder = file_path.rsplit('/', 1)[0] + '/'
    output_name = txt_cfg.get("out_put_name_file", "Merged_Output") + ".csv"
    merged_csv_key = wip_folder + output_name
    csv_buffer = io.StringIO()

    if header:
        csv_buffer.write(header)
    if detail_header:
        csv_buffer.write(detail_header)

    for row in all_detail_rows:
        csv_buffer.write(row)

    csv_buffer.write(f"Trailer ,{total_pkgs},{total_details}\n")
    
    data_time_formate=output_name.split('.')[-2]
    
    # Replace date format in output_name
    current_date = datetime.now().strftime(data_time_formate)
   
    output_name = output_name.replace(data_time_formate, current_date)
   
    merged_csv_key = wip_folder + output_name
    
    # Upload merged CSV directly to S3 (final location)
    s3.put_object(
        Bucket=bucket,
        Key=merged_csv_key,
        Body=csv_buffer.getvalue(),
        ContentType="text/csv"
    )
    

    # ---------------------------
    # SEND EMAIL WITH ATTACHMENT
    # ---------------------------
    email_cfg = config.get("emailConfig", {})
    subject = email_cfg.get("subject", "").replace("{date}", datetime.now().strftime("%m%d%Y"))
    body_text=f"""embed above text file"""
    tmp_download_path = f"/tmp/{output_name}"
   
    if not os.path.exists("/tmp"):
          os.makedirs("/tmp", exist_ok=True)
    send_email_with_attachment(
        sender_email=email_cfg.get("fromEmail"),
        recipient_emails=email_cfg.get("toEmails", []),
        cc_emails=email_cfg.get("ccEmails", []),
        subject=subject,
        body_text=body_text,
        attachment_file_path=tmp_download_path,   # Filename for local download (email function will handle)
        bucket_name=bucket,
        object_key=merged_csv_key,          # S3 key where file is stored
        attachment_flag=True,
        report_attachment_flag="true",
        multiple_files=None,
        attached_multiple_file=False,
        gmailtrigger=email_cfg.get("gmailtrigger", False),
        context=None
    )
    try:
       if tmp_download_path and os.path.isfile(tmp_download_path):
          os.remove(tmp_download_path)
          
    except Exception as e:
          log_message('failed', f'Failed to download file from S3: {e}')
          


    # ---------------------------
    # COPY FINAL OUTPUT TO CUSTOMER FOLDER
    # ---------------------------

    cust_key = f"jhi/checks/output/to_cust/{output_name}"

    s3.copy_object(
        Bucket=bucket,
        CopySource={"Bucket": bucket, "Key": merged_csv_key},
        Key=cust_key
    )

    return cust_key
    #return merged_csv_key

def download_file_from_s3(bucket_name, object_key, destination_path):
    """
    Downloads a file from an S3 bucket to a local destination path.

    Args:
        bucket_name (str): Name of the S3 bucket.
        object_key (str): Key/path of the object in S3.
        destination_path (str): Local path to save the downloaded file.

    Logs a message on success or failure.
    """
    try:
        s3.download_file(bucket_name, object_key, destination_path)
    except ClientError as e:
        log_message('failed', f'Failed to download file from S3: {e}')

def discover_files_by_prefix(bucket_name, base_path, prefixes):
    """Discover files in S3 that start with the given prefixes"""
    try:
        discovered_files = []
        
        # List all objects in the bucket with the base path
        if base_path:
            prefix_filter = base_path if base_path.endswith('/') else f"{base_path}/"
        else:
            prefix_filter = ""
        
        log_message('INFO', f'Discovering files in S3 bucket: {bucket_name}, path: {prefix_filter}')
        
        paginator = s3.get_paginator('list_objects_v2')
        pages = paginator.paginate(Bucket=bucket_name, Prefix=prefix_filter)
        
        for page in pages:
            if 'Contents' in page:
                for obj in page['Contents']:
                    file_key = obj['Key']
                    file_name = file_key.split('/')[-1]  # Get just the filename
                    
                    # Check if this file matches any of the prefixes
                    for prefix in prefixes:
                        if file_name.startswith(prefix):
                            discovered_files.append({
                                'object_key': file_key,
                                'file_name': file_name,
                                'prefix': prefix
                            })
                            log_message('INFO', f'Found matching file: {file_name} (matches prefix: {prefix})')
                            break  # Don't add the same file multiple times
        
        log_message('INFO', f'File discovery completed - found {len(discovered_files)} files matching prefixes: {prefixes}')
        return discovered_files
        
    except Exception as e:
        log_message('ERROR', f'Failed to discover files by prefix: {str(e)}')
        return []

def send_email_with_attachment(
    sender_email,
    recipient_emails,
    cc_emails,
    subject,
    body_text,
    attachment_file_path,
    bucket_name,
    object_key,
    attachment_flag,
    report_attachment_flag="",
    multiple_files=None,
    attached_multiple_file=False,
    gmailtrigger=False,
    context=None
):
    # Note: gmailtrigger only affects filename prefix logic, not whether email is sent
    if not gmailtrigger:
        log_message('INFO', 'Email trigger disabled - emails will be sent without Gmail-specific filename prefixes')
    
    # Create an SES client
    ses_client = boto3.client('ses')
    input_file_name = os.getenv('inputFileName', '')
    property_config_name = os.getenv('propertyConfigName', '')
    
    log_message('INFO', f'Starting email preparation for file: {input_file_name}')
    
    # Enhanced Splunk logging examples as requested
    if property_config_name:
        log_message('logging', f'Property config name: {property_config_name}')
    
    # Enhanced Splunk logging for email configuration
    log_message('INFO', f'Email configuration - Gmail trigger: {gmailtrigger}, Multiple files: {attached_multiple_file}')
    log_message('INFO', f'Email details - From: {sender_email}, To: {recipient_emails}, Subject: {subject}')
    
    # Log startFileNames if provided
    if attached_multiple_file and multiple_files:
        start_file_names = [f.get('start_file_name', '') for f in multiple_files]
        log_message('INFO', f'Processing multiple files with start names: {start_file_names}')
        log_message('INFO', f'Multiple attachment mode enabled - processing {len(multiple_files)} files')
    elif not attached_multiple_file:
        log_message('INFO', 'Single attachment mode enabled - processing one report file')

    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = ', '.join(recipient_emails)
    msg['Cc'] = ', '.join(cc_emails) if cc_emails else None

    # Attach body text
    msg.attach(MIMEText(body_text, 'plain'))

    # Handle single attachment if flag is True
    if attachment_flag and not attached_multiple_file:
        log_message('INFO', f'Processing single attachment - Bucket: {bucket_name}, Object: {object_key}')
        if object_key:
            try:
                # Download from S3
                log_message('INFO', f'Downloading single file from S3: {object_key}')
                download_file_from_s3(bucket_name, object_key, attachment_file_path)

                # Correctly extract full file name with extension
                if report_attachment_flag:
                    file_name = object_key.split('/')[-1]
                else:
                    file_name = object_key.split('/')[-1].split('.')[-2]

                # Attach the file
                with open(attachment_file_path, 'rb') as file:
                    part = MIMEApplication(file.read(), Name=file_name)
                    part['Content-Disposition'] = f'attachment; filename="{file_name}"'
                    msg.attach(part)
                
                log_message('SUCCESS', f'Single attachment added: {file_name}')
                log_message('INFO', f'Single attachment processing completed successfully')
            except Exception as e:
                log_message('ERROR', f'Failed to attach single file: {str(e)}')
                log_message('ERROR', f'Single attachment processing Failed: {str(e)}')
        else:
            # Log a warning but continue sending email
            log_message('WARNING', 'Attachment flag is True but no object_key was provided — sending without attachment')

    # Handle multiple attachments if flag is True
    elif attached_multiple_file and multiple_files:
        try:
            log_message('INFO', f'Processing {len(multiple_files)} multiple attachments')
            log_message('INFO', f'Multiple attachment mode - Gmail trigger: {gmailtrigger}')
            
            for i, file_info in enumerate(multiple_files):
                file_bucket = file_info.get('bucket_name', bucket_name)
                file_key = file_info.get('object_key', '')
                file_path = file_info.get('file_path', f'/tmp/attachment_{i}')
                start_file_name = file_info.get('start_file_name', input_file_name)
                
                log_message('INFO', f'Processing attachment {i+1}/{len(multiple_files)} - Start file: {start_file_name}, S3 key: {file_key}')
                
                if file_key:
                    # Download from S3
                    log_message('INFO', f'Downloading multiple file {i+1} from S3: {file_key}')
                    download_file_from_s3(file_bucket, file_key, file_path)
                    
                    # Extract file name with prefix if specified
                    base_file_name = file_key.split('/')[-1]
                    if gmailtrigger:
                        # Add prefix like <startfile name> for each attached file
                    #     file_name = f"<{start_file_name}>{base_file_name}"
                        if gmailtrigger and start_file_name and not base_file_name.startswith(start_file_name):
                            file_name = f"{start_file_name}.{base_file_name}"
                        else:
                            file_name = base_file_name 
                        log_message('INFO', f'Gmail trigger enabled - adding prefix to filename: {file_name}')
                    else:
                        file_name = base_file_name
                        log_message('INFO', f'Gmail trigger disabled - using base filename: {file_name}')
                    
                    # Attach the file
                    with open(file_path, 'rb') as file:
                        part = MIMEApplication(file.read(), Name=file_name)
                        part['Content-Disposition'] = f'attachment; filename="{file_name}"'
                        msg.attach(part)
                    
                    log_message('SUCCESS', f'Multiple attachment {i+1} added: {file_name}')
                else:
                    log_message('WARNING', f'No object_key provided for attachment {i+1}')
            
            log_message('INFO', f'Multiple attachment processing completed - {len(multiple_files)} files processed')
        except Exception as e:
            log_message('ERROR', f'Failed to attach multiple files: {str(e)}')

    # Send the email
    try:
        log_message('INFO', f'Attempting to send email via SES - From: {sender_email}, To: {recipient_emails}')
        if cc_emails:
            log_message('INFO', f'CC recipients: {cc_emails}')
        
        response = ses_client.send_raw_email(
            Source=sender_email,
            Destinations=recipient_emails + cc_emails if cc_emails else recipient_emails,
            RawMessage={'Data': msg.as_string()}
        )
        
        log_message('SUCCESS', f'Email sent successfully! Message ID: {response["MessageId"]}')
        log_message('SUCCESS', f'Email delivery completed - Recipients: {recipient_emails}')
        
        # Enhanced success logging with sample configuration message
        bucket_name = os.getenv('s3BucketInput', '')
        config_key = os.getenv('configKey', '')
        if bucket_name and config_key:
            log_message('INFO', f'Successfully created sample configuration: s3://{bucket_name}/{config_key}')
        
        return f"Email sent! Message ID: {response['MessageId']}"
    except NoCredentialsError:
        error_msg = "AWS credentials not available for SES"
        log_message('ERROR', error_msg)
        return "Credentials not available"
    except ClientError as e:
        error_code = e.response['Error']['Code']
        if error_code == 'AccessDenied':
            error_msg = "SES Access Denied - IAM role lacks ses:SendRawEmail permission"
            log_message('ERROR', error_msg)
            return "SES Access Denied - Check IAM permissions"
        elif error_code == 'MessageRejected':
            error_msg = f"Email message rejected by SES: {e}"
            log_message('ERROR', error_msg)
            return "Message rejected by SES"
        else:
            error_msg = f"SES error ({error_code}): {e}"
            log_message('ERROR', error_msg)
            return f"SES error: {error_code}"
    except Exception as e:
        error_msg = f"Unexpected error sending email: {e}"
        log_message('Failed', error_msg)  # Using 'Failed' status as requested
        return f"Unexpected error: {str(e)}"

# AWS Glue utility functions
def checkFile_isPresent(bucket_name, object_key, file_name, context):
          """Check if a file exists in S3 bucket"""
          try:
                    s3.head_object(Bucket=bucket_name, Key=object_key)
                    return True
          except ClientError as e:
                    if e.response['Error']['Code'] == 'NoSuchKey':
                              log_message("ERROR", f"File {file_name} not found in the bucket {bucket_name}: {str(e)}")
                              return False
                   
                    return False
          except Exception as e:
                    
                    log_message("ERROR", f"Failed tochecking filed is pressent onfiguration: {str(e)}")
                    return False

def create_sample_config(bucket_name, config_key, file_type):
          """Create a sample configuration file for the given file type"""
          try:
                   
                    sample_config = {
                              "report_types": {
                                        file_type: {
                                                  "data_sources": {
                                                            "input_xml_path": "",
                                                            "output_path": ""
                                                  },
                                                  "report_settings": {
                                                            "file_type": file_type,
                                                            "output_format": "csv",
                                                            "include_headers": True
                                                  },
                                                  "processing": {
                                                            "extract_all_elements": True,
                                                            "create_summary": True
                                                  },
                                                  "layout": {
                                                            "header_section": {
                                                                      "enabled": True,
                                                                      "fields": ["report_name", "generation_date"]
                                                            },
                                                            "data_section": {
                                                                      "enabled": True,
                                                                      "extract_all": True
                                                            },
                                                            "footer_section": {
                                                                      "enabled": True,
                                                                      "fields": ["total_records", "generation_time"]
                                                            }
                                                  }
                                        }
                              }
                    }
                   
                    # Upload sample config to S3
                    config_content = json.dumps(sample_config, indent=2)
                    s3.put_object(
                              Bucket=bucket_name,
                              Key=config_key,
                              Body=config_content,
                              ContentType='application/json'
                    )
                   
                    log_message("INFO", f"Successfully created sample configuration: s3://{bucket_name}/{config_key}")
                    return True
                   
          except Exception as e:
                    log_message("ERROR", f"Failed to create sample configuration: {str(e)}")
                    return False

def get_run_id():
     """Generate account_id using function"""
     try:
          account_id = boto3.client("sts").get_caller_identity()["Account"]
          run_id = "arn:dpm:glue:br_icsdev_dpmjayant_glue_reporting_checks:" + account_id
          return run_id
     except Exception as e:
          log_message("failed", f"Failed to create job: {str(e)}")
          raise


def log_message(status, message, context=None, event=None):
     """Function to log messages in splunk with enhanced logging"""
     
     try:
          # Safely extract from event or fallback to environment variables
          input_file_name = event.get('inputFileName', '') if event else os.getenv('inputFileName', '')
          property_config_name = event.get('propertyConfigName', '') if event else os.getenv('propertyConfigName', '')
          
          # Enhanced Splunk logging with proper format
          log_data = {
               'Status': status,
               'Message': str(message)
          }
          
          # Add optional fields if available
          if input_file_name:
               log_data['FileName'] = input_file_name
          if property_config_name:
               log_data['PropertyConfigName'] = property_config_name
          
          
          # Use the safe splunk.log_message format
          #splunk.log_message(log_data, context or get_run_id())
          # Only call splunk.log_message if we have a proper context object
          if context and hasattr(context, 'invoked_function_arn'):
              splunk.log_message(log_data, context)
          # else:
          #     print(f"INFO: Skipping Splunk logging - no valid context provided") 
              # If no proper context, just log to console (splunk logging will be skipped)
          
     except Exception as e:
          log_message("failed", f"log_message writing: {str(e)}")
          # Don't raise here as logging failure shouldn't stop the main process


def set_job_params_as_env_vars():
     """Set job parameters as environment variables for Lambda environment"""
     try:
          
          # In Lambda environment, we don't have sys.argv, so we'll set some default values
          # or use environment variables that are already set
          default_params = {
               'JOB_NAME': os.getenv('JOB_NAME', 'lambda_report_checks'),
               'JOB_RUN_ID': os.getenv('JOB_RUN_ID', 'lambda_run'),
               'JOB_ID': os.getenv('JOB_ID', 'lambda_job')
          }
          
          for key, value in default_params.items():
               if not os.getenv(key):
                    os.environ[key] = value
          
     except Exception as e:
          log_message("failed", f"Failed to read ENV file from S3: {str(e)}")
          # Don't raise here as this is not critical for Lambda execution

class ExcelFormatter:
     """Formats data as Excel-compatible text with proper column alignment"""

     def __init__(self):
          self.tab_char = '\t'# Tab character for Excel column separation

     def format_currency(self, value: float) -> str:
          """Format value as currency"""
          try:
                   return f"{float(value):.2f}"
          except:
                   return "0.00"

     def clean_masked_value(self, value: str) -> str:
          """Clean masked values (remove asterisks and extract numeric part)"""
          if not value:
                   return "0.00"
         
          # Remove asterisks and extract numeric part
          cleaned = str(value).replace('*', '')
         
          # If the cleaned value is empty or just whitespace, return 0.00
          if not cleaned.strip():
                   return "0.00"
         
          return cleaned

     def safe_float_convert(self, value: str, default: float = 0.0) -> float:
          """Safely convert a value to float, handling masked values"""
          if not value:
                   return default
         
          # Clean masked values first
          cleaned = self.clean_masked_value(value)
         
          try:
                   return float(cleaned)
          except (ValueError, TypeError):
                   return default

     def format_number(self, value: Any, width: int = 0) -> str:
          """Format number with specified width"""
          try:
                   num = float(value)
                   if width > 0:
                        return f"{num:>{width}.2f}"
                   return f"{num:.2f}"
          except:
                   return "0.00"

     def align_text(self, text: str, width: int, align: str = 'left') -> str:
          """Align text within specified width"""
          if align == 'right':
                   return text.rjust(width)
          elif align == 'center':
                   return text.center(width)
          else:# left
                   return text.ljust(width)

     def create_excel_row(self, columns: List[Dict], data: Dict) -> str:
          """Create an Excel-formatted row"""
          row_parts = []

          for col in columns:
                   value = self._get_column_value(col, data)
                   formatted_value = self._format_column_value(col, value)
                   row_parts.append(formatted_value)

          return self.tab_char.join(row_parts)

     def _get_column_value(self, col: Dict, data: Dict) -> Any:
          """Get value for a column from data"""
          source_field = col.get('source_field', '')

          if source_field == 'static':
                   return col.get('value', '')
          elif source_field == 'calculated':
                   return self._calculate_value(col, data)
          else:
                   value = data.get(source_field, '')
                   # For CUSTOM_FUND_CODE, pad with leading zeros to 7 characters
                   if source_field == 'CUSTOM_FUND_CODE' and value:
                        return value.zfill(7)
                   # For CUSTOM_CHECK_AMOUNT, clean masked values
                   elif source_field == 'CUSTOM_CHECK_AMOUNT' and value:
                        return self.clean_masked_value(value)
                   return value

     def _calculate_value(self, col: Dict, data: Dict) -> Any:
          """Calculate value based on column configuration"""
          calculation = col.get('calculation', '')

          if calculation == 'count_records':
                   return 1# Will be summed up
          elif calculation == 'sum_amounts':
                   # For Redemption_SWP reports, use CUSTOM_CHECK_AMOUNT
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'sum_sheets':
                   return float(data.get('DOC_SHEET_COUNT', 0))
          elif calculation == 'sum_images':
                   return float(data.get('DOC_IMAGE_COUNT', 0))
          elif calculation == 'sum_field':
                   field_name = col.get('field', '')
                   return float(data.get(field_name, 0))
          elif calculation == 'count_email_delivery':
                   return 1 if data.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'E' else 0
          elif calculation == 'count_print_delivery':
                   return 1 if data.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'P' else 0
          elif calculation == 'count_suppressed':
                   return 1 if data.get('DOC_SUPPRESSION_FLAG') == 'Y' else 0
          elif calculation == 'count_suppressed_amount':
                   # Count records with amount < $5.00
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   amount = self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 1 if amount < 5.00 else 0
          elif calculation == 'count_zero_check':
                   # Count records with zero check number
                   check_number = data.get('FN_CHECK_NUMBER', '')
                   return 1 if check_number == '0' or check_number == '' else 0
          elif calculation == 'count_suppressed_total':
                   # Count total suppressed records
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   amount = self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   check_number = data.get('CUSTOM_CHECK_NUMBER', '')
                   return 1 if amount < 5.00 or check_number == '0' or check_number == '' else 0
          elif calculation == 'count_packages':
                   return 1# Each record is a package
          elif calculation == 'count_check_pages':
                   return 2# Each check has 2 pages
          elif calculation == 'extract_group':
                   # Extract group from account number or other field
                   account = data.get('UC_ACCOUNT_NUMBER', '')
                   if len(account) >= 7:
                        return account[:7]
                   return account
          elif calculation == 'count_redemption_records':
                   # Count redemption records based on FN_CHECK_TYPE_DESC
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   return 1 if 'REDEMPTION' in check_type.upper() else 0
          elif calculation == 'sum_redemption_amounts':
                   # Sum amounts for redemption records
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   if 'REDEMPTION' in check_type.upper():
                        return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 0
          elif calculation == 'count_swp_records':
                   # Count SWP records based on FN_CHECK_TYPE_DESC
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   return 1 if 'SWP' in check_type.upper() else 0
          elif calculation == 'sum_swp_amounts':
                   # Sum amounts for SWP records
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   if 'SWP' in check_type.upper():
                        return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 0
          elif calculation == 'concatenate_fund_code_name':
                   # Concatenate CUSTOM_FUND_CODE and CUSTOM_FUND_DESC
                   fund_code = data.get('CUSTOM_FUND_CODE', '')
                   fund_desc = data.get('CUSTOM_FUND_DESC', '')
                   if fund_code and fund_desc:
                        return f"{fund_code} {fund_desc}"
                   elif fund_code:
                        return fund_code
                   elif fund_desc:
                        return fund_desc
                   return ""
          elif calculation == 'red_begin_check_number':
                   # Get the first RED check number for this fund
                   return self._get_red_begin_check_number(data)
          elif calculation == 'red_end_check_number':
                   # Get the last RED check number for this fund
                   return self._get_red_end_check_number(data)
          elif calculation == 'red_check_count':
                   # Count RED records for this fund
                   return self._get_red_check_count(data)
          elif calculation == 'total_red_dollar_value':
                   # Sum RED amounts for this fund
                   return self._get_total_red_dollar_value(data)
          elif calculation == 'swp_begin_check_number':
                   # Get the first SWP check number for this fund
                   return self._get_swp_begin_check_number(data)
          elif calculation == 'swp_end_check_number':
                   # Get the last SWP check number for this fund
                   return self._get_swp_end_check_number(data)
          elif calculation == 'swp_check_count':
                   # Count SWP records for this fund
                   return self._get_swp_check_count(data)
          elif calculation == 'total_swp_dollar_value':
                   # Sum SWP amounts for this fund
                   return self._get_total_swp_dollar_value(data)
          elif calculation == 'calculate_total_input':
                   # Count total number of input rows in the report
                   return 1 # Will be summed up for each record
          elif calculation == 'calculate_total_packages':
                   # Count total number of package entries (rows)
                   return 1 # Each record is a package
          elif calculation == 'calculate_total_check_page':
                   # Sum of Total Input + Total Packages
                   return 2 # Total Input (1) + Total Packages (1) = 2
          elif calculation == 'calculate_total_dollar_amount':
                   # Sum of all check amounts present in the report
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'calculate_total_input':
                   # Count total number of input rows in the report
                   return 1 # Will be summed up for each record
          elif calculation == 'calculate_total_packages':
                   # Count total number of package entries (rows)
                   return 1 # Each record is a package
          elif calculation == 'calculate_total_check_page':
                   # Sum of Total Input + Total Packages
                   return 2 # Total Input (1) + Total Packages (1) = 2
          elif calculation == 'calculate_total_dollar_amount':
                   # Sum of all check amounts present in the report
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'static':
                   return col.get('value', 0)
          else:
                   return 0

     def _get_red_begin_check_number(self, data: Dict) -> str:
          """Get the first RED check number for this fund"""
          return data.get('RED_BEGIN_CHECK_NUMBER', '')

     def _get_red_end_check_number(self, data: Dict) -> str:
          """Get the last RED check number for this fund"""
          return data.get('RED_END_CHECK_NUMBER', '')

     def _get_red_check_count(self, data: Dict) -> int:
          """Count RED records for this fund"""
          return data.get('RED_CHECK_COUNT', 0)

     def _get_total_red_dollar_value(self, data: Dict) -> float:
          """Sum RED amounts for this fund"""
          return data.get('TOTAL_RED_DOLLAR_VALUE', 0.0)

     def _get_swp_begin_check_number(self, data: Dict) -> str:
          """Get the first SWP check number for this fund"""
          return data.get('SWP_BEGIN_CHECK_NUMBER', '')

     def _get_swp_end_check_number(self, data: Dict) -> str:
          """Get the last SWP check number for this fund"""
          return data.get('SWP_END_CHECK_NUMBER', '')

     def _get_swp_check_count(self, data: Dict) -> int:
          """Count SWP records for this fund"""
          return data.get('SWP_CHECK_COUNT', 0)

     def _get_total_swp_dollar_value(self, data: Dict) -> float:
          """Sum SWP amounts for this fund"""
          return data.get('TOTAL_SWP_DOLLAR_VALUE', 0.0)

     def _format_column_value(self, col: Dict, value: Any) -> str:
          """Format column value according to column configuration"""
          width = col.get('width', 0)
          align = col.get('align', 'left')
          format_type = col.get('format', 'text')

          if format_type == 'currency':
                   formatted = self.format_currency(value)
          elif format_type == 'number':
                   formatted = self.format_number(value, width)
          else:
                   formatted = str(value)

          if width > 0:
                   return self.align_text(formatted, width, align)
          else:
                   return formatted

class CSVFormatter:
     """Formats data as CSV with proper column alignment and headers"""

     def __init__(self, config: Dict):
          self.config = config.get('export_settings', {}).get('csv', {})
          self.delimiter = self.config.get('delimiter', ',')
          self.quote_char = self.config.get('quote_char', '"')
          self.encoding = self.config.get('encoding', 'utf-8')
          self.include_headers = self.config.get('include_headers', True)

     def format_currency(self, value: float) -> str:
          """Format value as currency"""
          try:
                   return f"{float(value):.2f}"
          except:
                   return "0.00"

     def clean_masked_value(self, value: str) -> str:
          """Clean masked values (remove asterisks and extract numeric part)"""
          if not value:
                   return "0.00"
         
          # Remove asterisks and extract numeric part
          cleaned = str(value).replace('*', '')
         
          # If the cleaned value is empty or just whitespace, return 0.00
          if not cleaned.strip():
                   return "0.00"
         
          return cleaned

     def safe_float_convert(self, value: str, default: float = 0.0) -> float:
          """Safely convert a value to float, handling masked values"""
          if not value:
                   return default
         
          # Clean masked values first
          cleaned = self.clean_masked_value(value)
         
          try:
                   return float(cleaned)
          except (ValueError, TypeError):
                   return default

     def create_csv_headers(self, columns: List[Dict]) -> List[str]:
          """Create CSV headers from column configuration with single column layout and consistent alignment"""
          headers = []
          for col in columns:
                   header = col.get('csv_header', col.get('name', ''))
                   # Single column layout - labels and values in one column with consistent alignment
                   headers.append(header)
          return headers

     def create_csv_row(self, columns: List[Dict], data: Dict) -> List[str]:
          """Create a CSV row from data with single column layout and consistent alignment"""
          row_values = []

          for col in columns:
                   value = self._get_column_value(col, data)
                   formatted_value = self._format_column_value(col, value)
              
                   row_values.append(formatted_value)

          return row_values

     def _get_column_value(self, col: Dict, data: Dict) -> Any:
          """Get value for a column from data"""
          source_field = col.get('source_field', '')

          if source_field == 'static':
                   return col.get('value', '')
          elif source_field == 'calculated':
                   return self._calculate_value(col, data)
          else:
                   value = data.get(source_field, '')
                   # For CUSTOM_FUND_CODE, pad with leading zeros to 7 characters
                   if source_field == 'CUSTOM_FUND_CODE' and value:
                        return value.zfill(7)
                   # For CUSTOM_CHECK_AMOUNT, clean masked values
                   elif source_field == 'CUSTOM_CHECK_AMOUNT' and value:
                        return self.clean_masked_value(value)
                   return value

     def _calculate_value(self, col: Dict, data: Dict) -> Any:
          """Calculate value based on column configuration"""
          calculation = col.get('calculation', '')

          if calculation == 'count_records':
                   return 1# Will be summed up
          elif calculation == 'sum_amounts':
                   # For Redemption_SWP reports, use CUSTOM_CHECK_AMOUNT
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'sum_sheets':
                   return float(data.get('DOC_SHEET_COUNT', 0))
          elif calculation == 'sum_images':
                   return float(data.get('DOC_IMAGE_COUNT', 0))
          elif calculation == 'sum_field':
                   field_name = col.get('field', '')
                   return float(data.get(field_name, 0))
          elif calculation == 'count_email_delivery':
                   return 1 if data.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'E' else 0
          elif calculation == 'count_print_delivery':
                   return 1 if data.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'P' else 0
          elif calculation == 'count_suppressed':
                   return 1 if data.get('DOC_SUPPRESSION_FLAG') == 'Y' else 0
          elif calculation == 'count_suppressed_amount':
                   # Count records with amount < $5.00
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   amount = self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 1 if amount < 5.00 else 0
          elif calculation == 'count_zero_check':
                   # Count records with zero check number
                   check_number = data.get('FN_CHECK_NUMBER', '')
                   return 1 if check_number == '0' or check_number == '' else 0
          elif calculation == 'count_suppressed_total':
                   # Count total suppressed records
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   amount = self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   check_number = data.get('CUSTOM_CHECK_NUMBER', '')
                   return 1 if amount < 5.00 or check_number == '0' or check_number == '' else 0
          elif calculation == 'count_packages':
                   return 1# Each record is a package
          elif calculation == 'count_check_pages':
                   return 2# Each check has 2 pages
          elif calculation == 'extract_group':
                   # Extract group from account number or other field
                   account = data.get('UC_ACCOUNT_NUMBER', '')
                   if len(account) >= 7:
                        return account[:7]
                   return account
          elif calculation == 'count_redemption_records':
                   # Count redemption records based on FN_CHECK_TYPE_DESC
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   return 1 if 'REDEMPTION' in check_type.upper() else 0
          elif calculation == 'sum_redemption_amounts':
                   # Sum amounts for redemption records
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   if 'REDEMPTION' in check_type.upper():
                        return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 0
          elif calculation == 'count_swp_records':
                   # Count SWP records based on FN_CHECK_TYPE_DESC
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   return 1 if 'SWP' in check_type.upper() else 0
          elif calculation == 'sum_swp_amounts':
                   # Sum amounts for SWP records
                   check_type = data.get('CUSTOM_CHECK_TYPE', '')
                   if 'SWP' in check_type.upper():
                        return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
                   return 0
          elif calculation == 'concatenate_fund_code_name':
                   # Concatenate CUSTOM_FUND_CODE and CUSTOM_FUND_DESC
                   fund_code = data.get('CUSTOM_FUND_CODE', '')
                   fund_desc = data.get('CUSTOM_FUND_DESC', '')
                   if fund_code and fund_desc:
                        return f"{fund_code} {fund_desc}"
                   elif fund_code:
                        return fund_code
                   elif fund_desc:
                        return fund_desc
                   return ""
          elif calculation == 'calculate_total_input':
                   # Count total number of input rows in the report
                   return 1 # Will be summed up for each record
          elif calculation == 'calculate_total_packages':
                   # Count total number of package entries (rows)
                   return 1 # Each record is a package
          elif calculation == 'calculate_total_check_page':
                   # Sum of Total Input + Total Packages
                   return 2 # Total Input (1) + Total Packages (1) = 2
          elif calculation == 'calculate_total_dollar_amount':
                   # Sum of all check amounts present in the report
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'static':
                   return col.get('value', 0)
          else:
                   return 0

     def _get_red_begin_check_number(self, data: Dict) -> str:
          """Get the first RED check number for this fund"""
          return data.get('RED_BEGIN_CHECK_NUMBER', '')

     def _get_red_end_check_number(self, data: Dict) -> str:
          """Get the last RED check number for this fund"""
          return data.get('RED_END_CHECK_NUMBER', '')

     def _get_red_check_count(self, data: Dict) -> int:
          """Count RED records for this fund"""
          return data.get('RED_CHECK_COUNT', 0)

     def _get_total_red_dollar_value(self, data: Dict) -> float:
          """Sum RED amounts for this fund"""
          return data.get('TOTAL_RED_DOLLAR_VALUE', 0.0)

     def _get_swp_begin_check_number(self, data: Dict) -> str:
          """Get the first SWP check number for this fund"""
          return data.get('SWP_BEGIN_CHECK_NUMBER', '')

     def _get_swp_end_check_number(self, data: Dict) -> str:
          """Get the last SWP check number for this fund"""
          return data.get('SWP_END_CHECK_NUMBER', '')

     def _get_swp_check_count(self, data: Dict) -> int:
          """Count SWP records for this fund"""
          return data.get('SWP_CHECK_COUNT', 0)

     def _get_total_swp_dollar_value(self, data: Dict) -> float:
          """Sum SWP amounts for this fund"""
          return data.get('TOTAL_SWP_DOLLAR_VALUE', 0.0)

     def _format_column_value(self, col: Dict, value: Any) -> str:
          """Format column value according to column configuration"""
          format_type = col.get('format', 'text')

          if format_type == 'currency':
                   return self.format_currency(value)
          else:
                   return str(value)

class XMLDataExtractor:
     """Extracts data from XML files based on configuration"""

     def __init__(self):
          pass

     def extract_data(self, xml_file_path: str, config: Dict) -> List[Dict]:
          """Extract data from XML file based on configuration"""
          try:
                    # Try multiple encoding strategies
                    xml_content = None
                    # Parse S3 path
                    if xml_file_path.startswith('s3://'):
                              # Extract bucket and key from S3 URL
                              s3_path_parts = xml_file_path[5:].split('/', 1)
                              bucket = s3_path_parts[0]
                              key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
                    else:
                              # Use environment variables for S3 bucket and key
                              bucket = os.getenv('s3BucketInput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
                              key = xml_file_path

                    # Read file from S3
                    try:
                              response = s3.get_object(Bucket=bucket, Key=key)
                              raw_content = response['Body'].read()
                              log_message("INFO", f"Successfully read XML file from S3: s3://{bucket}/{key}")
                    except ClientError as e:
                              log_message("ERROR", f"Failed to read XML file from S3: {str(e)}")
                              return []

                    # Remove BOM if present
                    if raw_content.startswith(b'\xef\xbb\xbf'):
                              raw_content = raw_content[3:]
                    elif raw_content.startswith(b'\xff\xfe'):
                              raw_content = raw_content[2:]
                    elif raw_content.startswith(b'\xfe\xff'):
                              raw_content = raw_content[2:]

                    # Fix non-breaking spaces (common issue)
                    if b'\xc2\xa0' in raw_content[:200]:
                              raw_content = raw_content.replace(b'\xc2\xa0', b' ')

                    # Try to decode with different encodings
                    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                    for encoding in encodings:
                              try:
                                        xml_content = raw_content.decode(encoding)
                                        break
                              except UnicodeDecodeError:
                                        continue

                    if xml_content is None:
                              return []

                    # Clean the content
                    xml_content = xml_content.lstrip('\ufeff\ufffe\u200b')     # Remove BOM characters
                    xml_content = xml_content.replace('\xa0', ' ')                         # Replace non-breaking spaces
                    xml_content = xml_content.replace('\u00a0', ' ')
                   
                    # Remove control characters (except \n, \r, \t)
                    import re
                    xml_content = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]', '', xml_content)

                    # Strip leading/trailing whitespace
                    xml_content = xml_content.strip()

                    # Validate XML structure
                    if not xml_content.startswith('<?xml') and not xml_content.startswith('<'):
                              return []

                    # Parse the cleaned XML content
                    try:
                              root = ET.fromstring(xml_content)
                    except ET.ParseError as e:

                              # Try to fix common XML issues
                              try:
                                        # Try to add missing closing tags
                                        if '<Statements>' in xml_content and '</Statements>' not in xml_content:
                                                  xml_content = xml_content.replace('</File>', '</Statements>\n</File>')

                                        # Try parsing again
                                        root = ET.fromstring(xml_content)
                              except ET.ParseError as e2:
                                        return []

                    statements = []
                    # statement_path = config['xml_mapping']['statement_path']
                    statement_path = config.get('xml_mapping', {}).get('statement_path', './/Statement')



                    # Find all statement elements
                    statement_elements = root.findall(statement_path)

                    if len(statement_elements) == 0:
                              # Try alternative paths
                              alt_paths = ['Statement', '//Statement', './/Statement']
                              for alt_path in alt_paths:
                                        alt_elements = root.findall(alt_path)
                                        if len(alt_elements) > 0:
                                                  statement_elements = alt_elements
                                                  break

                    # Extract data for each statement
                    for i, statement in enumerate(statement_elements):
                              data = {}
                              #properties_path = config['xml_mapping']['properties_path']
                              properties_path = config.get('xml_mapping', {}).get('properties_path', 'Property')


                              # Find all property elements within this statement
                              prop_elements = statement.findall(properties_path)

                              for prop in prop_elements:
                                        name = prop.get('Name')
                                        value = prop.get('Value')
                                        if name:
                                                  data[name] = value

                              # Add data if found
                              if data:
                                        statements.append(data)
                    return statements

          except Exception as e:
                    import traceback
                    traceback.print_exc()
                    return []


     def extract_tilde_separated_data(self, file_path: str, config: Dict) -> List[Dict]:
          """Extract data from tilde-separated file based on configuration"""
          try:
                    # Parse S3 path
                    if file_path.startswith('s3://'):
                              # Extract bucket and key from S3 URL
                              s3_path_parts = file_path[5:].split('/', 1)
                              bucket = s3_path_parts[0]
                              key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
                    else:
                              # Use environment variables for S3 bucket and key
                              bucket = os.getenv('s3BucketInput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
                              key = file_path

                    # Read file from S3
                    try:
                              response = s3.get_object(Bucket=bucket, Key=key)
                              raw_content = response['Body'].read()
                              log_message("INFO", f"Successfully read tilde-separated file from S3: s3://{bucket}/{key}")
                    except ClientError as e:
                              log_message("ERROR", f"Failed to read tilde-separated file from S3: {str(e)}")
                              return []

                    # Decode content
                    try:
                              content = raw_content.decode('utf-8')
                    except UnicodeDecodeError:
                              try:
                                        content = raw_content.decode('latin-1')
                              except UnicodeDecodeError:
                                        content = raw_content.decode('cp1252')

                    # Parse lines for tilde-separated data extraction

                    # Parse lines
                    lines = content.strip().split('\n')
                    data = []
                    
                    # Get field mapping from config
                    field_mapping = config.get('input_mapping', {}).get('field_mapping', {})
                    delimiter = config.get('input_mapping', {}).get('delimiter', '~')
                    
                    for line_num, line in enumerate(lines, 1):
                              if not line.strip():
                                        continue
                                        
                              # Split by delimiter
                              fields = line.split(delimiter)
                              
                              # Create record based on field mapping
                              record = {}
                              for field_name, field_index in field_mapping.items():
                                        if field_index == "current_month_year":
                                              # Generate current month/year in MM/YYYY format
                                              from datetime import datetime
                                              current_date = datetime.now()
                                              record[field_name] = f"{current_date.month:02d}/{current_date.year}"
                                        elif isinstance(field_index, int) and field_index < len(fields):
                                              field_value = fields[field_index].strip()
                                              # For Fund field (12b1_fund_report), only include numeric values
                                              # For fund field (12b1_check_report), include all values
                                              if field_name == "Fund":
                                                    # Check if the value is numeric (contains only digits)
                                                    if field_value.isdigit():
                                                          record[field_name] = field_value
                                                    else:
                                                          # For combined extraction, don't skip the record entirely
                                                          # Just set the Fund field to empty and continue
                                                          record[field_name] = ""
                                              elif field_name == "fund":
                                                    # For 12b1_check_report fund field, include all values
                                                    record[field_name] = field_value
                                              else:
                                                    record[field_name] = field_value
                                        else:
                                              record[field_name] = ""
                              
                              if record:
                                        data.append(record)
                                        log_message("INFO", f"Processed line {line_num}: {record}")
                              else:
                                        log_message("INFO", f"Extracted {len(data)} records from tilde-separated file")
                    if data:
                         return data

          except Exception as e:
                    log_message("ERROR", f"Error extracting data from tilde-separated file {file_path}: {e}")
                    import traceback
                    traceback.print_exc()
                    return []

     def filter_data(self, data: List[Dict], filters: Dict) -> List[Dict]:
          """Filter data based on criteria"""
          if not filters:
                   return data

          filtered = []
          for record in data:
                   match = True
                   for field, expected_value in filters.items():
                        if record.get(field) != expected_value:
                             match = False
                             break
                   if match:
                        filtered.append(record)

          return filtered

class ReportGenerator:
     """Generates reports based on JSON configuration"""
     
     # Class-level flag to track if email has been sent for redemption_swp in current execution
     _redemption_swp_email_sent = False

     def __init__(self):  
          self.formatter = ExcelFormatter()
          self.extractor = XMLDataExtractor()
          # Dynamic mapping table for DOC_SPECIAL_HANDLING_CODE to SHCode
          self.special_handling_mapping = {}
          # Store all records for Redemption_SWP record-level calculations
          self.all_records = []
         
          # # Dynamic mapping table for DOC_SPECIAL_HANDLING_CODE to SHCode
          # # This will be loaded from configuration for Redemption_SWP reports
          # self.special_handling_mapping = {}

     def safe_float_convert(self, value: str, default: float = 0.0) -> float:
          """Safely convert a value to float, handling masked values"""
          if not value:
                   return default
         
          # Clean masked values first
          cleaned = str(value).replace('*', '')
         
          try:
                   return float(cleaned)
          except (ValueError, TypeError):
                   return default

     def clean_masked_value(self, value: str) -> str:
          """Clean masked values (remove asterisks and extract numeric part)"""
          if not value:
                   return "0.00"
         
          # Remove asterisks and extract numeric part
          cleaned = str(value).replace('*', '')
         
          # If the cleaned value is empty or just whitespace, return 0.00
          if not cleaned.strip():
                   return "0.00"
         
          return cleaned

     def format_currency(self, value: float) -> str:
          """Format value as currency"""
          try:
                   return f"{float(value):.2f}"
          except:
                   return "0.00"
         
     def load_config(self, config_file_path: str, file_type: str) -> Dict:
          """Load specific report configuration from unified config file"""
          try:
                    # Use S3 path for configuration
                    if config_file_path.startswith('s3://'):
                              # Extract bucket and key from S3 URL
                              s3_path_parts = config_file_path[5:].split('/', 1)
                              bucket = s3_path_parts[0]
                              key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
                    else:
                              # Use environment variables for S3 bucket and key
                              bucket = os.getenv('s3BucketInput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
                              key = os.getenv('configKey', 'jhi/checks/config/jhi_checks_reporting_config.json')

                    # Read configuration from S3
                    try:
                              response = s3.get_object(Bucket=bucket, Key=key)
                              config_content = response['Body'].read().decode('utf-8')
                              full_config = json.loads(config_content)
                              log_message("INFO", f"Successfully loaded configuration from S3: s3://{bucket}/{key}")
                    except ClientError as e:
                              log_message("ERROR", f"Failed to read configuration from S3: {str(e)}")
                             
                              # Try alternative configuration paths
                              alternative_configs = [
                                        f"jhi/checks/config/jhi_checks_reporting_config.json",
                                        f"jhi/checks/config/{file_type}_config.json",
                                        f"jhi/checks/config/default_config.json"
                              ]
                             
                              for alt_key in alternative_configs:
                                        try:
                                                  response = s3.get_object(Bucket=bucket, Key=alt_key)
                                                  config_content = response['Body'].read().decode('utf-8')
                                                  full_config = json.loads(config_content)
                                                  log_message("INFO", f"Successfully loaded alternative configuration: s3://{bucket}/{alt_key}")
                                                  break
                                        except ClientError as alt_e:
                                                  continue
                              else:
                                        log_message("ERROR", "No configuration files found in any alternative locations")
                                        return {}

                    config = full_config.get('report_types', {}).get(file_type, {})

                    # Load special handling mapping for Redemption_SWP reports
                    if file_type == 'redemption_swp' and 'layout' in config:
                              footer_config = config['layout'].get('footer_section', {})
                              if 'special_handling_mapping' in footer_config:
                                        self.special_handling_mapping = footer_config['special_handling_mapping']

                    return config

          except Exception as e:
                    return {}

     def generate_report(self, config_file_path: str) -> bool:
          """Generate report based on configuration file"""
          config = self.load_config(config_file_path)
          if not config:
                   return False

          return self.generate_report_from_config(config)
     
     def generate_report_from_config(self, config: Dict) -> bool:
            """Generate report based on configuration dictionary"""

            # Construct input XML path from environment
            bucket = os.getenv("s3BucketInput")
            input_path = os.getenv("inputPath")
            input_file_name = os.getenv("inputFileName")
            xml_path = f"s3://{bucket}/{input_path}{input_file_name}"


            # Check if this is a 12b1_compensation report that needs special handling FIRST
            report_type = config.get('workflow_metadata', {}).get('report_type')
            if report_type == '12b1_compensation':
                # For 12b1_compensation, skip the initial data extraction and use combined extraction
                return self._generate_12b1_compensation_separate_files(config, [], input_file_name, bucket)

            # Extract data - check if this is a tilde-separated file
            # For other reports, check if they use tilde-separated data
            use_tilde_separated = False
            tilde_config = None
            
            if config.get('input_file_type') == 'tilde_separated':
                use_tilde_separated = True
                tilde_config = config
          
            if use_tilde_separated:
                # Use tilde-separated data extraction
                input_file_path = tilde_config.get('input_mapping', {}).get('file_path', input_file_name)
                if not input_file_path.startswith('s3://'):
                    # For tilde-separated files, use the path directly without prepending input_path
                    # since the config already contains the full relative path
                    input_file_path = f"s3://{bucket}/{input_file_path}"
                data = self.extractor.extract_tilde_separated_data(input_file_path, tilde_config)
            else:
                # Use XML data extraction
                data = self.extractor.extract_data(xml_path, config)  
            
            if not data:
                return False

            # Determine output formats
            output_formats_config = config.get("workflow_metadata", {}).get(
                "output_formats", {"txt": {"is_type": True}}
            )
            active_formats = [fmt for fmt, settings in output_formats_config.items() if settings.get("is_type", False)]
            report_type = config.get('workflow_metadata', {}).get('report_type', 'unknown')
            if report_type != 'money_market_CWR':
                if not active_formats:
                    

                    return False

            success_count = 0
            total_formats = len(active_formats)
            # Track output paths for email attachment - especially important for redemption_swp
            output_paths = []
            primary_output_path = None  # For redemption_swp, use the txt format path

            for format_type in active_formats:
                format_settings = output_formats_config.get(format_type, {})

                report_postfix = format_settings.get("report_postfix", "_report")
                replace_ext = format_settings.get("replace_existing_ext", True)

                # Construct output file name
                
                if replace_ext:
                    base_name = format_settings.get("out_put_name_file")
                    #index_value=format_settings.get("corp_index_value")
                    #corp_value=input_file_name.split('_')[index_value]
                    #print(corp_value,'corp_value corp_value corp_valuecorp_valuecorp_value')
                    # if base_name:
                    #      # 1. Replace {corp} placeholder
                    #      if "{corp}" in base_name:
                    #           base_name = base_name.replace("{corp}", corp_value)
                    # Replace MMDDYYYY with today's date if present
                    data_time_formate=base_name.split('.')[-1]
                    if base_name and data_time_formate in base_name:
                         current_date = datetime.now().strftime(data_time_formate)
                         base_name = base_name.replace(data_time_formate, current_date)

                    # Build output name with postfix + format
                    output_file_name = f"{base_name}{report_postfix}.{format_type}"
                else:
                    base_name = os.path.splitext(input_file_name)[0]
                    output_file_name = f"{input_file_name}{report_postfix}.{format_type}"

                # Construct full output path
                report_path = os.getenv("reportPath", "jhi/checks/wip/wipid/report/")
                output_path = f"s3://{bucket}/{report_path}{output_file_name}"
                
                # Store output path for redemption_swp (use txt format as primary)
                report_type = config.get('workflow_metadata', {}).get('report_type', 'unknown')
                if report_type == 'redemption_swp' and format_type == 'txt':
                    primary_output_path = output_path
                output_paths.append(output_path)


                # Generate report
                if format_type == "txt":
                    success = self._generate_txt_report(config, data, output_path, format_settings)
                elif format_type == "csv":
                    success = self._generate_csv_report(config, data, output_path, format_settings)
                elif format_type == "xlsx":
                    success = self._generate_xlsx_report(config, data, output_path, format_settings)
                else:
                    continue

                if success:
                    success_count += 1

            # Send email if all reports are successfully generated (excluding 12b1_compensation which handles its own email)
            report_type = config.get('workflow_metadata', {}).get('report_type', 'unknown')
            should_send_email = (success_count == total_formats) and (report_type != '12b1_compensation')
            
            # For redemption_swp, ensure we use the primary output path (txt format)
            if report_type == 'redemption_swp' and primary_output_path:
                output_path = primary_output_path
            
            # Prevent duplicate email sending for redemption_swp reports
            if report_type == 'redemption_swp' and ReportGenerator._redemption_swp_email_sent:
                log_message("INFO", "Redemption_SWP email already sent in this execution - skipping duplicate email")
                should_send_email = False
            
            # Special handling for money_market_CWR - send email when all output formats are disabled
            if report_type == 'money_market_CWR':
                # Check if all output formats are disabled (is_type: false)
                output_formats_config = config.get("workflow_metadata", {}).get("output_formats", {})
                all_formats_disabled = all(not fmt_config.get("is_type", False) for fmt_config in output_formats_config.values())
                
                if all_formats_disabled:
                    # When all formats are disabled, send email if we have data (no report files generated)
                    # if len(data) == 0:
                    should_send_email = True
                    log_message('INFO', f'money_market_CWR: All output formats disabled, sending email with data ({len(data)} records)')
                    # else:
                    #     report_type = config.get('workflow_metadata', {}).get('report_type', 'unknown')
                    #     should_send_email = False
                    #     log_message('WARNING', f'money_market_CWR: All formats disabled but no data found, skipping email')
                    #     print(f"money_market_CWR: All formats disabled but no data, skipping email")
                else:
                    # Some formats are enabled, use normal logic
                    should_send_email = (success_count == total_formats)
                    log_message('INFO', f'money_market_CWR: Some formats enabled, using normal email logic')
            
            # Special logging for 12b1_compensation - email is handled separately
            if report_type == '12b1_compensation':
                log_message('INFO', f'12b1_compensation: Email handling is done separately after all three reports are generated')
            
            if should_send_email:
                email_config = config.get("emailConfig", {})
                from_email = email_config.get("fromEmail")
                to_emails = email_config.get("toEmails", [])
                current_date = datetime.now().strftime("%m%d%Y")
                subject = email_config.get("subject", "Report Generation Successful")
                subject_with_date = subject.replace("{date}", current_date)
                
                # Special debugging for 12b1_compensation
                report_type = config.get('workflow_metadata', {}).get('report_type', 'unknown')
                if report_type == '12b1_compensation':
                    log_message('INFO', f'12b1_compensation: Email config - from_email={from_email}, to_emails={to_emails}, subject={subject_with_date}')
                
                # Special handling for money_market_CWR - extract TLE properties and generate custom email body
                if report_type == 'money_market_CWR':
                    log_message('INFO', f'money_market_CWR: Processing TLE property extraction for email body')
                    
                    # Extract TLE property counts from the already extracted data
                    check_slip_count = 0
                    investment_slip_count = 0
                    
                    # Use the data that was already extracted (no need to re-extract)
                    if len(data) > 0:
                        for statement in data:
                            # Look for TLE properties with CUSTOM_CHECK_SUBTYPE
                            if 'CUSTOM_CHECK_SUBTYPE' in statement:
                                subtype_value = statement['CUSTOM_CHECK_SUBTYPE']
                                if subtype_value == 'CHECK_SLIP':
                                    check_slip_count += 1
                                elif subtype_value == 'INVESTMENT_SLIP':
                                    investment_slip_count += 1
                        total_slip_count = check_slip_count + investment_slip_count
                        log_message('INFO', f'money_market_CWR: TLE property counts - CHECK_SLIP: {check_slip_count}, INVESTMENT_SLIP: {investment_slip_count}')
                        total_packages = check_slip_count + investment_slip_count
                        # Generate custom email body for money_market_CWR
                    #     email_body_template = email_config.get("email_body_template", "Check Packages - {CHECK_SLIP_COUNT}\nInvestment Slip Packages - {INVESTMENT_SLIP_COUNT}")
                        email_body_template = email_config.get("email_body_template", "Check Packages - {CHECK_SLIP_COUNT}\n""Investment Slip Packages - {INVESTMENT_SLIP_COUNT}\n""Total Packages = {TOTAL_SLIP_COUNT}")
                        body_text = email_body_template.format(
                            CHECK_SLIP_COUNT=check_slip_count,
                            INVESTMENT_SLIP_COUNT=investment_slip_count,
                            TOTAL_SLIP_COUNT=total_slip_count
                        )
                        
                        log_message('INFO', f'money_market_CWR: Generated email body: {body_text}')
                    else:
                        # No data available - use default message
                        report_type = config.get('workflow_metadata', {}).get('report_type', 'unknown')
                        body_text = "No relevant data found in the input file."
                        log_message('WARNING', f'money_market_CWR: No data available for TLE property extraction')
                else:
                    # Default email body for other report types
                    body_text = f"""embed above text file"""

                

                if from_email and to_emails:
                    
                    # Get email configuration options first
                    gmailtrigger = email_config.get("gmailtrigger", False)
                    attached_multiple_file = email_config.get("attached_multiple_file", False)
                    start_file_names = email_config.get("startFileNames", [])
                    # Ensure gmailtrigger is properly converted to boolean for 12b1_compensation reports
                    if isinstance(gmailtrigger, str):
                        gmailtrigger = gmailtrigger.lower() in ['true', '1', 'yes']
                    elif not isinstance(gmailtrigger, bool):
                        gmailtrigger = bool(gmailtrigger)
                    
                    # Enhanced Splunk logging for email configuration
                    log_message('INFO', f'Email configuration loaded - Gmail trigger: {gmailtrigger} (type: {type(gmailtrigger)}), Multiple files: {attached_multiple_file}')
                    if start_file_names:
                        log_message('INFO', f'Start file names configured: {start_file_names}')
                    
                    # Special logging for 12b1_compensation reports
                    report_type = config.get('workflow_metadata', {}).get('report_type', 'unknown')
                    if report_type == '12b1_compensation':
                        log_message('INFO', f'12b1_compensation report detected - Email configuration: gmailtrigger={gmailtrigger}, attached_multiple_file={attached_multiple_file}')
                        
                        # Ensure 12b1_compensation always sends email regardless of file discovery issues
                        if not gmailtrigger:
                            log_message('WARNING', '12b1_compensation: gmailtrigger is False, but forcing email sending for this report type')
                            gmailtrigger = True  # Force email sending for 12b1_compensation
                        
                        log_message('INFO', f'12b1_compensation email sending will proceed: {gmailtrigger}')
                    
                    # Check if email should be sent based on gmailtrigger
                    if not gmailtrigger:
                        log_message('INFO', 'Email sending skipped - gmailtrigger is false')
                    else:
                        # Special logging for 12b1_compensation email sending
                        if report_type == '12b1_compensation':
                            log_message('INFO', f'12b1_compensation: Proceeding with email sending - gmailtrigger={gmailtrigger}')
                        if report_type == 'money_market_CWR':
                            output_path = ""  # avoid UnboundLocalError, no report file expected
                            bucket_name = os.getenv('s3BucketInput', '')
                            attachment_flag = False
                            attached_multiple_file=False
                            object_key = input_file_name

                        # Parse S3 path to extract bucket and object key
                        # For redemption_swp, ensure output_path is set (use primary_output_path if available)
                        if report_type == 'redemption_swp' and not output_path and primary_output_path:
                            output_path = primary_output_path
                            
                        if output_path and output_path.startswith("s3://"):
                            s3_path_parts = output_path[5:].split("/", 1)
                            bucket_name = s3_path_parts[0]
                            object_key = s3_path_parts[1] if len(s3_path_parts) > 1 else ""
                        else:
                            bucket_name = ""
                            object_key = ""

                        # Email body is already set above for money_market_CWR, use default for others
                        if report_type != 'money_market_CWR':
                            body_text = f"""embed above text file"""

                        cc_emails = email_config.get("ccEmails", [])
                        attachment_file_path = f"/tmp/{object_key.split('/')[-1]}" if object_key else ""
                        
                        # Prepare attachment logic based on configuration
                        multiple_files = None
                        attachment_flag = True  # Default to attaching files
                        
                        # Special handling for money_market_CWR - no attachments needed for pass notification
                        # if report_type == 'money_market_CWR':
                        #     attachment_flag = False
                        #     attached_multiple_file=False
                        #     log_message('INFO', f'money_market_CWR: Disabling attachments for pass notification email')
                        #     print(f"money_market_CWR: No attachments needed for pass notification")
                        
                        if attached_multiple_file:
                            # Multiple file attachment mode
                            if start_file_names:
                                # Discover files that match the prefixes in startFileNames
                                log_message('INFO', f'Preparing multiple attachments using prefix matching for: {start_file_names}')
                                
                                # Determine base path for file discovery
                                if object_key and '/' in object_key:
                                    # Extract directory path from the original object_key
                                    base_path = '/'.join(object_key.split('/')[:-1])  # Remove filename, keep directory
                                else:
                                    base_path = ""
                                
                                # Discover files matching the prefixes
                                discovered_files = discover_files_by_prefix(bucket_name, base_path, start_file_names)
                                
                                if discovered_files:
                                    # Create multiple file entries based on discovered files
                                    multiple_files = []
                                    for discovered_file in discovered_files:
                                        file_path = f"/tmp/{discovered_file['file_name']}"
                                        multiple_files.append({
                                            'bucket_name': bucket_name,
                                            'object_key': discovered_file['object_key'],
                                            'file_path': file_path,
                                            'start_file_name': discovered_file['file_name']
                                        })
                                        log_message('INFO', f'Added discovered file to multiple attachments: {discovered_file["file_name"]} (S3 path: {discovered_file["object_key"]})')
                                    
                                    log_message('INFO', f'Successfully discovered {len(discovered_files)} files matching prefixes')
                                else:
                                    # No files found matching the prefixes - special handling for 12b1_compensation
                                    log_message('WARNING', f'No files found matching prefixes: {start_file_names}')
                                    
                                    # Special handling for 12b1_compensation report
                                    if report_type == '12b1_compensation':
                                        log_message('INFO', '12b1_compensation: No matching files found, but proceeding with email sending using available report file')
                                        # For 12b1_compensation, use the generated report file even if no matching files found
                                        multiple_files = [{
                                            'bucket_name': bucket_name,
                                            'object_key': object_key,
                                            'file_path': attachment_file_path,
                                            'start_file_name': '12b1'  # Use the expected start file name
                                        }]
                                        log_message('INFO', f'12b1_compensation: Using generated report file as attachment: {object_key}')
                                    else:
                                        log_message('WARNING', 'Falling back to single file attachment')
                                        multiple_files = [{
                                            'bucket_name': bucket_name,
                                            'object_key': object_key,
                                            'file_path': attachment_file_path,
                                            'start_file_name': input_file_name
                                        }]
                            else:
                                # Fallback to single file if no startFileNames specified
                                log_message('WARNING', 'attached_multiple_file is true but no startFileNames provided - using single file')
                                multiple_files = [{
                                    'bucket_name': bucket_name,
                                    'object_key': object_key,
                                    'file_path': attachment_file_path,
                                    'start_file_name': input_file_name
                                }]
                        else:
                            # Single file attachment mode
                            log_message('INFO', 'Single file attachment mode - attaching only the generated report')
                            multiple_files = None  # Will use single attachment logic in send_email_with_attachment
                        
                        # Special debugging for 12b1_compensation email sending
                        if report_type == '12b1_compensation':
                            log_message('INFO', f'12b1_compensation: About to send email with params - sender={from_email}, recipients={to_emails}, subject={subject_with_date}')
                            log_message('INFO', f'12b1_compensation: Attachment params - bucket={bucket_name}, object_key={object_key}, multiple_files={multiple_files is not None}')
                        
                        email_result = send_email_with_attachment(
                            sender_email=from_email,
                            recipient_emails=to_emails,
                            cc_emails=cc_emails,
                            subject=subject_with_date,
                            body_text=body_text,
                            attachment_file_path=attachment_file_path,
                            bucket_name=bucket_name,
                            object_key=object_key,
                            attachment_flag=attachment_flag,
                            report_attachment_flag="true",
                            multiple_files=multiple_files,
                            attached_multiple_file=attached_multiple_file,
                            gmailtrigger=gmailtrigger,
                            context=None
                        )

                        # Log email result
                        if "SES Access Denied" in email_result or "Credentials not available" in email_result:
                            if cc_emails:
                                   log_message("WARNING", f"Email notification Failed - Report available at: {output_path}")
                        else:
                            log_message("SUCCESS", f"Email notification sent: {email_result}")
                            
                            # Mark redemption_swp email as sent to prevent duplicates
                            if report_type == 'redemption_swp':
                                ReportGenerator._redemption_swp_email_sent = True
                                log_message("INFO", "Redemption_SWP email sent flag set to prevent duplicates")
                            
                            # Special logging for 12b1_compensation email success
                            if report_type == '12b1_compensation':
                                log_message('SUCCESS', f'12b1_compensation: Email sent successfully - {email_result}')

            return success_count == total_formats    

     def _generate_txt_report(self, config: Dict, data: List[Dict], output_path: str, format_settings: Dict = None) -> bool:
         """Generate TXT report with preserved formatting"""
         if not output_path:
              return False
         if format_settings is None:
              format_settings = {}
         # Check if this is a Redemption_SWP report and handle specially
         report_type = config.get('workflow_metadata', {}).get('report_type')
         if report_type == 'redemption_swp':
              return self._generate_redemption_swp_txt_report(config, data, output_path)
         # Check if formatting should be preserved
         preserve_formatting = format_settings.get('preserve_formatting', True)
         # Generate report content
         if preserve_formatting:
              report_content = self._generate_report_content(config, data)
         else:
              # Generate simplified content if formatting preservation is disabled
              report_content = self._generate_simplified_report_content(config, data)
         # Parse S3 path
         if output_path.startswith('s3://'):
            s3_path_parts = output_path[5:].split('/', 1)
            bucket = s3_path_parts[0]
            key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
         else:
            bucket = os.getenv('s3BucketOutput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
            report_path = os.getenv('reportPath', 'jhi/checks/wip/wipid/report/')
            key = f"{report_path}{os.path.basename(output_path)}"

          # Upload to S3
         s3.put_object(
            Bucket=bucket,
            Key=key,
            Body=report_content.encode('utf-8'),
            ContentType='text/plain'
          )
        
         try:
             return True
         except Exception as e:
              return False

    
     def _generate_redemption_swp_txt_report(self, config: Dict, data: List[Dict], output_path: str) -> bool:
         """Generate specialized TXT report for Redemption_SWP matching XLSX logic exactly"""
         try:
            field_mapping = config.get('xml_mapping', {}).get('field_mapping', {})
            CUSTOM_CHECK_TYPE1 = field_mapping.get('CUSTOM_CHECK_TYPE1', 'CUSTOM_CHECK_TYPE1')
            CUSTOM_FUND_CODE = field_mapping.get('CUSTOM_FUND_CODE', 'CUSTOM_FUND_CODE')
            CUSTOM_FUND_DESC = field_mapping.get('CUSTOM_FUND_DESC', 'CUSTOM_FUND_DESC')
            CUSTOM_TRADE_DATE = field_mapping.get('TRADE_DATE', 'TRADE_DATE')
            CUSTOM_CHECK_AMOUNT = field_mapping.get('CUSTOM_CHECK_AMOUNT', 'CUSTOM_CHECK_AMOUNT')


              # Use the same logic as XLSX: filter and sort data exactly like XLSX version
              # Filter records to only include those with CUSTOM_CHECK_TYPE1 = "RED" or "SWP"
            filtered_data = [record for record in data
                               if record.get(CUSTOM_CHECK_TYPE1) in ['REDEMPTION', 'SWP']]
                
                # Sort by combined fund code + desc (for grouping), then by check type (RED first, then SWP) - same as XLSX
            filtered_data.sort(key=lambda x: (
                f"{x.get(CUSTOM_FUND_CODE, '')} {x.get(CUSTOM_FUND_DESC, '')}".strip(),
                x.get(CUSTOM_CHECK_TYPE1, '')
            ))
            self.all_records = filtered_data
            # Create output lines
            lines = []
              
            column_headers = (
                    config.get("layout", {})
                    .get("column_headers", {})
                    .get("row", [])
               )
             
            # Add header information - match XLSX layout exactly
            current_date = datetime.now().strftime("%m/%d/%Y")
            #   trade_date = "01/15/2025"
            #trade_date = data[0].get('TRADE_DATE',"")
            trade_date = data[0].get(CUSTOM_TRADE_DATE,"")
             
            lines.append(f"{current_date}||||||||||")
            lines.append(f"Trade Date: {trade_date}||||||||||")
            # Use the same pipe formatting as before
            lines.append("|".join(column_headers))
            # Respect optional spacing from config (if present)
            # Get column configuration from config - same as XLSX
            layout = config.get('layout', {})
            data_section = layout.get('data_section', {})
            columns = data_section.get('columns', [])
          
            # Group records by fund (fund_code + fund_desc) and create aggregated records
            from collections import defaultdict
            fund_groups = defaultdict(list)
            for record in filtered_data:
                fund_key = f"{record.get(CUSTOM_FUND_CODE, '')} {record.get(CUSTOM_FUND_DESC, '')}".strip()
                fund_groups[fund_key].append(record)
            
            # Sort fund keys to maintain order
            sorted_fund_keys = sorted(fund_groups.keys())
            
            # Generate one row per fund with aggregated data
            for fund_key in sorted_fund_keys:
                fund_records = fund_groups[fund_key]
                # Create aggregated record for this fund
                aggregated_record = self._create_aggregated_fund_record(fund_key, fund_records, config)
                if aggregated_record:
                    # Generate row using same column logic as XLSX
                    row_values = []
                    for col in columns:
                        value = self._get_column_value_for_xlsx(col, aggregated_record, config)
                        formatted_value = self._format_column_value_for_xlsx(col, value)
                        row_values.append(formatted_value)
                     
                    # Join with pipe separator for TXT format
                    line = '|'.join(str(v) for v in row_values)
                    lines.append(line)
             
            # Add totals row using same column structure
            if columns:
                total_values = []
                total_checks = 0
                total_dollar_value = 0.0
                
                # Calculate totals from aggregated records
                for fund_key in sorted_fund_keys:
                    fund_records = fund_groups[fund_key]
                    aggregated_record = self._create_aggregated_fund_record(fund_key, fund_records, config)
                    if aggregated_record:
                        total_checks += aggregated_record.get('CHECKS_PRINTED_TODAY', 0)
                        total_dollar_value += aggregated_record.get('DOLLAR_VALUE_TODAY', 0.0)
                
                for col in columns:
                    if col.get('name') == 'GROWTHFUND':
                        total_values.append('TOTAL')
                    elif col.get('name') == 'CHECKS PRINTED TODAY':
                        total_values.append(total_checks)
                    elif col.get('name') == 'DOLLAR VALUE TODAY':
                        total_values.append(f"{total_dollar_value:.2f}")
                    else:
                        total_values.append('')
                 
                total_line = '|'.join(str(v) for v in total_values)
                lines.append(total_line)
             
            # Add footer section - use specialized format for Redemption_SWP
            footer_lines = self._generate_redemption_swp_footer_section(config, data)
            if footer_lines:
                  lines.append('')  # Add blank line before footer
                  lines.extend(footer_lines)
             
              # Parse S3 path
            if output_path.startswith('s3://'):
                s3_path_parts = output_path[5:].split('/', 1)
                bucket = s3_path_parts[0]
                key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
            else:
                bucket = os.getenv('s3BucketOutput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
                report_path = os.getenv('reportPath', 'jhi/checks/wip/wipid/report/')
                key = f"{report_path}{os.path.basename(output_path)}"

             # Upload to S3
            s3.put_object(
                 Bucket=bucket,
                 Key=key,
                 Body='\n'.join(lines).encode('utf-8'),
                 ContentType='text/plain'
             )
             
            return True
             
         except Exception as e:
              return False
              
     def _generate_redemption_swp_footer_section(self, config: Dict, data: List[Dict]) -> List[str]:
         """Generate footer section for Redemption_SWP reports with pipe-separated format"""
         layout = config['layout']
         footer_config = layout.get('footer_section', {})
         
         lines = []
         
         # Categories from configuration
         categories = footer_config.get('categories', [])
         
         # Calculate dynamic footer data
         category_data = self._calculate_dynamic_footer_data(config,data)
         
         # Add the special header lines first
     #     lines.append("RETURN TO JANUS||||||||||                                                                                                                                                                                                                                                                                   ")
         lines.append("RETURN TO JANUS|NUMBER RECEIVED|||||||||                                                                                                                                                                                                                                                                    ")
         
         # Generate footer lines with pipe-separated format
         for i, category in enumerate(categories):
             data_value = category_data[i] if i < len(category_data) else 0
             
             # Map category names to the expected format
             display_name = category
          #    if category == "RETURNTOJANUS":
          #        display_name = "JANUS3"
             if category == "STATESTREETBANK":
                 display_name = "STATE STREET BANK"
             elif category == "DONOTMAIL":
                 display_name = "DO NOT MAIL"
             elif category == "PULLREPORTCHECKS":
                 display_name = "PULL REPORT CHECKS"
             elif category == "SUB-TOTAL":
                 # For Redemption_SWP, display both SUB-TOTALs:
                 # First SUB-TOTAL (index 6) = sum of JANUS, STATESTREETBANK, DONOTMAIL, VOID, PULLREPORTCHECKS
                 # Second SUB-TOTAL (index 12) = sum of OVERNIGHT W-SIG, OVERNIGHT WO-SIG, OVERNIGHT SAT-W-SIG, OVERNIGHT SAT-WO-SIG
                 # data_value is already correctly set to the appropriate SUB-TOTAL value
                 display_name = "SUB-TOTAL"
             elif category == "SPECIALHANDLING":
                 display_name = "SPECIAL HANDLING"
             elif category == "OVERNIGHTW-SIG":
                 display_name = "OVERNIGHT W-SIG"
             elif category == "OVERNIGHTWO-SIG":
                 display_name = "OVERNIGHT WO-SIG"
             elif category == "OVERNIGHTSAT-W-SIG":
                 display_name = "OVERNIGHT SAT-W-SIG"
             elif category == "OVERNIGHTSAT-WO-SIG":
                 display_name = "OVERNIGHT SAT-WO-SIG"
             elif category == "SUPPRESSEDTEST":
                 display_name = "SUPPRESSED TEST"
             
            # Ensure data_value is an integer (handle None, empty strings, space strings, etc.)
             try:
                 # Handle None, empty strings, space strings, and other non-numeric values
                 if data_value is None:
                     int_value = 0
                 elif isinstance(data_value, str):
                     # Strip whitespace and check if it's empty
                     stripped = data_value.strip()
                     if stripped == '':
                         int_value = 0
                     else:
                         # Try to convert string to int
                         int_value = int(stripped)
                 elif isinstance(data_value, (int, float)):
                     int_value = int(data_value)
                 else:
                     # For any other type, default to 0
                     int_value = 0
             except (ValueError, TypeError, AttributeError):
                 int_value = 0
             formatted_value = f"{int_value:07d}"
          
             # Create the line with pipe separators and proper spacing
             # Each line should have the category name, value, and then many pipes with spaces
             line = f"{display_name}|{formatted_value}|||||||||                                                                                                                                                                                                                                                                                        "
             lines.append(line)
         
         return lines

              
     def _generate_csv_report(self, config: Dict, data: List[Dict], output_path: str, format_settings: Dict = None) -> bool:
          """Generate CSV report with structured layout matching TXT format"""
          if not output_path:
                   return False

          if format_settings is None:
                   format_settings = {}

          # Only create directories for local paths, not S3 paths
          if not output_path.startswith('s3://'):
               os.makedirs(os.path.dirname(output_path), exist_ok=True)

          try:
                   # Check if this is a 12b1_compensation report and handle specially
                   report_type = config.get('workflow_metadata', {}).get('report_type')
                   if report_type == '12b1_compensation':
                        return self._generate_12b1_compensation_csv(config, data, output_path)
              
                   # Generate structured content like TXT format
                   report_content = self._generate_report_content(config, data)
         
                   # Convert structured content to CSV format with single column layout and consistent alignment
                   lines = report_content.split('\n')
                   csv_lines = []
         
                   for line in lines:
                        if line.strip() == "":
                             # Empty lines become empty CSV rows
                             csv_lines.append([])
                        elif '\t' in line:
                             # Tab-separated lines - convert to single column layout with consistent alignment
                             cells = line.split('\t')
                             # Create label-value pairs in single column with consistent alignment
                             csv_row = []
                             for i, cell_value in enumerate(cells):
                                  if i == 0:
                                        # First column contains the label - consistently aligned
                                        csv_row.append(cell_value)
                                  else:
                                        # Subsequent columns contain values - consistently aligned
                                        if cell_value.strip():# Only add non-empty values
                                             csv_row.append(cell_value)
                             csv_lines.append(csv_row)
                        else:
                             # Regular lines become single-column CSV rows with consistent alignment
                             csv_lines.append([line])
              
                   # Enhanced Logic for Redemption_SWP Report - CSV
                   if report_type == 'redemption_swp':
                        csv_lines = self._apply_redemption_swp_enhanced_logic_csv(csv_lines, data)     
                   # Parse S3 path or use environment variables
                   if output_path.startswith('s3://'):
                        s3_path_parts = output_path[5:].split('/', 1)
                        bucket = s3_path_parts[0]
                        key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
                   else:
                        bucket = os.getenv('s3BucketOutput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
                        report_path = os.getenv('reportPath', 'jhi/checks/wip/wipid/report/')
                        key = f"{report_path}{os.path.basename(output_path)}"
              
                   # Convert CSV data to string
                   csv_content = io.StringIO()
                   writer = csv.writer(csv_content, delimiter=',', quotechar='"')
                   for row in csv_lines:
                        writer.writerow(row)
                   csv_string = csv_content.getvalue()
              
                   # Upload to S3
                   s3.put_object(
                        Bucket=bucket,
                        Key=key,
                        Body=csv_string.encode('utf-8'),
                        ContentType='text/csv'
                   )
              
                   log_message("SUCCESS", f"CSV Report uploaded to S3: s3://{bucket}/{key}")
                   return True
          except Exception as e:
                   log_message("ERROR", f"Error uploading CSV report to S3: {str(e)}")
                   return False

     def _generate_xlsx_report(self, config: Dict, data: List[Dict], output_path: str, format_settings: Dict = None) -> bool:
          """Generate XLSX report with proper cell separation"""
          if not output_path:
                   return False

          if not XLSX_AVAILABLE:
                   return False

          if format_settings is None:
                   format_settings = {}

          # Only create directories for local paths, not S3 paths
          if not output_path.startswith('s3://'):
               os.makedirs(os.path.dirname(output_path), exist_ok=True)

          try:
                   wb = Workbook()
                   ws = wb.active
                   ws.title = config.get('export_settings', {}).get('xlsx', {}).get('sheet_name', 'Report')
         
                   current_row = 1
         
                   # For Redemption_SWP reports, ensure static text is placed FIRST in row 3 (A3, B3, C3)
                   report_type = config.get('workflow_metadata', {}).get('report_type')
                   if report_type == 'redemption_swp':
                        # Place static text in row 3 (A3, B3, C3) - these cells must remain fixed
                        self._apply_redemption_swp_enhanced_logic_xlsx(ws, data, current_row)
                   elif report_type == '12b1_compensation':
                        # Handle 12b1_compensation with three child reports
                        current_row = self._generate_12b1_compensation_xlsx(ws, config, data, current_row)
                        # Save and return early since we handle all content in the specialized method
                        wb.save(output_path)
                        return True
         
                   # Generate header only if header section exists
                   if 'header' in config.get('layout', {}):
                        header = self._generate_header(config)
                        header_lines = header.split('\n')
                        for line in header_lines:
                             if line.strip():
                                  cell = ws.cell(row=current_row, column=1, value=line)
                                  # Check if this is a subtitle (second non-empty line after title)
                                  header_config = config.get('layout', {}).get('header', {})
                                  subtitle = header_config.get('subtitle', '')
                                  if subtitle and line.strip() == subtitle:
                                        # Subtitle formatting - slightly smaller font
                                        cell.font = Font(bold=True, size=9)
                                  else:
                                        # Title formatting - normal size
                                        cell.font = Font(bold=True, size=10)
                                  cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
         
                   # Generate summary section with proper cell separation
                   current_row = self._generate_xlsx_summary_section(ws, config, data, current_row)
         
                   # Generate totals section with proper cell separation - position depends on report type
                   if 'totals_section' in config.get('layout', {}):
                        report_type = config.get('workflow_metadata', {}).get('report_type')
                        if report_type == 'replacement':
                             # For replacement report, totals appear before data section
                             current_row = self._generate_xlsx_totals_section(ws, config, data, current_row)
                        elif report_type == 'redemption_swp':
                             pass
                        else:
                             # Default behavior - totals before data
                             current_row = self._generate_xlsx_totals_section(ws, config, data, current_row)
         
                   # Generate data section with proper column separation
                   current_row = self._generate_xlsx_data_section(ws, config, data, current_row)

                   # Generate totals section for redemption_swp report (after data sections)
                   if 'totals_section' in config.get('layout', {}) and config.get('workflow_metadata', {}).get('report_type') == 'redemption_swp':
                        current_row = self._generate_xlsx_totals_section(ws, config, data, current_row)
         
                   # Generate footer if exists with proper cell separation
                   if 'footer_section' in config['layout']:
                        current_row = self._generate_xlsx_footer_section(ws, config, data, current_row)
         
                   # Auto-fit columns
                   for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                             try:
                                  if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                             except:
                                  pass
                        # Set appropriate width for each column
                        if column_letter == 'A':# First column (labels) - wider for readability
                             adjusted_width = min(max_length + 8, 70)
                        else:# Data columns - appropriate width for values
                             adjusted_width = min(max_length + 5, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
              
                        # Set row height to accommodate wrapped text
                        for cell in column:
                             if cell.value and len(str(cell.value)) > 40:# Long values need more height
                                  ws.row_dimensions[cell.row].height = 35
                             elif cell.value and len(str(cell.value)) > 20:# Medium values
                                  ws.row_dimensions[cell.row].height = 25
                             elif cell.value:# Regular rows with proper spacing
                                  ws.row_dimensions[cell.row].height = 20
                   # Parse S3 path or use environment variables
                   if output_path.startswith('s3://'):
                        s3_path_parts = output_path[5:].split('/', 1)
                        bucket = s3_path_parts[0]
                        key = s3_path_parts[1] if len(s3_path_parts) > 1 else ''
                   else:
                        bucket = os.getenv('s3BucketOutput', 'br-icsdev-dpmjayant-dataingress-us-east-1-s3')
                        report_path = os.getenv('reportPath', 'jhi/checks/wip/wipid/report/')
                        key = f"{report_path}{os.path.basename(output_path)}"
              
                   # Save workbook to bytes
                   xlsx_buffer = io.BytesIO()
                   wb.save(xlsx_buffer)
                   xlsx_buffer.seek(0)
              
                   # Upload to S3
                   s3.put_object(
                        Bucket=bucket,
                        Key=key,
                        Body=xlsx_buffer.getvalue(),
                        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                   )
              
                   log_message("SUCCESS", f"XLSX Report uploaded to S3: s3://{bucket}/{key}")
                   return True
          except Exception as e:
                   log_message("ERROR", f"Error uploading XLSX report to S3: {str(e)}")
                   return False

     def _generate_report_content(self, config: Dict, data: List[Dict]) -> str:
          """Generate the complete report content"""
          content_parts = []
         
          # Check if this is a 12b1_compensation report and handle specially
          report_type = config.get('workflow_metadata', {}).get('report_type')
          if report_type == '12b1_compensation':
                   return self._generate_12b1_compensation_content(config, data)

          # Generate header only if header section exists
          if 'header' in config.get('layout', {}):
                   header = self._generate_header(config)
                   content_parts.append(header)

          # Generate summary section if exists
          if 'summary_section' in config.get('layout', {}):
                   summary = self._generate_summary_section(config, data)
                   content_parts.append(summary)

          # Generate totals section if exists - position depends on report type
          if 'totals_section' in config.get('layout', {}):
                   if report_type == 'replacement':
                        # For replacement report, totals appear before data section
                        totals = self._generate_totals_section(config, data)
                        content_parts.append(totals)
                   elif report_type == 'redemption_swp':
                        # For Redemption report, totals should appear at the bottom after all tables
                        # This will be handled after data sections
                        pass
                   elif report_type == 'dividend_checks':
                        # For dividend report, totals should appear after data section
                        # This will be handled after data sections
                        pass
                   else:
                        # Default behavior - totals before data
                        totals = self._generate_totals_section(config, data)
                        content_parts.append(totals)

          # Generate data section if exists
          if 'data_section' in config.get('layout', {}):
                   data_section = self._generate_data_section(config, data)
                   content_parts.append(data_section)

          # Generate matrix sections if exist
          if 'matrix_sections' in config.get('layout', {}):
                   matrix_content = self._generate_matrix_text_section(config, data)
                   content_parts.append(matrix_content)

          # Generate totals section for redemption_swp and dividend_checks reports (after data sections)
          if 'totals_section' in config.get('layout', {}) and report_type in ['redemption_swp', 'dividend_checks']:
                   # Add blank line before totals section
                   content_parts.append("")# One blank line above totals
                   totals = self._generate_totals_section(config, data)
                   content_parts.append(totals)

          # Generate footer if exists
          if 'footer_section' in config.get('layout', {}):
                   footer = self._generate_footer_section(config, data)
                   content_parts.append(footer)

          return '\n'.join(content_parts)

     def _generate_12b1_compensation_content(self, config: Dict, data: List[Dict]) -> str:
          """Generate specialized content for 12b1_compensation reports with three child reports"""
          content_parts = []
         
          # Generate 12b1_RECONCILIATION_SUPPRESSION report
          if '12b1_RECONCILIATION_SUPPRESSION' in config:
                   reconciliation_content = self._generate_12b1_reconciliation_suppression_content(
                        config['12b1_RECONCILIATION_SUPPRESSION'], data)
                   content_parts.append(reconciliation_content)
                   content_parts.append("")  # Add spacing between reports
         
          # Generate 12b1_fund_report
          if '12b1_fund_report' in config:
                   fund_content = self._generate_12b1_fund_report_content(
                        config['12b1_fund_report'], data)
                   content_parts.append(fund_content)
                   content_parts.append("")  # Add spacing between reports
         
          # Generate 12b1_check_report
          if '12b1_check_report' in config:
                   check_content = self._generate_12b1_check_report_content(
                        config['12b1_check_report'], data)
                   content_parts.append(check_content)
         
          return '\n'.join(content_parts)

     def _filter_data_for_check_report(self, data: List[Dict], config: Dict) -> List[Dict]:
          """Filter data for 12b1_check_report to only include records with identifiers 0014 and 0002"""
          
          # Filter records to only include those with identifiers 0014 and 0002
          filtered_data = []
          for record in data:
               identifier = record.get('identifier', '')
               if identifier in ['0014', '0002']:
                    filtered_data.append(record)
                    # print(f"DEBUG: INCLUDED - Identifier: '{identifier}'")
               # else:
                    # print(f"DEBUG: FILTERED OUT - Identifier: '{identifier}' (not 0014 or 0002)")
          
          return filtered_data

     def _extract_combined_data_for_12b1_reports(self, config: Dict, input_file_name: str, bucket: str) -> List[Dict]:
          """Extract data with dynamic field mapping from JSON configuration for 12b1 reports"""
          try:
               # Get field mappings from JSON configuration for 12b1_check_report
               suppression_config = config.get('12b1_RECONCILIATION_SUPPRESSION', {})
               suppression_input = suppression_config.get('input_mapping', {})
               suppression_mapping = suppression_input.get('field_mapping', {})
               
               check_report_config = config.get('12b1_check_report', {})
               check_input = check_report_config.get('input_mapping', {})
               check_mapping  = check_input.get('field_mapping', {})
               
               # Create combined field mapping that includes all fields needed by both reports
               # Use JSON configuration for 12b1_check_report, add fund report fields
               combined_field_mapping = {
                    # Fields from JSON configuration for 12b1_check_report
                    **suppression_mapping,
                    **check_mapping,
                    # Additional fields needed by 12b1_fund_report
                    "Count": check_mapping.get("identifier", 0),  # Use same field as identifier
                    "Fund": check_mapping.get("fund", 7)  # Use same field as fund
               }
               
               # Create combined config for data extraction using JSON configuration
               combined_config = {
                    "input_file_type": "tilde_separated",
                    "input_mapping": {
                         "file_path": check_input.get("file_path", "jhi/checks/output/MUSU5928.TXT.asc"),
                         "delimiter": check_input.get("delimiter", "~"),
                         "field_mapping": combined_field_mapping
                    }
               }
               
               # print(f"DEBUG: JSON field mapping: {field_mapping}")
               
               # Extract data with combined field mapping
               input_file_path = combined_config['input_mapping']['file_path']
               if not input_file_path.startswith('s3://'):
                    input_file_path = f"s3://{bucket}/{input_file_path}"
               
               data = self.extractor.extract_tilde_separated_data(input_file_path, combined_config)
               
               return data
               
          except Exception as e:
               log_message('ERROR', f'12b1_compensation: Failed to extract combined data: {str(e)}')
               return []

     def _generate_12b1_compensation_separate_reports(self, config: Dict, data: List[Dict],input_file_name:str) -> Dict[str, str]:
          """Generate separate content for each 12b1_compensation report type"""
          separate_reports = {}
         
          try:
               # Generate 12b1_RECONCILIATION_SUPPRESSION report
               if '12b1_RECONCILIATION_SUPPRESSION' in config:
                    reconciliation_content = self._generate_12b1_reconciliation_suppression_content(
                         config['12b1_RECONCILIATION_SUPPRESSION'], data,input_file_name)
                    separate_reports['RECONCILIATION_SUPPRESSION'] = reconciliation_content
               else:
                    log_message('WARNING', '12b1_compensation: 12b1_RECONCILIATION_SUPPRESSION configuration not found')
         
               # Generate 12b1_fund_report
               if '12b1_fund_report' in config:
                    fund_content = self._generate_12b1_fund_report_content(
                         config['12b1_fund_report'], data)
                    separate_reports['FUND'] = fund_content
               else:
                    log_message('WARNING', '12b1_compensation: 12b1_fund_report configuration not found')
         
               # Generate 12b1_check_report
               if '12b1_check_report' in config:
                    # Check if 12b1_check_report has its own tilde-separated configuration
                    check_report_config = config['12b1_check_report']
                    if check_report_config.get('input_file_type') == 'tilde_separated':
                         # Filter the existing data to only include relevant records (0014 and 0002)
                         check_data = self._filter_data_for_check_report(data, check_report_config)
                         check_content = self._generate_12b1_check_report_content(
                              check_report_config, check_data)
                    else:
                         # Use the general data for XML-based reports
                         check_content = self._generate_12b1_check_report_content(
                              check_report_config, data)
                    separate_reports['FUNDACCT'] = check_content
               else:
                    log_message('WARNING', '12b1_compensation: 12b1_check_report configuration not found')
               
               log_message('INFO', f'12b1_compensation: Generated {len(separate_reports)} reports: {list(separate_reports.keys())}')
               
          except Exception as e:
               log_message('ERROR', f'12b1_compensation: Failed to generate reports: {str(e)}')
         
          return separate_reports

     def _generate_12b1_compensation_separate_files(self, config: Dict, data: List[Dict], input_file_name: str, bucket: str) -> bool:
          """Generate separate files for each 12b1_compensation report type"""
          try:
               # Get the current date for file naming
               current_date = datetime.now().strftime("%m%d%Y")
               
               # Extract data with combined field mapping for both reports in single pass
               combined_data = self._extract_combined_data_for_12b1_reports(config, input_file_name, bucket)
               
               if not combined_data:
                    log_message('ERROR', '12b1_compensation: No combined data extracted for reports')
                    return False
               
               # Generate separate content for each report type
               log_message('INFO', f'12b1_compensation: Starting report generation with {len(combined_data)} data records')
               
               separate_reports = self._generate_12b1_compensation_separate_reports(config, combined_data,input_file_name)
               
               if not separate_reports:
                    log_message('WARNING', '12b1_compensation: No reports generated - checking if this is expected')
                    
                    # For 12b1_compensation, we should still attempt to send email even if no reports are generated
                    # This ensures stakeholders are notified about the status
                    log_message('INFO', '12b1_compensation: Proceeding with email notification despite no reports generated')
                    
                    # Create empty reports to ensure email is still sent
                    separate_reports = {
                        'RECONCILIATION_SUPPRESSION': 'No data available for RECONCILIATION_SUPPRESSION report',
                        'FUND': 'No data available for FUND report', 
                        'FUNDACCT': 'No data available for FUNDACCT report'
                    }
               
               # File naming mapping
               file_mappings = {
                    'RECONCILIATION_SUPPRESSION': f"12b1.chck.RECONSUPPRESS.{current_date}.txt",
                    'FUND': f"12b1.chck.FUND.{current_date}.txt", 
                    'FUNDACCT': f"12b1.chck.FUNDACCT.{current_date}.txt"
               }
               
               success_count = 0
               total_reports = len(separate_reports)
               
               # Generate each report as a separate file
               for report_type, content in separate_reports.items():
                    if report_type in file_mappings:
                         output_file_name = file_mappings[report_type]
                         
                         # Construct full output path
                         report_path = os.getenv("reportPath", "jhi/checks/wip/wipid/report/")
                         output_path = f"s3://{bucket}/{report_path}{output_file_name}"
                         
                         
                         # Write the content to S3
                         success = self._write_content_to_s3(content, output_path)
                         
                         if success:
                              success_count += 1
               # Send email after all three reports are generated (success or failure)
               log_message('INFO', f'12b1_compensation: {success_count}/{total_reports} reports generated - proceeding with email')
               
               # Get email configuration
               email_config = config.get("emailConfig", {})
               from_email = email_config.get("fromEmail")
               to_emails = email_config.get("toEmails", [])
               current_date = datetime.now().strftime("%m%d%Y")
               subject = email_config.get("subject", "JHI 12b1 {date}")
               subject_with_date = subject.replace("{date}", current_date)
               
               if from_email and to_emails:
                    # Get email configuration options
                    gmailtrigger = email_config.get("gmailtrigger", True)
                    attached_multiple_file = email_config.get("attached_multiple_file", True)
                    start_file_names = email_config.get("startFileNames", ["12b1"])
                    
                    # Ensure gmailtrigger is properly converted to boolean
                    if isinstance(gmailtrigger, str):
                        gmailtrigger = gmailtrigger.lower() in ['true', '1', 'yes']
                    elif not isinstance(gmailtrigger, bool):
                        gmailtrigger = bool(gmailtrigger)
                    
                    log_message('INFO', f'12b1_compensation: Email config - from_email={from_email}, to_emails={to_emails}, subject={subject_with_date}')
                    log_message('INFO', f'12b1_compensation: Email settings - gmailtrigger={gmailtrigger}, attached_multiple_file={attached_multiple_file}')
                    
                    # Email is always sent for 12b1_compensation after all reports are attempted
                    if success_count == total_reports:
                        log_message('SUCCESS', f'12b1_compensation: All {total_reports} reports generated successfully')
                    else:
                        log_message('WARNING', f'12b1_compensation: Only {success_count}/{total_reports} reports succeeded - sending email anyway')
                    
                    # Prepare multiple file attachments for all three reports
                    multiple_files = []
                    report_path = os.getenv("reportPath", "jhi/checks/wip/wipid/report/")
                    
                    for report_type, content in separate_reports.items():
                        if report_type in file_mappings:
                            output_file_name = file_mappings[report_type]
                            object_key = f"{report_path}{output_file_name}"
                            file_path = f"/tmp/{output_file_name}"
                            
                            multiple_files.append({
                                'bucket_name': bucket,
                                'object_key': object_key,
                                'file_path': file_path,
                                'start_file_name': '12b1'
                            })
                            log_message('INFO', f'12b1_compensation: Added to email attachments - {output_file_name}')
                    
                    # Email body
                    body_text = f"""embed above text file"""
                    cc_emails = email_config.get("ccEmails", [])
                    
                    log_message('INFO', f'12b1_compensation: Sending email with {len(multiple_files)} attachments')
                    
                    # Send email with all three report attachments
                    email_result = send_email_with_attachment(
                        sender_email=from_email,
                        recipient_emails=to_emails,
                        cc_emails=cc_emails,
                        subject=subject_with_date,
                        body_text=body_text,
                        attachment_file_path="",  # Not used for multiple files
                        bucket_name=bucket,
                        object_key="",  # Not used for multiple files
                        attachment_flag=True,
                        report_attachment_flag="true",
                        multiple_files=multiple_files,
                        attached_multiple_file=True,
                        gmailtrigger=gmailtrigger,
                        context=None
                    )
                    
                    # Log email result
                    if email_result:
                        if "SES Access Denied" in str(email_result) or "Credentials not available" in str(email_result):
                            log_message("WARNING", f"12b1_compensation: Email notification failed - {email_result}")
                        else:
                            log_message("SUCCESS", f"12b1_compensation: Email notification sent: {email_result}")
                    else:
                        log_message("ERROR", "12b1_compensation: Email notification returned empty result")
               else:
                    log_message('ERROR', '12b1_compensation: Email configuration incomplete - missing from_email or to_emails')
               
               return success_count == total_reports
               
          except Exception as e:
               log_message('ERROR', f'12b1_compensation: Error generating separate files: {str(e)}')
               return False

     def _write_content_to_s3(self, content: str, output_path: str) -> bool:
          """Write content to S3 path"""
          try:
               from urllib.parse import urlparse
               
               # Parse S3 URL
               parsed_url = urlparse(output_path)
               bucket_name = parsed_url.netloc
               object_key = parsed_url.path.lstrip('/')
               s3.put_object(
                    Bucket=bucket_name,
                    Key=object_key,
                    Body=content.encode('utf-8'),
                    ContentType='text/plain'
               )
               
               return True
               
          except Exception as e:
               return False

     def _generate_12b1_standardized_header(self, config: Dict, data: List[Dict], format_type: str = 'txt') -> str:
          """Generate standardized header for 12b1 reports across all formats"""
          lines = []
          layout = config.get('layout', {})
          header_config = layout.get('header', {})
          
          title = header_config.get('title', '')
          date_position = header_config.get('date_position', 'right')

          if header_config.get('date') == 'dynamic':
                from datetime import datetime
                current_date = datetime.now().strftime('%m/%d/%Y')
          else:
                current_date = ''

          if title:
                if format_type == 'txt':
                    # Combine title and date in same line with spacing
                    # Adjust width (e.g. 80) to control distance between title and date
                    line_width = 80
                    line = f" {title:<{line_width - len(current_date)}}{current_date}"
                    lines.append(line)
                elif format_type == 'csv':
                    lines.append([title, current_date])
         
          # Generate subtitle
          subtitle = header_config.get('subtitle', '')
          if subtitle:
                   if format_type == 'txt':
                        lines.append(subtitle)
                   elif format_type == 'csv':
                        lines.append([subtitle])
         
          # Add spacing
          spacing = header_config.get('spacing', 1)
          for _ in range(spacing):
                   if format_type == 'txt':
                        lines.append("")
                   elif format_type == 'csv':
                        lines.append([""])
         
          if format_type == 'txt':
                   return '\n'.join(lines)
          elif format_type == 'csv':
                   return lines
          return ""

     def _generate_12b1_standardized_table_headers(self, columns: List[Dict], format_type: str = 'txt') -> str:
          """Generate standardized table headers across all formats"""
          if format_type == 'txt':
                   header_parts = []
                   for col in columns:
                        name = col.get('name', '')
                        width = col.get('width', 0)
                        align = col.get('align', 'left')
                   
                        if width > 0:
                             formatted = self.formatter.align_text(name, width, align)
                        else:
                             formatted = name
                        header_parts.append(formatted)
                   return self.formatter.tab_char.join(header_parts)
          elif format_type == 'csv':
                   return [col.get('name', '') for col in columns]
          return ""

     def _generate_12b1_standardized_data_row(self, columns: List[Dict], record: Dict, format_type: str = 'txt') -> str:
          """Generate standardized data row across all formats"""
          if format_type == 'txt':
                   return self.formatter.create_excel_row(columns, record)
          elif format_type == 'csv':
                   return [self._get_column_value_for_csv(col, record) for col in columns]
          return ""

     def _generate_12b1_suppression_section(self, section: Dict, data: List[Dict]) -> str:
          """Generate a suppression section for 12b1_compensation reports with standardized formatting"""
          lines = []
         
          # Add section title with standardized styling
          title = section.get('title', '')
          if title:
                   lines.append(title)
         
          # Add subtitle if provided
          subtitle = section.get('subtitle', '')
          if subtitle:
                   lines.append("") # Empty line separator
                   lines.append(subtitle)
         
          # Add date if provided
          date = section.get('date', '')
          if date:
                   lines.append("") # Empty line separator
                   lines.append(date)
         
          # Generate column headers with standardized alignment
          columns = section.get('columns', [])
          if columns:
                   # Use standardized header generation
                   header_line = self._generate_12b1_standardized_table_headers(columns, 'txt')
                   lines.append(header_line)
              
                   # Apply filtering based on CUSTOM_CHECK_NUMBER for 12b1_compensation reports
                   filtered_data = self._filter_12b1_data_by_check_number(data, section)
              
                   # Generate data rows for this section using standardized formatting
                   for record in filtered_data:
                        row = self._generate_12b1_standardized_data_row(columns, record, 'txt')
                        lines.append(row)
         
          # Add dynamic summary if configured
          summary_config = section.get('summary', {})
          if summary_config:
                   summary_lines = self._generate_12b1_section_summary(section, filtered_data)
                   lines.extend(summary_lines)
         
          # Add spacing between sections
          lines.append("")
         
          return '\n'.join(lines)

     def _generate_12b1_section_summary(self, section: Dict, data: List[Dict]) -> List[str]:
          """Generate summary for 12b1_compensation suppression sections"""
          lines = []
          summary_config = section.get('summary', {})
         
          if not summary_config:
                   return lines
         
          # Calculate summary values
          calculation = summary_config.get('calculation', '')
          if calculation == 'sum_amounts':
                   # Use CUSTOM_CHECK_AMOUNT for 12b1_compensation report to match the column display
                   # For 12b1_compensation report, sum all amounts in the filtered data (not just < $5.00)
                   total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
                   formatted_amount = self.formatter.format_currency(total_amount)
              
                   # Generate summary line
                   summary_line = summary_config.get('line', '')
                   if '{sum_amount}' in summary_line:
                        summary_line = summary_line.replace('{sum_amount}', formatted_amount)
              
                   lines.append(summary_line)
              
          elif calculation == 'count_records':
                   total_count = len(data)
              
                   # Generate summary line
                   summary_line = summary_config.get('line', '')
                   if '{count_rows}' in summary_line:
                        summary_line = summary_line.replace('{count_rows}', str(total_count))
              
                   lines.append(summary_line)
         
          # Add title after if configured
          title_after = summary_config.get('title_after', '')
          if title_after:
                   lines.append("")
                   lines.append(title_after)
         
          return lines

     def _filter_12b1_data_by_check_number(self, data: List[Dict], section: Dict) -> List[Dict]:
          """Filter data based on CUSTOM_CHECK_NUMBER for 12b1_compensation suppression sections"""
          # Determine filter type based on section title or filter configuration
          title = section.get('title', '').lower()
          subtitle = section.get('subtitle', '').lower()
          filter_type = section.get('filter', '')
         
          if 'zero' in title or 'zero' in subtitle or filter_type == 'zero_check_number':
                   # Table 2: Show only records where CUSTOM_CHECK_NUMBER = 0
                   return [record for record in data if self._is_zero_check_number(record.get('CUSTOM_CHECK_NUMBER', ''))]
          elif filter_type == 'non_zero_check_number' or filter_type == 'amount_less_than_5':
                   # Table 1: Show only records where CUSTOM_CHECK_NUMBER ≠ 0
                   return [record for record in data if not self._is_zero_check_number(record.get('CUSTOM_CHECK_NUMBER', ''))]
          else:
                   # Default: Show only records where CUSTOM_CHECK_NUMBER ≠ 0
                   return [record for record in data if not self._is_zero_check_number(record.get('CUSTOM_CHECK_NUMBER', ''))]

     def _is_zero_check_number(self, check_number: str) -> bool:
          """Check if a check number is considered zero"""
          if not check_number:
                   return True
          try:
                   # Convert to float to handle numeric strings
                   num_value = float(check_number)
                   return num_value == 0.0
          except (ValueError, TypeError):
                   # If it's not a number, check if it's '0' or empty
                   return check_number.strip() == '0' or check_number.strip() == ''

     def _generate_12b1_reconciliation_suppression_content(self, config: Dict, data: List[Dict],input_file_name:str) -> str:
          """Generate content for 12b1_RECONCILIATION_SUPPRESSION report with standardized styling"""
          
          bucket = os.getenv("s3BucketInput")
          input_path = os.getenv("inputPath")
          input_file_name = os.getenv("inputFileName")
          xml_path = f"s3://{bucket}/{input_path}{input_file_name}"
          xml_data= self.extractor.extract_data(xml_path, config)
          
          DOC_IMAGE_COUNT = config.get('DOC_IMAGE_COUNT', 'DOC_IMAGE_COUNT')
          CUSTOM_CHECK_AMOUNT = config.get('CUSTOM_CHECK_AMOUNT', 'CUSTOM_CHECK_AMOUNT')
          
          input_file_name_data=input_file_name.split('_')
          layout = config.get('layout', {})
          lines = []
          
          total_packages = len(xml_data)
          total_check_pages = sum(int(item.get(DOC_IMAGE_COUNT, 0)) for item in xml_data) 
          dollar_amount = sum(self.safe_float_convert(rec.get('CUSTOM_CHECK_AMOUNT', 0), default=0.0) for rec in xml_data)
          dollar_amount_fmt = self.format_currency(dollar_amount)

          # --- HEADER (reuse existing standardized header if needed) ---
          if 'header' in layout:
               header = self._generate_12b1_standardized_header(config, data, 'txt')
               header = header.rstrip()
               lines.append(header)

          # --- DATA SECTION (our new logic) ---
          results = self._generate_12b1_reconciliation_suppression_data_section(data)

          total_input = results["total_input_count"]
          suppressed_amt = results["suppressed_amount_count"]
          zero_chk = results["zero_check_count"]
          suppressed_total = suppressed_amt + zero_chk
          corp_no = input_file_name_data[0]
          cycle = input_file_name_data[2]
          
          lines.append("")
          lines.append(f" Reconciliation Report")
          lines.append(f"Corp: {corp_no:>42}")
          lines.append(f"Cycle: {cycle:>41}")
          lines.append("")
          lines.append(f"{'Total Input (codes 0014/0019):'} {total_input:>17}")
          lines.append(f"{'  Suppressed Amount < $5.00:'} {suppressed_amt:>19}")
          lines.append(f"{'  Suppressed Zero Check Number:'}{zero_chk:>17}")
          lines.append(f"{'Minus Suppressed Total:'} {suppressed_total:>24}")
          lines.append(f"{'Total Packages:'}{total_packages:>33}")
          lines.append(f"{'Total Check Pages:'}{total_check_pages:>30}")
          lines.append(f"{'Dollar Amount:'}{dollar_amount_fmt:>34}")
          lines.append("")

          # --- Suppression sections ---
          suppression_sections = layout.get('suppression_sections', [])

          for section in suppression_sections:
               lines.append("-------------")
               lines.append("")
               if 'header' in layout:
                    header = self._generate_12b1_standardized_header(config, data, 'txt')
                    header = header.rstrip()
                    lines.append(header)
               title = section.get("title", "")
               subtitle = section.get("subtitle", "")
               # lines.append(title)
               if subtitle:
                    lines.append("")
                    lines.append(f" {subtitle}")
                    lines.append("")
               # lines.append("Fund  Group  Inst  Check#  Amount  Name")
               lines.append(f"{'Fund':<8}{'Group':<10}{'Inst':<10}{'Check #':<15}{'Amount':>8}  {'Name'}")

               if "Compensation Amount < $5.00" in title or " < $5.00" in subtitle:
                    for r in results["suppressed_amount_rows"]:
                         if r['checkNumber']!='0000000000000':
                              lines.append(f"{r['Suppression']:6}  {r['Group']:7}  {r['Inst']:7}  {r['checkNumber']:10}  "
                                        f"{r['checkAmount']:>10}  {r['Name']}")
                    lines.append("")     
                    lines.append(f"Total Suppressed - Compensation Amount < $5.00 = {suppressed_amt}")
                    lines.append("")
               elif "ZeroCheckNumber" in title or "Zero Check" in subtitle:
                    for r in results["zero_check_rows"]:
                         lines.append(f"{r['Suppression']:6}  {r['Group']:7}  {r['Inst']:7}  {r['checkNumber']:10}  "
                                   f"{r['checkAmount']:>10}  {r['Name']}")
                    lines.append("")
                    lines.append(f" Total Suppressed - Zero Check Number = {zero_chk}")
                    lines.append("")
                    lines.append("")
                    lines.append("-------------")

          return "\n".join(lines)

     def _generate_12b1_fund_report_content(self, config: Dict, data: List[Dict]) -> str:
          """Generate content for 12b1_fund_report matching the provided layout style"""
          content_parts = []
          layout = config.get('layout', {})
         
          # Generate title (Row 1)
          title = "Janus 12B1 / Service Fee Check Application"
          content_parts.append(title)
         
          # Generate date report line (Row 2)
          from datetime import datetime
          current_date = datetime.now().strftime('%m/%Y')
          date_line = f"{current_date} Fund Report"
          content_parts.append(date_line)
         
          # Generate headers (Row 3)
          headers = f"{'Fund #':<15}{'Count':<10}"
          content_parts.append(headers)
         
          # Generate data section with fund codes and counts
          if 'data_section' in layout:
                   data_content = self._generate_12b1_fund_report_data_section(config, data)
                   content_parts.append(data_content)
         
          return '\n'.join(content_parts)

     def _generate_12b1_fund_report_data_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate data section for 12b1_fund_report matching the provided layout"""
          # Check if this is tilde-separated data (new format) or XML data (legacy format)
          if data and 'Fund' in data[0] and 'Count' in data[0]:
                   # New tilde-separated format - use Fund and Count fields directly
                   # Apply business logic: only process records where Count = "0002" and group by Fund
                   # First, filter records where Count = "0002"
                   filtered_records = []
                   for i, record in enumerate(data):
                            fund_code = record.get('Fund', '')
                            count = record.get('Count', '')
                            # Filter: only process records where Count = "0002"
                            if count == "0002":
                                   filtered_records.append(record)
                   # Group filtered records by Fund and count occurrences
                   fund_counts = {}
                   for record in filtered_records:
                            fund_code = record.get('Fund', '')
                            if fund_code:
                                   fund_counts[fund_code] = fund_counts.get(fund_code, 0) + 1
                   # Sort fund codes numerically for consistent output
                   sorted_funds = sorted(fund_counts.keys(), key=lambda x: int(x) if x.isdigit() else float('inf'))
                   # Generate grouped data rows
                   data_lines = []
                   for fund_code in sorted_funds:
                            count = fund_counts[fund_code]
                            data_lines.append(f"{fund_code:<15}{count:<10}")
                   return '\n'.join(data_lines)
          else:
                   # Legacy XML format - group data by fund code and count records
                   fund_counts = {}
                   for record in data:
                            fund_code = record.get('UC_FUND_CODE', '')
                            if fund_code:
                                 fund_counts[fund_code] = fund_counts.get(fund_code, 0) + 1
                   # Sort fund codes numerically
                   sorted_funds = sorted(fund_counts.keys(), key=lambda x: int(x) if x.isdigit() else float('inf'))
                   # Generate data rows
                   data_lines = []
                   for fund_code in sorted_funds:
                            count = fund_counts[fund_code]
                            data_lines.append(f"{fund_code:<15}{count:<10}")
                   return '\n'.join(data_lines)
 
 
     def _generate_12b1_check_report_content(self, config: Dict, data: List[Dict]) -> str:
          """Generate content for 12b1_check_report matching the provided layout style"""
          content_parts = []
          layout = config.get('layout', {})
         
          # Generate title (Row 1)
          title = "Janus 12B1 / Service Fee Check Application"
          content_parts.append(title)
         
          # Generate date report line (Row 2)
          from datetime import datetime
          current_date = datetime.now().strftime('%m/%Y')
          date_line = f"{current_date} Check Report"
          content_parts.append(date_line)
         
          # Generate empty row (Row 3)
          content_parts.append("")
          headers = (
          f"{'Check No.':<15}"   # shorter column for check number
          f"{'Check Amt':<15}"   # right align amounts for clarity
          f"{'Fund':<10}"        # fund code or name
          f"{'Account No.':<20}" # full account number, fixed width
          )
          content_parts.append(headers)
          # Generate data section with check data
          if 'data_section' in layout:
                   data_content = self._generate_12b1_check_report_data_section(config, data)
                   content_parts.append(data_content)
         
          return '\n'.join(content_parts)

     def _generate_12b1_check_report_data_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate data section for 12b1_check_report using JSON configuration mapping"""
          
          # Get column configuration from JSON
          layout = config.get('layout', {})
          data_section = layout.get('data_section', {})
          columns = data_section.get('columns', [])
          
          
          # Check if this is tilde-separated data (new format) or XML data (legacy format)
          if data and 'identifier' in data[0]:
               
               # Implement dynamic mapping between 0002 and 0014 entries using JSON config
               grouped_transactions = self._create_dynamic_0002_0014_mapping(data, config)
               
               # Generate data rows using JSON column configuration
               data_lines = []
               for transaction in grouped_transactions:
                    check_number = transaction.get('check_number', '')
                    check_amount = transaction.get('check_amount', '')
                    related_0002_records = transaction.get('related_0002_records', [])
                    
                    # Format check amount as currency using JSON configuration
                    if check_amount:
                         try:
                              check_amt_formatted = f"{float(check_amount):.2f}"
                         except (ValueError, TypeError):
                              check_amt_formatted = str(check_amount)
                    else:
                         check_amt_formatted = ""
                    
                    # Use JSON column widths for formatting
                    check_width = next((col.get('width', 20) for col in columns if col.get('name') == 'Check No.'), 20)
                    amount_width = next((col.get('width', 15) for col in columns if col.get('name') == 'Check Amt.'), 15)
                    fund_width = next((col.get('width', 10) for col in columns if col.get('name') == 'Fund'), 10)
                    account_width = next((col.get('width', 20) for col in columns if col.get('name') == 'Account No.'), 20)
                    
                    # Process each unique fund/account combination and show it twice (once for 0002, once for 0014)
                    for i, record in enumerate(related_0002_records):
                         fund = record.get('fund', '')
                         account_number = record.get('account_number', '')
                         
                         # First occurrence shows check number and amount (0002 record)
                         if i == 0:
                              data_lines.append(f"{check_number:<{check_width}}{check_amt_formatted:<{amount_width}}{fund:<{fund_width}}{account_number:<{account_width}}")
                         else:
                              # Subsequent records show empty check number and amount
                              data_lines.append(f"{'':<{check_width}}{'':<{amount_width}}{fund:<{fund_width}}{account_number:<{account_width}}")
                         
                         # Second occurrence shows empty check number and amount (0014 record)
                         data_lines.append(f"{'':<{check_width}}{'':<{amount_width}}{fund:<{fund_width}}{account_number:<{account_width}}")
               
               return '\n'.join(data_lines)
          else:
               # Legacy XML data format - use original logic
               # Group data by check number to handle the special formatting
               check_groups = {}
               for record in data:
                    check_number = record.get('FN_CHECK_NUMBER', '')
                    if check_number not in check_groups:
                         check_groups[check_number] = []
                    check_groups[check_number].append(record)
               
               # Generate data rows with special formatting
               data_lines = []
               for check_number, records in check_groups.items():
                    # First record shows check number and amount
                    first_record = records[0]
                    check_amt = first_record.get('FN_NET_AMOUNT', '')
                    fund = first_record.get('UC_FUND_CODE', '')
                    account = first_record.get('UC_ACCOUNT_NUMBER', '')
                    
                    # Format check amount as currency
                    if check_amt:
                         try:
                              check_amt_formatted = f"{float(check_amt):.2f}"
                         except (ValueError, TypeError):
                              check_amt_formatted = str(check_amt)
                    else:
                         check_amt_formatted = ""
                    
                    data_lines.append(f"{check_number:<20}{check_amt_formatted:<15}{fund:<10}{account:<20}")
                    
                    # Subsequent records show empty check number and amount
                    for record in records[1:]:
                         fund = record.get('UC_FUND_CODE', '')
                         account = record.get('UC_ACCOUNT_NUMBER', '')
                         data_lines.append(f"{'':<20}{'':<15}{fund:<10}{account:<20}")
               
               return '\n'.join(data_lines)

     def _create_dynamic_0002_0014_mapping(self, data: List[Dict], config: Dict = None) -> List[Dict]:
          """Create dynamic mapping between 0002 and 0014 entries for 12b1_check_report
          Uses 3-index value to properly group 0002 and 0014 records"""
          
          # Group records by 3-index value (index 3 in the tilde-separated data)
          groups_by_index = {}
          
          for record in data:
               identifier = record.get('identifier', '')
               index_3 = record.get('index_3', '')  # The 3-index value
               
               # Use fallback grouping if index_3 is empty
               if not index_3:
                    # Group by check number as fallback
                    check_number = record.get('check_number', '')
                    if identifier == '0014' and check_number:
                         index_3 = f"check_{check_number}"
                    else:
                         index_3 = "unknown"
               
               if index_3 not in groups_by_index:
                    groups_by_index[index_3] = {'0002': [], '0014': []}
               
               if identifier == '0002':
                    groups_by_index[index_3]['0002'].append(record)
               elif identifier == '0014':
                    groups_by_index[index_3]['0014'].append(record)
          
          
          # Create grouped transactions based on 3-index value
          grouped_transactions = []
          
          for index_3, group_data in groups_by_index.items():
               records_0002 = group_data['0002']
               records_0014 = group_data['0014']
               
               # Process each 0014 record in this group
               for record_0014 in records_0014:
                    check_number = record_0014.get('check_number', '')
                    check_amount = record_0014.get('check_amount', '')
                    
                    # Skip if check number is empty or zero
                    if not check_number or check_number == '0' or check_number == '0000000000000':
                         continue

                    try:
                         amount_value = float(check_amount) if check_amount else 0.0
                         if amount_value <= 5.0:
                              continue
                    except (ValueError, TypeError):
                         continue

                    
                    # Group 0002 records by fund/account combination to avoid duplicates
                    unique_fund_accounts = {}
                    for record_0002 in records_0002:
                         fund = record_0002.get('fund', '')
                         account_number = record_0002.get('account_number', '')
                         key = f"{fund}|{account_number}"
                         
                         if key not in unique_fund_accounts:
                              unique_fund_accounts[key] = record_0002
                    
                    # Convert back to list
                    unique_0002_records = list(unique_fund_accounts.values())
                    
                    # Create transaction with unique 0002 records from the same group
                    if unique_0002_records:
                         grouped_transaction = {
                              'check_number': check_number,
                              'check_amount': check_amount,
                              'related_0002_records': unique_0002_records,
                              'index_3': index_3
                         }
                         grouped_transactions.append(grouped_transaction)
          
          return grouped_transactions

     def _generate_12b1_compensation_csv(self, config: Dict, data: List[Dict], output_path: str) -> bool:
          """Generate CSV report for 12b1_compensation with standardized styling across all formats"""
          try:
                   csv_lines = []
              
                   # Generate 12b1_RECONCILIATION_SUPPRESSION report with standardized styling
                   if '12b1_RECONCILIATION_SUPPRESSION' in config:
                        reconciliation_lines = self._generate_12b1_reconciliation_suppression_csv(
                             config['12b1_RECONCILIATION_SUPPRESSION'], data)
                        csv_lines.extend(reconciliation_lines)
                        csv_lines.append([""])  # Add spacing between reports
              
                   # Generate 12b1_fund_report with standardized styling
                   if '12b1_fund_report' in config:
                        fund_lines = self._generate_12b1_fund_report_csv(
                             config['12b1_fund_report'], data)
                        csv_lines.extend(fund_lines)
                        csv_lines.append([""])  # Add spacing between reports
              
                   # Generate 12b1_check_report with standardized styling
                   if '12b1_check_report' in config:
                        check_lines = self._generate_12b1_check_report_csv(
                             config['12b1_check_report'], data)
                        csv_lines.extend(check_lines)
              
                   # Write CSV file
                   with open(output_path, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f, delimiter=',', quotechar='"')
                        for row in csv_lines:
                             writer.writerow(row)
              
                   return True
          except Exception as e:
                   return False

     def _generate_12b1_compensation_xlsx(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX content for 12b1_compensation with three child reports"""
          current_row = start_row
         
          # Generate 12b1_RECONCILIATION_SUPPRESSION report
          if '12b1_RECONCILIATION_SUPPRESSION' in config:
                   current_row = self._generate_12b1_reconciliation_suppression_xlsx(
                        ws, config['12b1_RECONCILIATION_SUPPRESSION'], data, current_row)
                   current_row += 2  # Add spacing between reports
         
          # Generate 12b1_fund_report
          if '12b1_fund_report' in config:
                   current_row = self._generate_12b1_fund_report_xlsx(
                        ws, config['12b1_fund_report'], data, current_row)
                   current_row += 2  # Add spacing between reports
         
          # Generate 12b1_check_report
          if '12b1_check_report' in config:
                   current_row = self._generate_12b1_check_report_xlsx(
                        ws, config['12b1_check_report'], data, current_row)
         
          return current_row
     
     def _generate_12b1_reconciliation_suppression_data_section(self, data: List[Dict]) -> Dict[str, Any]:
          """
          Extract suppression data for 12b1_RECONCILIATION_SUPPRESSION report.
          Returns counts and row details for the two suppression sections.
          """
          # Helper conversion
          def to_float(v):
               try:
                    return float(str(v).replace('*', '').strip())
               except Exception:
                    return 0.0

          total_input_count = 0
          suppressed_amount_rows = []
          zero_check_rows = []

          for rec in data:
               record_type = rec.get("RecordType", "").strip()
               if record_type != "0014":
                    continue
               total_input_count += 1

               check_amount = to_float(rec.get("checkAmount", 0))
               check_number = str(rec.get("checkNumber", "")).strip()
               name = rec.get("Name", "")
               fund = rec.get("Suppression", "")
               group = rec.get("Group", "")
               inst = rec.get("Inst", "")


               # Suppressed zero check number
               if check_number in ("", "0", "0000000000000"):
                    zero_check_rows.append({
                         "Suppression": fund,
                         "Group": group,
                         "Inst": inst,
                         "checkNumber": check_number,
                         "checkAmount": f"{check_amount:.2f}",
                         "Name": name
                    })

               # Suppress ONLY when amount < 5 AND check number is valid (not zero or blank)
               if check_amount < 5.00 and check_number not in ("", "0", "0000000000000"):
                    suppressed_amount_rows.append({
                         "Suppression": fund,
                         "Group": group,
                         "Inst": inst,
                         "checkNumber": check_number,
                         "checkAmount": f"{check_amount:.2f}",
                         "Name": name
                    })

          return {
               "total_input_count": total_input_count,
               "suppressed_amount_count": len(suppressed_amount_rows),
               "zero_check_count": len(zero_check_rows),
               "suppressed_amount_rows": suppressed_amount_rows,
               "zero_check_rows": zero_check_rows,
          }

     def _generate_12b1_reconciliation_suppression_xlsx(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX content for 12b1_RECONCILIATION_SUPPRESSION report with standardized styling"""
          current_row = start_row
          layout = config.get('layout', {})
         
          # Generate standardized header
          if 'header' in layout:
                   current_row = self._generate_12b1_standardized_xlsx_header(ws, config, data, current_row)
         
          # Generate summary section
          if 'summary_section' in layout:
                   current_row = self._generate_xlsx_summary_section(ws, config, data, current_row)
         
          # Generate totals section
          if 'totals_section' in layout:
                   current_row = self._generate_xlsx_totals_section(ws, config, data, current_row)
         
          # Generate suppression sections with standardized formatting
          suppression_sections = layout.get('suppression_sections', [])
          for section in suppression_sections:
                   current_row = self._generate_12b1_suppression_section_xlsx(ws, section, data, current_row)
         
          return current_row

     def _generate_12b1_fund_report_xlsx(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX content for 12b1_fund_report matching the provided layout style"""
          current_row = start_row
          layout = config.get('layout', {})
         
          # Generate title (Row 1)
          title = "Janus 12B1 / Service Fee Check Application"
          cell = ws.cell(row=current_row, column=1, value=title)
          cell.font = Font(bold=True, size=12)
          current_row += 1
         
          # Generate date report line (Row 2)
          from datetime import datetime
          current_date = datetime.now().strftime('%m/%Y')
          date_line = f"{current_date} Fund Report"
          cell = ws.cell(row=current_row, column=1, value=date_line)
          cell.font = Font(bold=True, size=10)
          current_row += 1
         
          # Generate headers (Row 3)
          headers = ["Fund #(fund code)", "Count"]
          for col_idx, header in enumerate(headers, 1):
                   cell = ws.cell(row=current_row, column=col_idx, value=header)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='center')
          current_row += 1
         
          # Generate data section with fund codes and counts
          if 'data_section' in layout:
                   current_row = self._generate_12b1_fund_report_xlsx_data(ws, config, data, current_row)
         
          return current_row

     def _generate_12b1_fund_report_xlsx_data(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX data section for 12b1_fund_report matching the provided layout"""
          current_row = start_row
          
          # Check if this is tilde-separated data (new format) or XML data (legacy format)
          if data and 'Fund' in data[0] and 'Count' in data[0]:
                   # New tilde-separated format - use Fund and Count fields directly
                   # Apply business logic: only process records where Count = "0002" and group by Fund
                   # First, filter records where Count = "0002"
                   filtered_records = []
                   for i, record in enumerate(data):
                            fund_code = record.get('Fund', '')
                            count = record.get('Count', '')
                            # Filter: only process records where Count = "0002"
                            if count == "0002":
                                   filtered_records.append(record)
                   # Group filtered records by Fund and count occurrences
                   fund_counts = {}
                   for record in filtered_records:
                            fund_code = record.get('Fund', '')
                            if fund_code:
                                   fund_counts[fund_code] = fund_counts.get(fund_code, 0) + 1
                   # Sort fund codes numerically for consistent output
                   sorted_funds = sorted(fund_counts.keys(), key=lambda x: int(x) if x.isdigit() else float('inf'))
                   # Generate data rows
                   for fund_code in sorted_funds:
                            count = fund_counts[fund_code]
                            ws.cell(row=current_row, column=1, value=fund_code)
                            ws.cell(row=current_row, column=2, value=count)
                            current_row += 1
          else:
                   # Legacy XML format - group data by fund code and count records
                   fund_counts = {}
                   for record in data:
                            fund_code = record.get('UC_FUND_CODE', '')
                            if fund_code:
                                 fund_counts[fund_code] = fund_counts.get(fund_code, 0) + 1
                   # Sort fund codes numerically
                   sorted_funds = sorted(fund_counts.keys(), key=lambda x: int(x) if x.isdigit() else float('inf'))
                   # Generate data rows
                   for fund_code in sorted_funds:
                            count = fund_counts[fund_code]
                            ws.cell(row=current_row, column=1, value=fund_code)
                            ws.cell(row=current_row, column=2, value=count)
                            current_row += 1
         
          return current_row

     def _generate_12b1_check_report_xlsx(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX content for 12b1_check_report matching the provided layout style"""
          current_row = start_row
          layout = config.get('layout', {})
         
          # Generate title (Row 1)
          title = "Janus 12B1 / Service Fee Check Application"
          cell = ws.cell(row=current_row, column=1, value=title)
          cell.font = Font(bold=True, size=12)
          current_row += 1
         
          # Generate date report line (Row 2)
          from datetime import datetime
          current_date = datetime.now().strftime('%m/%Y')
          date_line = f"{current_date} Check Report"
          cell = ws.cell(row=current_row, column=1, value=date_line)
          cell.font = Font(bold=True, size=10)
          current_row += 1
         
          # Generate empty row (Row 3)
          current_row += 1
         
          # Generate headers (Row 4)
          headers = ["Check No.", "Check Amt", "Fund", "Account No."]
          for col_idx, header in enumerate(headers, 1):
                   cell = ws.cell(row=current_row, column=col_idx, value=header)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='center')
          current_row += 1
         
          # Generate data section with check data
          if 'data_section' in layout:
                   current_row = self._generate_12b1_check_report_xlsx_data(ws, config, data, current_row)
         
          return current_row

     def _generate_12b1_check_report_xlsx_data(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX data section for 12b1_check_report using JSON configuration"""
          current_row = start_row
          
          # Get column configuration from JSON
          layout = config.get('layout', {})
          data_section = layout.get('data_section', {})
          columns = data_section.get('columns', [])
          
          
          # Check if data is from section-based extraction
          if data and 'identifier' in data[0]:
               # Data is from tilde-separated format - use JSON-driven mapping with grouping
               grouped_transactions = self._create_dynamic_0002_0014_mapping(data, config)
               for transaction in grouped_transactions:
                    check_number = transaction.get('check_number', '')
                    check_amount = transaction.get('check_amount', '')
                    related_0002_records = transaction.get('related_0002_records', [])
                    
                    # Format check amount as currency using JSON configuration
                    if check_amount:
                         try:
                              check_amt_value = float(check_amount)
                         except (ValueError, TypeError):
                              check_amt_value = 0
                    else:
                         check_amt_value = 0
                    
                    # Process each unique fund/account combination and show it twice (once for 0002, once for 0014)
                    for i, record in enumerate(related_0002_records):
                         fund = record.get('fund', '')
                         account_number = record.get('account_number', '')
                         
                         # First occurrence shows check number and amount (0002 record)
                         col_idx = 1
                         for col in columns:
                              if col.get('name') == 'Check No.':
                                   ws.cell(row=current_row, column=col_idx, value=check_number if i == 0 else "")
                              elif col.get('name') == 'Check Amt.':
                                   ws.cell(row=current_row, column=col_idx, value=check_amt_value if i == 0 else "")
                              elif col.get('name') == 'Fund':
                                   ws.cell(row=current_row, column=col_idx, value=fund)
                              elif col.get('name') == 'Account No.':
                                   ws.cell(row=current_row, column=col_idx, value=account_number)
                              col_idx += 1
                         current_row += 1
                         
                         # Second occurrence shows empty check number and amount (0014 record)
                         col_idx = 1
                         for col in columns:
                              if col.get('name') == 'Check No.':
                                   ws.cell(row=current_row, column=col_idx, value="")
                              elif col.get('name') == 'Check Amt.':
                                   ws.cell(row=current_row, column=col_idx, value="")
                              elif col.get('name') == 'Fund':
                                   ws.cell(row=current_row, column=col_idx, value=fund)
                              elif col.get('name') == 'Account No.':
                                   ws.cell(row=current_row, column=col_idx, value=account_number)
                              col_idx += 1
                         current_row += 1
          elif data and 'check_number' in data[0]:
               # Data is from section-based extraction - use corrected logic
               for record in data:
                    check_number = record.get('check_number', '')
                    check_amount = record.get('check_amount', '')
                    fund = record.get('fund', '')
                    account_number = record.get('account_number', '')
                    
                    # Format check amount as currency
                    if check_amount:
                         try:
                              check_amt_value = float(check_amount)
                         except (ValueError, TypeError):
                              check_amt_value = 0
                    else:
                         check_amt_value = 0
                    
                    ws.cell(row=current_row, column=1, value=check_number)
                    ws.cell(row=current_row, column=2, value=check_amt_value)
                    ws.cell(row=current_row, column=3, value=fund)
                    ws.cell(row=current_row, column=4, value=account_number)
                    current_row += 1
          else:
               # Legacy data format - use original logic
               # Group data by check number to handle the special formatting
               check_groups = {}
               for record in data:
                    check_number = record.get('FN_CHECK_NUMBER', '')
                    if check_number not in check_groups:
                         check_groups[check_number] = []
                    check_groups[check_number].append(record)
               
               # Generate data rows with special formatting
               for check_number, records in check_groups.items():
                    # First record shows check number and amount
                    first_record = records[0]
                    check_amt = first_record.get('FN_NET_AMOUNT', '')
                    fund = first_record.get('UC_FUND_CODE', '')
                    account = first_record.get('UC_ACCOUNT_NUMBER', '')
                    
                    # Format check amount as currency
                    if check_amt:
                         try:
                              check_amt_value = float(check_amt)
                         except (ValueError, TypeError):
                              check_amt_value = 0
                    else:
                         check_amt_value = 0
                    
                    ws.cell(row=current_row, column=1, value=check_number)
                    ws.cell(row=current_row, column=2, value=check_amt_value)
                    ws.cell(row=current_row, column=3, value=fund)
                    ws.cell(row=current_row, column=4, value=account)
                    current_row += 1
                    
                    # Subsequent records show empty check number and amount
                    for record in records[1:]:
                         fund = record.get('UC_FUND_CODE', '')
                         account = record.get('UC_ACCOUNT_NUMBER', '')
                         ws.cell(row=current_row, column=1, value="")
                         ws.cell(row=current_row, column=2, value="")
                         ws.cell(row=current_row, column=3, value=fund)
                         ws.cell(row=current_row, column=4, value=account)
                         current_row += 1
          
          return current_row

     def _generate_12b1_suppression_section_xlsx(self, ws, section: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX content for 12b1 suppression sections with standardized styling"""
          current_row = start_row
         
          # Add section title with consistent styling
          title = section.get('title', '')
          if title:
                   cell = ws.cell(row=current_row, column=1, value=title)
                   cell.font = Font(bold=True, size=12)
                   cell.alignment = Alignment(horizontal='left', vertical='top')
                   current_row += 1
         
          # Add subtitle if provided with consistent styling
          subtitle = section.get('subtitle', '')
          if subtitle:
                   cell = ws.cell(row=current_row, column=1, value=subtitle)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top')
                   current_row += 1
         
          # Generate column headers with consistent styling
          columns = section.get('columns', [])
          if columns:
                   for col_idx, col in enumerate(columns, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=col.get('name', ''))
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        # Add background color for headers
                        from openpyxl.styles import PatternFill
                        cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                   current_row += 1
              
                   # Apply filtering based on CUSTOM_CHECK_NUMBER for 12b1_compensation reports
                   filtered_data = self._filter_12b1_data_by_check_number(data, section)
              
                   # Generate data rows for this section with consistent styling
                   for record in filtered_data:
                        for col_idx, col in enumerate(columns, 1):
                             value = self._get_column_value_for_xlsx(col, record)
                             cell = ws.cell(row=current_row, column=col_idx, value=value)
                             cell.font = Font(size=9)
                             cell.alignment = Alignment(horizontal='left', vertical='center')
                             if col.get('format') == 'currency':
                                  cell.number_format = '$#,##0.00'
                                  cell.alignment = Alignment(horizontal='right', vertical='center')
                        current_row += 1
         
          # Add dynamic summary if configured with consistent styling
          summary_config = section.get('summary', {})
          if summary_config:
                   summary_lines = self._generate_12b1_section_summary(section, filtered_data)
                   for line in summary_lines:
                        cell = ws.cell(row=current_row, column=1, value=line)
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top')
                        current_row += 1
         
          return current_row

     def _generate_12b1_reconciliation_suppression_csv(self, config: Dict, data: List[Dict]) -> List[List[str]]:
          """Generate CSV content for 12b1_RECONCILIATION_SUPPRESSION report with standardized styling"""
          csv_lines = []
          layout = config.get('layout', {})
         
          # Generate standardized header
          if 'header' in layout:
                   header_lines = self._generate_12b1_standardized_header(config, data, 'csv')
                   csv_lines.extend(header_lines)
         
          # Generate summary section
          if 'summary_section' in layout:
                   summary = self._generate_summary_section(config, data)
                   summary_lines = summary.split('\n')
                   for line in summary_lines:
                        if line.strip():
                             csv_lines.append([line])
         
          # Generate totals section
          if 'totals_section' in layout:
                   totals = self._generate_totals_section(config, data)
                   totals_lines = totals.split('\n')
                   for line in totals_lines:
                        if line.strip():
                             csv_lines.append([line])
         
          # Generate suppression sections with standardized formatting
          suppression_sections = layout.get('suppression_sections', [])
         
          for section in suppression_sections:
                   section_lines = self._generate_12b1_suppression_section_csv(section, data)
                   csv_lines.extend(section_lines)
         
          return csv_lines

     def _generate_12b1_fund_report_csv(self, config: Dict, data: List[Dict]) -> List[List[str]]:
          """Generate CSV content for 12b1_fund_report matching the provided layout style"""
          csv_lines = []
          layout = config.get('layout', {})
         
          # Generate title (Row 1)
          title = "Janus 12B1 / Service Fee Check Application"
          csv_lines.append([title])
         
          # Generate date report line (Row 2)
          from datetime import datetime
          current_date = datetime.now().strftime('%m/%Y')
          date_line = f"{current_date} Fund Report"
          csv_lines.append([date_line])
         
          # Generate headers (Row 3)
          headers = "Fund #(fund code),Count"
          csv_lines.append([headers])
         
          # Generate data section with fund codes and counts
          if 'data_section' in layout:
                   data_content = self._generate_12b1_fund_report_data_section(config, data)
                   data_lines = data_content.split('\n')
                   for line in data_lines:
                        if line.strip():
                             csv_lines.append([line])
         
          return csv_lines

     def _generate_12b1_check_report_csv(self, config: Dict, data: List[Dict]) -> List[List[str]]:
          """Generate CSV content for 12b1_check_report matching the provided layout style"""
          csv_lines = []
          layout = config.get('layout', {})
         
          # Generate title (Row 1)
          title = "Janus 12B1 / Service Fee Check Application"
          csv_lines.append([title])
         
          # Generate date report line (Row 2)
          from datetime import datetime
          current_date = datetime.now().strftime('%m/%Y')
          date_line = f"{current_date} Check Report"
          csv_lines.append([date_line])
         
          # Generate empty row (Row 3)
          csv_lines.append([""])
         
          # Generate headers (Row 4)
          headers = "Check No.,Check Amt,Fund,Account No."
          csv_lines.append([headers])
         
          # Generate data section with check data using JSON configuration
          if 'data_section' in layout:
               # Get column configuration from JSON
               data_section = layout.get('data_section', {})
               columns = data_section.get('columns', [])
               
               
               # Check if data is from section-based extraction
               if data and 'identifier' in data[0]:
                    # Data is from tilde-separated format - use JSON-driven mapping with grouping
                    grouped_transactions = self._create_dynamic_0002_0014_mapping(data, config)
                    for transaction in grouped_transactions:
                         check_number = transaction.get('check_number', '')
                         check_amount = transaction.get('check_amount', '')
                         related_0002_records = transaction.get('related_0002_records', [])
                         
                         # Format check amount as currency using JSON configuration
                         if check_amount:
                              try:
                                   check_amt_formatted = f"{float(check_amount):.2f}"
                              except (ValueError, TypeError):
                                   check_amt_formatted = str(check_amount)
                         else:
                              check_amt_formatted = ""
                         
                         # Process each unique fund/account combination and show it twice (once for 0002, once for 0014)
                         for i, record in enumerate(related_0002_records):
                              fund = record.get('fund', '')
                              account_number = record.get('account_number', '')
                              
                              # First occurrence shows check number and amount (0002 record)
                              if i == 0:
                                   csv_lines.append([f"{check_number},{check_amt_formatted},{fund},{account_number}"])
                              else:
                                   csv_lines.append([f",,{fund},{account_number}"])
                              
                              # Second occurrence shows empty check number and amount (0014 record)
                              csv_lines.append([f",,{fund},{account_number}"])
               elif data and 'check_number' in data[0]:
                    # Data is from section-based extraction - use corrected logic
                    for record in data:
                         check_number = record.get('check_number', '')
                         check_amount = record.get('check_amount', '')
                         fund = record.get('fund', '')
                         account_number = record.get('account_number', '')
                         
                         # Format check amount as currency
                         if check_amount:
                              try:
                                   check_amt_formatted = f"{float(check_amount):.2f}"
                              except (ValueError, TypeError):
                                   check_amt_formatted = str(check_amount)
                         else:
                              check_amt_formatted = ""
                         
                         csv_lines.append([f"{check_number},{check_amt_formatted},{fund},{account_number}"])
               else:
                    # Legacy data format - use original logic
                   data_content = self._generate_12b1_check_report_data_section(config, data)
                   data_lines = data_content.split('\n')
                   for line in data_lines:
                        if line.strip():
                             csv_lines.append([line])
         
          return csv_lines

     def _generate_12b1_suppression_section_csv(self, section: Dict, data: List[Dict]) -> List[List[str]]:
          """Generate CSV content for 12b1 suppression sections with standardized formatting"""
          csv_lines = []
         
          # Add section title with standardized styling
          title = section.get('title', '')
          if title:
                   csv_lines.append([title])
         
          # Add subtitle if provided
          subtitle = section.get('subtitle', '')
          if subtitle:
                   csv_lines.append([""]) # Empty row separator
                   csv_lines.append([subtitle])
         
          # Add date if provided
          date = section.get('date', '')
          if date:
                   csv_lines.append([""]) # Empty row separator
                   csv_lines.append([date])
         
          # Generate column headers with standardized alignment
          columns = section.get('columns', [])
          if columns:
                   # Use standardized header generation
                   header_row = self._generate_12b1_standardized_table_headers(columns, 'csv')
                   csv_lines.append(header_row)
              
                   # Apply filtering based on CUSTOM_CHECK_NUMBER for 12b1_compensation reports
                   filtered_data = self._filter_12b1_data_by_check_number(data, section)
              
                   # Generate data rows for this section using standardized formatting
                   for record in filtered_data:
                        row = self._generate_12b1_standardized_data_row(columns, record, 'csv')
                        csv_lines.append(row)
         
          # Add dynamic summary if configured
          summary_config = section.get('summary', {})
          if summary_config:
                   summary_lines = self._generate_12b1_section_summary(section, filtered_data)
                   for line in summary_lines:
                        csv_lines.append([line])
         
          # Add spacing between sections
          csv_lines.append([""])
         
          return csv_lines

     def _generate_12b1_standardized_xlsx_header(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate standardized XLSX header for 12b1 reports with consistent styling"""
          current_row = start_row
          layout = config.get('layout', {})
          header_config = layout.get('header', {})
         
          # Generate title with consistent styling
          title = header_config.get('title', '')
          if title:
                   cell = ws.cell(row=current_row, column=1, value=title)
                   cell.font = Font(bold=True, size=14)
                   cell.alignment = Alignment(horizontal='left', vertical='top')
                   current_row += 1
         
          # Generate subtitle with consistent styling
          subtitle = header_config.get('subtitle', '')
          if subtitle:
                   cell = ws.cell(row=current_row, column=1, value=subtitle)
                   cell.font = Font(bold=True, size=12)
                   cell.alignment = Alignment(horizontal='left', vertical='top')
                   current_row += 1
         
          # Generate date if dynamic with consistent styling
          if header_config.get('date') == 'dynamic':
                   from datetime import datetime
                   current_date = datetime.now().strftime('%m/%d/%Y')
                   cell = ws.cell(row=current_row, column=1, value=f"Date: {current_date}")
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top')
                   current_row += 1
         
          # Add spacing with consistent formatting
          spacing = header_config.get('spacing', 1)
          current_row += spacing
         
          return current_row

     def _get_column_value_for_csv(self, col: Dict, data: Dict) -> str:
          """Get column value for CSV generation"""
          source_field = col.get('source_field', '')
          calculation = col.get('calculation', '')
          format_type = col.get('format', 'text')
         
          if calculation == 'extract_group':
                   # Extract group from dealer number (first 2 digits)
                   dealer_number = data.get('FN_DEALER_NUMBER', '')
                   if dealer_number and len(dealer_number) >= 2:
                        return dealer_number[:2]
                   return ''
          elif source_field == 'calculated':
                   # Handle calculated fields
                   if calculation == 'extract_group':
                        dealer_number = data.get('FN_DEALER_NUMBER', '')
                        if dealer_number and len(dealer_number) >= 2:
                             return dealer_number[:2]
                        return ''
                   else:
                        return ''
          else:
                   # Direct field mapping
                   value = data.get(source_field, '')
              
                   if format_type == 'currency':
                        try:
                             float_value = float(value) if value else 0.0
                             return self.formatter.format_currency(float_value)
                        except (ValueError, TypeError):
                             return str(value)
                   else:
                        return str(value)

     def _generate_header(self, config: Dict) -> str:
          """Generate report header with exact current design formatting"""
          layout = config['layout']
          header_config = layout.get('header', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          lines = []
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks':
                   # Title without parentheses
                   title = header_config.get('title', 'JANUS FUNDS: DIVIDEND CHECKS')
                   lines.append(f"        {title}")
              
                   # Centered subtitle
                   subtitle = header_config.get('subtitle', 'AUDIT REPORT')
                   if subtitle:
                        # Center the subtitle with proper spacing
                        lines.append("")
                        lines.append(f"             {subtitle}")
              
                   # Fields with specific alignment
                   fields = header_config.get('fields', [])
                   for field in fields:
                        name = field.get('name', '')
                        position = field.get('position', 'left')
                        alignment = field.get('alignment', 'left')
                   
                        # Handle dynamic date replacement
                        if '{CURRENT_DATE}' in name:
                             current_date = datetime.now().strftime('%m/%d/%Y')
                             name = name.replace('{CURRENT_DATE}', current_date)
                   
                        if position == 'left' and alignment == 'left':
                             lines.append(f"           {name}")
                        elif position == 'right' and alignment == 'right':
                             # Right-align the date field
                             lines.append(f"           {name}")
              
                   # Add spacing
                   spacing = header_config.get('spacing', 1)
                   for _ in range(spacing):
                        lines.append("")
              
                   return '\n'.join(lines)

          # Original logic for other report types
          # Title and date - preserve exact current design
          title = header_config.get('title', config.get('workflow_metadata', {}).get('report_name', 'Report'))
          subtitle = header_config.get('subtitle', '')
         
          # Handle dynamic dates for Redemption_SWP report
          if header_config.get('date') == 'dynamic':
                   date = datetime.now().strftime('%m/%d/%Y')
          else:
                   date = header_config.get('date', datetime.now().strftime('%m/%d/%Y'))
              
          if header_config.get('trade_date') == 'dynamic':
                   # For dynamic trade date, we'll need to extract from data
                   # For now, use current date - this will be enhanced to extract from data source
                   trade_date = f"Trade Date: {datetime.now().strftime('%m/%d/%Y')}"
          else:
                   trade_date = header_config.get('trade_date', datetime.now().strftime('%m/%d/%Y'))

          # Check if this report should have a single header line
          single_header = header_config.get('single_header', True)# Default to single header

          if header_config.get('date_position') == 'right':
                   # Exact formatting: title left-aligned to 50 chars, date right-aligned to 30 chars
                   line = f"{title:<50} {date:>15}\n"
                   lines.append(line)
         
                   # Add subtitle if provided, separated by a line
                   if subtitle:
                        lines.append("")# Empty line separator
                        lines.append(subtitle)
         
                   # Only add second line if single_header is false
                   if not single_header and trade_date and trade_date != date:
                        lines.append(trade_date)
          else:
                   # For reports without date_position (like Redemption), use the original logic
                   line = f"{date}"
                   lines.append(line)
         
                   # Add subtitle if provided, separated by a line
                   if subtitle:
                        lines.append("")# Empty line separator
                        lines.append(subtitle)
         
                   # Only add trade_date if single_header is false and trade_date exists
                   if not single_header and trade_date and trade_date != date:
                        lines.append(trade_date)

          # Add spacing based on configuration
          spacing = header_config.get('spacing', 2)# Default to 2 if not specified
          for _ in range(spacing):
                   lines.append("")

          return '\n'.join(lines)

     def _generate_matrix_text_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate matrix section for TXT/CSV output"""
          layout = config['layout']
          matrix_sections = layout.get('matrix_sections', [])

          all_lines = []

          for section in matrix_sections:
                   lines = []
         
                   # Section title
                   title = section.get('title', '')
                   if title:
                        lines.append(title)
                        lines.append("")
         
                   # Headers
                   headers = section.get('headers', [])
                   if headers:
                        header_line = self.formatter.tab_char.join([f"{h:30s}" if i == 0 else f"{h:15s}" for i, h in enumerate(headers)])
                        lines.append(header_line)
         
                   # Rows
                   rows_config = section.get('rows', [])
                   for row_config in rows_config:
                        row_label = row_config.get('label', '')
                        cell_values = row_config.get('values', [])
                        indent_level = row_config.get('indent', 0)
              
                        # Apply indentation
                        if indent_level > 0:
                             row_label = '' * indent_level + row_label
              
                        # Build row
                        row_parts = [f"{row_label:30s}"]
              
                        for value_config in cell_values:
                             if isinstance(value_config, dict):
                                  calc_type = value_config.get('calculation', '')
                                  value = self._calculate_matrix_value(calc_type, data, value_config)
                                  if isinstance(value, (int, float)):
                                        row_parts.append(f"{value:>15}")
                                  else:
                                        row_parts.append(f"{str(value):>15}")
                             else:
                                  row_parts.append(f"{str(value_config):>15}")
              
                        lines.append(self.formatter.tab_char.join(row_parts))
         
                   lines.append("")
                   all_lines.extend(lines)

          return '\n'.join(all_lines)

     def _generate_data_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate data section with exact current design formatting"""
          layout = config.get('layout', {})
          data_config = layout.get('data_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          lines = []
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks' and data_config and 'columns' in data_config:
                   # Group data by fund code
                   fund_groups = {}
                   for record in data:
                        fund_code = record.get('CUSTOM_FUND_CODE', '')
                        if fund_code not in fund_groups:
                             fund_groups[fund_code] = []
                        fund_groups[fund_code].append(record)
              
                   # Column headers with proper formatting
                   columns = data_config['columns']
              
                   # Create header row with proper alignment
                   header_line = ""
                   for col in columns:
                        name = col.get('name', '')
                        width = col.get('width', 0)
                        align = col.get('align', 'left')
                   
                        if width > 0:
                             formatted = self.formatter.align_text(name, width, align)
                        else:
                             formatted = name
                   
                        header_line += formatted + " "
              
                   lines.append(header_line.rstrip())
              
                   # Add separator line
                   separator_line = ""
                   for col in columns:
                        width = col.get('width', 0)
                        if width > 0:
                             separator_line += "-" * width + " "
                        else:
                             separator_line += "-" * len(col.get('name', '')) + " "
              
                   lines.append(separator_line.rstrip())
              
                   # Data rows for each fund
                   for fund_code in sorted(fund_groups.keys()):
                        fund_records = fund_groups[fund_code]
                   
                        # Calculate totals for this fund
                        total_statements = len(fund_records)
                        total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', '0')) for record in fund_records)
                   
                        # Create fund record
                        fund_record = {
                             'CUSTOM_FUND_CODE': fund_code,
                             'Client Statements': total_statements,
                             'Client Amount': total_amount
                        }
                   
                        # Generate row with proper formatting
                        row_line = ""
                        for col in columns:
                             source_field = col.get('source_field', '')
                             calculation = col.get('calculation', '')
                        
                             if calculation == 'count_client_statements':
                                  value = fund_record.get('Client Statements', 0)
                             elif calculation == 'sum_client_amounts':
                                  value = fund_record.get('Client Amount', 0)
                             elif calculation == 'sum_custom_check_amounts':
                                  value = fund_record.get('Client Amount', 0)
                             else:
                                  value = fund_record.get(source_field, '')
                        
                             # Format the value
                             if col.get('format') == 'currency':
                                  try:
                                       value = f"{float(value):.2f}"
                                  except (ValueError, TypeError):
                                       value = "0.00"
                             elif col.get('format') == 'number':
                                  value = str(value)
                             else:
                                  value = str(value)
                        
                             width = col.get('width', 0)
                             align = col.get('align', 'left')
                        
                             if width > 0:
                                  formatted = self.formatter.align_text(value, width, align)
                             else:
                                  formatted = value
                        
                             row_line += formatted + " "
                   
                        lines.append(row_line.rstrip())
              
                   # Note: Totals section is handled separately, not in data section
              
                   return '\n'.join(lines)

          # Handle regular data_section if it exists
          if data_config and 'columns' in data_config:
                   # Column headers - exact current design formatting
                   columns = data_config['columns']
                   header_parts = []
         
                   for col in columns:
                        name = col.get('name', '')
                        width = col.get('width', 0)
                        align = col.get('align', 'left')
              
                        if width > 0:
                             formatted = self.formatter.align_text(name, width, align)
                        else:
                             formatted = name
              
                        header_parts.append(formatted)
         
                   # Use tab character for Excel-compatible formatting (current design)
                   lines.append(self.formatter.tab_char.join(header_parts))
              
                   # Add informational row for Redemption_SWP report
                   info_config = layout.get('informational_row', {})
                   if info_config.get('enabled', False):
                        info_row = self._generate_informational_row(info_config, data)
                        if info_row:
                             lines.append(info_row)

                   # Data rows - exact current design formatting
                   report_type = config.get('workflow_metadata', {}).get('report_type')
              
                   if report_type == 'redemption_swp':
                        # For Redemption_SWP reports, display aggregated records by fund
                        # Get mapped XML field names from config
                        field_mapping = config.get('xml_mapping', {}).get('field_mapping', {})
                        CUSTOM_CHECK_TYPE1 = field_mapping.get('CUSTOM_CHECK_TYPE1', 'CUSTOM_CHECK_TYPE1')
                        CUSTOM_FUND_CODE = field_mapping.get('CUSTOM_FUND_CODE', 'CUSTOM_FUND_CODE')
                        CUSTOM_FUND_DESC = field_mapping.get('CUSTOM_FUND_DESC', 'CUSTOM_FUND_DESC')
                        
                        # Filter records to only include those with CUSTOM_CHECK_TYPE1 = "RED" or "SWP"
                        filtered_data = [record for record in data
                                   if record.get(CUSTOM_CHECK_TYPE1) in ['REDEMPTION', 'SWP']]
                   
                        # Group records by fund (fund_code + fund_desc)
                        from collections import defaultdict
                        fund_groups = defaultdict(list)
                        for record in filtered_data:
                            fund_key = f"{record.get(CUSTOM_FUND_CODE, '')} {record.get(CUSTOM_FUND_DESC, '')}".strip()
                            fund_groups[fund_key].append(record)
                        
                        # Sort fund keys to maintain order
                        sorted_fund_keys = sorted(fund_groups.keys())
                        
                        # Generate one row per fund with aggregated data
                        for fund_key in sorted_fund_keys:
                            fund_records = fund_groups[fund_key]
                            # Create aggregated record for this fund
                            aggregated_record = self._create_aggregated_fund_record(fund_key, fund_records, config)
                            if aggregated_record:
                                # Generate row for this aggregated fund record
                                row = self.formatter.create_excel_row(data_config['columns'], aggregated_record)
                                lines.append(row)
                   else:
                        # For other reports, use original logic
                        for record in data:
                             row = self.formatter.create_excel_row(data_config['columns'], record)
                             lines.append(row)
              
                   # Enhanced Logic for Redemption_SWP Report
                   if report_type == 'redemption_swp':
                        lines = self._apply_redemption_swp_enhanced_logic(lines, data)

          # Add spacing - always 1 blank line after data (current design)
          if lines:
                   lines.append("")

          return '\n'.join(lines)

     def _generate_informational_row(self, info_config: Dict, data: List[Dict]) -> str:
          """Generate informational row for Redemption_SWP report"""
          columns = info_config.get('columns', [])
          if not columns:
                   return ""
         
          row_parts = []
         
          for col in columns:
                   name = col.get('name', '')
                   calculation = col.get('calculation', '')
                   format_type = col.get('format', 'text')
              
                   if calculation == 'format_fund_display':
                        # Format fund code and name with proper padding
                        fund_info = self._format_fund_display(data)
                        row_parts.append(fund_info)
                   elif calculation == 'count_total_checks':
                        # Count total checks (redemption + SWP)
                        total_checks = len(data)
                        row_parts.append(str(total_checks))
                   elif calculation == 'sum_total_amounts':
                        # Sum total amounts (currently blank as per requirements)
                        # JHI wants this populated but currently blank
                        row_parts.append("") # Currently blank as per requirements
                   else:
                        row_parts.append("")
         
          return self.formatter.tab_char.join(row_parts)

     def _format_fund_display(self, data: List[Dict]) -> str:
          """Format fund code and name display for informational row"""
          # For Redemption_SWP reports, we need to show individual fund codes per row
          # not concatenated values. This function should return a single fund code.
          # The actual fund display logic should be handled in the data section generation.
          return ""

     def _apply_redemption_swp_enhanced_logic(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Apply enhanced logic for Redemption_SWP report: move 11 lines to the top of the table"""
          if not lines:
                   return lines
         
          # 1. Move the 11 lines to the top of the table
          enhanced_lines = self._move_11_lines_to_top(lines, data)
         
          return enhanced_lines

     def _create_aggregated_fund_record(self, fund_key: str, fund_records: List[Dict], config: Dict = None) -> Dict:
          """Create aggregated record for a fund in Redemption_SWP report"""
          # Get mapped XML field names if config is provided
          if config and 'xml_mapping' in config and 'field_mapping' in config['xml_mapping']:
               field_mapping = config['xml_mapping']['field_mapping']
               CUSTOM_FUND_CODE = field_mapping.get('CUSTOM_FUND_CODE', 'CUSTOM_FUND_CODE')
               CUSTOM_FUND_DESC = field_mapping.get('CUSTOM_FUND_DESC', 'CUSTOM_FUND_DESC')
               CUSTOM_CHECK_TYPE1 = field_mapping.get('CUSTOM_CHECK_TYPE1', 'CUSTOM_CHECK_TYPE1')
               CUSTOM_CHECK_NUMBER = field_mapping.get('CUSTOM_CHECK_NUMBER', 'CUSTOM_CHECK_NUMBER')
               CUSTOM_CHECK_AMOUNT = field_mapping.get('CUSTOM_CHECK_AMOUNT', 'CUSTOM_CHECK_AMOUNT')
          
          # Get fund code and description from first record
          if not fund_records:
               return None
          
          first_record = fund_records[0]
          fund_code = first_record.get(CUSTOM_FUND_CODE, '')
          fund_desc = first_record.get(CUSTOM_FUND_DESC, '')
          
          # Separate REDEMPTION and SWP records
          red_records = [r for r in fund_records if r.get(CUSTOM_CHECK_TYPE1) == 'REDEMPTION']
          swp_records = [r for r in fund_records if r.get(CUSTOM_CHECK_TYPE1) == 'SWP']
          
          # Initialize values
          red_begin_check = ""
          red_end_check = ""
          red_count = 0
          red_total_amount = 0.0
          swp_begin_check = ""
          swp_end_check = ""
          swp_count = 0
          swp_total_amount = 0.0
          
          # Calculate RED values at fund level
          if red_records:
               red_check_numbers = [r.get(CUSTOM_CHECK_NUMBER, '') for r in red_records if r.get(CUSTOM_CHECK_NUMBER)]
               if red_check_numbers:
                    # Filter out empty strings and sort
                    red_check_numbers = [cn for cn in red_check_numbers if cn]
                    if red_check_numbers:
                         red_check_numbers.sort()
                         red_begin_check = red_check_numbers[0]
                         red_end_check = red_check_numbers[-1]
               red_count = len(red_records)
               red_total_amount = sum(self.safe_float_convert(r.get(CUSTOM_CHECK_AMOUNT, 0)) for r in red_records)
          
          # Calculate SWP values at fund level
          if swp_records:
               swp_check_numbers = [r.get(CUSTOM_CHECK_NUMBER, '') for r in swp_records if r.get(CUSTOM_CHECK_NUMBER)]
               if swp_check_numbers:
                    # Filter out empty strings and sort
                    swp_check_numbers = [cn for cn in swp_check_numbers if cn]
                    if swp_check_numbers:
                         swp_check_numbers.sort()
                         swp_begin_check = swp_check_numbers[0]
                         swp_end_check = swp_check_numbers[-1]
               swp_count = len(swp_records)
               swp_total_amount = sum(self.safe_float_convert(r.get(CUSTOM_CHECK_AMOUNT, 0)) for r in swp_records)
          
          # CHECKS PRINTED TODAY = total count of checks for this fund (both RED and SWP)
          checks_printed_today = red_count + swp_count
          
          # DOLLAR VALUE TODAY = sum of all check amounts for this fund
          dollar_value_today = red_total_amount + swp_total_amount
          
          # Create aggregated record
          aggregated_record = {
               'CUSTOM_FUND_CODE': fund_code,
               'CUSTOM_FUND_DESC': fund_desc,
               'CUSTOM_CHECK_AMOUNT': dollar_value_today,  # For "DOLLAR VALUE TODAY" column
               'CHECKS_PRINTED_TODAY': checks_printed_today,
               'DOLLAR_VALUE_TODAY': dollar_value_today,
               # RED values: fund-level aggregated
               'RED_BEGIN_CHECK_NUMBER': red_begin_check,
               'RED_END_CHECK_NUMBER': red_end_check,
               'RED_CHECK_COUNT': red_count,
               'TOTAL_RED_DOLLAR_VALUE': red_total_amount,
               # SWP values: fund-level aggregated
               'SWP_BEGIN_CHECK_NUMBER': swp_begin_check,
               'SWP_END_CHECK_NUMBER': swp_end_check,
               'SWP_CHECK_COUNT': swp_count,
               'TOTAL_SWP_DOLLAR_VALUE': swp_total_amount
          }
          
          return aggregated_record

     def _create_individual_record(self, record: Dict, config: Dict = None) -> Dict:
          """Create enhanced record for individual Redemption_SWP record"""
          # Get mapped XML field names if config is provided (for Redemption_SWP report)
          if config and 'xml_mapping' in config and 'field_mapping' in config['xml_mapping']:
               field_mapping = config['xml_mapping']['field_mapping']
               CUSTOM_FUND_CODE = field_mapping.get('CUSTOM_FUND_CODE', 'CUSTOM_FUND_CODE')
               CUSTOM_FUND_DESC = field_mapping.get('CUSTOM_FUND_DESC', 'CUSTOM_FUND_DESC')
               CUSTOM_CHECK_TYPE1 = field_mapping.get('CUSTOM_CHECK_TYPE1', 'CUSTOM_CHECK_TYPE1')
               CUSTOM_CHECK_NUMBER = field_mapping.get('CUSTOM_CHECK_NUMBER', 'CUSTOM_CHECK_NUMBER')
               CUSTOM_CHECK_AMOUNT = field_mapping.get('CUSTOM_CHECK_AMOUNT', 'CUSTOM_CHECK_AMOUNT')
          else:
               # Fallback to default field names if config not provided
               CUSTOM_FUND_CODE = 'CUSTOM_FUND_CODE'
               CUSTOM_FUND_DESC = 'CUSTOM_FUND_DESC'
               CUSTOM_CHECK_TYPE1 = 'CUSTOM_CHECK_TYPE1'
               CUSTOM_CHECK_NUMBER = 'CUSTOM_CHECK_NUMBER'
               CUSTOM_CHECK_AMOUNT = 'CUSTOM_CHECK_AMOUNT'
          
          fund_code = record.get(CUSTOM_FUND_CODE, '')
          fund_desc = record.get(CUSTOM_FUND_DESC, '')
          check_type1 = record.get(CUSTOM_CHECK_TYPE1, '')
          check_number = record.get(CUSTOM_CHECK_NUMBER, '')
          check_amount = self.safe_float_convert(record.get(CUSTOM_CHECK_AMOUNT, 0))
          
          # Create fund key for grouping: CUSTOM_FUND_CODE + " " + CUSTOM_FUND_DESC
          fund_key = f"{fund_code} {fund_desc}".strip()
          
          # Initialize all values
          red_begin_check = ""
          red_end_check = ""
          red_count = 0
          red_total_amount = 0.0
          swp_begin_check = ""
          swp_end_check = ""
          swp_count = 0
          swp_total_amount = 0.0
          checks_printed_today = 0

          # Group records by fund_key (CUSTOM_FUND_CODE + " " + CUSTOM_FUND_DESC) and calculate totals at fund level
          if hasattr(self, "all_records") and isinstance(self.all_records, list):
               # Get all records for this fund (grouped by fund_key)
               fund_records = [
                    r for r in self.all_records
                    if f"{r.get(CUSTOM_FUND_CODE, '')} {r.get(CUSTOM_FUND_DESC, '')}".strip() == fund_key
               ]
               
               # Separate REDEMPTION and SWP records for this fund
               red_records = [r for r in fund_records if r.get(CUSTOM_CHECK_TYPE1) == 'REDEMPTION']
               swp_records = [r for r in fund_records if r.get(CUSTOM_CHECK_TYPE1) == 'SWP']
               
               # Calculate RED values at fund level
               if red_records:
                    red_check_numbers = [r.get(CUSTOM_CHECK_NUMBER, '') for r in red_records if r.get(CUSTOM_CHECK_NUMBER)]
                    if red_check_numbers:
                         red_check_numbers.sort()
                         red_begin_check = red_check_numbers[0]
                         red_end_check = red_check_numbers[-1]
                    red_count = len(red_records)
                    red_total_amount = sum(self.safe_float_convert(r.get(CUSTOM_CHECK_AMOUNT, 0)) for r in red_records)
               
               # Calculate SWP values at fund level
               if swp_records:
                    swp_check_numbers = [r.get(CUSTOM_CHECK_NUMBER, '') for r in swp_records if r.get(CUSTOM_CHECK_NUMBER)]
                    if swp_check_numbers:
                         swp_check_numbers.sort()
                         swp_begin_check = swp_check_numbers[0]
                         swp_end_check = swp_check_numbers[-1]
                    swp_count = len(swp_records)
                    swp_total_amount = sum(self.safe_float_convert(r.get(CUSTOM_CHECK_AMOUNT, 0)) for r in swp_records)
               
               # CHECKS PRINTED TODAY = total count of checks for this fund (both RED and SWP)
               checks_printed_today = red_count + swp_count
         
        #   # Set values based on the individual record's check type
        #   # For individual records:
        #   # - Redemption records: show this record's amount in RED column, show fund-level SWP totals in SWP columns
        #   # - SWP records: show this record's amount in SWP column, show fund-level RED totals in RED columns
        #   if check_type1 == 'REDEMPTION':
        #            # For Redemption records, show this record's amount in RED column
        #            # SWP values remain from fund-level calculation (already set above)
        #            pass  # red_total_amount will be set to check_amount below
        #   elif check_type1 == 'SWP':
        #            # For SWP records, show this record's amount in SWP column
        #            # RED values remain from fund-level calculation (already set above)
        #            pass  # swp_total_amount will be set to check_amount below
          
          # Create enhanced record
          enhanced_record = {
                   'CUSTOM_FUND_CODE': fund_code,
                   'CUSTOM_FUND_DESC': fund_desc,
                   'CUSTOM_CHECK_TYPE': check_type1,
                   'CUSTOM_CHECK_AMOUNT': check_amount,
                   'CHECKS_PRINTED_TODAY': checks_printed_today,
                   # RED values: fund-level aggregated (begin/end/count) and record-specific amount for RED records
                   'RED_BEGIN_CHECK_NUMBER': red_begin_check,
                   'RED_END_CHECK_NUMBER': red_end_check,
                   'RED_CHECK_COUNT': red_count,
                   # For Redemption records, show this record's amount; for SWP records, show fund-level RED total
                   'TOTAL_RED_DOLLAR_VALUE': check_amount if check_type1 == 'REDEMPTION' else red_total_amount,
                   # SWP values: fund-level aggregated (begin/end/count) and record-specific amount for SWP records
                   'SWP_BEGIN_CHECK_NUMBER': swp_begin_check,
                   'SWP_END_CHECK_NUMBER': swp_end_check,
                   'SWP_CHECK_COUNT': swp_count,
                   # For SWP records, show this record's amount; for Redemption records, show fund-level SWP total
                   'TOTAL_SWP_DOLLAR_VALUE': check_amount if check_type1 == 'SWP' else swp_total_amount
          }
          
          return enhanced_record

     def _calculate_group_totals(self, grouped_data: Dict[str, Dict[str, List[Dict]]]) -> Dict[str, Dict[str, Dict]]:
          """Calculate totals for each group in Redemption_SWP report"""
          group_totals = {}
         
          for fund_code, type_data in grouped_data.items():
                   group_totals[fund_code] = {}
              
                   for check_type, records in type_data.items():
                        total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in records)
                        count = len(records)
                   
                        group_totals[fund_code][check_type] = {
                             'count': count,
                             'total_amount': total_amount,
                             'records': records
                        }
         
          return group_totals

     def _get_dynamic_special_handling_counts(self, data: List[Dict]) -> Dict[str, int]:
          """Dynamically extract and count DOC_SPECIAL_HANDLING_CODE values from XML data"""
          special_handling_counts = {}
         
          for record in data:
                   handling_code = record.get('DOC_SPECIAL_HANDLING_CODE', '')
                   if handling_code:
                        # Count occurrences of each handling code
                        special_handling_counts[handling_code] = special_handling_counts.get(handling_code, 0) + 1
         
          return special_handling_counts

     def _map_special_handling_to_shcode(self, handling_code: str) -> Dict:
          """Map DOC_SPECIAL_HANDLING_CODE to corresponding SHCode and description"""
          if handling_code in self.special_handling_mapping:
                   return self.special_handling_mapping[handling_code]
          else:
                   # Return unknown mapping for codes not in our table
                   return {
                        'code': f'UNKNOWN-{handling_code}',
                        'description': f'Unknown Special Handling Code: {handling_code}',
                        'shcode': handling_code
                   }

     def _get_special_handling_summary(self, data: List[Dict]) -> Dict:
          """Get detailed summary of special handling codes found in the data"""
          special_handling_counts = self._get_dynamic_special_handling_counts(data)
          summary = {
                   'total_records': len(data),
                   'special_handling_codes_found': list(special_handling_counts.keys()),
                   'code_details': {},
                   'unknown_codes': []
          }
         
          for handling_code, count in special_handling_counts.items():
                   mapping = self._map_special_handling_to_shcode(handling_code)
                   summary['code_details'][handling_code] = {
                        'count': count,
                        'mapping': mapping
                   }
              
                   if 'UNKNOWN' in mapping['code']:
                        summary['unknown_codes'].append({
                             'code': handling_code,
                             'count': count
                        })
         
          return summary

     def add_special_handling_mapping(self, shcode: str, code: str, description: str) -> None:
          """Dynamically add a new special handling code mapping"""
          self.special_handling_mapping[shcode] = {
                   'code': code,
                   'description': description,
                   'shcode': shcode
          }

     def get_special_handling_mappings(self) -> Dict:
          """Get all current special handling mappings"""
          return self.special_handling_mapping.copy()

     def _get_suppressedtest(self, config: Dict, data: List[Dict]) -> Dict[str, int]:
          """
          Reads a suppression report from S3 and counts lines that are suppressed,
          either by 'Suppressed' status or by containing '"isSuppressed": true'.
          Uses configuration values for TLE field name, path, and regex pattern.
          """
          suppressedtest_counts = 0

          # Get S3 config
          # config is already the redemption_swp config, not the root config
          layout = config.get("layout", {})
          suppressedtest_config = layout.get("suppressedtest", {})
          # Get TLE field name dynamically from configuration
          tle_field_name = suppressedtest_config.get('TLE', 'UC_FILE_NAME')
          if not tle_field_name:
               return {"count": 0}

          # Get file name from XML data using the TLE field name from config
          first_file_name = data[0].get(tle_field_name, "").strip()
          first_file_name = first_file_name.split(".txt")[0]
          if not first_file_name:
               return {"count": 0}

          # Get path from configuration
          path_prefix = suppressedtest_config.get("path", "")
          if not path_prefix:
               return {"count": 0}

          # Ensure path ends with / for proper S3 prefix matching
          if not path_prefix.endswith('/'):
               path_prefix = f"{path_prefix}/"

          bucket_name = os.getenv("s3BucketInput")
          if not bucket_name:
               return {"count": 0}

          # --- Fetch all matching files from S3 ---
          # Use paginator to handle large directories (S3 doesn't support wildcards in prefix)
          # Updated to find ALL files matching the pattern, not just the first one
          target_keys = []
          try:
               paginator = s3.get_paginator('list_objects_v2')
               # Build filename pattern regex (matches files starting with first_file_name and ending with .txt)
               filename_pattern = re.compile(rf"{re.escape(first_file_name)}.*\.rpt$", re.IGNORECASE)
               # Iterate through all pages of objects in the path
               for page in paginator.paginate(Bucket=bucket_name, Prefix=path_prefix):
                    if 'Contents' not in page:
                         continue
                    for obj in page['Contents']:
                         # Extract just the filename from the full S3 key
                         obj_key = obj['Key']
                         filename = obj_key.split('/')[-1]
                         
                         # Check if the filename matches the pattern
                         if filename_pattern.search(filename):
                              target_keys.append(obj_key)
               
               # Sort keys to ensure consistent processing order
               target_keys.sort()
               
          except Exception as e:
               return {"count": 0}

          if not target_keys:
               return {"count": 0}


          # --- Apply regex from configuration to count suppressed lines ---
          regex_pattern = suppressedtest_config.get("regex-match", "")
          if not regex_pattern:
               return {"count": 0}
          
          # Compile and use the regex pattern from configuration
          # Use MULTILINE flag so that ^ matches the start of each line, not just the start of the string
          pattern = re.compile(regex_pattern, re.MULTILINE)
          
          # Process all matching files and combine suppressed counts
          total_suppressedtest_counts = 0
          processed_files = []
          
          for target_key in target_keys:
               try:
                    # Read the file from S3
                    obj = s3.get_object(Bucket=bucket_name, Key=target_key)
                    file_text = obj["Body"].read().decode("utf-8")
                    file_size = len(file_text)
                    line_count = len(file_text.splitlines())
                    
                    # Count lines that match the pattern
                    # The regex pattern starts with ^ so it's designed to match from the start of each line
                    matching_lines = pattern.finditer(file_text)
                    file_suppressedtest_counts = sum(1 for _ in matching_lines)
                    total_suppressedtest_counts += file_suppressedtest_counts
                    
                    processed_files.append({
                         "file": target_key,
                         "count": file_suppressedtest_counts
                    })
                    
                    
               except Exception as e:
                    continue
          

          return {"files": processed_files, "count": total_suppressedtest_counts}

     def _calculate_dynamic_footer_data(self, config: Dict,data: List[Dict]) -> List:
          """Calculate dynamic footer data based on CUSTOM_PULL_TYPE and DOC_SPECIAL_HANDLING_CODE for Redemption_SWP report"""
          # Count occurrences of each CUSTOM_PULL_TYPE
          field_mapping = config.get('xml_mapping', {}).get('field_mapping', {})
          CUSTOM_PULL_TYPE = field_mapping.get('CUSTOM_PULL_TYPE', 'CUSTOM_PULL_TYPE')
          pull_type_counts = {}
          for record in data:
                   pull_type = record.get(CUSTOM_PULL_TYPE, '')
                   if pull_type:
                        clean_pull_type = pull_type.strip()
                        pull_type_counts[clean_pull_type] = pull_type_counts.get(clean_pull_type, 0) + 1
         
          # Get dynamic special handling counts
          special_handling_counts = self._get_dynamic_special_handling_counts(data)
         
          # Define the categories in order
          categories = [
                   "RETURNTOJANUS",
                   "JANUS",
                   "STATESTREETBANK",
                   "DONOTMAIL",
                   "VOID",
                   "PULLREPORTCHECKS",
                   "SUB-TOTAL",
                   "SPECIALHANDLING",
                   "OVERNIGHTW-SIG",
                   "OVERNIGHTWO-SIG",
                   "OVERNIGHTSAT-W-SIG",
                   "OVERNIGHTSAT-WO-SIG",
                   "SUB-TOTAL",
                   "NORMAL",
                   "TOTAL",
                   "SUPPRESSEDTEST"
          ]
         
          # Initialize category data with zeros
          category_data = [0] * len(categories)
         
          # Map individual pull types to their counts (handle actual XML values)
          category_data[1] = pull_type_counts.get('JANUS', 0) # JANUS
          # category_data[2] = pull_type_counts.get('STATW STREET BANK', 0) # STATESTREETBANK (actual XML value)
          category_data[2] = pull_type_counts.get('STATE STREET BANK', pull_type_counts.get('STANDARD', 0))

          # category_data[3] = pull_type_counts.get('DO NOT MAIL', 0) # DONOTMAIL (actual XML value)
          category_data[3] = (pull_type_counts.get('DO NOT MAIL', 0) + pull_type_counts.get('JANUS DO NOT MAIL', 0))
          category_data[4] = pull_type_counts.get('VOID', 0) # VOID
          category_data[5] = pull_type_counts.get('PULLREPORTCHECKS', 0) # PULLREPORTCHECKS
         
          # Calculate first SUB-TOTAL (sum of JANUS, STATESTREETBANK, DONOTMAIL, VOID, PULLREPORTCHECKS)
          category_data[6] = (category_data[1] + category_data[2] + category_data[3] + 
                             category_data[4] + category_data[5])  # SUB-TOTAL
         
          # Map special handling types dynamically based on DOC_SPECIAL_HANDLING_CODE
          # Initialize special handling counts to 0
          overnight_w_sig_count = 0
          overnight_wo_sig_count = 0
          overnight_sat_w_sig_count = 0
          overnight_sat_wo_sig_count = 0
         
          # Process each special handling code found in the data
          for handling_code, count in special_handling_counts.items():
                   mapping = self._map_special_handling_to_shcode(handling_code)
              
                   # Map to the appropriate category based on the code
                   if mapping['code'] == 'OVERNIGHTW-SIG':
                        overnight_w_sig_count = count
                   elif mapping['code'] == 'OVERNIGHTWO-SIG':
                        overnight_wo_sig_count = count
                   elif mapping['code'] == 'OVERNIGHTSAT-W-SIG':
                        overnight_sat_w_sig_count = count
                   elif mapping['code'] == 'OVERNIGHTSAT-WO-SIG':
                        overnight_sat_wo_sig_count = count
          # Assign the dynamically calculated counts
          category_data[8] = overnight_w_sig_count # OVERNIGHTW-SIG
          category_data[9] = overnight_wo_sig_count # OVERNIGHTWO-SIG
          category_data[10] = overnight_sat_w_sig_count # OVERNIGHTSAT-W-SIG
          category_data[11] = overnight_sat_wo_sig_count # OVERNIGHTSAT-WO-SIG
         
          # Calculate second SUB-TOTAL (sum of OVERNIGHTW-SIG, OVERNIGHTWO-SIG, OVERNIGHTSAT-W-SIG, OVERNIGHTSAT-WO-SIG)
          category_data[12] = (category_data[8] + category_data[9] + category_data[10] + category_data[11])  # SUB-TOTAL (second occurrence)
         
          suppressedtest=self._get_suppressedtest(config, data)
          # category_data[15] = suppressedtest #pull_type_counts.get('SUPPRESSEDTEST', 0) # SUPPRESSEDTEST
          category_data[15] =suppressedtest.get('count',0) if isinstance(suppressedtest,dict) else suppressedtest
         
          # Calculate RETURNTOJANUS (sum of JANUS, STATESTREETBANK, DONOTMAIL, VOID, PULLREPORTCHECKS)
          # Note: This excludes SUB-TOTAL since SUB-TOTAL is now calculated, not a raw value
          return_to_janus = (category_data[1] + category_data[2] + category_data[3] +
                         category_data[4] + category_data[5])
          category_data[0] = return_to_janus
         
          # Calculate SPECIALHANDLING dynamically (sum of all special handling codes found)
          #special_handling = (category_data[8] + category_data[9] + category_data[10] + category_data[11])
          category_data[7] = ''
         
          # Calculate TOTAL (total count of all records in the data)
          total_all_records = len(data)
          category_data[14] = total_all_records  # TOTAL
          
          # Calculate NORMAL = TOTAL - (first SUB-TOTAL + second SUB-TOTAL)
          category_data[13] = total_all_records - (category_data[6] + category_data[12])  # NORMAL
         
          return category_data

     def _move_11_lines_to_top(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Move the 11 lines to the top of the table in Redemption_SWP report"""
          if not lines:
                   return lines
         
          # Create the 11 lines that need to be moved to the top
          top_lines = self._create_11_top_lines(data)
         
          # Find where the actual data table starts (after headers)
          data_start_index = self._find_data_table_start(lines)
         
          if data_start_index == -1:
                   # If we can't find the data table start, just prepend the 11 lines
                   return top_lines + lines
         
          # Split the lines into header section and data section
          header_section = lines[:data_start_index]
          data_section = lines[data_start_index:]
         
          # Combine: header section + 11 top lines + data section
          return header_section + top_lines + data_section

     def _create_11_top_lines(self, data: List[Dict]) -> List[str]:
          """Create the 11 lines that need to be moved to the top"""
          lines = []
         
          # Line 1: Date
          current_date = datetime.now().strftime('%m/%d/%Y')
          lines.append(current_date)
         
          # Line 2: Trade Date
          trade_date = f"Trade Date: {datetime.now().strftime('%m/%d/%Y')}"
          lines.append(trade_date)
         
          # Line 3: Fund Code and Fund Name information
          lines.append("Fund Code and Fund Name from RRSU0224 & RRSU0001.")
         
          # Line 4: Fund Code length information
          lines.append("Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits.")
         
          # Line 5: Blank field instruction
          lines.append("always blank. JHI would like BR to populate this")
         
          # Line 6: Another blank field instruction
          lines.append("always blank. JHI would like BR to populate this")
         
          # Line 7: Specific order information
          lines.append("Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order).")
         
          # Line 8: Total checks information
          total_checks = len(data)
          lines.append(f"This is a total of both Redemption and SWP checks in the Fund. Total: {total_checks}")
         
          # Line 9: Total dollar value information
          total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
          lines.append(f"Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field. Calculated Total: {total_amount:.2f}")
         
          # Line 10: Empty line for spacing
          lines.append("")
         
          # Line 11: Table headers (this will be the consolidated header line)
          consolidated_header = self._create_consolidated_static_line(data)
          lines.append(consolidated_header)
         
          return lines

     def _find_data_table_start(self, lines: List[str]) -> int:
          """Find where the actual data table starts"""
          for i, line in enumerate(lines):
                   # Look for table headers or data rows
                   if line.strip() and ('GROWTH' in line.upper() or 'FUND' in line.upper() or
                                  'CHECKS' in line.upper() or 'TOTAL' in line.upper()):
                        return i
          return -1

     def _insert_static_informational_lines(self, data: List[Dict]) -> List[str]:
          """Insert static informational lines for Redemption_SWP report"""
          static_lines = []
         
          # Above "GROWTH FUND" - Fund Code and Fund Name information
          fund_info_line = "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order)."
          static_lines.append(fund_info_line)
         
          # Above "CHECKS PRINTED TODAY" - Total checks information
          total_checks = len(data)
          checks_info_line = f"This is a total of both Redemption and SWP checks in the Fund. Total: {total_checks}"
          static_lines.append(checks_info_line)
         
          # Above "DOLLAR VALUE TODAY" - Total dollar value information
          total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
          dollar_info_line = f"Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field. Calculated Total: {total_amount:.2f}"
          static_lines.append(dollar_info_line)
         
          return static_lines

     def _insert_static_informational_lines_aligned(self, data: List[Dict], lines: List[str], header_start_index: int) -> List[str]:
          """Insert static informational lines above specific headers with proper column alignment"""
          static_lines = []
         
          # Find the header line to determine column structure
          header_line = None
          for i in range(header_start_index, len(lines)):
                   if lines[i].strip() and ('GROWTHFUND' in lines[i].upper() or 'GROWTH FUND' in lines[i].upper()):
                        header_line = lines[i]
                        break
         
          if header_line:
                   # Determine column structure from header line
                   if '\t' in header_line:
                        # Tab-separated format
                        columns = header_line.split('\t')
                        num_columns = len(columns)
                   else:
                        # Single column or space-separated
                        num_columns = 1
          else:
                   # Default to single column if no header found
                   num_columns = 1
         
          # Create consolidated static line with text distributed across columns A, B, C
          if num_columns >= 3:
                   # Multi-column format - distribute text across columns A, B, C
                   column_a_text = "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order)."
                   column_b_text = "This is a total of both Redemption and SWP checks in the Fund."
                   column_c_text = "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
              
                   # Create consolidated line with tab separation
                   consolidated_line = f"{column_a_text}\t{column_b_text}\t{column_c_text}"
              
                   # Add empty columns if there are more than 3 columns
                   for _ in range(num_columns - 3):
                        consolidated_line += '\t'
              
                   static_lines.append(consolidated_line)
          else:
                   # Fallback to single column format
                   static_texts = [
                        "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order).",
                        "This is a total of both Redemption and SWP checks in the Fund.",
                        "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
                   ]
              
                   for text in static_texts:
                        static_lines.append(text)
         
          return static_lines

     def _apply_row_reordering(self, lines: List[str]) -> List[str]:
          """Apply row reordering: exchange row 4 and row 5"""
          if len(lines) < 5:
                   return lines
         
          # Create a copy of the lines
          reordered_lines = lines.copy()
         
          # Exchange row 4 (index 3) and row 5 (index 4)
          reordered_lines[3], reordered_lines[4] = reordered_lines[4], reordered_lines[3]
         
          return reordered_lines

     def _position_consolidated_text_after_row_4(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Position consolidated static text immediately after row 4"""
          if len(lines) < 4:
                   return lines
         
          # Create consolidated static text line
          consolidated_line = self._create_consolidated_static_line(data)
         
          # Find the position after row 4 (index 4)
          # Insert the consolidated line at position 5 (index 4)
          positioned_lines = lines[:4] + [consolidated_line] + lines[4:]
         
          return positioned_lines

     def _create_consolidated_static_line(self, data: List[Dict]) -> str:
          """Create consolidated static line with text distributed across columns A, B, C"""
          column_a_text = "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order)."
          column_b_text = "This is a total of both Redemption and SWP checks in the Fund."
          column_c_text = "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
         
          # Create consolidated line with tab separation for TXT format
          return f"{column_a_text}\t{column_b_text}\t{column_c_text}"

     def _apply_redemption_swp_enhanced_logic_xlsx(self, ws, data: List[Dict], start_row: int) -> int:
          """Apply enhanced logic for Redemption_SWP report in XLSX format"""
          # For Redemption_SWP reports, we need to ensure the static text is placed in row 3 (A3, B3, C3)
          # These cells must remain fixed and static
         
          # Cell A3: Fund Code and Fund Name information
          cell_a3 = ws.cell(row=3, column=1, value="Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order).")
          cell_a3.font = Font(bold=True, size=10)
          cell_a3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
          # Cell B3: Total checks information
          cell_b3 = ws.cell(row=3, column=2, value="This is a total of both Redemption and SWP checks in the Fund.")
          cell_b3.font = Font(bold=True, size=10)
          cell_b3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
          # Cell C3: Total dollar value information
          cell_c3 = ws.cell(row=3, column=3, value="Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field.")
          cell_c3.font = Font(bold=True, size=10)
          cell_c3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
          # Return the current row (no additional rows added)
          return start_row

     def _apply_redemption_swp_enhanced_logic_csv(self, csv_lines: List[List[str]], data: List[Dict]) -> List[List[str]]:
          """Apply enhanced logic for Redemption_SWP report in CSV format - move 11 lines to top"""
          if not csv_lines:
                   return csv_lines
         
          # Convert CSV lines to string format for processing
          lines = []
          for row in csv_lines:
                   if row:
                        lines.append(','.join(str(cell) for cell in row))
                   else:
                        lines.append("")
         
          # Apply the enhanced logic to move 11 lines to top
          enhanced_lines = self._move_11_lines_to_top_csv(lines, data)
         
          # Convert back to CSV format
          enhanced_csv_lines = []
          for line in enhanced_lines:
                   if line.strip():
                        enhanced_csv_lines.append(line.split(','))
                   else:
                        enhanced_csv_lines.append([])
         
          return enhanced_csv_lines

     def _move_11_lines_to_top_csv(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Move the 11 lines to the top of the table in CSV format"""
          if not lines:
                   return lines
         
          # Create the 11 lines that need to be moved to the top (CSV format)
          top_lines = self._create_11_top_lines_csv(data)
         
          # Find where the actual data table starts (after headers)
          data_start_index = self._find_data_table_start(lines)
         
          if data_start_index == -1:
                   # If we can't find the data table start, just prepend the 11 lines
                   return top_lines + lines
         
          # Split the lines into header section and data section
          header_section = lines[:data_start_index]
          data_section = lines[data_start_index:]
         
          # Combine: header section + 11 top lines + data section
          return header_section + top_lines + data_section

     def _create_11_top_lines_csv(self, data: List[Dict]) -> List[str]:
          """Create the 11 lines that need to be moved to the top (CSV format)"""
          lines = []
         
          # Line 1: Date
          current_date = datetime.now().strftime('%m/%d/%Y')
          lines.append(current_date)
         
          # Line 2: Trade Date
          trade_date = f"Trade Date: {datetime.now().strftime('%m/%d/%Y')}"
          lines.append(trade_date)
         
          # Line 3: Fund Code and Fund Name information
          lines.append("Fund Code and Fund Name from RRSU0224 & RRSU0001.")
         
          # Line 4: Fund Code length information
          lines.append("Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits.")
         
          # Line 5: Blank field instruction
          lines.append("always blank. JHI would like BR to populate this")
         
          # Line 6: Another blank field instruction
          lines.append("always blank. JHI would like BR to populate this")
         
          # Line 7: Specific order information
          lines.append("Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order).")
         
          # Line 8: Total checks information
          total_checks = len(data)
          lines.append(f"This is a total of both Redemption and SWP checks in the Fund. Total: {total_checks}")
         
          # Line 9: Total dollar value information
          total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
          lines.append(f"Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field. Calculated Total: {total_amount:.2f}")
         
          # Line 10: Empty line for spacing
          lines.append("")
         
          # Line 11: Table headers (this will be the consolidated header line for CSV)
          consolidated_header = self._create_consolidated_static_line_csv(data)
          lines.append(consolidated_header)
         
          return lines

     def _apply_redemption_swp_enhanced_logic_csv_strings(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Apply enhanced logic for Redemption_SWP report in CSV string format"""
          if not lines:
                   return lines
         
          # 1. Detect three consecutive blank lines and insert static informational lines above headers
          enhanced_lines = []
          i = 0
          while i < len(lines):
                   current_line = lines[i]
              
                   # Check if current line is blank/empty
                   if not current_line.strip():
                        # Count consecutive blank lines starting from current position
                        blank_count = 0
                        j = i
                        while j < len(lines) and not lines[j].strip():
                             blank_count += 1
                             j += 1
                   
                        # If we found 3 or more consecutive blank lines, insert static text above headers
                        if blank_count >= 3:
                             # Insert static informational lines above the relevant headers with proper alignment
                             enhanced_lines.extend(self._insert_static_informational_lines_csv_aligned(data, lines, j))
                             # Skip the blank lines we just processed
                             i = j
                             continue
              
                   enhanced_lines.append(current_line)
                   i += 1
         
          # 2. Apply row reordering (exchange row 4 and row 5)
          enhanced_lines = self._apply_row_reordering(enhanced_lines)
         
          # 3. Position consolidated static text immediately after row 4
          enhanced_lines = self._position_consolidated_text_after_row_4_csv(enhanced_lines, data)
         
          return enhanced_lines

     def _position_consolidated_text_after_row_4_csv(self, lines: List[str], data: List[Dict]) -> List[str]:
          """Position consolidated static text immediately after row 4 for CSV format"""
          if len(lines) < 4:
                   return lines
         
          # Create consolidated static text line for CSV
          consolidated_line = self._create_consolidated_static_line_csv(data)
         
          # Find the position after row 4 (index 4)
          # Insert the consolidated line at position 5 (index 4)
          positioned_lines = lines[:4] + [consolidated_line] + lines[4:]
         
          return positioned_lines

     def _create_consolidated_static_line_csv(self, data: List[Dict]) -> str:
          """Create consolidated static line with text distributed across columns A, B, C for CSV format"""
          column_a_text = "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order)."
          column_b_text = "This is a total of both Redemption and SWP checks in the Fund."
          column_c_text = "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
         
          # Create consolidated line with comma separation for CSV format
          return f"{column_a_text},{column_b_text},{column_c_text}"

     def _insert_static_informational_lines_csv_aligned(self, data: List[Dict], lines: List[str], header_start_index: int) -> List[str]:
          """Insert static informational lines above specific headers with proper CSV column alignment"""
          static_lines = []
         
          # Find the header line to determine column structure
          header_line = None
          for i in range(header_start_index, len(lines)):
                   if lines[i].strip() and ('GROWTHFUND' in lines[i].upper() or 'GROWTH FUND' in lines[i].upper()):
                        header_line = lines[i]
                        break
         
          if header_line:
                   # Determine column structure from header line
                   if ',' in header_line:
                        # Comma-separated format
                        columns = header_line.split(',')
                        num_columns = len(columns)
                   else:
                        # Single column
                        num_columns = 1
          else:
                   # Default to single column if no header found
                   num_columns = 1
         
          # Create consolidated static line with text distributed across columns A, B, C
          if num_columns >= 3:
                   # Multi-column format - distribute text across columns A, B, C
                   column_a_text = "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order)."
                   column_b_text = "This is a total of both Redemption and SWP checks in the Fund."
                   column_c_text = "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
              
                   # Create consolidated line with comma separation
                   consolidated_line = f"{column_a_text},{column_b_text},{column_c_text}"
              
                   # Add empty columns if there are more than 3 columns
                   for _ in range(num_columns - 3):
                        consolidated_line += ','
              
                   static_lines.append(consolidated_line)
          else:
                   # Fallback to single column format
                   static_texts = [
                        "Fund Code and Fund Name from RRSU0224 & RRSU0001. Fund Code length: BR needs to add leading zeros so the fund code is always 7 characters in length. All JHI fund codes are 2 digits. Specific order – list in Fund Code order; all Redemptions first (Fund Code order), followed by all SWP (Fund Code order).",
                        "This is a total of both Redemption and SWP checks in the Fund.",
                        "Total of both Redemption and SWP checks in the Fund. Currently always blank. JHI would like BR to populate this field."
                   ]
              
                   for text in static_texts:
                        static_lines.append(text)
         
          return static_lines

     def _generate_simplified_report_content(self, config: Dict, data: List[Dict]) -> str:
          """Generate simplified report content without complex formatting"""
          content_parts = []

          # Generate basic header
          report_name = config.get('workflow_metadata', {}).get('report_name', 'Report')
          content_parts.append(report_name)
          content_parts.append("")

          # Generate data section only
          data_section = config['layout'].get('data_section', {})
          if 'columns' in data_section:
                   columns = data_section['columns']
         
                   # Simple column headers
                   headers = [col.get('name', '') for col in columns]
                   content_parts.append('\t'.join(headers))
         
                   # Data rows
                   for record in data:
                        row_values = []
                        for col in columns:
                             source_field = col.get('source_field', '')
                             if source_field == 'static':
                                  value = col.get('value', '')
                             elif source_field == 'calculated':
                                  value = self._calculate_simple_value(col, record)
                             else:
                                  value = record.get(source_field, '')
                             row_values.append(str(value))
                        content_parts.append('\t'.join(row_values))

          return '\n'.join(content_parts)

     def _calculate_simple_value(self, col: Dict, data: Dict) -> Any:
          """Calculate simple value for simplified reports"""
          calculation = col.get('calculation', '')

          if calculation == 'count_records':
                   return 1
          elif calculation == 'sum_amounts':
                   # For Redemption_SWP reports, use CUSTOM_CHECK_AMOUNT
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          else:
                   return 0

     def _generate_xlsx_data_section(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX data section with proper cell separation - dynamically adapts to JSON structure"""
          layout = config['layout']
          current_row = start_row
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          # Check for data_section first
          data_config = layout.get('data_section', {})
          if data_config and 'columns' in data_config:
                   # Special handling for dividend reports
                   if report_type == 'dividend_checks':
                        # Group data by fund code
                        fund_groups = {}
                        for record in data:
                             fund_code = record.get('CUSTOM_FUND_CODE', '')
                             if fund_code not in fund_groups:
                                  fund_groups[fund_code] = []
                             fund_groups[fund_code].append(record)
                   
                        # Generate column headers
                        columns = data_config.get('columns', [])
                        for col_idx, col in enumerate(columns, 1):
                             name = col.get('name', '')
                             cell = ws.cell(row=current_row, column=col_idx, value=name)
                             cell.font = Font(bold=True, size=10)
                             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1
                   
                        # Generate data rows for each fund
                        for fund_code in sorted(fund_groups.keys()):
                             fund_records = fund_groups[fund_code]
                        
                             # Calculate totals for this fund
                             total_statements = len(fund_records)
                             total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', '0')) for record in fund_records)
                        
                             # Create fund record
                             fund_record = {
                                  'CUSTOM_FUND_CODE': fund_code,
                                  'Client Statements': total_statements,
                                  'Client Amount': total_amount
                             }
                        
                             # Generate row
                             for col_idx, col in enumerate(columns, 1):
                                  source_field = col.get('source_field', '')
                                  calculation = col.get('calculation', '')
                             
                                  if calculation == 'count_client_statements':
                                        value = fund_record.get('Client Statements', 0)
                                  elif calculation == 'sum_client_amounts':
                                        value = fund_record.get('Client Amount', 0)
                                  elif calculation == 'sum_custom_check_amounts':
                                        value = fund_record.get('Client Amount', 0)
                                  else:
                                        value = fund_record.get(source_field, '')
                             
                                  # Format the value
                                  if col.get('format') == 'currency':
                                        try:
                                             value = f"{float(value):.2f}"
                                        except (ValueError, TypeError):
                                             value = "0.00"
                                  elif col.get('format') == 'number':
                                        value = str(value)
                                  else:
                                        value = str(value)
                             
                                  cell = ws.cell(row=current_row, column=col_idx, value=value)
                                  cell.font = Font(size=10)
                             
                                  # Set alignment based on column configuration
                                  align = col.get('align', 'left')
                                  if align == 'right':
                                        cell.alignment = Alignment(horizontal='right', vertical='top')
                                  else:
                                        cell.alignment = Alignment(horizontal='left', vertical='top')
                        
                             current_row += 1
                   
                        # Add note if specified
                        table_style = data_config.get('table_style', {})
                        note = table_style.get('note', '')
                        if note:
                             current_row += 1
                             cell = ws.cell(row=current_row, column=1, value=note)
                             cell.font = Font(size=10, italic=True)
                             cell.alignment = Alignment(horizontal='right', vertical='top')
                             current_row += 1
                   
                        return current_row
              
                   # Original logic for other report types
                   # Generate column headers
                   columns = data_config.get('columns', [])
                   for col_idx, col in enumerate(columns, 1):
                        name = col.get('name', '')
                        cell = ws.cell(row=current_row, column=col_idx, value=name)
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                   current_row += 1
              
                   # Add informational row for Redemption_SWP report
                   info_config = layout.get('informational_row', {})
                   if info_config.get('enabled', False):
                        current_row = self._generate_xlsx_informational_row(ws, info_config, data, current_row)
              
                   # Generate data rows
                   if report_type == 'redemption_swp':
                        # For Redemption_SWP reports, display aggregated records by fund
                        # Get mapped XML field names from config
                        field_mapping = config.get('xml_mapping', {}).get('field_mapping', {})
                        CUSTOM_CHECK_TYPE1 = field_mapping.get('CUSTOM_CHECK_TYPE1', 'CUSTOM_CHECK_TYPE1')
                        CUSTOM_FUND_CODE = field_mapping.get('CUSTOM_FUND_CODE', 'CUSTOM_FUND_CODE')
                        CUSTOM_FUND_DESC = field_mapping.get('CUSTOM_FUND_DESC', 'CUSTOM_FUND_DESC')
                        
                        # Filter records to only include those with CUSTOM_CHECK_TYPE1 = "RED" or "SWP"
                        filtered_data = [record for record in data
                                   if record.get(CUSTOM_CHECK_TYPE1) in ['REDEMPTION', 'SWP']]
                   
                        # Group records by fund (fund_code + fund_desc)
                        from collections import defaultdict
                        fund_groups = defaultdict(list)
                        for record in filtered_data:
                            fund_key = f"{record.get(CUSTOM_FUND_CODE, '')} {record.get(CUSTOM_FUND_DESC, '')}".strip()
                            fund_groups[fund_key].append(record)
                        
                        # Sort fund keys to maintain order
                        sorted_fund_keys = sorted(fund_groups.keys())
                        
                        # Generate one row per fund with aggregated data
                        for fund_key in sorted_fund_keys:
                            fund_records = fund_groups[fund_key]
                            # Create aggregated record for this fund
                            aggregated_record = self._create_aggregated_fund_record(fund_key, fund_records, config)
                            if aggregated_record:
                                # Generate row for this aggregated fund record
                                for col_idx, col in enumerate(columns, 1):
                                    value = self._get_column_value_for_xlsx(col, aggregated_record, config)
                                    formatted_value = self._format_column_value_for_xlsx(col, value)
                                    cell = ws.cell(row=current_row, column=col_idx, value=formatted_value)
                                    cell.font = Font(size=10)
                                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                                current_row += 1
                   else:
                        # For other reports, use original logic
                        for record in data:
                             for col_idx, col in enumerate(columns, 1):
                                  value = self._get_column_value_for_xlsx(col, record)
                                  formatted_value = self._format_column_value_for_xlsx(col, value)
                                  cell = ws.cell(row=current_row, column=col_idx, value=formatted_value)
                                  cell.font = Font(size=10)
                                  cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
              
                   # Enhanced Logic for Redemption_SWP Report - XLSX
                   # Note: Static text placement is handled earlier in the generation process

          # Check for suppression_sections (used in 12b1 compensation reports)
          suppression_sections = layout.get('suppression_sections', [])
          for section in suppression_sections:
                   if 'columns' in section:
                        # Add section title
                        title = section.get('title', '')
                        if title:
                             cell = ws.cell(row=current_row, column=1, value=title)
                             cell.font = Font(bold=True, size=10)
                             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
              
                        # Add subtitle if provided
                        subtitle = section.get('subtitle', '')
                        if subtitle:
                             cell = ws.cell(row=current_row, column=1, value=subtitle)
                             cell.font = Font(bold=True, size=9)
                             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
              
                        # Apply filtering based on CUSTOM_CHECK_NUMBER for 12b1_compensation reports
                        filtered_data = self._filter_12b1_data_by_check_number(data, section)
                   
                        # Generate columns for this section
                        current_row = self._generate_xlsx_columns_section(ws, section, filtered_data, current_row)
              
                        # Add dynamic summary if configured
                        summary_config = section.get('summary', {})
                        if summary_config:
                             current_row = self._generate_xlsx_dynamic_summary(ws, section, filtered_data, current_row)
              
                        current_row += 1# Add spacing between sections

          # NEW: Check for matrix_sections (for complex layouts like the image)
          matrix_sections = layout.get('matrix_sections', [])
          for section in matrix_sections:
                   current_row = self._generate_xlsx_matrix_section(ws, section, data, current_row)
                   current_row += 1# Add spacing between sections

          return current_row

     def _generate_xlsx_columns_section(self, ws, section_config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX columns section with proper cell separation"""
          current_row = start_row
          columns = section_config.get('columns', [])

          if not columns:
                   return current_row

          # Generate column headers
          for col_idx, col in enumerate(columns, 1):
                   name = col.get('name', '')
                   cell = ws.cell(row=current_row, column=col_idx, value=name)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

          current_row += 1
         
          # Add informational row for Redemption_SWP report
          # Check if this is a data_section and if informational row is enabled
          layout = section_config.get('layout', {}) if hasattr(section_config, 'get') else {}
          if not layout: # If section_config doesn't have layout, try to get it from parent
                   # This is a workaround - we need to pass the full config to access layout
                   pass
         
          # For now, we'll add the informational row logic in the main data section method
          # This will be handled in _generate_xlsx_data_section
         
          # Generate data rows with proper cell separation
          for record in data:
                   for col_idx, col in enumerate(columns, 1):
                        value = self._get_column_value_for_xlsx(col, record)
                        formatted_value = self._format_column_value_for_xlsx(col, value)
              
                        cell = ws.cell(row=current_row, column=col_idx, value=formatted_value)
                        cell.font = Font(size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   current_row += 1

          return current_row

     def _generate_xlsx_informational_row(self, ws, info_config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX informational row for Redemption_SWP report"""
          current_row = start_row
          columns = info_config.get('columns', [])
         
          if not columns:
                   return current_row
         
          for col_idx, col in enumerate(columns, 1):
                   name = col.get('name', '')
                   calculation = col.get('calculation', '')
                   format_type = col.get('format', 'text')
              
                   if calculation == 'format_fund_display':
                        # Format fund code and name with proper padding
                        fund_info = self._format_fund_display(data)
                        cell = ws.cell(row=current_row, column=col_idx, value=fund_info)
                   elif calculation == 'count_total_checks':
                        # Count total checks (redemption + SWP)
                        total_checks = len(data)
                        cell = ws.cell(row=current_row, column=col_idx, value=total_checks)
                   elif calculation == 'sum_total_amounts':
                        # Sum total amounts (currently blank as per requirements)
                        # JHI wants this populated but currently blank
                        cell = ws.cell(row=current_row, column=col_idx, value="") # Currently blank
                   else:
                        cell = ws.cell(row=current_row, column=col_idx, value="")
              
                   # Format the cell
                   cell.font = Font(size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
          return current_row + 1

     def _generate_xlsx_summary_section(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX summary section with proper cell separation"""
          layout = config.get('layout', {})
          summary_config = layout.get('summary_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          if not summary_config:
                   return start_row

          current_row = start_row
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks':
                   # Title with proper alignment
                   title = summary_config.get('title', 'GENERAL AUDIT COUNTS')
                   if title:
                        cell = ws.cell(row=current_row, column=1, value=f"{title}")
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1
              
                   # Subtitle with proper alignment
                   subtitle = summary_config.get('subtitle', '')
                   if subtitle:
                        # Use current date for document date
                        current_date = datetime.now().strftime('%m/%d/%Y')
                        cell = ws.cell(row=current_row, column=1, value=f"{subtitle}     {current_date}")
                        cell.font = Font(size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1
              
                   # Data block with proper alignment
                   data_block = summary_config.get('data_block', {})
                   if data_block:
                        current_row += 1
                        label = data_block.get('label', 'R06686 Data')
                        # Use current date for record date
                        record_date = datetime.now().strftime('%m/%d/%Y')
                        cell = ws.cell(row=current_row, column=1, value=f"{label} Record Date:    {record_date}")
                        cell.font = Font(size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1
                    
                        # Left-aligned fields with right-aligned dates
                        fields = data_block.get('fields', [])
                        right_aligned = data_block.get('right_aligned', [])
                    
                        for i, field in enumerate(fields):
                             name = field.get('name', '')
                             calculation = field.get('calculation', '')
                             
                             # Calculate value based on calculation type
                             if calculation == 'count_input_records':
                                  value = len(data)
                             elif calculation == 'count_total_checks':
                                  value = len(data) # Total checks = number of records
                             elif calculation == 'sum_total_check_amounts':
                                  value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', '0')) for record in data)
                             else:
                                  value = field.get('value', '')
                             
                             # Format the value
                             if field.get('format') == 'currency':
                                  try:
                                       value = f"{float(value):.2f}"
                                  except (ValueError, TypeError):
                                       value = "0.00"
                             else:
                                  value = str(value)
                             
                             # Format based on field type for exact alignment
                             if i == 0:  # First field (Input Record Count)
                                  cell = ws.cell(row=current_row, column=1, value=f"{name}    {value}        Payable Date:   {record_date}")
                             else:  # Other fields
                                  cell = ws.cell(row=current_row, column=1, value=f"{name}          {value}")
                             cell.font = Font(size=10)
                             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
              
                   # Add spacing
                   spacing = summary_config.get('spacing', 2)
                   for _ in range(spacing):
                        current_row += 1
              
                   return current_row

          # Original logic for other report types
          # Title
          title = summary_config.get('title', '')
          if title:
                   cell = ws.cell(row=current_row, column=1, value=title)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                   current_row += 1

          # Fields with proper cell separation
          fields = summary_config.get('fields', [])
          for field in fields:
                   if "row" in field:
                        # Handle grouped fields in one row
                        for col_idx, subfield in enumerate(field["row"], start=1):
                             name = subfield.get('name', '')
                             value = subfield.get('value', '')
                             cell_a = ws.cell(row=current_row, column=col_idx * 2 - 1, value=name)
                             cell_a.font = Font(bold=True, size=10)
                             cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             cell_b = ws.cell(row=current_row, column=col_idx * 2, value=value)
                             cell_b.font = Font(size=10)
                             cell_b.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                             current_row += 1
                   else:
                        name = field.get('name', '')
                   
                        # Handle dynamic field fetching
                        if 'source_field' in field:
                             # Get value from first record in data
                             source_field = field.get('source_field', '')
                             value = data[0].get(source_field, '') if data else ''
                        else:
                             # Fallback to static value for backward compatibility
                             value = field.get('value', '')
                   
                        # Put name in column A, value in column B
                        cell_a = ws.cell(row=current_row, column=1, value=name)
                        cell_a.font = Font(bold=True, size=10)
                        cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)               
                        cell_b = ws.cell(row=current_row, column=2, value=value)
                        cell_b.font = Font(size=10)
                        cell_b.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1

          # Add blank row after section if it had content
          if summary_config.get('fields'):
                   current_row += 1

          return current_row
     
     def _generate_xlsx_totals_section(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX totals section with proper cell separation"""
          layout = config.get('layout', {})
          totals_config = layout.get('totals_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          if not totals_config:
                   return start_row

          current_row = start_row
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks':
                   # Calculate totals for dividend reports
                   total_records = len(data)
                   total_amount = sum(self.safe_float_convert(record.get('FN_NET_AMOUNT', 0)) for record in data)
              
                   # Title
                   title = totals_config.get('title', 'TOTALS')
                   if title:
                        cell = ws.cell(row=current_row, column=1, value=title)
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        current_row += 1
              
                   # Fields
                   fields = totals_config.get('fields', [])
                   for field in fields:
                        if 'row' in field:
                             # Handle row-based fields
                             row_fields = field.get('row', [])
                             row_parts = []
                        
                             for col_idx, row_field in enumerate(row_fields, 1):
                                  name = row_field.get('name', '')
                                  calculation = row_field.get('calculation', '')
                                  position = row_field.get('position', 'left')
                             
                                  if calculation == 'sum_client_statements':
                                        value = total_records
                                  elif calculation == 'sum_client_amounts':
                                        value = total_amount
                                  elif calculation == 'sum_custom_check_amounts':
                                        value = total_amount
                                  else:
                                        value = row_field.get('value', '')
                             
                                  # Format the value
                                  if row_field.get('format') == 'currency':
                                        try:
                                             value = f"{float(value):.2f}"
                                        except (ValueError, TypeError):
                                             value = "0.00"
                                  elif row_field.get('format') == 'number':
                                        value = str(value)
                                  else:
                                        value = str(value)
                             
                                  # Set cell value and alignment
                                  cell = ws.cell(row=current_row, column=col_idx, value=f"{name} {value}")
                                  cell.font = Font(size=10)
                             
                                  if position == 'right':
                                        cell.alignment = Alignment(horizontal='right', vertical='top')
                                  else:
                                        cell.alignment = Alignment(horizontal='left', vertical='top')
                        
                             current_row += 1
              
                   # Add spacing
                   spacing = totals_config.get('spacing', 1)
                   for _ in range(spacing):
                        current_row += 1
              
                   return current_row

          # Original logic for other report types
          # Calculate totals
          total_records = len(data)
          total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
          total_sheets = sum(float(record.get('DOC_SHEET_COUNT', 0)) for record in data)
          total_images = sum(float(record.get('DOC_IMAGE_COUNT', 0)) for record in data)
          email_delivery_count = sum(1 for record in data if record.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'E')
          print_delivery_count = sum(1 for record in data if record.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'P')
          suppressed_count = sum(1 for record in data if record.get('DOC_SUPPRESSION_FLAG') == 'Y')

          # Additional calculations for different report types
          # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT and CUSTOM_CHECK_NUMBER for consistency
          suppressed_amount_count = sum(1 for record in data if self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) < 5.00)
          zero_check_count = sum(1 for record in data if record.get('CUSTOM_CHECK_NUMBER', '') == '0' or record.get('CUSTOM_CHECK_NUMBER', '') == '')
          suppressed_total_count = sum(1 for record in data if self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) < 5.00 or record.get('CUSTOM_CHECK_NUMBER', '') == '0' or record.get('CUSTOM_CHECK_NUMBER', '') == '')
          packages_count = len(data)# Each record is a package
          check_pages_count = len(data) * 2# Each check has 2 pages

          # Fields with proper cell separation
          fields = totals_config.get('fields', [])
          for field in fields:
                   name = field.get('name', '')
                   calculation = field.get('calculation', '')
                   format_type = field.get('format', 'text')
         
                   if calculation == 'count_records':
                        value = total_records
                   elif calculation == 'sum_amounts':
                        value = total_amount
                   elif calculation == 'sum_sheets':
                        value = total_sheets
                   elif calculation == 'sum_images':
                        value = total_images
                   elif calculation == 'count_email_delivery':
                        value = email_delivery_count
                   elif calculation == 'count_print_delivery':
                        value = print_delivery_count
                   elif calculation == 'count_suppressed':
                        value = suppressed_count
                   elif calculation == 'count_suppressed_amount':
                        value = suppressed_amount_count
                   elif calculation == 'count_zero_check':
                        value = zero_check_count
                   elif calculation == 'count_suppressed_total':
                        value = suppressed_total_count
                   elif calculation == 'count_packages':
                        value = packages_count
                   elif calculation == 'count_check_pages':
                        value = check_pages_count
                   elif calculation == 'count_records * 2':
                        value = total_records * 2
                   elif calculation == 'count_redemption_records':
                        # Count redemption records
                        value = sum(1 for record in data if 'REDEMPTION' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'sum_redemption_amounts':
                        # Sum redemption amounts
                        value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data
                                  if 'REDEMPTION' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'count_swp_records':
                        # Count SWP records
                        value = sum(1 for record in data if 'SWP' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'sum_swp_amounts':
                        # Sum SWP amounts
                        value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data
                                  if 'SWP' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'calculate_total_input':
                        # Count total number of input rows in the report
                        value = total_records
                   elif calculation == 'calculate_total_packages':
                        # Count total number of package entries (rows)
                        value = packages_count
                   elif calculation == 'calculate_total_check_page':
                        # Sum of Total Input + Total Packages
                        value = total_records + packages_count
                   elif calculation == 'calculate_total_dollar_amount':
                        # Sum of all check amounts present in the report
                        value = total_amount
                   else:
                        value = 0

                   if format_type == 'currency':
                        formatted_value = self.formatter.format_currency(value)
                   else:
                        formatted_value = str(value)

                   # Put name in column A, value in column B
                   cell_a = ws.cell(row=current_row, column=1, value=name)
                   cell_a.font = Font(bold=True, size=10)
                   cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   cell_b = ws.cell(row=current_row, column=2, value=formatted_value)
                   cell_b.font = Font(size=10)
                   cell_b.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   current_row += 1

          # Add blank row after section if it had content
          if totals_config.get('fields'):
                   current_row += 1

          return current_row

     def _generate_xlsx_footer_section(self, ws, config: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX footer section with proper cell separation"""
          layout = config['layout']
          footer_config = layout.get('footer_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type')

          current_row = start_row

          # Categories and data with proper cell separation
          categories = footer_config.get('categories', [])
         
          # Use dynamic data for Redemption_SWP reports, static data for others
          if report_type == 'redemption_swp':
                   category_data = self._calculate_dynamic_footer_data(config,data)
          else:
                   category_data = footer_config.get('category_data', [])

          for i, category in enumerate(categories):
                   data_value = category_data[i] if i < len(category_data) else ""
         
                   # For Redemption_SWP, display both SUB-TOTALs:
                   # First SUB-TOTAL (index 6) = sum of JANUS, STATESTREETBANK, DONOTMAIL, VOID, PULLREPORTCHECKS
                   # Second SUB-TOTAL (index 12) = sum of OVERNIGHT W-SIG, OVERNIGHT WO-SIG, OVERNIGHT SAT-W-SIG, OVERNIGHT SAT-WO-SIG
                   # data_value is already correctly set to the appropriate SUB-TOTAL value
         
                   # Put category name in column A, value in column B
                   cell_a = ws.cell(row=current_row, column=1, value=category)
                   cell_a.font = Font(bold=True, size=10)
                   cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   cell_b = ws.cell(row=current_row, column=2, value=data_value)
                   cell_b.font = Font(size=10)
                   cell_b.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   current_row += 1

          return current_row

     def _get_column_value_for_xlsx(self, col: Dict, data: Dict, config: Dict = None) -> Any:
          """Get value for a column from data for XLSX generation"""
          source_field = col.get('source_field', '')

          if source_field == 'static':
                   return col.get('value', '')
          elif source_field == 'calculated':
                   return self._calculate_value_for_xlsx(col, data, config)
          else:
                   value = data.get(source_field, '')
                   # For CUSTOM_FUND_CODE, pad with leading zeros to 7 characters
                   if source_field == 'CUSTOM_FUND_CODE' and value:
                        return value.zfill(7)
                   # For CUSTOM_CHECK_AMOUNT, clean masked values
                   elif source_field == 'CUSTOM_CHECK_AMOUNT' and value:
                        return self.clean_masked_value(value)
                   return value

     def _calculate_value_for_xlsx(self, col: Dict, data: Dict, config: Dict = None) -> Any:
          """Calculate value based on column configuration for XLSX"""
          calculation = col.get('calculation', '')

          if calculation == 'count_records':
                   return 1
          elif calculation == 'sum_amounts':
                   # For Redemption_SWP reports, use CUSTOM_CHECK_AMOUNT
                   return self.safe_float_convert(data.get('CUSTOM_CHECK_AMOUNT', 0))
          elif calculation == 'extract_group':
                   account = data.get('UC_ACCOUNT_NUMBER', '')
                   if len(account) >= 7:
                        return account[:7]
                   return account
          elif calculation == 'concatenate_fund_code_name' :
                   # Get mapped XML field names if config is provided (for Redemption_SWP report)
                   if config and 'xml_mapping' in config and 'field_mapping' in config['xml_mapping']:
                        field_mapping = config['xml_mapping']['field_mapping']
                        CUSTOM_FUND_CODE = field_mapping.get('CUSTOM_FUND_CODE', 'CUSTOM_FUND_CODE')
                        CUSTOM_FUND_DESC = field_mapping.get('CUSTOM_FUND_DESC', 'CUSTOM_FUND_DESC')
                        # Get values using mapped field names from the original record
                        # Note: data here is the enhanced_record which should have the standard field names
                        # But we need to check if we're working with original data or enhanced data
                        fund_code = data.get(CUSTOM_FUND_CODE)
                        fund_desc = data.get(CUSTOM_FUND_DESC)
               #     else:
               #          # Fallback to default field names
               #          fund_code = data.get('CUSTOM_FUND_CODE', '')
               #          fund_desc = data.get('CUSTOM_FUND_DESC', '')
                   
                   # Concatenate CUSTOM_FUND_CODE and CUSTOM_FUND_DESC
                   if fund_code and fund_desc:
                        return f"{fund_code}{fund_desc}"
                   elif fund_code:
                        return fund_code
                   elif fund_desc:
                        return fund_desc
                   return ""
          elif calculation == 'red_begin_check_number':
                   return data.get('RED_BEGIN_CHECK_NUMBER', '')
          elif calculation == 'red_end_check_number':
                   return data.get('RED_END_CHECK_NUMBER', '')
          elif calculation == 'red_check_count':
                   return data.get('RED_CHECK_COUNT', 0)
          elif calculation == 'total_red_dollar_value':
                   return data.get('TOTAL_RED_DOLLAR_VALUE', 0.0)
          elif calculation == 'swp_begin_check_number':
                   return data.get('SWP_BEGIN_CHECK_NUMBER', '')
          elif calculation == 'swp_end_check_number':
                   return data.get('SWP_END_CHECK_NUMBER', '')
          elif calculation == 'swp_check_count':
                   return data.get('SWP_CHECK_COUNT', 0)
          elif calculation == 'total_swp_dollar_value':
                   return data.get('TOTAL_SWP_DOLLAR_VALUE', 0.0)
          elif calculation == 'checks_printed_today':
                   return data.get('CHECKS_PRINTED_TODAY', 0)
          else:
                   return 0

     def _format_column_value_for_xlsx(self, col: Dict, value: Any) -> str:
          """Format column value for XLSX"""
          format_type = col.get('format', 'text')

          if format_type == 'currency':
                   try:
                        return f"{float(value):.2f}"
                   except:
                        return "0.00"
          else:
                   return str(value)

     def _generate_footer_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate footer section"""
          layout = config['layout']
          footer_config = layout.get('footer_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type')

          lines = []

          # Categories
          categories = footer_config.get('categories', [])
         
          # Use dynamic data for Redemption_SWP reports, static data for others
          if report_type == 'redemption_swp':
                   category_data = self._calculate_dynamic_footer_data(config,data)
          else:
                   category_data = footer_config.get('category_data', [])

          for i, category in enumerate(categories):
                   data_value = category_data[i] if i < len(category_data) else ""
                   
                   # For Redemption_SWP, display both SUB-TOTALs:
                   # First SUB-TOTAL (index 6) = sum of JANUS, STATESTREETBANK, DONOTMAIL, VOID, PULLREPORTCHECKS
                   # Second SUB-TOTAL (index 12) = sum of OVERNIGHT W-SIG, OVERNIGHT WO-SIG, OVERNIGHT SAT-W-SIG, OVERNIGHT SAT-WO-SIG
                   # data_value is already correctly set to the appropriate SUB-TOTAL value
                   
                   line = f"{category:<30} {data_value:>15}"
                   lines.append(line)

          return '\n'.join(lines)

     def _generate_txt_dynamic_summary(self, section: Dict, data: List[Dict]) -> List[str]:
          """Generate TXT dynamic summary section after table completion"""
          lines = []
          summary_config = section.get('summary', {})

          if not summary_config:
                   return lines

          # Calculate the summary value based on configuration
          calculation = summary_config.get('calculation', '')
          summary_line = summary_config.get('line', '')
          title_after = summary_config.get('title_after', '')

          if summary_line:
                   # Calculate the value
                   if calculation == 'sum_amounts':
                        # For 12b1_compensation report, sum all amounts in the filtered data (not just < $5.00)
                        # Use CUSTOM_CHECK_AMOUNT for consistency
                        total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
                        formatted_line = summary_line.replace('{sum_amount}', f"{total_amount:.2f}")
                   elif calculation == 'count_records':
                        # Count records with zero check number
                        count = sum(1 for record in data
                             if record.get('FN_CHECK_NUMBER', '') == '0' or record.get('FN_CHECK_NUMBER', '') == '')
                        formatted_line = summary_line.replace('{count_rows}', str(count))
                   else:
                        formatted_line = summary_line
         
                   # Add the summary line
                   lines.append(formatted_line)

          # Add title after if specified (with one-line gap)
          if title_after:
                   lines.append("")# One-line gap
                   lines.append(title_after)

          return lines

     def _generate_xlsx_dynamic_summary(self, ws, section: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX dynamic summary section after table completion"""
          current_row = start_row
          summary_config = section.get('summary', {})

          if not summary_config:
                   return current_row

          # Calculate the summary value based on configuration
          calculation = summary_config.get('calculation', '')
          summary_line = summary_config.get('line', '')
          title_after = summary_config.get('title_after', '')

          if summary_line:
                   # Calculate the value
                   if calculation == 'sum_amounts':
                        # For 12b1_compensation report, sum all amounts in the filtered data (not just < $5.00)
                        # Use CUSTOM_CHECK_AMOUNT for consistency
                        total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
                        formatted_line = summary_line.replace('{sum_amount}', f"{total_amount:.2f}")
                   elif calculation == 'count_records':
                        # Count records with zero check number
                        count = sum(1 for record in data
                             if record.get('FN_CHECK_NUMBER', '') == '0' or record.get('FN_CHECK_NUMBER', '') == '')
                        formatted_line = summary_line.replace('{count_rows}', str(count))
                   else:
                        formatted_line = summary_line
         
                   # Add the summary line
                   cell = ws.cell(row=current_row, column=1, value=formatted_line)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                   current_row += 1

          # Add title after if specified (with one-line gap)
          if title_after:
                   current_row += 1# One-line gap
                   cell = ws.cell(row=current_row, column=1, value=title_after)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                   current_row += 1

          return current_row

     def _generate_xlsx_matrix_section(self, ws, section: Dict, data: List[Dict], start_row: int) -> int:
          """Generate XLSX matrix section - for complex tabular layouts with calculated values"""
          current_row = start_row

          # Section title
          title = section.get('title', '')
          if title:
                   cell = ws.cell(row=current_row, column=1, value=title)
                   cell.font = Font(bold=True, size=11)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                   current_row += 1

          # Get matrix configuration
          headers = section.get('headers', [])
          rows_config = section.get('rows', [])

          # Generate header row
          for col_idx, header in enumerate(headers, 1):
                   cell = ws.cell(row=current_row, column=col_idx, value=header)
                   cell.font = Font(bold=True, size=10)
                   cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

          current_row += 1

          # Generate data rows
          for row_config in rows_config:
                   row_label = row_config.get('label', '')
                   cell_values = row_config.get('values', [])
                   is_bold = row_config.get('bold', False)
                   is_total = row_config.get('is_total', False)
                   indent_level = row_config.get('indent', 0)
         
                   # Apply indentation to label
                   if indent_level > 0:
                        row_label = '' * indent_level + row_label
         
                   # Column 1: Label
                   cell = ws.cell(row=current_row, column=1, value=row_label)
                   cell.font = Font(bold=is_bold or is_total, size=10)
                   cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
         
                   # Remaining columns: Values
                   for col_idx, value_config in enumerate(cell_values, 2):
                        # Calculate or retrieve value
                        if isinstance(value_config, dict):
                             calc_type = value_config.get('calculation', '')
                             value = self._calculate_matrix_value(calc_type, data, value_config)
                        else:
                             value = value_config
              
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.font = Font(bold=is_bold or is_total, size=10)
                        cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
         
                   current_row += 1

          return current_row

     def _calculate_matrix_value(self, calc_type: str, data: List[Dict], config: Dict) -> Any:
          """Calculate value for matrix cells based on calculation type"""
          if calc_type == 'count_records':
                   filters = config.get('filters', {})
                   return self._count_filtered_records(data, filters)

          elif calc_type == 'sum_field':
                   field_name = config.get('field', '')
                   filters = config.get('filters', {})
                   return self._sum_filtered_field(data, field_name, filters)

          elif calc_type == 'multiply':
                   base_calc = config.get('base', {})
                   multiplier = config.get('multiplier', 1)
                   base_value = self._calculate_matrix_value(base_calc.get('calculation', ''), data, base_calc)
                   return base_value * multiplier

          elif calc_type == 'subtract':
                   minuend = config.get('minuend', {})
                   subtrahend = config.get('subtrahend', {})
                   minuend_value = self._calculate_matrix_value(minuend.get('calculation', ''), data, minuend)
                   subtrahend_value = self._calculate_matrix_value(subtrahend.get('calculation', ''), data, subtrahend)
                   return minuend_value - subtrahend_value

          elif calc_type == 'static':
                   return config.get('value', 0)

          elif calc_type == 'sum_amounts':
                   filters = config.get('filters', {})
                   # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT for consistency
                   return self._sum_filtered_field(data, 'CUSTOM_CHECK_AMOUNT', filters)

          elif calc_type == 'sum_sheets':
                   filters = config.get('filters', {})
                   return self._sum_filtered_field(data, 'DOC_SHEET_COUNT', filters)

          elif calc_type == 'sum_images':
                   filters = config.get('filters', {})
                   return self._sum_filtered_field(data, 'DOC_IMAGE_COUNT', filters)

          else:
                   return 0

     def _count_filtered_records(self, data: List[Dict], filters: Dict) -> int:
          """Count records that match the given filters"""
          if not filters:
                   return len(data)

          count = 0
          for record in data:
                   match = True
                   for field, expected_value in filters.items():
                        record_value = record.get(field, '')
              
                        # Handle different comparison types
                        if isinstance(expected_value, dict):
                             operator = expected_value.get('operator', '==')
                             compare_value = expected_value.get('value', '')
                   
                             # Convert to appropriate types for comparison
                             try:
                                  if operator in ['<', '>', '<=', '>=']:
                                        record_num = float(record_value) if record_value else 0
                                        compare_num = float(compare_value) if compare_value else 0
                             
                                        if operator == '<':
                                             if not (record_num < compare_num):
                                                  match = False
                                                  break
                                        elif operator == '>':
                                             if not (record_num > compare_num):
                                                  match = False
                                                  break
                                        elif operator == '<=':
                                             if not (record_num <= compare_num):
                                                  match = False
                                                  break
                                        elif operator == '>=':
                                             if not (record_num >= compare_num):
                                                  match = False
                                                  break
                                  elif operator == '==':
                                        if record_value != compare_value:
                                             match = False
                                             break
                                  elif operator == '!=':
                                        if record_value == compare_value:
                                             match = False
                                             break
                                  elif operator == 'in':
                                        if record_value not in compare_value:
                                             match = False
                                             break
                             except (ValueError, TypeError):
                                  # If conversion fails, treat as string comparison
                                  if operator == '==':
                                        if record_value != compare_value:
                                             match = False
                                             break
                                  elif operator == '!=':
                                        if record_value == compare_value:
                                             match = False
                                             break
                        else:
                             if record_value != expected_value:
                                  match = False
                                  break
                   if match:
                        count += 1
          return count

     def _sum_filtered_field(self, data: List[Dict], field_name: str, filters: Dict) -> float:
          """Sum a field value for records that match the given filters"""
          if not field_name:
                   return 0.0
          total = 0.0
          for record in data:
                   match = True
         
                   # Check filters
                   for filter_field, expected_value in filters.items():
                        if record.get(filter_field, '') != expected_value:
                             match = False
                             break
                   if match:
                        try:
                             total += float(record.get(field_name, 0))
                        except:
                             pass

          return total

     def _generate_summary_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate summary section with exact current design formatting"""
          layout = config.get('layout', {})
          summary_config = layout.get('summary_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          if not summary_config:
                   return ""

          lines = []
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks':
                   # Title with proper alignment
                   title = summary_config.get('title', 'GENERAL AUDIT COUNTS')
                   lines.append(f"{title}")
              
                   # Subtitle with proper alignment - use current date
                   subtitle = summary_config.get('subtitle', '')
                   if subtitle:
                        # Use current date for document date
                        current_date = datetime.now().strftime('%m/%d/%Y')
                        lines.append(f"{subtitle}                          {current_date}")
              
                   # Data block with proper alignment
                   data_block = summary_config.get('data_block', {})
                   if data_block:
                        lines.append("")
                        label = data_block.get('label', 'R06686 Data')
                   
                        # Use current date for record date
                        record_date = datetime.now().strftime('%m/%d/%Y')
                        lines.append(f"{label:<41}{'Record Date:':<15}{record_date:>5}") 
                        
                        # Left-aligned fields with right-aligned dates
                        fields = data_block.get('fields', [])
                        right_aligned = data_block.get('right_aligned', [])
                    
                        for i, field in enumerate(fields):
                             name = field.get('name', '')
                             calculation = field.get('calculation', '')
                             
                             # Calculate value based on calculation type
                             if calculation == 'count_input_records':
                                  value = len(data)
                             elif calculation == 'count_total_checks':
                                  value = len(data) # Total checks = number of records
                             elif calculation == 'sum_total_check_amounts':
                                  value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', '0')) for record in data)
                             else:
                                  value = field.get('value', '')
                             
                             # Format the value
                             if field.get('format') == 'currency':
                                  try:
                                       value = f"{float(value):.2f}"
                                  except (ValueError, TypeError):
                                       value = "0.00"
                             else:
                                  value = str(value)
                             
                             LABEL_WIDTH = 28
                             VALUE_WIDTH = 10

                             if i == 0:
                                   # First row (Input Record Count) + Payable Date
                                   lines.append(
                                        f"{name:<{LABEL_WIDTH}}{value:>{VALUE_WIDTH}}   "
                                        f"{'Payable Date:':<15}{record_date:>10}"
                                   )
                             else:
                                   # Align other fields consistently
                                   lines.append(f"{name:<{LABEL_WIDTH}}{value:>{VALUE_WIDTH}}")
              
                   # Add spacing
                   spacing = summary_config.get('spacing', 2)
                   for _ in range(spacing):
                        lines.append("")
              
                   return '\n'.join(lines)

          # Original logic for other report types
          # Title - "Reconciliation Report"
          title = summary_config.get('title', '')
          if title:
                   lines.append(title)

          # Fields - exact current design formatting
          fields = summary_config.get('fields', [])
          for field in fields:
                   name = field.get('name', '')
                   position = field.get('position', 'left')
              
                   # Handle dynamic field fetching
                   if 'source_field' in field:
                        # Get value from first record in data
                        source_field = field.get('source_field', '')
                        
                        # Special handling for replacement report: extract from inputKeyList[i]
                        if report_type == 'replacement':
                             # Get input_key_value from config (this is inputKeyList[i], not from JSON config)
                             input_key_value = config.get('data_sources', {}).get('input_key_value', '')
                             if input_key_value:
                                  # Extract filename from path if it's a full path (e.g., "path/to/A_B_C" -> "A_B_C")
                                  # If input_key_value is already just the filename, this will still work
                                  key_filename = input_key_value.split("/")[-1]
                                  # Remove file extension if present (e.g., "A_B_C.xml" -> "A_B_C")
                                  base_name = os.path.splitext(key_filename)[0]
                                  # Split by "_" to get parts
                                  parts = base_name.split('_')
                                  if len(parts) >= 3:
                                       # Use index 0 for Corp (UC_CORP), index 2 for Cycle (CYCLE)
                                       if source_field == 'UC_CORP':
                                            value = parts[0] if len(parts) > 0 else ''
                                       elif source_field == 'CYCLE':
                                            value = parts[2] if len(parts) > 2 else ''
                                       else:
                                            # Fallback to original logic for other fields
                                            value = data[0].get(source_field, '') if data else ''
                                  else:
                                       # Not enough parts, fallback to original logic
                                       value = data[0].get(source_field, '') if data else ''
                             else:
                                  # No input_key_value in config, fallback to original logic
                                  value = data[0].get(source_field, '') if data else ''
                        else:
                             # For non-replacement reports, use original logic
                             value = data[0].get(source_field, '') if data else ''
                   else:
                        # Fallback to static value for backward compatibility
                        value = field.get('value', '')
         
                   new_space=20
                   if position == 'left':
                        # Exact formatting: name left-aligned to 20 chars, value follows
                        line = f"{name:<20} {value}"
                   else:             
                        if len(name)==4:
                            line = f"{name}{value:>{new_space}}"
                        else:
                            new_space = new_space+4-len(name)
                            line = f"{name}{value:>{new_space}}"         
                   lines.append(line)

          # Add spacing - always 2 blank lines after summary (current design)
          if lines:
                   lines.append("")
                   lines.append("")

          return '\n'.join(lines)
     
     def _generate_totals_section(self, config: Dict, data: List[Dict]) -> str:
          """Generate totals section with exact current design formatting"""
          layout = config.get('layout', {})
          totals_config = layout.get('totals_section', {})
          report_type = config.get('workflow_metadata', {}).get('report_type', '')

          if not totals_config:
                   return ""

          lines = []
         
          # Special handling for dividend reports
          if report_type == 'dividend_checks':
                   # Calculate totals for dividend reports
                   total_records = len(data)
                   total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
              
                   # Title
                   title = totals_config.get('title', 'TOTALS')
                   lines.append(title)
              
                   # Fields
                   fields = totals_config.get('fields', [])
                   for field in fields:
                        if 'row' in field:
                             # Handle row-based fields
                             row_fields = field.get('row', [])
                             row_parts = []
                             # Add statements count
                             statements_count = total_records
                             totals_line = f"{'Total:':<25}{statements_count:<5}{'':<2}{total_amount:>10.2f}"
                             lines.append(totals_line)
              
                   # Add spacing
                   spacing = totals_config.get('spacing', 1)
                   for _ in range(spacing):
                        lines.append("")
              
                   return '\n'.join(lines)

          # Original logic for other report types
          # Calculate totals
          total_records = len(data)
          total_amount = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data)
          total_sheets = sum(float(record.get('DOC_SHEET_COUNT', 0)) for record in data)
          total_images = sum(float(record.get('DOC_IMAGE_COUNT', 0)) for record in data)
          email_delivery_count = sum(1 for record in data if record.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'E')
          print_delivery_count = sum(1 for record in data if record.get('UC_DOCUMENT_DELIVERY_PREFERENCE') == 'P')
          suppressed_count = sum(1 for record in data if record.get('DOC_SUPPRESSION_FLAG') == 'Y')

          # Additional calculations for different report types
          # For 12b1_compensation report, use CUSTOM_CHECK_AMOUNT and CUSTOM_CHECK_NUMBER for consistency
          suppressed_amount_count = sum(1 for record in data if ( self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) < 5.00 and record.get('CUSTOM_CHECK_NUMBER', '') not in ['0', '', None]))
          #sum(1 for record in data if self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) < 5.00)
          zero_check_count = sum(1 for record in data if record.get('CUSTOM_CHECK_NUMBER', '') == '0' or record.get('CUSTOM_CHECK_NUMBER', '') == '')
          suppressed_total_count = sum(1 for record in data if self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) < 5.00 or record.get('CUSTOM_CHECK_NUMBER', '') == '0' or record.get('CUSTOM_CHECK_NUMBER', '') == '')
          packages_count = len(data)# Each record is a package
          check_pages_count = len(data) * 2# Each check has 2 pages

          # Fields - exact current design formatting
          fields = totals_config.get('fields', [])
          for field in fields:
                   name = field.get('name', '')
                   calculation = field.get('calculation', '')
                   format_type = field.get('format', 'text')
         
                   if calculation == 'count_records':
                        value = total_records
                   elif calculation == 'sum_amounts':
                        value = total_amount
                   elif calculation == 'sum_sheets':
                        value = total_sheets
                   elif calculation == 'sum_images':
                        value = total_images
                   elif calculation == 'count_email_delivery':
                        value = email_delivery_count
                   elif calculation == 'count_print_delivery':
                        value = print_delivery_count
                   elif calculation == 'count_suppressed':
                        value = suppressed_count
                   elif calculation == 'count_suppressed_amount':
                        value = suppressed_amount_count
                   elif calculation == 'count_zero_check':
                        value = zero_check_count
                   elif calculation == 'count_suppressed_total':
                        value = suppressed_total_count
                   elif calculation == 'count_packages':
                        value = packages_count
                   elif calculation == 'count_check_pages':
                        value = check_pages_count
                   elif calculation == 'count_records * 2':
                        value = total_records * 2
                   elif calculation == 'count_redemption_records':
                        # Count redemption records
                        value = sum(1 for record in data if 'REDEMPTION' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'sum_redemption_amounts':
                        # Sum redemption amounts
                        value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data
                                  if 'REDEMPTION' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'count_swp_records':
                        # Count SWP records
                        value = sum(1 for record in data if 'SWP' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'sum_swp_amounts':
                        # Sum SWP amounts
                        value = sum(self.safe_float_convert(record.get('CUSTOM_CHECK_AMOUNT', 0)) for record in data
                                  if 'SWP' in record.get('CUSTOM_CHECK_TYPE', '').upper())
                   elif calculation == 'calculate_total_input':
                        # Count total number of input rows in the report
                        value = total_records
                   elif calculation == 'calculate_total_packages':
                        # Count total number of package entries (rows)
                        value = packages_count
                   elif calculation == 'calculate_total_check_page':
                        # Sum of Total Input + Total Packages
                        value = total_records + packages_count
                   elif calculation == 'calculate_total_dollar_amount':
                        # Sum of all check amounts present in the report
                        value = total_amount
                   else:
                        value = 0

                   if format_type == 'currency':
                        formatted_value = self.formatter.format_currency(value)
                   else:
                        formatted_value = str(value)

                   # Exact formatting: name left-aligned to 30 chars, value right-aligned to 15 chars
                   line = f"{name:<30} {formatted_value:>15}"
                   lines.append(line)

          # Add spacing - always 1 blank line after totals (current design)
          if lines:
                   lines.append("")

          return '\n'.join(lines)

class JSONDrivenReportManager:
     """Main manager for JSON-driven report generation"""

     def __init__(self, bucket=None, config_key=None):
          self.generator = ReportGenerator()
          self.config_dir = 'config'
          self._patterns_cache = None
          self.bucket = bucket
          self.config_key = config_key
    
     def _load_full_config(self) -> Dict:
          """Load full configuration including patterns from S3"""
          if self._patterns_cache is not None:
               return self._patterns_cache
          
          try:
               bucket = self.bucket
               key = self.config_key
               key=key[0]

               response = s3.get_object(Bucket=bucket, Key=key)
               config_content = response['Body'].read().decode('utf-8')
               full_config = json.loads(config_content)
               
               self._patterns_cache = full_config
               return full_config
          except Exception as e:
               log_message("ERROR", f"Failed to load full config: {str(e)}")
               return {}

     def detect_file_type(self, filenames):
        """Detect file type for a single filename OR a list of filenames."""
        import re

        # Ensure filenames is always a list
        if isinstance(filenames, str):
            filenames = [filenames]

        # Load patterns
        full_config = self._load_full_config()
        patterns = full_config.get("patterns", {})

        matched_type = None
        for filename in filenames:
            current_match = None
            for file_type, pattern_list in patterns.items():
                for pattern in pattern_list:
                    if re.search(pattern, filename, re.IGNORECASE):
                        current_match = file_type
                        break
                if current_match:
                    break
            if current_match is None:
                # If any filename doesn't match → whole list is invalid
                return "unknown"
            # First file sets the expected type
            if matched_type is None:
                matched_type = current_match
            else:
                # All must match same file_type
                if matched_type != current_match:
                    return "unknown"
        return matched_type


     def process_file_by_type(self, xml_file_path: str, file_type: str) -> bool:
          
          """Process file based on detected type"""
                    # Use S3 configuration path
          s3_bucket = os.getenv(
                    "s3BucketInput", "br-icsdev-dpmjayant-dataingress-us-east-1-s3"
          )
          config_key = os.getenv(
                    "configKey", "jhi/checks/config/jhi_checks_reporting_config.json"
          )
          config_path = f"s3://{s3_bucket}/{config_key}"
          config = self.generator.load_config(config_path, file_type)
          

          if not config:
                    return False

          # Add the input XML path to the config
          if "data_sources" not in config:
                    config["data_sources"] = {}
          config["data_sources"]["input_xml_path"] = xml_file_path

          return self.generator.generate_report_from_config(config)
          # type: ignore #print("Inside procees_file_by_type_6")


     def generate_all_reports(self) -> Dict[str, bool]:
          """Generate all reports based on configuration files"""
          results = {}

          if not os.path.exists(self.config_dir):
                   return results

          config_files = [f for f in os.listdir(self.config_dir) if f.endswith('.json')]

          if not config_files:
                   return results


          for config_file in config_files:
                   config_path = os.path.join(self.config_dir, config_file)
                   report_type = config_file.replace('_config.json', '').replace('_report', '')
         
                   success = self.generator.generate_report(config_path)
                   results[report_type] = success
          return results

     def generate_single_report(self, report_type: str) -> bool:
          """Generate a single report by type"""
          # Try different naming patterns
          possible_configs = [
                   f"{report_type}_config.json",
                   f"{report_type}_report_config.json",
                   f"{report_type}.json"
          ]

          config_path = None
          for config_file in possible_configs:
                   test_path = os.path.join(self.config_dir, config_file)
                   if os.path.exists(test_path):
                        config_path = test_path
                        break

          if not config_path:
                   return False

          return self.generator.generate_report(config_path)

def get_key_list(bucket_name, path, xml_regex):
          keyList = []
          valueList = []
          seen_keys = set()  # Track seen keys to prevent duplicates
          try:
                    paginator = s3.get_paginator('list_objects_v2')
                    regex_pattern = re.compile(xml_regex)
                    count = 0
                    for page in paginator.paginate(Bucket=bucket_name, Prefix=path):
                              if 'Contents' not in page:
                                        continue
                              for obj in page['Contents']:
                                        count += 1
                                        filename = obj['Key'].split('/')[-1]
                                        if regex_pattern.search(filename):
                                                  # Only add if we haven't seen this key before
                                                  if obj['Key'] not in seen_keys:
                                                              keyList.append(obj['Key'])
                                                              seen_keys.add(obj['Key'])
                    return keyList
          except Exception as e:
                    message=f"Error listing objects in bucket {bucket_name} with prefix {path}: {str(e)}"
                    log_message("Failed", message)
                    raise Exception(message)

def execute_report_generation(input_key, bucket, input_path, report_path, config_key, manager, overall_success, transaction_id=""):
          try:
                    # Construct full S3 path to input XML
                    s3_input_path = f"s3://{bucket}/{input_key}"
                    input_file_name = input_key.split("/")[-1]
                    # Detect file type
                    file_type = manager.detect_file_type(input_file_name)
                    log_message("INFO", f"Detected file type: {file_type}")
                    

                    # Check if config file exists first
                    config_exists = checkFile_isPresent(bucket, config_key, config_key.split('/')[-1], None)
                   
                    # if not config_exists:
                             
                    #           # Try to create a sample configuration file
                    #           if create_sample_config(bucket, config_key, file_type):
                    #                 print(f"INFO: Sample configuration created successfully")
                    #           else:
                    #                 print(f"WARNING: Failed to create sample configuration, will use default config")
                   
                    # Load config
                    config = manager.generator.load_config(f"s3://{bucket}/{config_key}", file_type)
                    if not config:
                             
                              # Create a default configuration for the file type
                              default_config = {
                                        "data_sources": {
                                                  "input_xml_path": f"s3://{bucket}/{input_key}",
                                                  "output_path": f"s3://{bucket}/{report_path}"
                                        },
                                        "report_settings": {
                                                  "file_type": file_type,
                                                  "output_format": "csv",
                                                  "include_headers": True
                                        },
                                        "processing": {
                                                  "extract_all_elements": True,
                                                  "create_summary": True
                                        }
                              }
                             
                              log_message("INFO", f"Using default configuration for file type {file_type}")
                              config = default_config

                    # Set environment variables for use in generate_report_from_config
                    os.environ["s3BucketInput"] = bucket
                    os.environ["inputPath"] = input_path
                    os.environ["inputFileName"] = input_file_name
                    os.environ["reportPath"] = report_path
                   
                    # Store input_key (from inputKeyList[i]) in config for replacement report processing
                    if "data_sources" not in config:
                        config["data_sources"] = {}
                    # Store the original input_key value (inputKeyList[i]) for replacement reports
                    config["data_sources"]["input_key_value"] = input_key
                    # Also store input_file_name for backward compatibility
                    config["data_sources"]["input_file_name"] = input_file_name
                    # Generate report
                    manager.generator.generate_report_from_config(config)
                    
                    files_list=list_csv_files(bucket, input_key)
                    if files_list:
                        merge_and_rename_csv_files(files_list, bucket, config_key, manager, transaction_id="")        
          except Exception as e:
                    log_message("ERROR", f"Exception in execute_report_generation: {str(e)}")
                    overall_success = False

inputFileName=""

import json
import traceback
import os


def lambda_handler(event, context):
    """AWS Glue job main function (list-based params)"""
    try:

        # Set up Glue job parameters (needed for logging/env)
        set_job_params_as_env_vars()

        # Get input parameters from event
        inputFileName = event.get('inputFileName', '')
        bucket = event.get('bucket', '')
        transactionId=event.get('transactionId')
        # Set transaction_id as environment variable for use in processing functions
        if transactionId:
            os.environ["TRANSACTION_ID"] = str(transactionId)
            os.environ["transactionId"] = str(transactionId)

        # Parse JSON lists - handle both string and list inputs
        def safe_json_parse(value, default='[]'):
            """Safely parse JSON, handling both string and list inputs"""
            if isinstance(value, list):
                return value
            elif isinstance(value, str):
                try:
                    return json.loads(value)
                except (json.JSONDecodeError, TypeError):
                    return json.loads(default)
            else:
                return json.loads(default)
        
        inputPathList = safe_json_parse(event.get('inputPathList', '[]'))        
        inputPatternList = safe_json_parse(event.get('inputPatternList', '[]'))
        inputKeyList = safe_json_parse(event.get('inputKeyList', '[]'))        
        reportPathList = safe_json_parse(event.get('reportPathList', '[]'))
        configKeyList = safe_json_parse(event.get('configKeyList', '[]'))
        # Initialize manager
        # manager = JSONDrivenReportManager()
        manager = JSONDrivenReportManager(bucket=bucket, config_key=configKeyList)        
        # Reset redemption_swp email sent flag at start of each lambda execution
        ReportGenerator._redemption_swp_email_sent = False
        # Helper: pick singletons or indexed element
        def pick(lst, idx):
            """Return lst[idx] if available; else broadcast lst[0] if len==1; else error."""
            if not lst:
                raise ValueError("A required list is empty.")
            if len(lst) == 1:
                return lst[0]
            return lst[idx]
        
        list_len = max(len(inputPathList), len(inputPatternList), len(reportPathList), len(configKeyList))
        overall_success = True
        processed_any = False
        # Track processed files to prevent duplicate processing, especially for redemption_swp
        processed_files = set()
        if not inputKeyList:
            for i in range(list_len):
                try:
                    input_path = pick(inputPathList, i)
                    input_pattern = pick(inputPatternList, i)
                    report_path = pick(reportPathList, i)
                    config_key = pick(configKeyList, i)
                    discovered_keys = get_key_list(bucket, input_path, input_pattern)
                    # If no keys found with the specific pattern, try to discover files of all supported types
                    if not discovered_keys:
                        full_config = manager._load_full_config()
                        patterns = full_config.get('patterns', {})
                        all_patterns = []
                        if patterns:
                            for file_type, pattern_list in patterns.items():
                                all_patterns.extend(pattern_list)
                        for pattern in all_patterns:
                            discovered_keys = get_key_list(bucket, input_path, pattern)
                            if discovered_keys:
                                break                        
                        if not discovered_keys:
                            log_message("INFO", f"No keys discovered for path={input_path} with any pattern")
                            continue
                    # Deduplicate discovered_keys to prevent processing the same file twice
                    # This is especially important for redemption_swp reports
                    discovered_keys = list(dict.fromkeys(discovered_keys)) 
                    processed_any = True
                    for j, key in enumerate(discovered_keys):
                        # Skip if this file has already been processed (prevents duplicate reports/emails)
                        if key in processed_files:
                            log_message("INFO", f"Skipping duplicate file: {key} (already processed)")
                            continue
                        processed_files.add(key)
                        execute_report_generation(key, bucket, input_path, report_path, config_key, manager, overall_success, transactionId)
                except Exception as e:
                    log_message("ERROR", f"Exception processing group {i+1}: {str(e)}", context)
                    overall_success = False
        else:
            for i in range(len(inputKeyList)):
                try:
                    # Skip if this file has already been processed (prevents duplicate reports/emails)
                    if inputKeyList[i] in processed_files:
                        log_message("INFO", f"Skipping duplicate file: {inputKeyList[i]} (already processed)")
                        continue
                    processed_files.add(inputKeyList[i])
                    execute_report_generation(inputKeyList[i], bucket, inputPathList[i], reportPathList[i], configKeyList[i], manager, overall_success, transactionId)
                    processed_any = True
                except Exception as e:
                    log_message("ERROR", f"Exception processing provided key {i+1}: {str(e)}", context)
                    overall_success = False

        if not processed_any:
            log_message("Failed", "No inputs to process across all list groups", context)
            return {'statusCode': 400, 'body': json.dumps('No inputs processed')}
        return {'statusCode': 200, 'body': json.dumps('Reports processed successfully')}

    except Exception as e:
        log_message("Failed", f"Lambda job Failed with error: {str(e)}", context)
        return {'statusCode': 500, 'body': json.dumps(f'Lambda job Failed: {str(e)}')}
